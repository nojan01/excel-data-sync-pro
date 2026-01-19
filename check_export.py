#!/usr/bin/env python3
"""Prüfe Table-Konsistenz in Export-Datei"""

from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries

# Export-Datei prüfen
try:
    wb = load_workbook('/Users/nojan/Desktop/Export_2025-08-31 DEFENCE&SPACE MVMS Master Asset List GER.xlsx')
    ws = wb['DEFENCE&SPACE Aug-2025']

    print('=== EXPORT DATEI ===')
    
    if not ws.tables:
        print('KEINE TABLES GEFUNDEN!')
    else:
        for table_name in ws.tables:
            table = ws.tables[table_name]
            print(f'Table: {table_name}')
            print(f'  ref: {table.ref}')
            print(f'  autoFilter: {table.autoFilter}')
            if table.autoFilter:
                print(f'  autoFilter.ref: {table.autoFilter.ref}')
            print(f'  tableColumns: {len(table.tableColumns)}')
            
            # Prüfe ob Column-Namen mit Header-Zellen übereinstimmen
            min_col, min_row, max_col, max_row = range_boundaries(table.ref)
            
            print(f'  Header-Vergleich (erste 10):')
            for i in range(min(10, len(table.tableColumns))):
                col_name = table.tableColumns[i].name
                cell_val = ws.cell(row=min_row, column=min_col + i).value
                match = '✅' if col_name == cell_val else '❌'
                print(f'    {i+1}: tableCol="{col_name}" vs cell="{cell_val}" {match}')

    wb.close()
except Exception as e:
    print(f'Fehler: {e}')
    import traceback
    traceback.print_exc()
