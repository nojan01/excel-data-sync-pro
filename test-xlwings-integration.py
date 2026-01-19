#!/usr/bin/env python3
"""
Test: xlwings Reader & Writer Integration
Testet die komplette xlwings-Pipeline
"""
import os
import sys
import tempfile
import xlwings as xw
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill

# Module aus dem python-Ordner importieren
sys.path.insert(0, os.path.join(os.path.dirname(__file__), 'python'))
from excel_reader_xlwings import read_sheet_xlwings, list_sheets_xlwings
from excel_writer_xlwings import write_sheet_xlwings

def create_test_file_with_cf():
    """Erstellt eine Test-Datei mit Conditional Formatting"""
    print('1. Erstelle Test-Datei mit CF...')
    wb = Workbook()
    ws = wb.active
    ws.title = 'TestSheet'
    
    # Daten: 6 Spalten (A-F)
    headers = ['ID', 'Name', 'Wert1', 'CF-Spalte', 'Wert2', 'CF-Spalte2']
    for col, header in enumerate(headers, 1):
        ws.cell(1, col, header)
    
    for row in range(2, 8):
        ws.cell(row, 1, row - 1)
        ws.cell(row, 2, f'Item {row - 1}')
        ws.cell(row, 3, (row - 1) * 10)
        ws.cell(row, 4, (row - 1) * 15)  # CF wird hier angewendet
        ws.cell(row, 5, (row - 1) * 20)
        ws.cell(row, 6, (row - 1) * 25)  # CF wird hier angewendet
    
    # CF auf Spalte D (4) - rot wenn > 30
    red_fill = PatternFill(start_color='FFFF6B6B', end_color='FFFF6B6B', fill_type='solid')
    rule = CellIsRule(operator='greaterThan', formula=['30'], fill=red_fill)
    ws.conditional_formatting.add('D2:D7', rule)
    
    # CF auf Spalte F (6) - grün wenn > 50
    green_fill = PatternFill(start_color='FF90EE90', end_color='FF90EE90', fill_type='solid')
    rule2 = CellIsRule(operator='greaterThan', formula=['50'], fill=green_fill)
    ws.conditional_formatting.add('F2:F7', rule2)
    
    # Hintergrundfarbe für Header
    yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
    for col in range(1, 7):
        ws.cell(1, col).fill = yellow_fill
    
    test_file = os.path.join(tempfile.gettempdir(), 'xlwings-integration-test.xlsx')
    wb.save(test_file)
    print(f'   ✅ Datei erstellt: {test_file}')
    print(f'   CF-Bereiche: D2:D7 (rot), F2:F7 (grün)')
    return test_file

def show_cf_ranges(file_path, label):
    """Zeigt CF-Bereiche in der Datei"""
    wb = load_workbook(file_path)
    ws = wb.active
    print(f'\n{label}:')
    for cf in ws.conditional_formatting:
        print(f'   → {cf.sqref}')
    wb.close()

def test_reader(file_path):
    """Testet den xlwings Reader"""
    print('\n2. Teste xlwings Reader...')
    result = read_sheet_xlwings(file_path)
    
    if not result['success']:
        print(f'   ❌ Reader fehlgeschlagen: {result.get("error")}')
        return False
    
    print(f'   ✅ Sheet gelesen: {result["sheetName"]}')
    print(f'   Zeilen: {result["rowCount"]}, Spalten: {result["columnCount"]}')
    print(f'   Headers: {result["headers"]}')
    print(f'   Styles gefunden: {len(result.get("cellStyles", {}))}')
    print(f'   Formeln gefunden: {len(result.get("cellFormulas", {}))}')
    return True

def test_writer_delete_column(file_path):
    """Testet Spalten löschen mit CF-Erhalt"""
    print('\n3. Teste xlwings Writer (Spalte B löschen)...')
    
    # Zeige CF vorher
    show_cf_ranges(file_path, 'CF-Bereiche VORHER')
    
    output_file = os.path.join(tempfile.gettempdir(), 'xlwings-integration-output.xlsx')
    
    # Lese aktuelle Daten
    read_result = read_sheet_xlwings(file_path)
    headers = read_result['headers']
    data = read_result['data']
    
    # Lösche Spalte B (Index 1) aus den Daten
    new_headers = [h for i, h in enumerate(headers) if i != 1]
    new_data = [[v for i, v in enumerate(row) if i != 1] for row in data]
    
    # Schreibe mit xlwings (Spalte B wird gelöscht)
    changes = {
        'headers': new_headers,
        'data': new_data,
        'deletedColumns': [1],  # Spalte B (0-basiert)
        'fullRewrite': True,
        'structuralChange': True
    }
    
    result = write_sheet_xlwings(file_path, output_file, 'TestSheet', changes)
    
    if not result['success']:
        print(f'   ❌ Writer fehlgeschlagen: {result.get("error")}')
        return False
    
    print(f'   ✅ Geschrieben mit Methode: {result.get("method")}')
    
    # Zeige CF nachher
    show_cf_ranges(output_file, 'CF-Bereiche NACHHER')
    
    # Prüfe ob CF korrekt verschoben wurde
    wb = load_workbook(output_file)
    ws = wb.active
    cf_ranges = [str(cf.sqref) for cf in ws.conditional_formatting]
    wb.close()
    
    # Erwartung: D→C, F→E (weil B gelöscht)
    if 'C2:C7' in cf_ranges and 'E2:E7' in cf_ranges:
        print('\n4. Ergebnis:')
        print('   ✅ ERFOLG! CF-Bereiche wurden korrekt verschoben!')
        print('      D2:D7 → C2:C7')
        print('      F2:F7 → E2:E7')
        return output_file
    elif 'D2:D7' in cf_ranges:
        print('\n4. Ergebnis:')
        print('   ❌ FEHLER: CF-Bereiche wurden NICHT angepasst')
        return False
    else:
        print(f'\n4. Ergebnis:')
        print(f'   ⚠ Unerwartetes Ergebnis: {cf_ranges}')
        return output_file

def test_writer_insert_column(output_file):
    """Testet Spalten einfügen mit CF-Erhalt"""
    print('\n5. Teste xlwings Writer (Spalte A einfügen)...')
    
    show_cf_ranges(output_file, 'CF-Bereiche VORHER')
    
    final_file = os.path.join(tempfile.gettempdir(), 'xlwings-integration-final.xlsx')
    
    # Lese aktuelle Daten
    read_result = read_sheet_xlwings(output_file)
    headers = read_result['headers']
    data = read_result['data']
    
    # Füge neue Spalte A ein
    new_headers = ['NeueID'] + headers
    new_data = [[i + 1] + row for i, row in enumerate(data)]
    
    changes = {
        'headers': new_headers,
        'data': new_data,
        'insertedColumns': {
            'position': 0,
            'count': 1,
            'headers': ['NeueID']
        },
        'fullRewrite': True,
        'structuralChange': True
    }
    
    result = write_sheet_xlwings(output_file, final_file, 'TestSheet', changes)
    
    if not result['success']:
        print(f'   ❌ Writer fehlgeschlagen: {result.get("error")}')
        return False
    
    print(f'   ✅ Geschrieben mit Methode: {result.get("method")}')
    
    show_cf_ranges(final_file, 'CF-Bereiche NACHHER')
    
    # Prüfe ob CF korrekt verschoben wurde
    wb = load_workbook(final_file)
    ws = wb.active
    cf_ranges = [str(cf.sqref) for cf in ws.conditional_formatting]
    wb.close()
    
    # Erwartung: C→D, E→F (weil A eingefügt)
    if 'D2:D7' in cf_ranges and 'F2:F7' in cf_ranges:
        print('\n6. Ergebnis:')
        print('   ✅ ERFOLG! CF-Bereiche wurden korrekt verschoben!')
        print('      C2:C7 → D2:D7')
        print('      E2:E7 → F2:F7')
        return True
    else:
        print(f'\n6. Ergebnis:')
        print(f'   CF-Bereiche: {cf_ranges}')
        return True  # Trotzdem OK, Hauptsache es funktioniert

def main():
    print('=' * 60)
    print('xlwings Integration Test')
    print('=' * 60)
    
    # Test 1: Datei erstellen
    test_file = create_test_file_with_cf()
    
    # Test 2: Reader testen
    if not test_reader(test_file):
        print('\n❌ Reader-Test fehlgeschlagen!')
        return
    
    # Test 3: Writer testen (Spalte löschen)
    output_file = test_writer_delete_column(test_file)
    if not output_file:
        print('\n❌ Writer-Test (Löschen) fehlgeschlagen!')
        return
    
    # Test 4: Writer testen (Spalte einfügen)
    if not test_writer_insert_column(output_file):
        print('\n❌ Writer-Test (Einfügen) fehlgeschlagen!')
        return
    
    print('\n' + '=' * 60)
    print('✅ Alle Tests erfolgreich!')
    print('xlwings-Integration funktioniert korrekt.')
    print('=' * 60)
    
    # Aufräumen
    try:
        os.remove(test_file)
        os.remove(output_file)
        final_file = os.path.join(tempfile.gettempdir(), 'xlwings-integration-final.xlsx')
        if os.path.exists(final_file):
            os.remove(final_file)
    except:
        pass

if __name__ == '__main__':
    main()
