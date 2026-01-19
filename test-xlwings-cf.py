#!/usr/bin/env python3
"""
Test: xlwings Spalten löschen mit CF-Erhalt
"""
import xlwings as xw
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
import os

# 1. Erstelle Test-Datei mit openpyxl (mit CF)
print('1. Erstelle Test-Datei mit Conditional Formatting...')
wb = Workbook()
ws = wb.active
ws.title = 'Test'

# Daten: 6 Spalten (A-F)
for col in range(1, 7):
    ws.cell(1, col, f'Header{col}')
for row in range(2, 6):
    for col in range(1, 7):
        ws.cell(row, col, col * 10 + row)

# CF auf Spalte D (4)
red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
rule = CellIsRule(operator='greaterThan', formula=['30'], fill=red_fill)
ws.conditional_formatting.add('D2:D5', rule)

# CF auf Spalte F (6)
green_fill = PatternFill(start_color='FF00FF00', end_color='FF00FF00', fill_type='solid')
rule2 = CellIsRule(operator='greaterThan', formula=['50'], fill=green_fill)
ws.conditional_formatting.add('F2:F5', rule2)

test_file = '/tmp/xlwings-cf-test.xlsx'
wb.save(test_file)
print(f'   CF auf D2:D5 (rot) und F2:F5 (grün) gesetzt')
print(f'   Datei: {test_file}')

# 2. Zeige CF vor dem Löschen
print('\n2. CF-Bereiche VORHER:')
for cf in ws.conditional_formatting:
    print(f'   → {cf.sqref}')

# 3. Lösche Spalte B mit xlwings (Excel)
print('\n3. Lösche Spalte B mit xlwings/Excel...')
app = xw.App(visible=False)
app.display_alerts = False
try:
    book = app.books.open(test_file)
    sheet = book.sheets['Test']
    # Lösche Spalte B
    sheet.range('B:B').delete()
    book.save()
    book.close()
    print('   ✅ Spalte B gelöscht und gespeichert')
finally:
    app.quit()

# 4. Prüfe CF in der gespeicherten Datei
print('\n4. CF-Bereiche NACHHER:')
wb2 = load_workbook(test_file)
ws2 = wb2.active

for cf in ws2.conditional_formatting:
    print(f'   → {cf.sqref}')

# 5. Erwartung prüfen
print('\n5. Ergebnis:')
cf_ranges = [str(cf.sqref) for cf in ws2.conditional_formatting]
expected = ['C2:C5', 'E2:E5']  # D→C, F→E (weil B gelöscht)

if 'C2:C5' in cf_ranges and 'E2:E5' in cf_ranges:
    print('   ✅ ERFOLG! CF-Bereiche wurden korrekt verschoben!')
    print(f'      D2:D5 → C2:C5')
    print(f'      F2:F5 → E2:E5')
elif 'D2:D5' in cf_ranges:
    print('   ❌ FEHLER: CF-Bereiche wurden NICHT angepasst')
else:
    print(f'   ? Unerwartetes Ergebnis: {cf_ranges}')

# Aufräumen
os.remove(test_file)
print('\n✅ Test abgeschlossen!')
