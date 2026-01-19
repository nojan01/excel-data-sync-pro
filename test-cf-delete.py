#!/usr/bin/env python3
"""Test: Wie verhält sich unsere manuelle CF-Anpassung bei delete_cols?"""

import sys
sys.path.insert(0, '/Users/nojan/Documents/GitHub/mvms-tool-electron/python')

from openpyxl import load_workbook, Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
import tempfile
import os

# Importiere unsere Funktion
from excel_writer import adjust_conditional_formatting

print("=== Test: Manuelle CF-Anpassung bei delete_cols ===\n")

# Erstelle Test-Workbook
wb = Workbook()
ws = wb.active

# Daten einfügen (Spalten A-I)
for col in range(1, 10):
    ws.cell(1, col, f"Header{col}")
    for row in range(2, 6):
        ws.cell(row, col, row * col)

# CF auf Spalte D (Index 4)
green = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
rule = CellIsRule(operator="greaterThan", formula=["10"], fill=green)
ws.conditional_formatting.add("D2:D5", rule)

# Noch eine CF auf Spalte F
rule2 = CellIsRule(operator="lessThan", formula=["5"], fill=green)
ws.conditional_formatting.add("F2:F5", rule2)

print("VORHER:")
for r, _ in ws.conditional_formatting._cf_rules.items():
    print(f"  {r.sqref}")

# Speichern
tmp = tempfile.mktemp(suffix=".xlsx")
wb.save(tmp)

# Neu laden
wb2 = load_workbook(tmp)
ws2 = wb2.active

print("\n--- Lösche Spalte B (Index 1, 0-basiert) ---")

# UNSERE manuelle CF-Anpassung ZUERST
deleted_cols = [1]  # Spalte B = Index 1
adjust_conditional_formatting(ws2, deleted_cols, None)

# Dann delete_cols
ws2.delete_cols(2)  # 1-basiert

print("\nNACHHER:")
for r, _ in ws2.conditional_formatting._cf_rules.items():
    print(f"  {r.sqref}")

# Speichern und nochmal laden
tmp2 = tempfile.mktemp(suffix=".xlsx")
wb2.save(tmp2)

wb3 = load_workbook(tmp2)
ws3 = wb3.active
print("\nNACH SPEICHERN UND NEULADEN:")
for r, _ in ws3.conditional_formatting._cf_rules.items():
    print(f"  {r.sqref}")

# Aufräumen
os.unlink(tmp)
os.unlink(tmp2)

print("\n" + "="*50)
print("ERWARTET:")
print("  D2:D5 -> C2:C5 (Spalte D wurde zu C)")
print("  F2:F5 -> E2:E5 (Spalte F wurde zu E)")
print("="*50)
