#!/usr/bin/env python3
import sys
sys.path.insert(0, 'python')
from excel_writer_xlwings import write_sheet_xlwings
from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill
import tempfile
import os

# Test-Datei erstellen
test_file = os.path.join(tempfile.gettempdir(), 'bridge-test.xlsx')
output_file = os.path.join(tempfile.gettempdir(), 'bridge-test-output.xlsx')

wb = Workbook()
ws = wb.active
ws.title = 'Test'
for col in range(1, 6):
    ws.cell(1, col, f'Col{col}')
for row in range(2, 5):
    for col in range(1, 6):
        ws.cell(row, col, col * 10 + row)

# CF auf Spalte D
red_fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
rule = CellIsRule(operator='greaterThan', formula=['30'], fill=red_fill)
ws.conditional_formatting.add('D2:D4', rule)
wb.save(test_file)
print(f'Test-Datei: {test_file}')

# CF vorher
wb = load_workbook(test_file)
ws = wb.active
print('CF vorher:', [str(cf.sqref) for cf in ws.conditional_formatting])
wb.close()

# Simuliere Frontend-Daten (Spalte B gelöscht)
changes = {
    'headers': ['Col1', 'Col3', 'Col4', 'Col5'],  # Ohne Col2
    'data': [
        [12, 32, 42, 52],
        [13, 33, 43, 53],
        [14, 34, 44, 54],
    ],
    'deletedColumns': [1],  # Spalte B (0-basiert)
    'fullRewrite': True,
    'structuralChange': True
}

result = write_sheet_xlwings(test_file, output_file, 'Test', changes)
print(f'Result: {result}')

# CF nachher
wb = load_workbook(output_file)
ws = wb.active
print('CF nachher:', [str(cf.sqref) for cf in ws.conditional_formatting])
print('Headers nachher:', [ws.cell(1, c).value for c in range(1, 5)])
wb.close()

os.remove(test_file)
os.remove(output_file)
