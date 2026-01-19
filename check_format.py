#!/usr/bin/env python3
from openpyxl import load_workbook

# Original
orig = load_workbook('/Users/nojan/Desktop/2025-08-31 DEFENCE&SPACE MVMS Master Asset List GER.xlsx')
ws_orig = orig['DEFENCE&SPACE Aug-2025']

# Export  
exp = load_workbook('/Users/nojan/Desktop/DEBUG_Export.xlsx')
ws_exp = exp.active

# Spaltenbreiten - zeige mehr Details
print('=== Spaltenbreiten Vergleich ===')
print('Original: A, B, C(Country-deleted), D, E, F')
print('Export:   A, B, C(=orig D),        D, E, F')
print()
for i, col_letter in enumerate(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']):
    orig_width = ws_orig.column_dimensions[col_letter].width
    exp_width = ws_exp.column_dimensions[col_letter].width
    
    # Nach Löschen Spalte C: Export C sollte = Original D sein
    expected_orig_col = col_letter
    if col_letter >= 'C':
        # Export C = Original D, Export D = Original E, etc.
        expected_orig_col = chr(ord(col_letter) + 1)
    
    expected_width = ws_orig.column_dimensions[expected_orig_col].width if expected_orig_col <= 'Z' else None
    match = '✅' if abs((exp_width or 0) - (expected_width or 0)) < 0.01 else '❌'
    
    print(f'{col_letter}: orig={orig_width}, exp={exp_width}, expected(from {expected_orig_col})={expected_width} {match}')

orig.close()
exp.close()
