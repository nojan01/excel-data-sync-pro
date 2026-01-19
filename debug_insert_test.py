#!/usr/bin/env python3
"""Test: Spalte einfügen mit Formatierung kopieren"""
import sys
sys.path.insert(0, 'python')
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from copy import copy

# Test mit der echten Test-Datei
wb = load_workbook('/Users/nojan/Desktop/test-styles-exceljs.xlsx')
ws = wb.active

print("=== VOR Insert ===")
# Zeige Spalten D und E (4 und 5) Formatierung
for col in [4, 5]:
    cell = ws.cell(row=2, column=col)
    fill_color = cell.fill.start_color.rgb if cell.fill and cell.fill.start_color else 'None'
    print(f"Spalte {col} ({get_column_letter(col)}), Zeile 2: fill={fill_color}")

# Simuliere Insert bei Position 3 (0-basiert = Spalte D, 1-basiert = 4)
insert_at = 4  # Excel 1-basiert

# Speichere Formatierung von sourceColumn 3 (0-basiert) = Spalte D (4)
source_excel_col = 4  # 1-basiert
source_format = {}
print(f"\nSpeichere Format von Spalte {source_excel_col} ({get_column_letter(source_excel_col)}):")
for row in range(1, min(5, ws.max_row + 1)):
    cell = ws.cell(row=row, column=source_excel_col)
    source_format[row] = {
        'fill': copy(cell.fill) if cell.fill else None,
        'font': copy(cell.font) if cell.font else None,
    }
    fill_rgb = cell.fill.start_color.rgb if cell.fill and cell.fill.start_color else 'None'
    print(f"  Zeile {row}: fill.start_color.rgb = {fill_rgb}")

# Füge Spalte ein
ws.insert_cols(insert_at, 1)

print("\n=== NACH Insert (vor Format-Kopie) ===")
for col in [4, 5, 6]:
    cell = ws.cell(row=2, column=col)
    fill_color = cell.fill.start_color.rgb if cell.fill and cell.fill.start_color else 'None'
    print(f"Spalte {col} ({get_column_letter(col)}), Zeile 2: fill={fill_color}")

# Wende Formatierung auf neue Spalte an
print(f"\nWende Format auf Spalte {insert_at} ({get_column_letter(insert_at)}) an...")
for row, fmt in source_format.items():
    cell = ws.cell(row=row, column=insert_at)
    if fmt['fill']:
        cell.fill = fmt['fill']
    if fmt['font']:
        cell.font = fmt['font']

print("\n=== NACH Format-Kopie ===")
for col in [4, 5, 6]:
    cell = ws.cell(row=2, column=col)
    fill_color = cell.fill.start_color.rgb if cell.fill and cell.fill.start_color else 'None'
    print(f"Spalte {col} ({get_column_letter(col)}), Zeile 2: fill={fill_color}")

# Speichern
wb.save('/Users/nojan/Desktop/DEBUG_Insert_Test.xlsx')
print("\n✅ Datei gespeichert: ~/Desktop/DEBUG_Insert_Test.xlsx")
wb.close()
