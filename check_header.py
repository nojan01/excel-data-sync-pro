from openpyxl import load_workbook

# Vergleiche Original und Export
print("=== ORIGINAL ===")
wb = load_workbook('/Users/nojan/Desktop/2025-08-31 DEFENCE&SPACE MVMS Master Asset List GER.xlsx')
ws = wb.active
print(f'Spalten: {ws.max_column}')
for col in range(ws.max_column - 3, ws.max_column):
    cell = ws.cell(row=1, column=col+1)
    print(f'Spalte {col} ({cell.value}): patternType={cell.fill.patternType}, fgColor.type={cell.fill.fgColor.type if cell.fill.fgColor else None}')
wb.close()

print()
print("=== EXPORT ===")
wb = load_workbook('/Users/nojan/Desktop/Export_2025-08-31 DEFENCE&SPACE MVMS Master Asset List GER.xlsx')
ws = wb.active
print(f'Spalten: {ws.max_column}')
for col in range(ws.max_column - 3, ws.max_column):
    cell = ws.cell(row=1, column=col+1)
    print(f'Spalte {col} ({cell.value}): patternType={cell.fill.patternType}, fgColor.type={cell.fill.fgColor.type if cell.fill.fgColor else None}')
wb.close()
