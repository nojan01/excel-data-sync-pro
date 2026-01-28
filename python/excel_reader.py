#!/usr/bin/env python3
"""
Excel Reader für Excel Data Sync Pro
Verwendet openpyxl für bessere Kompatibilität mit Excel-Formaten
"""

import json
import sys
import os
from datetime import datetime, date
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule, FormulaRule, ColorScaleRule, DataBarRule


def argb_to_hex(argb):
    """Konvertiert ARGB (z.B. 'FF00FF00') zu Hex ('#00FF00')"""
    if not argb or argb == '00000000':
        return None
    if len(argb) == 8:
        return '#' + argb[2:].upper()
    elif len(argb) == 6:
        return '#' + argb.upper()
    return None


def get_fill_color(cell):
    """Extrahiert die Hintergrundfarbe einer Zelle"""
    try:
        fill = cell.fill
        if fill and fill.patternType and fill.patternType != 'none':
            fg_color = fill.fgColor
            if fg_color:
                if fg_color.type == 'rgb' and fg_color.rgb:
                    return argb_to_hex(fg_color.rgb)
                elif fg_color.type == 'indexed':
                    # Indexed colors - simplified mapping for common colors
                    indexed_colors = {
                        0: '#000000', 1: '#FFFFFF', 2: '#FF0000', 3: '#00FF00',
                        4: '#0000FF', 5: '#FFFF00', 6: '#FF00FF', 7: '#00FFFF',
                        8: '#000000', 9: '#FFFFFF', 10: '#FF0000', 11: '#00FF00',
                        12: '#0000FF', 13: '#FFFF00', 14: '#FF00FF', 15: '#00FFFF',
                    }
                    return indexed_colors.get(fg_color.indexed)
                elif fg_color.type == 'theme':
                    # Theme colors würden Theme-Auflösung erfordern
                    # Für jetzt: None zurückgeben (wird später verbessert)
                    pass
    except Exception:
        pass
    return None


def get_font_info(cell, default_font=None):
    """Extrahiert Font-Informationen - nur wenn sie vom Default abweichen"""
    try:
        font = cell.font
        if font:
            info = {}
            # Nur speichern wenn vom Default abweichend
            default_name = default_font.get('name', 'Calibri') if default_font else 'Calibri'
            default_size = default_font.get('size', 11) if default_font else 11
            
            if font.bold:
                info['bold'] = True
            if font.italic:
                info['italic'] = True
            if font.underline:
                info['underline'] = True
            if font.color and font.color.rgb and font.color.rgb != 'FF000000':
                color = argb_to_hex(font.color.rgb)
                if color and color != '#000000':
                    info['color'] = color
            # Font name/size nur wenn vom Default abweichend
            if font.size and font.size != default_size:
                info['size'] = font.size
            if font.name and font.name != default_name:
                info['name'] = font.name
            return info if info else None
    except Exception:
        pass
    return None


def serialize_value(value):
    """Konvertiert Python-Werte zu JSON-kompatiblen Werten"""
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.strftime('%d.%m.%Y %H:%M:%S')
    if isinstance(value, date):
        return value.strftime('%d.%m.%Y')
    if isinstance(value, (int, float)):
        return value
    return str(value)


def read_sheet(file_path, sheet_name=None, options=None):
    """
    Liest ein Excel-Sheet und gibt Daten + Metadaten zurück
    
    Args:
        file_path: Pfad zur Excel-Datei
        sheet_name: Name des Sheets (None = aktives Sheet)
        options: Dict mit Optionen (extractStyles, etc.)
    
    Returns:
        Dict mit headers, data, styles, etc.
    """
    options = options or {}
    extract_styles = options.get('extractStyles', True)
    
    try:
        # Workbook laden
        # WICHTIG: read_only=False ist nötig für vollständige Style-Extraktion
        # aber wir können data_only=True nicht verwenden, da wir Formeln brauchen
        wb = load_workbook(file_path, data_only=False, read_only=False)
        
        # Sheet auswählen
        if sheet_name:
            if sheet_name not in wb.sheetnames:
                return {'success': False, 'error': f'Sheet "{sheet_name}" nicht gefunden'}
            ws = wb[sheet_name]
        else:
            ws = wb.active
        
        # Dimensionen ermitteln
        max_row = ws.max_row or 1
        max_col = ws.max_column or 1
        
        # Headers (erste Zeile)
        headers = []
        for col in range(1, max_col + 1):
            cell = ws.cell(row=1, column=col)
            headers.append(serialize_value(cell.value))
        
        # Daten (ab Zeile 2)
        data = []
        for row_idx in range(2, max_row + 1):
            row_data = []
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                row_data.append(serialize_value(cell.value))
            data.append(row_data)
        
        result = {
            'success': True,
            'headers': headers,
            'data': data,
            'sheetName': ws.title,
            'rowCount': max_row - 1,  # Ohne Header
            'columnCount': max_col
        }
        
        # Styles extrahieren wenn gewünscht
        if extract_styles:
            cell_styles = {}
            cell_fonts = {}
            number_formats = {}
            
            for row_idx in range(1, max_row + 1):
                for col_idx in range(1, max_col + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    # Key ist 0-basiert (row-1, col-1)
                    key = f"{row_idx - 1}-{col_idx - 1}"
                    
                    # Fill Color
                    fill_color = get_fill_color(cell)
                    if fill_color:
                        cell_styles[key] = fill_color
                    
                    # Font Info - nur wenn vom Default abweichend
                    font_info = get_font_info(cell, {'name': 'Arial', 'size': 10})
                    if font_info:
                        cell_fonts[key] = font_info
                    
                    # Number Format - nur wenn nicht Standard
                    if cell.number_format and cell.number_format != 'General':
                        number_formats[key] = cell.number_format
            
            result['cellStyles'] = cell_styles
            result['cellFonts'] = cell_fonts
            result['numberFormats'] = number_formats
            # Default Font für die GUI
            result['defaultFont'] = {'name': 'Arial', 'size': 10}
        
        # Merged Cells
        merged = []
        for merged_range in ws.merged_cells.ranges:
            merged.append(str(merged_range))
        result['mergedCells'] = merged
        
        # AutoFilter
        if ws.auto_filter and ws.auto_filter.ref:
            result['autoFilterRange'] = ws.auto_filter.ref
        
        # Hidden Columns
        hidden_cols = []
        for col_idx in range(1, max_col + 1):
            col_dim = ws.column_dimensions.get(get_column_letter(col_idx))
            if col_dim and col_dim.hidden:
                hidden_cols.append(col_idx - 1)  # 0-basiert
        result['hiddenColumns'] = hidden_cols
        
        # Hidden Rows
        hidden_rows = []
        for row_idx in range(2, max_row + 1):  # Ohne Header
            row_dim = ws.row_dimensions.get(row_idx)
            if row_dim and row_dim.hidden:
                hidden_rows.append(row_idx - 2)  # 0-basiert, ohne Header
        result['hiddenRows'] = hidden_rows
        
        # Column Widths
        col_widths = {}
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            col_dim = ws.column_dimensions.get(col_letter)
            if col_dim and col_dim.width:
                col_widths[col_idx - 1] = col_dim.width
        result['columnWidths'] = col_widths
        
        # Formeln erkennen
        cell_formulas = {}
        
        for row_idx in range(1, max_row + 1):
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.data_type == 'f' and cell.value and str(cell.value).startswith('='):
                    key = f"{row_idx - 1}-{col_idx - 1}"
                    formula = str(cell.value)
                    cell_formulas[key] = formula
        
        result['cellFormulas'] = cell_formulas
        
        wb.close()
        return result
        
    except Exception as e:
        return {'success': False, 'error': str(e)}


def list_sheets(file_path):
    """Listet alle Sheets in einer Excel-Datei"""
    try:
        wb = load_workbook(file_path, read_only=True)
        sheets = wb.sheetnames
        wb.close()
        return {'success': True, 'sheets': sheets}
    except Exception as e:
        return {'success': False, 'error': str(e)}


def main():
    """Hauptfunktion - liest Befehle von stdin"""
    # Auf Windows: Stelle sicher dass stdin/stdout UTF-8 verwenden
    import io
    if sys.platform == 'win32':
        sys.stdin = io.TextIOWrapper(sys.stdin.buffer, encoding='utf-8')
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    
    if len(sys.argv) < 2:
        print(json.dumps({'success': False, 'error': 'Kein Befehl angegeben'}))
        sys.exit(1)
    
    command = sys.argv[1]
    
    if command == 'list_sheets':
        if len(sys.argv) < 3:
            print(json.dumps({'success': False, 'error': 'Kein Dateipfad angegeben'}))
            sys.exit(1)
        result = list_sheets(sys.argv[2])
        print(json.dumps(result, ensure_ascii=False))
    
    elif command == 'read_sheet':
        if len(sys.argv) < 3:
            print(json.dumps({'success': False, 'error': 'Kein Dateipfad angegeben'}))
            sys.exit(1)
        file_path = sys.argv[2]
        sheet_name = sys.argv[3] if len(sys.argv) > 3 else None
        options = json.loads(sys.argv[4]) if len(sys.argv) > 4 else {}
        result = read_sheet(file_path, sheet_name, options)
        print(json.dumps(result, ensure_ascii=False))
    
    else:
        print(json.dumps({'success': False, 'error': f'Unbekannter Befehl: {command}'}))
        sys.exit(1)


if __name__ == '__main__':
    main()
