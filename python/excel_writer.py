# Prüfe, ob openpyxl installiert ist, und gib eine verständliche Fehlermeldung aus
try:
    import openpyxl
except ImportError:
    import sys
    print("[Fehler] Das Python-Modul 'openpyxl' ist nicht installiert. Bitte führe im python-Verzeichnis 'pip3 install -r requirements.txt' aus.", file=sys.stderr)
    sys.exit(1)
#!/usr/bin/env python3
"""
Excel Writer für Excel Data Sync Pro
Verwendet openpyxl für bessere Kompatibilität mit Excel-Formaten

Der große Vorteil von openpyxl: 
- Öffnet die Original-Datei und modifiziert nur die geänderten Zellen
- Behält ALLE Formatierungen, bedingte Formatierungen, Tabellen, etc.

WICHTIG: openpyxl's delete_cols() aktualisiert CF-Bereiche NICHT automatisch!
Für strukturelle Änderungen (Spalten löschen/einfügen) nutzen wir xlwings wenn
Microsoft Excel installiert ist - das erhält ALLE Formatierungen.
"""

import json
import sys
import os
import re
from datetime import datetime, date
from copy import copy

# ============================================================================
# MONKEY-PATCH: openpyxl PatternFill um extLst zu ignorieren
# Manche Excel-Dateien haben erweiterte Formatierungen die openpyxl nicht kennt
# WICHTIG: Muss VOR dem Import von openpyxl erfolgen!
# ============================================================================
import openpyxl.styles.fills as _fills_module
from openpyxl.styles.colors import Color
from openpyxl.descriptors.base import Typed

# Patch die Typed Descriptor Klasse um None-Werte für Color mit Default zu ersetzen
_original_typed_set = Typed.__set__

def _patched_typed_set(self, instance, value):
    """Gepatchter Typed.__set__ der None für Color-Typen durch Default ersetzt"""
    if value is None and hasattr(self, 'expected_type') and self.expected_type == Color:
        # Statt None einen transparenten Default-Color setzen
        value = Color(rgb='00000000')
    _original_typed_set(self, instance, value)

Typed.__set__ = _patched_typed_set

_OriginalPatternFill = _fills_module.PatternFill
_original_init = _OriginalPatternFill.__init__

def _patched_init(self, patternType=None, fgColor=None, bgColor=None, 
                  fill_type=None, start_color=None, end_color=None, **kwargs):
    """Gepatchter __init__ der unbekannte kwargs wie extLst ignoriert"""
    _original_init(self, patternType=patternType, fgColor=fgColor, bgColor=bgColor,
                   fill_type=fill_type, start_color=start_color, end_color=end_color)

_OriginalPatternFill.__init__ = _patched_init

# Patch auch from_tree um extLst child nodes zu entfernen
_original_from_tree = _OriginalPatternFill.from_tree.__func__

@classmethod  
def _patched_from_tree(cls, node):
    """Gepatchte from_tree die extLst child nodes entfernt"""
    for child in list(node):
        tag = child.tag.split('}')[-1] if '}' in child.tag else child.tag
        if tag == 'extLst':
            node.remove(child)
        # Wenn fgColor oder bgColor leer ist (keine Attribute), entferne es auch
        elif tag in ('fgColor', 'bgColor') and not child.attrib:
            node.remove(child)
    return _original_from_tree(cls, node)

_OriginalPatternFill.from_tree = _patched_from_tree
# ============================================================================

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import range_boundaries, coordinate_from_string
from openpyxl.styles import PatternFill, Font, Alignment, Border
from openpyxl.styles.colors import Color
from openpyxl.formatting.formatting import ConditionalFormattingList

# ============================================================================
# FIX 1: Sichere __copy__ für ALLE Serialisable-Unterklassen
# Serialisable.__copy__ macht einen XML-Round-Trip (to_tree → from_tree),
# der bei Nested-Deskriptoren mit "Nested.from_tree() missing 1 required
# positional argument: 'node'" crasht.
# Lösung: Direkt aus __dict__ kopieren statt XML-Serialisierung.
# Patch auf Serialisable selbst = schützt ALLE Unterklassen automatisch.
# ============================================================================
from openpyxl.descriptors.serialisable import Serialisable as _Serialisable

_original_serialisable_copy = _Serialisable.__copy__

def _safe_serialisable_copy(self):
    """Sicherer __copy__ via object.__new__ + __dict__ — KEIN Konstruktor,
    KEIN XML-Round-Trip. Verhindert 'Nested.from_tree() missing node' Fehler.
    Listen/Dicts werden shallow-kopiert um Shared-Mutation zu verhindern."""
    cp = object.__new__(self.__class__)
    for key, val in self.__dict__.items():
        if isinstance(val, list):
            cp.__dict__[key] = list(val)
        elif isinstance(val, dict):
            cp.__dict__[key] = dict(val)
        else:
            cp.__dict__[key] = val
    return cp

_Serialisable.__copy__ = _safe_serialisable_copy
# ============================================================================

# ============================================================================
# FIX 2: Defensiver Patch für ALLE Nested-Deskriptor from_tree Methoden
# Manche Excel-Dateien erzeugen beim load_workbook()-Parsing Situationen,
# in denen Nested.from_tree() OHNE node-Argument aufgerufen wird.
# Patch: from_tree gibt None zurück statt zu crashen wenn node fehlt.
# ============================================================================
import openpyxl.descriptors.nested as _nested_mod

def _make_safe_from_tree(original_from_tree):
    """Wraps a Nested from_tree so that a missing/None node returns None."""
    def _safe_from_tree(self, node=None):
        if node is None:
            return None
        return original_from_tree(self, node)
    return _safe_from_tree

# Patch Nested und ALLE Unterklassen
for _ncls_name in ('Nested', 'NestedValue', 'NestedText', 'NestedFloat',
                    'NestedInteger', 'NestedString', 'NestedBool',
                    'NestedNoneSet', 'NestedSet', 'NestedMinMax', 'EmptyTag'):
    _ncls = getattr(_nested_mod, _ncls_name, None)
    if _ncls and 'from_tree' in vars(_ncls):
        _ncls.from_tree = _make_safe_from_tree(_ncls.from_tree)

# Patch auch Serialisable.from_tree um robuster zu sein:
# Wenn ein Nested-Deskriptor None zurückgibt, Attribut überspringen
_orig_serialisable_from_tree = _Serialisable.from_tree.__func__

@classmethod
def _safe_serialisable_from_tree(cls, node):
    """Robuste from_tree die None-Werte von Nested-Deskriptoren toleriert."""
    try:
        return _orig_serialisable_from_tree(cls, node)
    except TypeError as _ft_err:
        if 'from_tree' in str(_ft_err) and ('missing' in str(_ft_err) or 'node' in str(_ft_err)):
            # Fallback: Attribute nur aus XML-Attributen extrahieren (ohne Kinder)
            import sys
            sys.stderr.write(f"[PATCH] Serialisable.from_tree TypeError abgefangen: {_ft_err} — verwende Fallback\n")
            attrib = dict(node.attrib)
            # Namespaced keys entfernen
            for key in list(attrib):
                if key.startswith('{'):
                    del attrib[key]
            try:
                return cls(**attrib)
            except Exception:
                return cls()
        raise

_Serialisable.from_tree = _safe_serialisable_from_tree
# ============================================================================

def _safe_load_workbook(path, rich_text=True):
    """load_workbook() mit Fallback: bei Nested.from_tree Fehler ohne rich_text laden."""
    try:
        return load_workbook(path, rich_text=rich_text)
    except TypeError as e:
        err_str = str(e)
        if 'extLst' in err_str:
            raise
        if rich_text and ('from_tree' in err_str or 'Nested' in err_str or 'node' in err_str):
            sys.stderr.write(f"[LOAD] rich_text=True fehlgeschlagen ({e}), Fallback ohne rich_text\n")
            return load_workbook(path, rich_text=False)
        raise

# Standard Theme-Farben (Office Default Theme)
# Diese werden verwendet wenn Theme-Farben nicht aufgelöst werden können
# ACHTUNG: Die Reihenfolge ist wichtig! Excel speichert Theme-Index 0-9
THEME_COLORS = [
    'FFFFFF',  # 0: lt1 - Light 1 (Background 1, usually white)
    '000000',  # 1: dk1 - Dark 1 (Text 1, usually black)
    'E7E6E6',  # 2: lt2 - Light 2 (Background 2)
    '44546A',  # 3: dk2 - Dark 2 (Text 2)
    '4472C4',  # 4: accent1 - Blue
    'ED7D31',  # 5: accent2 - Orange
    '70AD47',  # 6: accent3 - GREEN (not gray!)
    'FFC000',  # 7: accent4 - Gold
    '5B9BD5',  # 8: accent5 - Light Blue
    '7030A0',  # 9: accent6 - Purple
]


def _check_zip_drawings(zip_path, label):
    """Diagnose-Helper: Prüft ob ein ZIP drawing-relevante Dateien enthält."""
    import zipfile
    import re
    import sys
    try:
        with zipfile.ZipFile(zip_path, 'r') as zf:
            names = zf.namelist()
            drawings = [n for n in names if 'drawing' in n.lower() or 'media/' in n.lower()]
            has_drawing_el = False
            for n in names:
                if n.startswith('xl/worksheets/') and n.endswith('.xml') and '/_rels/' not in n:
                    content = zf.read(n).decode('utf-8', errors='replace')
                    if re.search(r'<drawing[\s>]', content):
                        has_drawing_el = True
                        break
            sys.stderr.write(f"[CHECKPOINT {label}] drawings/media files: {drawings}, <drawing> in sheet XML: {has_drawing_el}\n")
    except Exception as e:
        sys.stderr.write(f"[CHECKPOINT {label}] Fehler: {e}\n")


def fix_xlsx_relationships(xlsx_path):
    """
    Repariert openpyxl-gespeicherte XLSX-Dateien.
    
    openpyxl hat mehrere Probleme:
    1. Schreibt absolute Pfade in Relationships (z.B. Target="/xl/tables/table1.xml")
       statt relative Pfade (Target="../tables/table1.xml")
    2. Schreibt XML-Dateien ohne XML-Header (<?xml version="1.0"?>)
    3. Fügt headerRowCount="1" zu Tables hinzu, was Probleme verursachen kann
    4. Setzt xmlns an falsche Position (muss am Anfang des table-Elements sein)
    
    Dies führt dazu, dass Excel die Datei als beschädigt erkennt und Tables/AutoFilter entfernt.
    
    Verwendet ZIP-to-ZIP Re-Zip um sicherzustellen, dass KEINE Dateien verloren gehen
    (insbesondere drawings, media, embeddings etc. die openpyxl ohne Pillow nicht korrekt
    verarbeitet).
    """
    import zipfile
    import tempfile
    import shutil
    
    XML_HEADER = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
    
    # Erstelle temporäre Kopie
    temp_dir = tempfile.mkdtemp()
    temp_xlsx = os.path.join(temp_dir, 'fixed.xlsx')
    
    try:
        # Extrahiere die XLSX
        with zipfile.ZipFile(xlsx_path, 'r') as zf:
            zf.extractall(temp_dir)
        
        fixed_count = 0
        
        # Durchsuche alle XML-Dateien
        for root, dirs, files in os.walk(temp_dir):
            for f in files:
                if not f.endswith('.xml') and not f.endswith('.rels'):
                    continue
                    
                full_path = os.path.join(root, f)
                
                with open(full_path, 'r', encoding='utf-8') as fp:
                    content = fp.read()
                
                original_content = content
                
                # FIX 1: Füge XML-Header hinzu wenn er fehlt
                if not content.startswith('<?xml'):
                    content = XML_HEADER + content
                
                # FIX 2: Konvertiere absolute Pfade zu relativen (nur für .rels Dateien)
                if f.endswith('.rels'):
                    rel_root = root.replace(temp_dir, '').lstrip(os.sep)
                    
                    if 'worksheets/_rels' in rel_root or 'worksheets\\_rels' in rel_root:
                        content = content.replace('Target="/xl/tables/', 'Target="../tables/')
                        content = content.replace('Target="/xl/drawings/', 'Target="../drawings/')
                        content = content.replace('Target="/xl/printerSettings/', 'Target="../printerSettings/')
                    elif '_rels' in rel_root:
                        content = content.replace('Target="/xl/', 'Target="')
                
                # FIX 3: Repariere Table-XML (table*.xml Dateien)
                if f.startswith('table') and f.endswith('.xml') and 'tables' in root:
                    # openpyxl setzt xmlns am Ende der Attribute, aber es muss am Anfang sein
                    # Außerdem fügt es headerRowCount="1" hinzu, was Probleme macht
                    import re
                    
                    # Entferne headerRowCount="1" - das Original hat es nicht
                    content = re.sub(r'\s+headerRowCount="1"', '', content)
                    
                    # Stelle sicher, dass xmlns direkt nach <table kommt
                    # Pattern: <table ...andere attribute... xmlns="...">
                    # Ziel:    <table xmlns="..." ...andere attribute...>
                    match = re.search(r'<table\s+([^>]*?)xmlns="([^"]+)"([^>]*)>', content)
                    if match:
                        before_xmlns = match.group(1).strip()
                        xmlns_value = match.group(2)
                        after_xmlns = match.group(3).strip()
                        
                        # Nur umordnen wenn xmlns nicht schon am Anfang ist
                        if before_xmlns:
                            all_attrs = f'{before_xmlns} {after_xmlns}'.strip()
                            new_table_tag = f'<table xmlns="{xmlns_value}" {all_attrs}>'
                            content = content[:match.start()] + new_table_tag + content[match.end():]
                
                # FIX 4: Repariere leere inlineStr Zellen in sheet*.xml
                # openpyxl schreibt <c r="X1" t="inlineStr"></c> ohne <is> Element
                # xlsx-populate erwartet aber <is><t>...</t></is> bei t="inlineStr"
                # Lösung: Entferne t="inlineStr" bei leeren Zellen
                if f.startswith('sheet') and f.endswith('.xml') and 'worksheets' in root:
                    import re
                    # Pattern: <c ... t="inlineStr"></c> oder <c ... t="inlineStr"/>
                    # Diese leeren inlineStr-Zellen müssen repariert werden
                    content = re.sub(
                        r'<c\s+([^>]*?)t="inlineStr"([^>]*?)></c>',
                        r'<c \1\2/>',
                        content
                    )
                    content = re.sub(
                        r'<c\s+([^>]*?)t="inlineStr"([^>]*?)/>', 
                        r'<c \1\2/>',
                        content
                    )
                    # Auch leere Rows entfernen: <row r="2"></row> -> entfernen
                    content = re.sub(r'<row r="\d+"></row>', '', content)
                
                if content != original_content:
                    fixed_count += 1
                    with open(full_path, 'w', encoding='utf-8') as fp:
                        fp.write(content)
        
        if fixed_count > 0:
            # ZIP-to-ZIP Re-Zip: Starte vom INPUT-ZIP und ersetze modifizierte Dateien.
            # Dadurch bleiben ALLE Einträge erhalten (auch drawings, media, embeddings etc.
            # die openpyxl ohne Pillow nicht schreibt und die beim temp_dir-Walk fehlen würden).
            temp_files = {}
            for root, dirs, files in os.walk(temp_dir):
                dirs[:] = [d for d in dirs if d != '__MACOSX']
                for f in files:
                    if f == 'fixed.xlsx' or f == '.DS_Store' or f.startswith('._'):
                        continue
                    full_path = os.path.join(root, f)
                    arc_name = os.path.relpath(full_path, temp_dir).replace('\\', '/')
                    temp_files[arc_name] = full_path
            
            written = set()
            with zipfile.ZipFile(temp_xlsx, 'w', zipfile.ZIP_DEFLATED) as new_zf:
                # Phase 1: Alle Einträge aus dem INPUT-ZIP durchgehen
                with zipfile.ZipFile(xlsx_path, 'r') as orig_zf:
                    for item in orig_zf.infolist():
                        if item.filename.endswith('/'):
                            continue
                        name = item.filename
                        if name.startswith('__MACOSX') or name.endswith('.DS_Store') or \
                           name.split('/')[-1].startswith('._'):
                            continue
                        if name in temp_files:
                            new_zf.write(temp_files[name], name)
                        else:
                            data = orig_zf.read(name)
                            info = item
                            info.compress_type = zipfile.ZIP_DEFLATED
                            new_zf.writestr(info, data)
                        written.add(name)
                
                # Phase 2: Neue Dateien aus temp_dir die NICHT im Input waren
                for arc_name, full_path in temp_files.items():
                    if arc_name not in written:
                        new_zf.write(full_path, arc_name)
                        written.add(arc_name)
            
            # Ersetze Original mit reparierter Version
            shutil.copy2(temp_xlsx, xlsx_path)
    
    finally:
        # Cleanup
        shutil.rmtree(temp_dir, ignore_errors=True)


def restore_table_xml_from_original(output_path, original_path, table_changes=None):
    """
    Kopiert die Table-XML aus der Original-Datei und passt nur ref/tableColumns an.
    
    openpyxl verliert wichtige XML-Attribute wie xr:uid, xmlns:mc, xmlns:xr etc.
    Diese Funktion stellt die Original-Struktur wieder her und passt nur die
    notwendigen Felder an.
    
    Args:
        output_path: Pfad zur Export-Datei (wird modifiziert)
        original_path: Pfad zur Original-Datei
        table_changes: Dict mit {table_name: {'ref': new_ref, 'columns': [col_names]}}
                       Wenn None oder leer, werden alle Tables vom Original kopiert.
    """
    import zipfile
    import tempfile
    import shutil
    import re
    import sys
    
    # Prüfe ob original_path gültig ist
    if not original_path:
        sys.stderr.write(f"[restore_table_xml] Übersprungen: original_path leer\n")
        return
    
    # Wenn original_path == output_path: Nutze Backup von restore_external_links
    if os.path.normpath(original_path) == os.path.normpath(output_path):
        backup_candidate = getattr(restore_external_links_from_original, '_backup_original_path', None)
        if backup_candidate and os.path.exists(backup_candidate):
            original_path = backup_candidate
            sys.stderr.write(f"[restore_table_xml] Verwende Backup: {backup_candidate}\n")
        else:
            sys.stderr.write(f"[restore_table_xml] Übersprungen: original==output und kein Backup\n")
            return
    
    if not os.path.exists(original_path):
        sys.stderr.write(f"[restore_table_xml] Original existiert nicht: {original_path}\n")
        return
    
    sys.stderr.write(f"[restore_table_xml] Starte Wiederherstellung von {original_path}\n")
    
    # Bei table_changes=None: Leeres Dict verwenden (alle Tables werden kopiert)
    if table_changes is None:
        table_changes = {}
    
    temp_dir = tempfile.mkdtemp()
    temp_xlsx = os.path.join(temp_dir, 'restored.xlsx')
    orig_temp_dir = tempfile.mkdtemp()
    
    try:
        # Extrahiere beide XLSX-Dateien
        with zipfile.ZipFile(output_path, 'r') as zf:
            zf.extractall(temp_dir)
        with zipfile.ZipFile(original_path, 'r') as zf:
            zf.extractall(orig_temp_dir)
        
        fixed_count = 0
        
        # Finde alle table*.xml Dateien
        tables_dir = os.path.join(temp_dir, 'xl', 'tables')
        orig_tables_dir = os.path.join(orig_temp_dir, 'xl', 'tables')
        
        
        if os.path.exists(tables_dir) and os.path.exists(orig_tables_dir):
            for f in os.listdir(tables_dir):
                if not f.startswith('table') or not f.endswith('.xml'):
                    continue
                
                
                export_table_path = os.path.join(tables_dir, f)
                orig_table_path = os.path.join(orig_tables_dir, f)
                
                if not os.path.exists(orig_table_path):
                    continue
                
                # Lies beide Dateien
                with open(export_table_path, 'r', encoding='utf-8') as fp:
                    export_content = fp.read()
                with open(orig_table_path, 'r', encoding='utf-8') as fp:
                    orig_content = fp.read()
                
                # Extrahiere table name aus Export
                name_match = re.search(r'name="([^"]+)"', export_content)
                if not name_match:
                    continue
                table_name = name_match.group(1)
                
                # Prüfe ob wir Änderungen für diese Table haben
                if table_name not in table_changes:
                    # Keine Änderungen - kopiere einfach das Original
                    with open(export_table_path, 'w', encoding='utf-8') as fp:
                        fp.write(orig_content)
                    fixed_count += 1
                    continue
                
                changes = table_changes[table_name]
                new_ref = changes.get('ref')
                new_columns = changes.get('columns', [])
                
                # Starte mit dem Original-Content
                new_content = orig_content
                
                # Aktualisiere ref in <table> und <autoFilter>
                if new_ref:
                    # Table ref
                    new_content = re.sub(r'(<table[^>]*\s)ref="[^"]+"', f'\\1ref="{new_ref}"', new_content)
                    # AutoFilter ref
                    new_content = re.sub(r'(<autoFilter[^>]*\s)ref="[^"]+"', f'\\1ref="{new_ref}"', new_content)
                
                # filterColumn/sortState bereinigen bei Spaltenänderungen
                # Wenn sich die Spaltenanzahl geändert hat, werden die colId-Werte
                # ungültig → Excel verwirft die gesamte Tabelle
                if new_columns:
                    orig_col_count_m = re.search(r'<tableColumns\s+count="(\d+)"', orig_content)
                    orig_col_count = int(orig_col_count_m.group(1)) if orig_col_count_m else 0
                    if len(new_columns) != orig_col_count:
                        # Spaltenanzahl hat sich geändert → filterColumn-Einträge entfernen
                        # (colId-Werte sind nicht mehr gültig)
                        new_content = re.sub(
                            r'<filterColumn\s[^>]*colId="[^"]*"[^>]*/>\s*', '', new_content)
                        new_content = re.sub(
                            r'<filterColumn\s[^>]*colId="[^"]*"[^>]*>.*?</filterColumn>\s*',
                            '', new_content, flags=re.DOTALL)
                        # sortState/sortCondition auch entfernen
                        new_content = re.sub(
                            r'<sortState[^>]*>.*?</sortState>\s*',
                            '', new_content, flags=re.DOTALL)
                        # Leere autoFilter bereinigen
                        new_content = re.sub(
                            r'(<autoFilter\s[^>]*?)>\s*</autoFilter>', r'\1/>', new_content)
                
                # Aktualisiere tableColumns
                if new_columns:
                    # Finde den tableColumns-Block
                    tc_match = re.search(r'<tableColumns[^>]*>.*?</tableColumns>', new_content, re.DOTALL)
                    if tc_match:
                        # Extrahiere die Original-Columns
                        orig_columns = re.findall(r'<tableColumn\s[^/]*(?:/>|>.*?</tableColumn>)', tc_match.group(0), re.DOTALL)
                        
                        # Erstelle ein Dict: orig_name -> Liste von (index, xml) für Duplikate
                        orig_by_name = {}
                        for idx, orig_col in enumerate(orig_columns):
                            name_match = re.search(r'name="([^"]+)"', orig_col)
                            if name_match:
                                orig_name = name_match.group(1)
                                if orig_name not in orig_by_name:
                                    orig_by_name[orig_name] = []
                                orig_by_name[orig_name].append((idx, orig_col))
                        
                        # Zähler für bereits verwendete Duplikate pro Name
                        used_count = {}
                        
                        # Baue neue tableColumns
                        new_tc_content = f'<tableColumns count="{len(new_columns)}">'
                        
                        for i, col_name in enumerate(new_columns):
                            matching_orig = None
                            
                            # Suche nach Original-Column mit gleichem Namen
                            if col_name in orig_by_name:
                                # Wie viele mit diesem Namen haben wir schon verwendet?
                                used = used_count.get(col_name, 0)
                                available = orig_by_name[col_name]
                                
                                if used < len(available):
                                    # Nimm die nächste verfügbare mit diesem Namen
                                    matching_orig = available[used][1]
                                    used_count[col_name] = used + 1
                            
                            if matching_orig:
                                # Nutze Original-Column und aktualisiere nur die ID und den Namen
                                col_xml = re.sub(r'id="\d+"', f'id="{i+1}"', matching_orig)
                                # Name auch aktualisieren (für den Fall dass er sich geändert hat)
                                safe_name = col_name.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;').replace('"', '&quot;')
                                col_xml = re.sub(r'name="[^"]+"', f'name="{safe_name}"', col_xml)
                                new_tc_content += col_xml
                            else:
                                # Neue Spalte ohne xr3:uid
                                # Escape special XML chars in name
                                safe_name = col_name.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;').replace('"', '&quot;')
                                new_tc_content += f'<tableColumn id="{i+1}" name="{safe_name}"/>'
                        
                        new_tc_content += '</tableColumns>'
                        new_content = new_content[:tc_match.start()] + new_tc_content + new_content[tc_match.end():]
                
                # Schreibe die reparierte Datei
                with open(export_table_path, 'w', encoding='utf-8') as fp:
                    fp.write(new_content)
                fixed_count += 1
        
        
        if fixed_count > 0:
            # ZIP-to-ZIP Re-Zip: Starte vom INPUT-ZIP und ersetze modifizierte Dateien.
            # Bewahrt ALLE Einträge (drawings, media, embeddings etc.)
            temp_files = {}
            for root, dirs, files in os.walk(temp_dir):
                dirs[:] = [d for d in dirs if d != '__MACOSX']
                for f in files:
                    if f == 'restored.xlsx' or f == '.DS_Store' or f.startswith('._'):
                        continue
                    full_path = os.path.join(root, f)
                    arc_name = os.path.relpath(full_path, temp_dir).replace('\\', '/')
                    temp_files[arc_name] = full_path
            
            written = set()
            with zipfile.ZipFile(temp_xlsx, 'w', zipfile.ZIP_DEFLATED) as new_zf:
                # Phase 1: Alle Einträge aus dem OUTPUT-ZIP durchgehen
                with zipfile.ZipFile(output_path, 'r') as input_zf:
                    for item in input_zf.infolist():
                        if item.filename.endswith('/'):
                            continue
                        name = item.filename
                        if name.startswith('__MACOSX') or name.endswith('.DS_Store') or \
                           name.split('/')[-1].startswith('._'):
                            continue
                        if name in temp_files:
                            new_zf.write(temp_files[name], name)
                        else:
                            data = input_zf.read(name)
                            info = item
                            info.compress_type = zipfile.ZIP_DEFLATED
                            new_zf.writestr(info, data)
                        written.add(name)
                
                # Phase 2: Neue Dateien aus temp_dir die NICHT im Input waren
                for arc_name, full_path in temp_files.items():
                    if arc_name not in written:
                        new_zf.write(full_path, arc_name)
                        written.add(arc_name)
            
            shutil.copy2(temp_xlsx, output_path)
    
    finally:
        shutil.rmtree(temp_dir, ignore_errors=True)
        shutil.rmtree(orig_temp_dir, ignore_errors=True)


# Worksheet-Element-Reihenfolge nach ECMA-376 (OpenXML Schema)
# Diese Elemente müssen am Ende von <worksheet> in genau dieser Reihenfolge stehen.
_WORKSHEET_END_ELEMENTS = [
    'drawing', 'legacyDrawing', 'legacyDrawingHF', 'drawingHF',
    'picture', 'oleObjects', 'controls', 'webPublishItems',
    'tableParts', 'extLst'
]


def _insert_ws_element(ws_content, element_xml, element_name):
    """
    Fügt ein Element in Worksheet-XML an der schema-konformen Position ein.
    
    Excel repariert Dateien mit falscher Element-Reihenfolge — dabei können
    Elemente (z.B. <drawing>) verloren gehen. Deshalb muss das Element
    VOR allen Elementen eingefügt werden, die im Schema DANACH kommen.
    
    WICHTIG: Sucht nur im TAIL-Bereich des Worksheets (nach </sheetData>),
    um Verwechslung mit verschachtelten Elementen wie <extLst> innerhalb
    von <sheetViews>, <conditionalFormatting> etc. zu vermeiden.
    """
    import re
    try:
        idx = _WORKSHEET_END_ELEMENTS.index(element_name)
    except ValueError:
        # Unbekanntes Element → sicher vor </worksheet> einfügen
        return ws_content.replace('</worksheet>', element_xml + '\n</worksheet>')
    
    # KRITISCH: Nur im TAIL suchen (nach </sheetData>), um verschachtelte
    # Elemente wie <extLst> innerhalb <sheetViews> nicht zu treffen.
    tail_start = 0
    sd_match = re.search(r'</sheetData>', ws_content)
    if sd_match:
        tail_start = sd_match.end()
    
    # Finde die früheste Position eines nachfolgenden Elements IM TAIL
    for after_elem in _WORKSHEET_END_ELEMENTS[idx + 1:]:
        # Suche nach <elementName (mit Leerzeichen oder >) um Verwechslung zu vermeiden
        pos_match = re.search(r'<' + re.escape(after_elem) + r'[\s>/]', ws_content[tail_start:])
        if pos_match:
            insert_pos = tail_start + pos_match.start()
            return ws_content[:insert_pos] + element_xml + '\n' + ws_content[insert_pos:]
    
    # Kein nachfolgendes Element gefunden → vor </worksheet>
    return ws_content.replace('</worksheet>', element_xml + '\n</worksheet>')


def _strip_slicers_from_zip(xlsx_path, source_bytes=None, return_bytes=False):
    """
    Entfernt ALLE Slicer-Infrastruktur aus einer fertiggestellten XLSX-Datei (ZIP-to-ZIP).
    
    Wird nach direct_xml_column_operations() aufgerufen, wenn Spalten gelöscht/eingefügt/
    verschoben wurden. Spaltenoperationen invalidieren SlicerCache-Spaltenreferenzen →
    orphaned Slicer-Einträge → Excel-Reparaturmodus.
    
    Entfernt werden:
    1. xl/slicerCaches/*.xml Dateien
    2. xl/slicers/*.xml Dateien
    3. Slicer-Overrides aus [Content_Types].xml
    4. Slicer-Relationships aus xl/_rels/workbook.xml.rels
    5. Slicer-Relationships aus xl/worksheets/_rels/sheet*.xml.rels
    6. Slicer-Shapes (mc:AlternateContent) aus xl/drawings/drawing*.xml
    7. Slicer-Relationships aus xl/drawings/_rels/drawing*.xml.rels
    8. Slicer-extLst aus xl/workbook.xml
    9. Slicer-extLst aus xl/worksheets/sheet*.xml
    """
    import zipfile
    import shutil
    import re
    import sys

    SLICER_URIS = [
        'http://schemas.microsoft.com/office/drawing/2010/slicer',
        'http://schemas.microsoft.com/office/drawing/2014/slicer',
    ]
    # extLst URIs für Slicer in workbook.xml und sheet XMLs
    SLICER_EXTLST_URIS = [
        '{A8765BA9-456A-4dab-B4F3-ACF838C121DE}',   # x14:slicerList (Sheet)
        '{79F54976-1DA5-4618-B6BA-64F0C8D81F0C}',   # x15:slicerCaches (Workbook)
    ]

    temp_output = xlsx_path + '.slicer_strip_tmp' if not return_bytes else None
    stripped_anything = False

    try:
        src_stream = source_bytes if source_bytes is not None else xlsx_path
        with zipfile.ZipFile(src_stream, 'r') as src_zip:
            namelist = src_zip.namelist()

            # Prüfe ob überhaupt Slicer-Artefakte vorhanden
            # Breit prüfen: Dateien, Content_Types, workbook.xml, Drawings, Rels
            has_slicer_files = any(
                n.startswith('xl/slicerCaches/') or n.startswith('xl/slicers/')
                for n in namelist
            )
            has_slicer_refs = False
            for check_path in ['[Content_Types].xml', 'xl/workbook.xml', 'xl/_rels/workbook.xml.rels']:
                if check_path in namelist:
                    check_content = src_zip.read(check_path).decode('utf-8').lower()
                    if 'slicer' in check_content:
                        has_slicer_refs = True
                        break
            if not has_slicer_refs:
                # Auch in Worksheet-XMLs und Drawing-XMLs prüfen
                for n in namelist:
                    if (re.match(r'xl/worksheets/sheet\d+\.xml$', n) or
                        re.match(r'xl/drawings/drawing\d+\.xml$', n) or
                        re.match(r'xl/worksheets/_rels/sheet\d+\.xml\.rels$', n) or
                        re.match(r'xl/drawings/_rels/drawing\d+\.xml\.rels$', n)):
                        check_content = src_zip.read(n).decode('utf-8').lower()
                        if 'slicer' in check_content:
                            has_slicer_refs = True
                            break

            if not has_slicer_files and not has_slicer_refs:
                sys.stderr.write(f"[SLICER_STRIP] Keine Slicer gefunden — übersprungen\n")
                if return_bytes:
                    if source_bytes is not None:
                        source_bytes.seek(0)
                    return source_bytes
                return False
            
            sys.stderr.write(f"[SLICER_STRIP] Slicer-Artefakte erkannt: files={has_slicer_files}, refs={has_slicer_refs}\n")

            # Dateien die komplett entfernt werden
            skip_files = set()
            for n in namelist:
                if n.startswith('xl/slicerCaches/') or n.startswith('xl/slicers/'):
                    skip_files.add(n)
                    sys.stderr.write(f"[SLICER_STRIP] Datei entfernt: {n}\n")

            if skip_files:
                stripped_anything = True

            # Modifizierte Dateien sammeln
            modified_files = {}

            # --- [Content_Types].xml: Slicer-Overrides entfernen ---
            if '[Content_Types].xml' in namelist:
                ct = src_zip.read('[Content_Types].xml').decode('utf-8')
                ct_orig = ct
                # Entferne Override-Einträge für slicerCaches und slicers
                ct = re.sub(
                    r'<Override\s[^>]*PartName="[^"]*slicer[^"]*"[^>]*/>\s*',
                    '', ct, flags=re.IGNORECASE)
                if ct != ct_orig:
                    modified_files['[Content_Types].xml'] = ct.encode('utf-8')
                    stripped_anything = True
                    sys.stderr.write(f"[SLICER_STRIP] [Content_Types].xml: Slicer-Overrides entfernt\n")

            # --- xl/_rels/workbook.xml.rels: Slicer-Rels entfernen ---
            wb_rels_path = 'xl/_rels/workbook.xml.rels'
            if wb_rels_path in namelist:
                wb_rels = src_zip.read(wb_rels_path).decode('utf-8')
                wb_rels_orig = wb_rels
                for rel_m in list(re.finditer(r'<Relationship\s[^>]*/>', wb_rels)):
                    rel_el = rel_m.group(0)
                    type_m = re.search(r'Type="([^"]+)"', rel_el)
                    target_m = re.search(r'Target="([^"]+)"', rel_el)
                    is_slicer = False
                    if type_m and 'slicer' in type_m.group(1).lower():
                        is_slicer = True
                    elif target_m and 'slicer' in target_m.group(1).lower():
                        is_slicer = True
                    if is_slicer:
                        wb_rels = wb_rels.replace(rel_el, '')
                        sys.stderr.write(f"[SLICER_STRIP] workbook.xml.rels: Slicer-Rel entfernt"
                                         f" Target={target_m.group(1) if target_m else '?'}\n")
                if wb_rels != wb_rels_orig:
                    wb_rels = re.sub(r'\n\s*\n', '\n', wb_rels)
                    modified_files[wb_rels_path] = wb_rels.encode('utf-8')
                    stripped_anything = True

            # --- xl/workbook.xml: Slicer-extLst UND Slicer-definedNames entfernen ---
            if 'xl/workbook.xml' in namelist:
                wb_xml = src_zip.read('xl/workbook.xml').decode('utf-8')
                wb_xml_orig = wb_xml
                # 1) Entferne <ext> Blöcke mit bekannten Slicer-URIs
                for uri in SLICER_EXTLST_URIS:
                    escaped_uri = re.escape(uri)
                    wb_xml = re.sub(
                        r'<ext\s[^>]*uri="' + escaped_uri + r'"[^>]*>.*?</ext>\s*',
                        '', wb_xml, flags=re.DOTALL | re.IGNORECASE)
                # 2) Catch-All: <ext> Blöcke die "slicer" enthalten (alternative URIs)
                pos = 0
                while True:
                    m = re.search(r'<ext\s[^>]*>.*?</ext>\s*', wb_xml[pos:], re.DOTALL)
                    if not m:
                        break
                    block = m.group(0)
                    abs_start = pos + m.start()
                    abs_end = pos + m.end()
                    if 'slicer' in block.lower():
                        wb_xml = wb_xml[:abs_start] + wb_xml[abs_end:]
                        sys.stderr.write(f"[SLICER_STRIP] workbook.xml: <ext> Block mit 'slicer' entfernt\n")
                    else:
                        pos = abs_end
                # 3) Entferne leere <extLst></extLst>
                wb_xml = re.sub(r'<extLst>\s*</extLst>', '', wb_xml)
                # 4) Slicer-definedNames entfernen (Name enthält "Slicer" — immer hidden)
                # Pattern: <definedName name="Slicer_..." ...>...</definedName>
                # oder: <definedName name="_xlnm.Slicer_..." ...>...</definedName>
                dn_removed = 0
                for dn_m in list(re.finditer(
                    r'<definedName\s[^>]*name="([^"]*)"[^>]*>.*?</definedName>\s*',
                    wb_xml, re.DOTALL)):
                    dn_name = dn_m.group(1)
                    if 'slicer' in dn_name.lower():
                        wb_xml = wb_xml[:dn_m.start()] + wb_xml[dn_m.end():]
                        dn_removed += 1
                        # Offset ändert sich — neu suchen
                        break  # wird durch while-Schleife unten erneut durchlaufen
                # Wiederhole bis keine Slicer-definedNames mehr gefunden
                while True:
                    dn_m = re.search(
                        r'<definedName\s[^>]*name="([^"]*[Ss]licer[^"]*)"[^>]*>.*?</definedName>\s*',
                        wb_xml, re.DOTALL)
                    if not dn_m:
                        break
                    wb_xml = wb_xml[:dn_m.start()] + wb_xml[dn_m.end():]
                    dn_removed += 1
                if dn_removed > 0:
                    sys.stderr.write(f"[SLICER_STRIP] workbook.xml: {dn_removed} Slicer-definedNames entfernt\n")
                    # Entferne leere <definedNames></definedNames> falls alle entfernt
                    wb_xml = re.sub(r'<definedNames>\s*</definedNames>', '', wb_xml)
                if wb_xml != wb_xml_orig:
                    modified_files['xl/workbook.xml'] = wb_xml.encode('utf-8')
                    stripped_anything = True
                    sys.stderr.write(f"[SLICER_STRIP] workbook.xml: Slicer-Artefakte entfernt\n")

            # --- Worksheet-Rels und Sheet-XMLs ---
            for n in namelist:
                # Worksheet _rels: Slicer-Rels entfernen
                if re.match(r'xl/worksheets/_rels/sheet\d+\.xml\.rels$', n):
                    ws_rels = src_zip.read(n).decode('utf-8')
                    ws_rels_orig = ws_rels
                    for rel_m in list(re.finditer(r'<Relationship\s[^>]*/>', ws_rels)):
                        rel_el = rel_m.group(0)
                        type_m = re.search(r'Type="([^"]+)"', rel_el)
                        target_m = re.search(r'Target="([^"]+)"', rel_el)
                        is_slicer = False
                        if type_m and 'slicer' in type_m.group(1).lower():
                            is_slicer = True
                        elif target_m and 'slicer' in target_m.group(1).lower():
                            is_slicer = True
                        if is_slicer:
                            ws_rels = ws_rels.replace(rel_el, '')
                            sys.stderr.write(f"[SLICER_STRIP] {n}: Slicer-Rel entfernt\n")
                    if ws_rels != ws_rels_orig:
                        ws_rels = re.sub(r'\n\s*\n', '\n', ws_rels)
                        modified_files[n] = ws_rels.encode('utf-8')
                        stripped_anything = True

                # Sheet XML: Slicer-extLst entfernen
                if re.match(r'xl/worksheets/sheet\d+\.xml$', n):
                    sheet_xml = src_zip.read(n).decode('utf-8')
                    sheet_xml_orig = sheet_xml
                    # 1) Bekannte Slicer-URIs
                    for uri in SLICER_EXTLST_URIS:
                        escaped_uri = re.escape(uri)
                        sheet_xml = re.sub(
                            r'<ext\s[^>]*uri="' + escaped_uri + r'"[^>]*>.*?</ext>\s*',
                            '', sheet_xml, flags=re.DOTALL | re.IGNORECASE)
                    # 2) Catch-All: <ext> Blöcke die "slicer" enthalten
                    ext_pos = 0
                    while True:
                        ext_m = re.search(r'<ext\s[^>]*>.*?</ext>\s*', sheet_xml[ext_pos:], re.DOTALL)
                        if not ext_m:
                            break
                        block = ext_m.group(0)
                        abs_start = ext_pos + ext_m.start()
                        abs_end = ext_pos + ext_m.end()
                        if 'slicer' in block.lower():
                            sheet_xml = sheet_xml[:abs_start] + sheet_xml[abs_end:]
                            sys.stderr.write(f"[SLICER_STRIP] {n}: <ext> Block mit 'slicer' entfernt\n")
                        else:
                            ext_pos = abs_end
                    # Entferne leere <extLst></extLst>
                    sheet_xml = re.sub(r'<extLst>\s*</extLst>', '', sheet_xml)
                    if sheet_xml != sheet_xml_orig:
                        modified_files[n] = sheet_xml.encode('utf-8')
                        stripped_anything = True
                        sys.stderr.write(f"[SLICER_STRIP] {n}: Slicer-extLst entfernt\n")

            # --- Drawing-XMLs: Slicer-Shapes entfernen ---
            for n in namelist:
                if re.match(r'xl/drawings/drawing\d+\.xml$', n):
                    drawing = src_zip.read(n).decode('utf-8')
                    drawing_orig = drawing
                    removed = 0

                    # Anchor-Blöcke mit Slicer-URIs entfernen
                    for anchor_tag in ['twoCellAnchor', 'oneCellAnchor', 'absoluteAnchor']:
                        for prefix in ['xdr:', '']:
                            pattern = f'<{prefix}{anchor_tag}[\\s>].*?</{prefix}{anchor_tag}>'
                            pos = 0
                            while True:
                                m = re.search(pattern, drawing[pos:], re.DOTALL)
                                if not m:
                                    break
                                block = m.group(0)
                                abs_start = pos + m.start()
                                abs_end = pos + m.end()
                                if any(uri in block for uri in SLICER_URIS):
                                    drawing = drawing[:abs_start] + drawing[abs_end:]
                                    removed += 1
                                else:
                                    pos = abs_end

                    # Standalone mc:AlternateContent mit Slicer
                    pos = 0
                    while True:
                        m = re.search(r'<mc:AlternateContent\b[^>]*>.*?</mc:AlternateContent>',
                                      drawing[pos:], re.DOTALL)
                        if not m:
                            break
                        block = m.group(0)
                        abs_start = pos + m.start()
                        abs_end = pos + m.end()
                        if any(uri in block for uri in SLICER_URIS):
                            drawing = drawing[:abs_start] + drawing[abs_end:]
                            removed += 1
                        else:
                            pos = abs_end

                    if drawing != drawing_orig:
                        modified_files[n] = drawing.encode('utf-8')
                        stripped_anything = True
                        sys.stderr.write(f"[SLICER_STRIP] {n}: {removed} Slicer-Shapes entfernt\n")

                # Drawing rels: Slicer-Rels entfernen
                if re.match(r'xl/drawings/_rels/drawing\d+\.xml\.rels$', n):
                    dr_rels = src_zip.read(n).decode('utf-8')
                    dr_rels_orig = dr_rels
                    for rel_m in list(re.finditer(r'<Relationship\s[^>]*/>', dr_rels)):
                        rel_el = rel_m.group(0)
                        type_m = re.search(r'Type="([^"]+)"', rel_el)
                        target_m = re.search(r'Target="([^"]+)"', rel_el)
                        is_slicer = False
                        if type_m and 'slicer' in type_m.group(1).lower():
                            is_slicer = True
                        elif target_m and 'slicer' in target_m.group(1).lower():
                            is_slicer = True
                        if is_slicer:
                            dr_rels = dr_rels.replace(rel_el, '')
                            sys.stderr.write(f"[SLICER_STRIP] {n}: Slicer-Rel entfernt\n")
                    if dr_rels != dr_rels_orig:
                        dr_rels = re.sub(r'\n\s*\n', '\n', dr_rels)
                        modified_files[n] = dr_rels.encode('utf-8')
                        stripped_anything = True

            # --- Prüfe ob Drawings nach Slicer-Stripping leer geworden sind ---
            # Wenn eine drawing*.xml keine Anchors mehr hat, müssen auch die
            # Sheet-Referenzen (<drawing r:id="..."/>) und Content_Types-Overrides
            # entfernt werden, sonst bleibt ein leeres Drawing-Artefakt.
            empty_drawings = set()  # ZIP-Pfade von leeren Drawings
            for n in list(modified_files.keys()):
                if re.match(r'xl/drawings/drawing\d+\.xml$', n):
                    drawing_content = modified_files[n].decode('utf-8')
                    has_anchors = bool(re.search(
                        r'<(?:xdr:)?(?:twoCellAnchor|oneCellAnchor|absoluteAnchor)[\s>]',
                        drawing_content))
                    if not has_anchors:
                        empty_drawings.add(n)
                        skip_files.add(n)  # Leere Drawing-Datei entfernen
                        del modified_files[n]
                        stripped_anything = True
                        sys.stderr.write(f"[SLICER_STRIP] {n}: Leeres Drawing entfernt\n")
                        # Auch zugehörige Rels-Datei entfernen
                        rels_path_for_drawing = n.replace('drawings/', 'drawings/_rels/') + '.rels'
                        if rels_path_for_drawing in namelist or rels_path_for_drawing in modified_files:
                            skip_files.add(rels_path_for_drawing)
                            modified_files.pop(rels_path_for_drawing, None)
                            sys.stderr.write(f"[SLICER_STRIP] {rels_path_for_drawing}: Leere Drawing-Rels entfernt\n")

            if empty_drawings:
                # Drawing-Overrides aus Content_Types entfernen
                ct_key = '[Content_Types].xml'
                if ct_key in modified_files:
                    ct = modified_files[ct_key].decode('utf-8')
                elif ct_key in namelist:
                    ct = src_zip.read(ct_key).decode('utf-8')
                else:
                    ct = None
                if ct:
                    ct_before = ct
                    for dp in empty_drawings:
                        part_name = '/' + dp  # e.g. /xl/drawings/drawing1.xml
                        escaped_pn = re.escape(part_name)
                        ct = re.sub(
                            r'<Override\s[^>]*PartName="' + escaped_pn + r'"[^>]*/>\s*',
                            '', ct)
                    if ct != ct_before:
                        modified_files[ct_key] = ct.encode('utf-8')
                        sys.stderr.write(f"[SLICER_STRIP] [Content_Types].xml: Drawing-Overrides für leere Drawings entfernt\n")

                # <drawing r:id="..."/> Referenzen aus Sheet-XMLs entfernen
                for n in namelist:
                    if not re.match(r'xl/worksheets/sheet\d+\.xml$', n):
                        continue
                    sheet_xml = modified_files.get(n)
                    if sheet_xml:
                        sheet_xml = sheet_xml.decode('utf-8')
                    else:
                        sheet_xml = src_zip.read(n).decode('utf-8')
                    sheet_orig = sheet_xml
                    # Finde den rId der Drawing-Referenz und prüfe ob die Ziel-Drawing leer ist
                    drawing_ref_m = re.search(r'<drawing\s+[^>]*r:id="([^"]+)"[^>]*/>', sheet_xml)
                    if not drawing_ref_m:
                        drawing_ref_m = re.search(r'<drawing\s+[^>]*r:id="([^"]+)"[^>]*>.*?</drawing>', sheet_xml, re.DOTALL)
                    if drawing_ref_m:
                        dr_rid = drawing_ref_m.group(1)
                        # Finde die Drawing-Datei über die Worksheet-Rels
                        ws_rels_path = n.replace('worksheets/', 'worksheets/_rels/') + '.rels'
                        ws_rels_content = None
                        if ws_rels_path in modified_files:
                            ws_rels_content = modified_files[ws_rels_path].decode('utf-8')
                        elif ws_rels_path in namelist:
                            ws_rels_content = src_zip.read(ws_rels_path).decode('utf-8')
                        if ws_rels_content:
                            target_m = re.search(
                                r'Id="' + re.escape(dr_rid) + r'"[^>]*Target="([^"]+)"',
                                ws_rels_content)
                            if target_m:
                                target = target_m.group(1)
                                # Resolve relative path: ../drawings/drawing1.xml → xl/drawings/drawing1.xml
                                if not target.startswith('/'):
                                    resolved = 'xl/worksheets/' + target
                                    rparts = resolved.split('/')
                                    norm = []
                                    for p in rparts:
                                        if p == '..':
                                            if norm:
                                                norm.pop()
                                        elif p != '.':
                                            norm.append(p)
                                    resolved = '/'.join(norm)
                                else:
                                    resolved = target.lstrip('/')
                                if resolved in empty_drawings:
                                    # Drawing-Referenz aus Sheet entfernen
                                    sheet_xml = re.sub(r'<drawing\s+[^>]*/>', '', sheet_xml)
                                    sheet_xml = re.sub(r'<drawing\s+[^>]*>.*?</drawing>', '', sheet_xml, flags=re.DOTALL)
                                    # Auch die Drawing-Rel aus Worksheet-Rels entfernen
                                    rel_pattern = r'<Relationship\s[^>]*Id="' + re.escape(dr_rid) + r'"[^>]*/>'
                                    ws_rels_content = re.sub(rel_pattern, '', ws_rels_content)
                                    ws_rels_content = re.sub(r'\n\s*\n', '\n', ws_rels_content)
                                    modified_files[ws_rels_path] = ws_rels_content.encode('utf-8')
                                    sys.stderr.write(f"[SLICER_STRIP] {n}: <drawing> Referenz entfernt (leere drawing)\n")
                    if sheet_xml != sheet_orig:
                        modified_files[n] = sheet_xml.encode('utf-8')

            if not stripped_anything:
                sys.stderr.write(f"[SLICER_STRIP] Keine Slicer-Artefakte gefunden — Datei unverändert\n")
                if return_bytes:
                    if source_bytes is not None:
                        source_bytes.seek(0)
                    return source_bytes
                return False

            # ZIP neu schreiben ohne Slicer-Dateien, mit modifizierten Dateien
            import io as _io
            dst_target = _io.BytesIO() if return_bytes else temp_output
            with zipfile.ZipFile(dst_target, 'w', zipfile.ZIP_DEFLATED) as dst_zip:
                for item in src_zip.infolist():
                    if item.filename.endswith('/'):
                        continue
                    if item.filename.startswith('__MACOSX') or item.filename.endswith('.DS_Store'):
                        continue
                    if item.filename in skip_files:
                        continue
                    if item.filename in modified_files:
                        item.compress_type = zipfile.ZIP_DEFLATED
                        dst_zip.writestr(item, modified_files[item.filename])
                    else:
                        dst_zip.writestr(item, src_zip.read(item.filename))

        if return_bytes:
            dst_target.seek(0)
            sys.stderr.write(f"[SLICER_STRIP] Slicer-Infrastruktur erfolgreich entfernt (in-memory)\n")
            return dst_target

        # Ersetze Original
        os.remove(xlsx_path)
        shutil.move(temp_output, xlsx_path)
        sys.stderr.write(f"[SLICER_STRIP] Slicer-Infrastruktur erfolgreich entfernt aus {xlsx_path}\n")
        return True

    except Exception as e:
        if temp_output and os.path.exists(temp_output):
            os.remove(temp_output)
        sys.stderr.write(f"[SLICER_STRIP] Fehler: {e}\n")
        import traceback
        traceback.print_exc(file=sys.stderr)
        if return_bytes:
            raise
        return False


def _strip_pivot_tables_for_sheet(xlsx_path, sheet_name, source_bytes=None, return_bytes=False):
    """
    Entfernt PivotTable-Infrastruktur für ein bestimmtes Sheet aus einer XLSX-Datei (ZIP-to-ZIP).

    Wird nach direct_xml_column_operations() aufgerufen, wenn Spalten gelöscht/
    verschoben wurden. Spaltenoperationen invalidieren die PivotTable <location ref>
    und pivotField-Definitionen → Excel meldet "Zellinformationen"-Reparaturfehler.

    Entfernt werden:
    1. pivotTable*.xml Dateien die vom Sheet referenziert werden
    2. pivotTable*.xml.rels Dateien
    3. PivotTable-Relationships aus xl/worksheets/_rels/sheet*.xml.rels
    4. PivotTable-Overrides aus [Content_Types].xml
    5. Verwaiste PivotCache-Dateien (Definition + Records) falls kein anderes PivotTable
       sie noch referenziert
    6. Verwaiste PivotCache-Referenzen aus xl/workbook.xml und xl/_rels/workbook.xml.rels
    """
    import zipfile, shutil, os, re
    from xml.etree import ElementTree as ET

    MAIN_NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
    RELS_NS = 'http://schemas.openxmlformats.org/package/2006/relationships'
    R_NS = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'

    temp_output = xlsx_path + '.pvt_tmp' if not return_bytes else None

    try:
        src_stream = source_bytes if source_bytes is not None else xlsx_path
        with zipfile.ZipFile(src_stream, 'r') as src_zip:
            namelist = set(src_zip.namelist())
            modified_files = {}
            skip_files = set()
            stripped_anything = False

            # --- Finde den Sheet-ZIP-Pfad ---
            wb_xml_raw = src_zip.read('xl/workbook.xml').decode('utf-8')
            wb_root = ET.fromstring(wb_xml_raw)

            sheet_rid = None
            for sheet_el in wb_root.iter(f'{{{MAIN_NS}}}sheet'):
                if sheet_el.get('name') == sheet_name:
                    sheet_rid = sheet_el.get(f'{{{R_NS}}}id')
                    break

            if not sheet_rid:
                sys.stderr.write(f"[PIVOT_STRIP] Sheet '{sheet_name}' nicht gefunden\n")
                if return_bytes:
                    if source_bytes is not None:
                        source_bytes.seek(0)
                    return source_bytes
                return False

            wb_rels_raw = src_zip.read('xl/_rels/workbook.xml.rels').decode('utf-8')
            wb_rels_root = ET.fromstring(wb_rels_raw)

            sheet_file = None
            for rel_el in wb_rels_root.iter(f'{{{RELS_NS}}}Relationship'):
                if rel_el.get('Id') == sheet_rid:
                    sheet_file = rel_el.get('Target')
                    break

            if not sheet_file:
                sys.stderr.write(f"[PIVOT_STRIP] Relationship {sheet_rid} nicht gefunden\n")
                if return_bytes:
                    if source_bytes is not None:
                        source_bytes.seek(0)
                    return source_bytes
                return False

            sheet_zip_path = 'xl/' + sheet_file.lstrip('/')
            parts = sheet_zip_path.split('/')
            normalized = []
            for p in parts:
                if p == '..':
                    if normalized:
                        normalized.pop()
                elif p != '.':
                    normalized.append(p)
            sheet_zip_path = '/'.join(normalized)

            sheet_rels_path = sheet_zip_path.replace(
                'worksheets/', 'worksheets/_rels/') + '.rels'

            if sheet_rels_path not in namelist:
                sys.stderr.write(f"[PIVOT_STRIP] Keine Sheet-Rels gefunden: {sheet_rels_path}\n")
                return False

            # --- Finde PivotTable-Relationships vom Sheet ---
            sheet_rels_xml = src_zip.read(sheet_rels_path).decode('utf-8')
            sheet_rels_orig = sheet_rels_xml

            PIVOT_TABLE_TYPE = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotTable'
            pivot_files_to_remove = []  # ZIP-Pfade der zu entfernenden pivotTable*.xml

            for rel_m in list(re.finditer(r'<Relationship\s[^>]*/>', sheet_rels_xml)):
                rel_el = rel_m.group(0)
                type_m = re.search(r'Type="([^"]+)"', rel_el)
                target_m = re.search(r'Target="([^"]+)"', rel_el)
                if type_m and type_m.group(1) == PIVOT_TABLE_TYPE:
                    target = target_m.group(1) if target_m else ''
                    # Resolve relative path
                    if not target.startswith('/'):
                        resolved = 'xl/worksheets/' + target
                        rparts = resolved.split('/')
                        norm = []
                        for p in rparts:
                            if p == '..':
                                if norm:
                                    norm.pop()
                            elif p != '.':
                                norm.append(p)
                        resolved = '/'.join(norm)
                    else:
                        resolved = target.lstrip('/')

                    pivot_files_to_remove.append(resolved)
                    sheet_rels_xml = sheet_rels_xml.replace(rel_el, '')
                    sys.stderr.write(f"[PIVOT_STRIP] {sheet_rels_path}: PivotTable-Rel entfernt → {resolved}\n")

            if not pivot_files_to_remove:
                sys.stderr.write(f"[PIVOT_STRIP] Keine PivotTables auf Sheet '{sheet_name}' gefunden\n")
                if return_bytes:
                    if source_bytes is not None:
                        source_bytes.seek(0)
                    return source_bytes
                return False

            sheet_rels_xml = re.sub(r'\n\s*\n', '\n', sheet_rels_xml)
            modified_files[sheet_rels_path] = sheet_rels_xml.encode('utf-8')
            stripped_anything = True

            # --- Sammle cacheIds der zu entfernenden PivotTables ---
            removed_cache_ids = set()
            for pt_path in pivot_files_to_remove:
                if pt_path in namelist:
                    pt_xml = src_zip.read(pt_path).decode('utf-8')
                    cache_id_m = re.search(r'cacheId="(\d+)"', pt_xml)
                    if cache_id_m:
                        removed_cache_ids.add(cache_id_m.group(1))
                    skip_files.add(pt_path)
                    # Auch zugehörige .rels entfernen
                    pt_rels = pt_path.replace('pivotTables/', 'pivotTables/_rels/') + '.rels'
                    if pt_rels in namelist:
                        skip_files.add(pt_rels)
                    sys.stderr.write(f"[PIVOT_STRIP] {pt_path}: PivotTable-Datei entfernt\n")

            # --- Prüfe ob die Caches noch von ANDEREN PivotTables referenziert werden ---
            surviving_cache_ids = set()
            for n in namelist:
                if n.startswith('xl/pivotTables/pivotTable') and n.endswith('.xml') and n not in skip_files:
                    pt_xml = src_zip.read(n).decode('utf-8')
                    cid_m = re.search(r'cacheId="(\d+)"', pt_xml)
                    if cid_m:
                        surviving_cache_ids.add(cid_m.group(1))

            orphaned_cache_ids = removed_cache_ids - surviving_cache_ids
            sys.stderr.write(f"[PIVOT_STRIP] cacheIds entfernt={removed_cache_ids}, "
                             f"überlebend={surviving_cache_ids}, verwaist={orphaned_cache_ids}\n")

            # --- Entferne verwaiste PivotCaches ---
            if orphaned_cache_ids:
                # Finde die rIds der verwaisten Caches aus workbook.xml
                orphaned_wb_rids = set()
                wb_xml = src_zip.read('xl/workbook.xml').decode('utf-8')
                wb_xml_orig = wb_xml
                for cid in orphaned_cache_ids:
                    # <pivotCache cacheId="0" r:id="rId17"/>
                    pc_m = re.search(
                        r'<pivotCache\s[^>]*cacheId="' + re.escape(cid) + r'"[^>]*/>',
                        wb_xml)
                    if pc_m:
                        rid_m = re.search(r'r:id="([^"]+)"', pc_m.group(0))
                        if rid_m:
                            orphaned_wb_rids.add(rid_m.group(1))
                        wb_xml = wb_xml.replace(pc_m.group(0), '')
                        sys.stderr.write(f"[PIVOT_STRIP] workbook.xml: pivotCache cacheId={cid} entfernt\n")

                # Leere <pivotCaches></pivotCaches> entfernen
                wb_xml = re.sub(r'<pivotCaches>\s*</pivotCaches>', '', wb_xml)

                if wb_xml != wb_xml_orig:
                    modified_files['xl/workbook.xml'] = wb_xml.encode('utf-8')

                # Entferne verwaiste Cache-Rels aus workbook.xml.rels
                wb_rels_xml = src_zip.read('xl/_rels/workbook.xml.rels').decode('utf-8')
                wb_rels_orig = wb_rels_xml
                for rid in orphaned_wb_rids:
                    # Finde die Relationship und das Target (um die Datei auch zu entfernen)
                    rel_m = re.search(
                        r'<Relationship\s[^>]*Id="' + re.escape(rid) + r'"[^>]*/>',
                        wb_rels_xml)
                    if rel_m:
                        target_m = re.search(r'Target="([^"]+)"', rel_m.group(0))
                        if target_m:
                            cache_def_path = 'xl/' + target_m.group(1).lstrip('/')
                            # Normalisiere Pfad
                            norm = []
                            for p in cache_def_path.split('/'):
                                if p == '..':
                                    if norm:
                                        norm.pop()
                                elif p != '.':
                                    norm.append(p)
                            cache_def_path = '/'.join(norm)

                            skip_files.add(cache_def_path)
                            sys.stderr.write(f"[PIVOT_STRIP] {cache_def_path}: Cache-Definition entfernt\n")
                            # Versuche auch die Records-Datei zu finden
                            cache_def_rels = cache_def_path.replace(
                                'pivotCache/', 'pivotCache/_rels/') + '.rels'
                            if cache_def_rels in namelist:
                                cache_rels_xml = src_zip.read(cache_def_rels).decode('utf-8')
                                rec_m = re.search(r'Target="([^"]+)"', cache_rels_xml)
                                if rec_m:
                                    rec_path = cache_def_path.rsplit('/', 1)[0] + '/' + rec_m.group(1)
                                    skip_files.add(rec_path)
                                    sys.stderr.write(f"[PIVOT_STRIP] {rec_path}: Cache-Records entfernt\n")
                                skip_files.add(cache_def_rels)

                        wb_rels_xml = wb_rels_xml.replace(rel_m.group(0), '')
                        sys.stderr.write(f"[PIVOT_STRIP] workbook.xml.rels: Cache-Rel {rid} entfernt\n")

                wb_rels_xml = re.sub(r'\n\s*\n', '\n', wb_rels_xml)
                if wb_rels_xml != wb_rels_orig:
                    modified_files['xl/_rels/workbook.xml.rels'] = wb_rels_xml.encode('utf-8')

            # --- Content_Types aufräumen ---
            ct_key = '[Content_Types].xml'
            ct = src_zip.read(ct_key).decode('utf-8')
            ct_orig = ct
            for skip_path in skip_files:
                part_name = '/' + skip_path
                escaped_pn = re.escape(part_name)
                ct = re.sub(
                    r'<Override\s[^>]*PartName="' + escaped_pn + r'"[^>]*/>\s*',
                    '', ct)
            if ct != ct_orig:
                modified_files[ct_key] = ct.encode('utf-8')
                sys.stderr.write(f"[PIVOT_STRIP] [Content_Types].xml: PivotTable-Overrides entfernt\n")

            if not stripped_anything:
                sys.stderr.write(f"[PIVOT_STRIP] Keine PivotTable-Artefakte gefunden\n")
                if return_bytes:
                    if source_bytes is not None:
                        source_bytes.seek(0)
                    return source_bytes
                return False

            # --- ZIP neu schreiben ---
            import io as _io
            dst_target = _io.BytesIO() if return_bytes else temp_output
            with zipfile.ZipFile(dst_target, 'w', zipfile.ZIP_DEFLATED) as dst_zip:
                for item in src_zip.infolist():
                    if item.filename.endswith('/'):
                        continue
                    if item.filename.startswith('__MACOSX') or item.filename.endswith('.DS_Store'):
                        continue
                    if item.filename in skip_files:
                        continue
                    if item.filename in modified_files:
                        item.compress_type = zipfile.ZIP_DEFLATED
                        dst_zip.writestr(item, modified_files[item.filename])
                    else:
                        dst_zip.writestr(item, src_zip.read(item.filename))

        if return_bytes:
            dst_target.seek(0)
            sys.stderr.write(f"[PIVOT_STRIP] PivotTable-Infrastruktur für Sheet '{sheet_name}' erfolgreich entfernt (in-memory)\n")
            return dst_target

        # Ersetze Original
        os.remove(xlsx_path)
        shutil.move(temp_output, xlsx_path)
        sys.stderr.write(f"[PIVOT_STRIP] PivotTable-Infrastruktur für Sheet '{sheet_name}' erfolgreich entfernt\n")
        return True

    except Exception as e:
        if temp_output and os.path.exists(temp_output):
            os.remove(temp_output)
        sys.stderr.write(f"[PIVOT_STRIP] Fehler: {e}\n")
        import traceback
        traceback.print_exc(file=sys.stderr)
        if return_bytes:
            raise
        return False


def _strip_slicer_shapes_from_drawings(drawings_dir):
    """
    Entfernt Slicer-Shapes aus drawing*.xml Dateien.
    
    Slicer-Shapes werden als <mc:AlternateContent> Blöcke gespeichert, die
    graphicData mit Slicer-URIs enthalten. Nach Spaltenoperationen sind die
    referenzierten SlicerCaches invalide → Excel entfernt den Shape →
    "Entfernter Teil: Zeichnungsform".
    
    Entfernt werden:
    - <mc:AlternateContent> Blöcke mit Slicer-URIs in drawing*.xml
    - Slicer-Relationships in drawing*.xml.rels
    """
    import re
    import os
    import sys
    
    # Slicer-URIs die in graphicData vorkommen
    SLICER_URIS = [
        'http://schemas.microsoft.com/office/drawing/2010/slicer',
        'http://schemas.microsoft.com/office/drawing/2014/slicer',
    ]
    
    for fname in os.listdir(drawings_dir):
        if not fname.startswith('drawing') or not fname.endswith('.xml'):
            continue
        if fname.endswith('.rels'):
            continue
        
        fpath = os.path.join(drawings_dir, fname)
        with open(fpath, 'r', encoding='utf-8') as f:
            content = f.read()
        
        original_content = content
        removed_count = 0
        
        # Finde und entferne mc:AlternateContent Blöcke mit Slicer-URIs.
        # Diese Blöcke haben die Form:
        # <mc:AlternateContent>
        #   <mc:Choice Requires="...">
        #     <xdr:graphicFrame>
        #       <a:graphic>
        #         <a:graphicData uri="http://schemas.microsoft.com/office/drawing/2010/slicer">
        #           <sle:slicer name="..."/>
        #         </a:graphicData>
        #       </a:graphic>
        #     </xdr:graphicFrame>
        #   </mc:Choice>
        #   <mc:Fallback/>
        # </mc:AlternateContent>
        #
        # PLUS den umgebenden <xdr:twoCellAnchor> Block, der den graphicFrame enthält.
        # Der Anchor enthält Position, Größe etc. und ist der Container-Block im Drawing.
        
        # Strategie: Finde alle twoCellAnchor-Blöcke die einen Slicer enthalten.
        # twoCellAnchor kann auch als oneCellAnchor oder absoluteAnchor vorkommen.
        for anchor_tag in ['twoCellAnchor', 'oneCellAnchor', 'absoluteAnchor']:
            # Pattern: <xdr:TAG ...>...</xdr:TAG> oder <TAG ...>...</TAG>
            # Verwende nicht-gierigen Match
            for prefix in ['xdr:', '']:
                pattern = f'<{prefix}{anchor_tag}[\\s>].*?</{prefix}{anchor_tag}>'
                # While-Schleife mit search_start: nach jeder Entfernung
                # bleibt search_start gleich; bei Nicht-Slicer überspringen wir den Block
                anchor_search_start = 0
                while True:
                    m = re.search(pattern, content[anchor_search_start:], re.DOTALL)
                    if not m:
                        break
                    block = m.group(0)
                    abs_start = anchor_search_start + m.start()
                    abs_end = anchor_search_start + m.end()
                    is_slicer = any(uri in block for uri in SLICER_URIS)
                    if is_slicer:
                        content = content[:abs_start] + content[abs_end:]
                        removed_count += 1
                        sys.stderr.write(f"[strip_slicer] {fname}: Slicer-Anchor ({prefix}{anchor_tag}) entfernt\n")
                        # anchor_search_start bleibt gleich
                    else:
                        anchor_search_start = abs_end  # Nicht-Slicer → weiter suchen
        
        # Falls noch mc:AlternateContent mit Slicer außerhalb von Anchors existiert
        # While-Schleife: nach jeder Entfernung String-Positionen neu berechnen
        search_start = 0
        while True:
            m = re.search(r'<mc:AlternateContent\b[^>]*>.*?</mc:AlternateContent>', content[search_start:], re.DOTALL)
            if not m:
                break
            block = m.group(0)
            abs_start = search_start + m.start()
            abs_end = search_start + m.end()
            is_slicer = any(uri in block for uri in SLICER_URIS)
            if is_slicer:
                content = content[:abs_start] + content[abs_end:]
                removed_count += 1
                sys.stderr.write(f"[strip_slicer] {fname}: Standalone mc:AlternateContent mit Slicer entfernt\n")
                # search_start bleibt gleich (nächste Suche ab gleicher Position)
            else:
                search_start = abs_end  # Nicht-Slicer → weiter suchen nach diesem Block
        
        if content != original_content:
            with open(fpath, 'w', encoding='utf-8') as f:
                f.write(content)
            sys.stderr.write(f"[strip_slicer] {fname}: {removed_count} Slicer-Shapes entfernt, "
                           f"{len(original_content)} → {len(content)} bytes\n")
        
        # Entferne Slicer-Relationships aus drawing*.xml.rels
        rels_dir = os.path.join(drawings_dir, '_rels')
        rels_file = os.path.join(rels_dir, f'{fname}.rels')
        if os.path.exists(rels_file):
            with open(rels_file, 'r', encoding='utf-8') as f:
                rels_content = f.read()
            
            original_rels = rels_content
            # Slicer-Relationship-Types
            slicer_rel_types = [
                'slicer',
                'slicerCache',
            ]
            
            for rel_match in list(re.finditer(r'<Relationship\s[^>]*/>', rels_content)):
                rel_el = rel_match.group(0)
                type_m = re.search(r'Type="([^"]+)"', rel_el)
                target_m = re.search(r'Target="([^"]+)"', rel_el)
                if type_m:
                    rel_type = type_m.group(1).lower()
                    is_slicer_rel = any(s in rel_type for s in slicer_rel_types)
                    if is_slicer_rel:
                        rels_content = rels_content.replace(rel_el, '')
                        sys.stderr.write(f"[strip_slicer] {fname}.rels: Slicer-Rel entfernt: "
                                       f"Target={target_m.group(1) if target_m else '?'}\n")
                elif target_m:
                    # Auch nach Target prüfen (slicers/ Pfad)
                    target = target_m.group(1).lower()
                    if 'slicer' in target:
                        rels_content = rels_content.replace(rel_el, '')
                        sys.stderr.write(f"[strip_slicer] {fname}.rels: Slicer-Rel (nach Target) entfernt: {target}\n")
            
            if rels_content != original_rels:
                # Bereinige leere Zeilen
                rels_content = re.sub(r'\n\s*\n', '\n', rels_content)
                with open(rels_file, 'w', encoding='utf-8') as f:
                    f.write(rels_content)


def restore_external_links_from_original(output_path, original_path, structural_change=False):
    """
    Kopiert die externalLinks-Dateien, slicerCaches und definedNames aus dem Original zurück.
    
    openpyxl verliert wichtige XML-Namespaces wie xmlns:mc, mc:Ignorable, xmlns:x14 etc.,
    vereinfacht definedNames (entfernt localSheetId Attribute) und verliert Slicers komplett.
    
    Args:
        structural_change: Wenn True, werden Worksheet-Rels per MERGE wiederhergestellt
                           (openpyxl's Rels behalten, nur fehlende ergänzen).
                           Wenn False, werden Rels per REPLACE aus dem Original kopiert.
                           MERGE ist nötig bei Spalten/Zeilen-Operationen, wo openpyxl
                           die Rels aktualisiert hat. REPLACE ist sicher bei reinen
                           Format-Änderungen (FALL 3b), wo openpyxl die Struktur nicht ändert.
    """
    import tempfile
    import shutil
    import zipfile
    import re
    
    sys.stderr.write(f"[restore_ext] START: output_path={output_path}, original_path={original_path}\n")
    
    if not original_path:
        sys.stderr.write(f"[restore_ext] SKIPPED: original_path leer\n")
        return
    
    if not os.path.exists(original_path):
        sys.stderr.write(f"[restore_ext] SKIPPED: original_path existiert nicht: {original_path}\n")
        return
    
    # Wenn original_path == output_path ("Speichern" statt "Speichern unter"),
    # müssen wir eine temporäre Backup-Kopie erstellen, da openpyxl die Datei
    # bereits überschrieben hat. Die Backup-Kopie enthält noch die Original-Daten
    # WENN sie VOR dem openpyxl-Save erstellt wurde (siehe _backup_original_path).
    backup_path = None
    # WINDOWS: normpath normalisiert Pfade (Slashes, Groß/Klein, trailing sep)
    if os.path.normpath(original_path) == os.path.normpath(output_path):
        sys.stderr.write(f"[restore_ext] original_path == output_path — verwende _backup_original_path\n")
        # Prüfe ob ein Backup vor dem Save erstellt wurde
        backup_candidate = getattr(restore_external_links_from_original, '_backup_original_path', None)
        if backup_candidate and os.path.exists(backup_candidate):
            original_path = backup_candidate
            backup_path = backup_candidate
            sys.stderr.write(f"[restore_ext] Verwende Backup: {backup_path}\n")
        else:
            sys.stderr.write(f"[restore_ext] SKIPPED: Kein Backup verfügbar und original==output\n")
            return
    
    sys.stderr.write(f"[restore_ext] Original existiert, starte Wiederherstellung...\n")
    
    temp_dir = None
    orig_temp_dir = None
    
    try:
        temp_dir = tempfile.mkdtemp()
        orig_temp_dir = tempfile.mkdtemp()
        temp_xlsx = os.path.join(temp_dir, 'restored.xlsx')
        
        with zipfile.ZipFile(output_path, 'r') as zf:
            zf.extractall(temp_dir)
        with zipfile.ZipFile(original_path, 'r') as zf:
            # DEBUG: Zeige alle Dateien im Original-ZIP die mit Bildern/Drawings zu tun haben
            all_names = zf.namelist()
            drawing_related = [n for n in all_names if any(k in n.lower() for k in ['draw', 'media', 'image', 'picture', 'vml', 'richdata', 'rdrichvalue'])]
            sys.stderr.write(f"[restore_ext] Original-ZIP Dateien (drawing/media-related): {drawing_related}\n")
            sys.stderr.write(f"[restore_ext] Original-ZIP alle xl/ Dateien: {[n for n in all_names if n.startswith('xl/')]}\n")
            zf.extractall(orig_temp_dir)
        
        # DEBUG: Inhalt von sheet1.xml.rels anzeigen
        orig_rels_file = os.path.join(orig_temp_dir, 'xl', 'worksheets', '_rels', 'sheet1.xml.rels')
        if os.path.exists(orig_rels_file):
            with open(orig_rels_file, 'r', encoding='utf-8') as f:
                rels_content = f.read()
            sys.stderr.write(f"[restore_ext] sheet1.xml.rels Inhalt: {rels_content[:500]}\n")
        
        # DEBUG: Prüfe ob <drawing> Element in original sheet1.xml vorhanden
        orig_sheet1 = os.path.join(orig_temp_dir, 'xl', 'worksheets', 'sheet1.xml')
        if os.path.exists(orig_sheet1):
            with open(orig_sheet1, 'r', encoding='utf-8') as f:
                sheet1_content = f.read()
            has_drawing = '<drawing ' in sheet1_content or '<drawing>' in sheet1_content
            has_legacy = '<legacyDrawing ' in sheet1_content
            has_picture = '<picture ' in sheet1_content
            sys.stderr.write(f"[restore_ext] Original sheet1.xml: <drawing>={has_drawing}, <legacyDrawing>={has_legacy}, <picture>={has_picture}\n")
        
        # DEBUG: Prüfe ob <drawing> Element in output sheet1.xml vorhanden
        dest_sheet1 = os.path.join(temp_dir, 'xl', 'worksheets', 'sheet1.xml')
        if os.path.exists(dest_sheet1):
            with open(dest_sheet1, 'r', encoding='utf-8') as f:
                dest_sheet1_content = f.read()
            has_drawing_dest = '<drawing ' in dest_sheet1_content or '<drawing>' in dest_sheet1_content
            sys.stderr.write(f"[restore_ext] Output sheet1.xml: <drawing>={has_drawing_dest}\n")
        
        ext_links_dir = os.path.join(temp_dir, 'xl', 'externalLinks')
        orig_ext_links_dir = os.path.join(orig_temp_dir, 'xl', 'externalLinks')
        
        fixed_count = 0
        
        if os.path.exists(orig_ext_links_dir) and os.path.exists(ext_links_dir):
            # externalLink-Dateien VOM ORIGINAL kopieren (überschreiben!).
            # openpyxl korrumpiert externalLink-XML beim Round-Trip
            # (fehlende Namespaces, falsche Struktur, beschädigte cached values).
            # Das Original enthält die korrekten, von Excel validierten Versionen.
            # Die Reihenfolge/Anzahl der Dateien bleibt gleich (openpyxl ändert sie nicht),
            # daher stimmen die Indizes in workbook.xml/definedNames weiterhin.
            for f in os.listdir(orig_ext_links_dir):
                if f.startswith('externalLink') and f.endswith('.xml'):
                    orig_file = os.path.join(orig_ext_links_dir, f)
                    dest_file = os.path.join(ext_links_dir, f)
                    shutil.copy2(orig_file, dest_file)
                    sys.stderr.write(f"[restore_ext] externalLink aus Original kopiert: {f}\n")
                    fixed_count += 1
            
            # _rels ebenfalls vom Original kopieren (Konsistenz mit externalLink-Dateien)
            orig_rels_dir = os.path.join(orig_ext_links_dir, '_rels')
            dest_rels_dir = os.path.join(ext_links_dir, '_rels')
            if os.path.exists(orig_rels_dir):
                if os.path.exists(dest_rels_dir):
                    shutil.rmtree(dest_rels_dir)
                shutil.copytree(orig_rels_dir, dest_rels_dir)
                sys.stderr.write(f"[restore_ext] externalLinks/_rels komplett aus Original kopiert\n")
                fixed_count += 1
        elif os.path.exists(orig_ext_links_dir) and not os.path.exists(ext_links_dir):
            # openpyxl hat externalLinks komplett verloren → alles kopieren
            shutil.copytree(orig_ext_links_dir, ext_links_dir)
            sys.stderr.write(f"[restore_ext] externalLinks komplett aus Original kopiert (fehlten)\n")
            fixed_count += 1
        
        # Kopiere slicerCaches aus dem Original (openpyxl verliert Slicers komplett)
        # Bei structural_change (Spaltenoperationen) NICHT kopieren:
        # SlicerCaches referenzieren Tabellenspalten die sich geändert haben →
        # invalide Referenzen → Excel kaskadiert: Cache → Slicer → Zeichnungsform.
        # Ohne die Dateien gibt es nichts zum Reparieren.
        orig_slicer_dir = os.path.join(orig_temp_dir, 'xl', 'slicerCaches')
        dest_slicer_dir = os.path.join(temp_dir, 'xl', 'slicerCaches')
        if os.path.exists(orig_slicer_dir):
            if structural_change:
                sys.stderr.write(f"[restore_ext] xl/slicerCaches: ÜBERSPRUNGEN (structural_change=True, invalide nach Spaltenoperation)\n")
            else:
                if not os.path.exists(dest_slicer_dir):
                    os.makedirs(dest_slicer_dir)
                for f in os.listdir(orig_slicer_dir):
                    if f.endswith('.xml'):
                        shutil.copy2(os.path.join(orig_slicer_dir, f), os.path.join(dest_slicer_dir, f))
                        fixed_count += 1
        
        # Kopiere slicers Ordner auch (falls vorhanden)
        # Bei structural_change: NICHT kopieren (gleicher Grund wie slicerCaches)
        orig_slicers_dir = os.path.join(orig_temp_dir, 'xl', 'slicers')
        dest_slicers_dir = os.path.join(temp_dir, 'xl', 'slicers')
        if os.path.exists(orig_slicers_dir):
            if structural_change:
                sys.stderr.write(f"[restore_ext] xl/slicers: ÜBERSPRUNGEN (structural_change=True)\n")
            else:
                if os.path.exists(dest_slicers_dir):
                    shutil.rmtree(dest_slicers_dir)
                shutil.copytree(orig_slicers_dir, dest_slicers_dir)
                fixed_count += 1
        
        # sharedStrings.xml NICHT vom Original kopieren!
        # openpyxl erzeugt beim Speichern eine NEUE sharedStrings.xml mit neu
        # nummerierten Indizes. Die Worksheet-Zellen referenzieren diese neuen
        # Indizes (<c t="s"><v>N</v></c>). Wenn wir die Original-sharedStrings.xml
        # zurückkopieren, zeigen die Indizes auf falsche Strings →
        # Excel meldet "Reparatur auf Dateiebene".
        # (openpyxl schreibt reguläre Strings als shared strings, NICHT inline!
        #  Nur CellRichText-Objekte werden als inlineStr geschrieben.)
        sys.stderr.write(f"[restore_ext] sharedStrings.xml: NICHT kopiert (openpyxl-Indizes beibehalten)\n")
        
        # Stelle workbook.xml SELEKTIV wieder her
        # NICHT blind kopieren! openpyxl aktualisiert externalReferences cached values.
        # Nur Namespace-Deklarationen und definedNames vom Original übernehmen.
        workbook_path = os.path.join(temp_dir, 'xl', 'workbook.xml')
        orig_workbook_path = os.path.join(orig_temp_dir, 'xl', 'workbook.xml')
        
        if os.path.exists(workbook_path) and os.path.exists(orig_workbook_path):
            with open(workbook_path, 'r', encoding='utf-8') as f:
                dest_wb_content = f.read()
            with open(orig_workbook_path, 'r', encoding='utf-8') as f:
                orig_wb_content = f.read()
            
            wb_modified = False
            
            # 1. Namespace-Deklarationen vom Original übernehmen
            orig_wb_root = re.search(r'(<workbook\b[^>]+>)', orig_wb_content)
            dest_wb_root = re.search(r'(<workbook\b[^>]+>)', dest_wb_content)
            if orig_wb_root and dest_wb_root and orig_wb_root.group(1) != dest_wb_root.group(1):
                dest_wb_content = dest_wb_content.replace(dest_wb_root.group(1), orig_wb_root.group(1), 1)
                wb_modified = True
                sys.stderr.write(f"[restore_ext] workbook.xml: Namespaces vom Original wiederhergestellt\n")
            
            # 2. definedNames NICHT vom Original übernehmen!
            # definedNames können externe Referenzen enthalten wie [7]Sheet1!$A$1
            # wobei [7] ein Index in <externalReferences> ist.
            # Wenn openpyxl die externalReferences anders nummeriert als das Original,
            # verweisen die Indizes auf falsche Workbooks → Excel-Reparaturmodus.
            # openpyxl's eigene definedNames haben korrekte Indizes zu seinen externalReferences.
            sys.stderr.write(f"[restore_ext] workbook.xml: definedNames von openpyxl beibehalten (Index-Konsistenz)\n")
            
            if wb_modified:
                with open(workbook_path, 'w', encoding='utf-8') as f:
                    f.write(dest_wb_content)
                fixed_count += 1
        
        # Stelle workbook.xml.rels SELEKTIV wieder her
        # NICHT blind kopieren! Nur fehlende Relationships ergänzen (z.B. slicerCaches).
        # externalLinks-Relationships NICHT überschreiben (openpyxl hat korrekte rIds).
        rels_path = os.path.join(temp_dir, 'xl', '_rels', 'workbook.xml.rels')
        orig_rels_path = os.path.join(orig_temp_dir, 'xl', '_rels', 'workbook.xml.rels')
        wb_rels_rid_mapping = {}  # Mapping: original rId → neuer rId (für workbook.xml extLst)
        if os.path.exists(rels_path) and os.path.exists(orig_rels_path):
            with open(rels_path, 'r', encoding='utf-8') as f:
                dest_rels_content = f.read()
            with open(orig_rels_path, 'r', encoding='utf-8') as f:
                orig_rels_content = f.read()
            
            # Sammle alle Targets die openpyxl bereits hat
            dest_targets = set(re.findall(r'Target="([^"]+)"', dest_rels_content))
            
            # Finde fehlende Relationships aus dem Original
            missing_rels = []
            for rel_match in re.finditer(r'<Relationship\s[^>]+/>', orig_rels_content):
                rel_el = rel_match.group(0)
                target_m = re.search(r'Target="([^"]+)"', rel_el)
                if target_m and target_m.group(1) not in dest_targets:
                    target_val = target_m.group(1)
                    # externalLinks NICHT ergänzen (openpyxl hat korrekte cached values)
                    if 'externalLinks/' in target_val:
                        sys.stderr.write(f"[restore_ext] workbook.xml.rels: externalLink übersprungen: {target_val}\n")
                        continue
                    # Worksheet-Relationships NICHT ergänzen wenn sie im Output fehlen!
                    # Diese wurden absichtlich entfernt (nicht-ausgewählte Sheets beim Export).
                    if re.match(r'worksheets/sheet\d+\.xml$', target_val):
                        sys.stderr.write(f"[restore_ext] workbook.xml.rels: Entferntes Worksheet übersprungen: {target_val}\n")
                        continue
                    # Bei structural_change: Slicer-Targets NICHT ergänzen
                    # (SlicerCaches sind nach Spaltenoperationen invalide)
                    if structural_change and 'slicer' in target_val.lower():
                        sys.stderr.write(f"[restore_ext] workbook.xml.rels: Slicer übersprungen: {target_val}\n")
                        continue
                    # Bei structural_change: PivotCache-Targets NICHT ergänzen
                    # PivotCaches referenzieren Spalten/Bereiche die sich geändert haben
                    if structural_change and 'pivot' in target_val.lower():
                        sys.stderr.write(f"[restore_ext] workbook.xml.rels: PivotCache übersprungen: {target_val}\n")
                        continue
                    # Nur ergänzen wenn die Zieldatei existiert (oder extern ist)
                    is_external = 'TargetMode="External"' in rel_el
                    if is_external:
                        missing_rels.append(rel_el)
                    else:
                        target_file_check = os.path.normpath(os.path.join(temp_dir, 'xl', target_val))
                        orig_target_check = os.path.normpath(os.path.join(orig_temp_dir, 'xl', target_val))
                        if os.path.exists(target_file_check):
                            missing_rels.append(rel_el)
                        elif os.path.exists(orig_target_check):
                            # Datei aus Original kopieren (z.B. slicerCache)
                            os.makedirs(os.path.dirname(target_file_check), exist_ok=True)
                            shutil.copy2(orig_target_check, target_file_check)
                            missing_rels.append(rel_el)
                            sys.stderr.write(f"[restore_ext] workbook.xml.rels: {target_val} aus Original kopiert\n")
            
            if missing_rels:
                # rId-Konflikte vermeiden: bestehende rIds sammeln
                existing_rids = set(re.findall(r'Id="(rId\d+)"', dest_rels_content))
                max_rid = 0
                for rid in existing_rids:
                    num = int(rid.replace('rId', ''))
                    if num > max_rid:
                        max_rid = num
                
                # Fehlende Rels mit konfliktfreien rIds einfügen
                renumbered_rels = []
                for rel_el in missing_rels:
                    rid_m = re.search(r'Id="(rId\d+)"', rel_el)
                    if rid_m and rid_m.group(1) in existing_rids:
                        # Konflikt: neuen rId vergeben
                        max_rid += 1
                        new_rid = f'rId{max_rid}'
                        old_rid = rid_m.group(1)
                        rel_el = rel_el.replace(f'Id="{old_rid}"', f'Id="{new_rid}"', 1)
                        existing_rids.add(new_rid)
                        wb_rels_rid_mapping[old_rid] = new_rid
                        sys.stderr.write(f"[restore_ext] workbook.xml.rels: rId-Konflikt {old_rid} → {new_rid}\n")
                    elif rid_m:
                        existing_rids.add(rid_m.group(1))
                        # KRITISCH: max_rid aktualisieren! Sonst erzeugt max_rid+1
                        # bei späteren Konflikten eine rId die schon vergeben ist.
                        num = int(rid_m.group(1).replace('rId', ''))
                        if num > max_rid:
                            max_rid = num
                    renumbered_rels.append(rel_el)
                
                # Vor </Relationships> einfügen
                insert_str = '\n'.join(renumbered_rels)
                dest_rels_content = dest_rels_content.replace('</Relationships>', insert_str + '\n</Relationships>')
                with open(rels_path, 'w', encoding='utf-8') as f:
                    f.write(dest_rels_content)
                fixed_count += 1
                sys.stderr.write(f"[restore_ext] workbook.xml.rels: {len(renumbered_rels)} fehlende Relationships ergänzt\n")
        
        # workbook.xml <extLst> aus Original wiederherstellen (nur REPLACE-Modus).
        # Bei structural_change (Spaltenoperationen) werden Tabellenspalten geändert.
        # SlicerCaches referenzieren diese Spalten → werden invalidiert → Excel
        # kaskadiert: ungültiger Cache → Slicer → Zeichnungsform entfernt.
        # Ohne extLst-Referenzen ignoriert Excel die orphaned SlicerCache-Dateien.
        if not structural_change and os.path.exists(workbook_path) and os.path.exists(orig_workbook_path):
            with open(workbook_path, 'r', encoding='utf-8') as f:
                dest_wb_content_2 = f.read()
            with open(orig_workbook_path, 'r', encoding='utf-8') as f:
                orig_wb_content_2 = f.read()
            
            # Finde die workbook-level <extLst> im Original (letzte vor </workbook>)
            orig_wb_end = orig_wb_content_2.rfind('</workbook>')
            if orig_wb_end >= 0:
                orig_wb_extlst_start = orig_wb_content_2.rfind('<extLst', 0, orig_wb_end)
                if orig_wb_extlst_start >= 0:
                    orig_wb_extlst_end = orig_wb_content_2.find('</extLst>', orig_wb_extlst_start)
                    if orig_wb_extlst_end >= 0:
                        after_wb_ext = orig_wb_content_2[orig_wb_extlst_end + len('</extLst>'):].strip()
                        if after_wb_ext.startswith('</workbook>'):
                            orig_wb_extlst = orig_wb_content_2[orig_wb_extlst_start:orig_wb_extlst_end + len('</extLst>')]
                            
                            # rIds mappen (workbook.xml.rels Renumbering)
                            if wb_rels_rid_mapping:
                                def _wb_rid_replacer(m):
                                    orig_rid = m.group(1)
                                    mapped = wb_rels_rid_mapping.get(orig_rid, orig_rid)
                                    return f'r:id="{mapped}"'
                                orig_wb_extlst = re.sub(r'r:id="([^"]+)"', _wb_rid_replacer, orig_wb_extlst)
                            
                            # Entferne bestehende workbook-level <extLst> im Dest
                            dest_wb_end = dest_wb_content_2.rfind('</workbook>')
                            if dest_wb_end >= 0:
                                dest_wb_extlst_start = dest_wb_content_2.rfind('<extLst', 0, dest_wb_end)
                                if dest_wb_extlst_start >= 0:
                                    dest_wb_extlst_end = dest_wb_content_2.find('</extLst>', dest_wb_extlst_start)
                                    if dest_wb_extlst_end >= 0:
                                        after_dest_wb = dest_wb_content_2[dest_wb_extlst_end + len('</extLst>'):].strip()
                                        if after_dest_wb.startswith('</workbook>'):
                                            dest_wb_content_2 = dest_wb_content_2[:dest_wb_extlst_start] + dest_wb_content_2[dest_wb_extlst_end + len('</extLst>'):]
                            
                            # Einfügen vor </workbook>
                            dest_wb_content_2 = dest_wb_content_2.replace('</workbook>', orig_wb_extlst + '\n</workbook>')
                            with open(workbook_path, 'w', encoding='utf-8') as f:
                                f.write(dest_wb_content_2)
                            fixed_count += 1
                            sys.stderr.write(f"[restore_ext] workbook.xml: <extLst> aus Original wiederhergestellt (slicerCaches etc.)\n")
                            if wb_rels_rid_mapping:
                                sys.stderr.write(f"[restore_ext] workbook.xml: rId-Mapping angewendet: {wb_rels_rid_mapping}\n")
        
        # [Content_Types].xml Merge wird NACH den Datei-Kopien durchgeführt (s.u.)
        # Grund: Override-Einträge prüfen ob die Datei existiert.
        # Wenn der Merge VOR dem Kopieren von drawings/media/richData läuft,
        # fehlen die Dateien noch → Override wird übersprungen → "Entfernter Teil: Zeichnungsform"
        
        # Kopiere xl/media aus Original (Bilder/Images - openpyxl kann diese verlieren)
        orig_media_dir = os.path.join(orig_temp_dir, 'xl', 'media')
        dest_media_dir = os.path.join(temp_dir, 'xl', 'media')
        if os.path.exists(orig_media_dir):
            media_files = os.listdir(orig_media_dir)
            sys.stderr.write(f"[restore_ext] xl/media im Original gefunden: {len(media_files)} Dateien: {media_files}\n")
            if os.path.exists(dest_media_dir):
                shutil.rmtree(dest_media_dir)
            shutil.copytree(orig_media_dir, dest_media_dir)
            fixed_count += 1
        else:
            sys.stderr.write(f"[restore_ext] xl/media im Original NICHT vorhanden\n")
        
        # Kopiere xl/drawings aus Original (Drawing-XML mit Bild-Positionen/Ankern)
        orig_drawings_dir = os.path.join(orig_temp_dir, 'xl', 'drawings')
        dest_drawings_dir = os.path.join(temp_dir, 'xl', 'drawings')
        if os.path.exists(orig_drawings_dir):
            drawings_files = os.listdir(orig_drawings_dir)
            sys.stderr.write(f"[restore_ext] xl/drawings im Original gefunden: {len(drawings_files)} Dateien: {drawings_files}\n")
            if os.path.exists(dest_drawings_dir):
                shutil.rmtree(dest_drawings_dir)
            shutil.copytree(orig_drawings_dir, dest_drawings_dir)
            fixed_count += 1
            
            # Bei structural_change: Slicer-Shapes aus drawing*.xml entfernen.
            # Grund: drawing1.xml enthält mc:AlternateContent-Blöcke mit Slicer-Shapes.
            # Diese referenzieren slicer*.xml → slicerCache*.xml → Tabellenspalten.
            # Nach Spaltenoperationen sind die SlicerCaches invalide (Spaltenreferenzen
            # stimmen nicht mehr). Excel validiert die Kette über das Drawing und
            # entfernt dann den Slicer-Shape → "Entfernter Teil: Zeichnungsform".
            # Lösung: mc:AlternateContent-Blöcke mit Slicer-URIs entfernen.
            # Andere Shapes (Bilder, Charts, Textboxen) bleiben erhalten.
            if structural_change:
                _strip_slicer_shapes_from_drawings(dest_drawings_dir)
        else:
            sys.stderr.write(f"[restore_ext] xl/drawings im Original NICHT vorhanden\n")
        
        # KRITISCH: Kopiere xl/richData aus Original (Excel 365 Zellbilder)
        # Moderne Excel 365 Zellbilder verwenden richData statt drawings:
        # - xl/richData/rdrichvalue.xml (Bild-Werte)
        # - xl/richData/rdRichValueStructure.xml (Struktur)
        # - xl/richData/rdRichValueTypes.xml (Typen)
        # - xl/richData/richValueRel.xml (Beziehungen zu media/)
        # - xl/richData/_rels/richValueRel.xml.rels (tatsächliche Datei-Referenzen)
        # openpyxl kennt richData NICHT und entfernt alles komplett!
        orig_richdata_dir = os.path.join(orig_temp_dir, 'xl', 'richData')
        dest_richdata_dir = os.path.join(temp_dir, 'xl', 'richData')
        if os.path.exists(orig_richdata_dir):
            richdata_files = []
            for root_dir, dirs, files_list in os.walk(orig_richdata_dir):
                for ff in files_list:
                    rel_path = os.path.relpath(os.path.join(root_dir, ff), orig_richdata_dir)
                    richdata_files.append(rel_path)
            sys.stderr.write(f"[restore_ext] xl/richData im Original gefunden: {len(richdata_files)} Dateien: {richdata_files}\n")
            if os.path.exists(dest_richdata_dir):
                shutil.rmtree(dest_richdata_dir)
            shutil.copytree(orig_richdata_dir, dest_richdata_dir)
            fixed_count += 1
        else:
            sys.stderr.write(f"[restore_ext] xl/richData im Original NICHT vorhanden\n")
        
        # KRITISCH: Kopiere xl/metadata.xml aus Original (benötigt für richData/vm-Attribute)
        # metadata.xml definiert die Value Metadata Typen die richData-Zellbilder referenzieren
        orig_metadata = os.path.join(orig_temp_dir, 'xl', 'metadata.xml')
        dest_metadata = os.path.join(temp_dir, 'xl', 'metadata.xml')
        if os.path.exists(orig_metadata):
            shutil.copy2(orig_metadata, dest_metadata)
            sys.stderr.write(f"[restore_ext] xl/metadata.xml aus Original kopiert\n")
            fixed_count += 1
        
        # Kopiere xl/printerSettings aus Original (Druckeinstellungen)
        # Worksheet-Rels aus dem Original referenzieren printerSettings-Dateien.
        # Wenn diese fehlen, erzeugt Excel "Reparatur auf Dateiebene".
        orig_printer_dir = os.path.join(orig_temp_dir, 'xl', 'printerSettings')
        dest_printer_dir = os.path.join(temp_dir, 'xl', 'printerSettings')
        if os.path.exists(orig_printer_dir):
            printer_files = os.listdir(orig_printer_dir)
            sys.stderr.write(f"[restore_ext] xl/printerSettings im Original gefunden: {len(printer_files)} Dateien\n")
            if os.path.exists(dest_printer_dir):
                shutil.rmtree(dest_printer_dir)
            shutil.copytree(orig_printer_dir, dest_printer_dir)
            fixed_count += 1
        
        # =====================================================================
        # Worksheet-Relationships: MERGE oder REPLACE je nach structural_change
        #
        # MERGE (structural_change=True):
        #   openpyxl hat Spalten/Zeilen eingefügt/gelöscht → es hat die Rels
        #   aktualisiert (neue rIds). Wir behalten openpyxl's Rels und ergänzen
        #   nur FEHLENDE Rels aus dem Original (z.B. drawings, printerSettings).
        #   Die rId-Referenzen im Worksheet-XML (tableParts, pageSetup, etc.)
        #   bleiben von openpyxl → intern konsistent.
        #
        # REPLACE (structural_change=False):
        #   Keine strukturellen Änderungen → openpyxl hat die Rels nur
        #   umnummeriert, aber keine neuen hinzugefügt. Wir ersetzen komplett
        #   aus dem Original. Die Worksheet-XML-Elemente (drawing, tableParts,
        #   pageSetup) werden ebenfalls aus dem Original übernommen → konsistent.
        # =====================================================================
        orig_ws_rels_dir = os.path.join(orig_temp_dir, 'xl', 'worksheets', '_rels')
        dest_ws_rels_dir = os.path.join(temp_dir, 'xl', 'worksheets', '_rels')
        ws_rid_mappings = {}  # Nur für MERGE: {rels_filename: {orig_rId: dest_rId}}
        
        if os.path.exists(orig_ws_rels_dir):
            if structural_change:
                # === MERGE: openpyxl's Rels behalten, nur fehlende ergänzen ===
                if not os.path.exists(dest_ws_rels_dir):
                    os.makedirs(dest_ws_rels_dir)
                
                rels_files = [f for f in os.listdir(orig_ws_rels_dir) if f.endswith('.rels')]
                sys.stderr.write(f"[restore_ext] xl/worksheets/_rels MERGE: {len(rels_files)} Dateien: {rels_files}\n")
                
                for rels_fn in rels_files:
                    orig_rels_fp = os.path.join(orig_ws_rels_dir, rels_fn)
                    dest_rels_fp = os.path.join(dest_ws_rels_dir, rels_fn)
                    
                    with open(orig_rels_fp, 'r', encoding='utf-8') as f:
                        orig_rels_xml = f.read()
                    
                    # Parse Original-Rels
                    orig_rels = {}
                    for m in re.finditer(r'(<Relationship\s[^>]*?Id="([^"]+)"[^>]*?Target="([^"]+)"[^>]*?/>)', orig_rels_xml):
                        orig_rels[m.group(2)] = {'target': m.group(3), 'el': m.group(1)}
                    
                    # Parse Dest-Rels (openpyxl's)
                    dest_rels = {}
                    if os.path.exists(dest_rels_fp):
                        with open(dest_rels_fp, 'r', encoding='utf-8') as f:
                            dest_rels_xml = f.read()
                        for m in re.finditer(r'<Relationship\s[^>]*?Id="([^"]+)"[^>]*?Target="([^"]+)"[^>]*?/>', dest_rels_xml):
                            dest_rels[m.group(1)] = m.group(2)
                    else:
                        dest_rels_xml = '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"></Relationships>'
                    
                    # Target→dest_rId Lookup (normalisiert)
                    target_to_dest_rid = {}
                    for rid, target in dest_rels.items():
                        norm = target.replace('\\', '/').lower()
                        target_to_dest_rid[norm] = rid
                    
                    # Mapping bauen und fehlende Rels finden
                    mapping = {}
                    missing_rels = []
                    existing_rids = set(dest_rels.keys())
                    max_rid_num = 0
                    for rid in existing_rids:
                        num_m = re.search(r'\d+', rid)
                        if num_m:
                            max_rid_num = max(max_rid_num, int(num_m.group(0)))
                    
                    for orig_rid, orig_info in orig_rels.items():
                        norm_target = orig_info['target'].replace('\\', '/').lower()
                        
                        # Bei structural_change: Slicer-Rels NICHT ergänzen!
                        # SlicerCaches/Slicers referenzieren Tabellenspalten die sich
                        # geändert haben → invalide Referenzen → Excel-Reparatur.
                        # Wenn wir die Rels ergänzen, werden auch die Slicer-Dateien
                        # aus dem Original kopiert → orphaned Slicer-Einträge.
                        if structural_change:
                            orig_el = orig_info['el']
                            type_m_sc = re.search(r'Type="([^"]+)"', orig_el)
                            is_slicer_rel = False
                            is_pivot_rel = False
                            if type_m_sc:
                                rel_type_val = type_m_sc.group(1).lower()
                                if 'slicer' in rel_type_val:
                                    is_slicer_rel = True
                                if 'pivot' in rel_type_val:
                                    is_pivot_rel = True
                            if 'slicer' in norm_target:
                                is_slicer_rel = True
                            if 'pivot' in norm_target:
                                is_pivot_rel = True
                            if is_slicer_rel:
                                sys.stderr.write(f"[restore_ext] MERGE {rels_fn}: Slicer-Rel übersprungen (structural_change): {orig_info['target']}\n")
                                continue
                            if is_pivot_rel:
                                sys.stderr.write(f"[restore_ext] MERGE {rels_fn}: PivotTable-Rel übersprungen (structural_change): {orig_info['target']}\n")
                                continue
                        
                        if norm_target in target_to_dest_rid:
                            # Rel existiert in dest → mapping
                            mapping[orig_rid] = target_to_dest_rid[norm_target]
                        else:
                            # Rel fehlt in dest → mit neuem rId ergänzen
                            max_rid_num += 1
                            new_rid = f'rId{max_rid_num}'
                            mapping[orig_rid] = new_rid
                            
                            # Neues Relationship-Element mit neuem rId
                            new_el = re.sub(r'Id="[^"]+"', f'Id="{new_rid}"', orig_info['el'])
                            missing_rels.append(new_el)
                            existing_rids.add(new_rid)
                            
                            # Zieldatei aus Original kopieren falls nötig
                            target_path = orig_info['target']
                            dest_file = os.path.normpath(os.path.join(temp_dir, 'xl', 'worksheets', target_path))
                            orig_file = os.path.normpath(os.path.join(orig_temp_dir, 'xl', 'worksheets', target_path))
                            if not os.path.exists(dest_file) and os.path.exists(orig_file):
                                os.makedirs(os.path.dirname(dest_file), exist_ok=True)
                                shutil.copy2(orig_file, dest_file)
                                sys.stderr.write(f"[restore_ext] MERGE {rels_fn}: Zieldatei kopiert: {target_path}\n")
                    
                    if missing_rels:
                        insert_str = '\n'.join(missing_rels)
                        dest_rels_xml = dest_rels_xml.replace('</Relationships>', insert_str + '\n</Relationships>')
                        with open(dest_rels_fp, 'w', encoding='utf-8') as f:
                            f.write(dest_rels_xml)
                        fixed_count += 1
                        sys.stderr.write(f"[restore_ext] MERGE {rels_fn}: {len(missing_rels)} fehlende Rels ergänzt\n")
                    elif not os.path.exists(dest_rels_fp):
                        # Original hat Rels, dest nicht → komplett schreiben
                        with open(dest_rels_fp, 'w', encoding='utf-8') as f:
                            f.write(orig_rels_xml)
                        fixed_count += 1
                    
                    # Bei structural_change: Slicer-Rels aus Worksheet-Rels entfernen.
                    # Grund: slicer-Rels verbinden die Kette sheet→slicer→slicerCache→Tabellenspalte.
                    # Nach Spaltenoperationen sind SlicerCaches invalide → Excel kaskadiert.
                    if structural_change and os.path.exists(dest_rels_fp):
                        with open(dest_rels_fp, 'r', encoding='utf-8') as f:
                            ws_rels_content = f.read()
                        ws_rels_orig = ws_rels_content
                        for rel_m in list(re.finditer(r'<Relationship\s[^>]*/>', ws_rels_content)):
                            rel_el = rel_m.group(0)
                            type_m = re.search(r'Type="([^"]+)"', rel_el)
                            if type_m:
                                rel_type = type_m.group(1).lower()
                                if 'slicer' in rel_type:
                                    ws_rels_content = ws_rels_content.replace(rel_el, '')
                                    target_m = re.search(r'Target="([^"]+)"', rel_el)
                                    sys.stderr.write(f"[strip_slicer] {rels_fn}: Slicer-Rel entfernt: {target_m.group(1) if target_m else '?'}\n")
                        if ws_rels_content != ws_rels_orig:
                            ws_rels_content = re.sub(r'\n\s*\n', '\n', ws_rels_content)
                            with open(dest_rels_fp, 'w', encoding='utf-8') as f:
                                f.write(ws_rels_content)
                    
                    ws_rid_mappings[rels_fn] = mapping
                    sys.stderr.write(f"[restore_ext] MERGE {rels_fn}: Mapping={mapping}\n")
            else:
                # === REPLACE: Rels komplett aus Original übernehmen ===
                rels_files = os.listdir(orig_ws_rels_dir)
                sys.stderr.write(f"[restore_ext] xl/worksheets/_rels REPLACE: {len(rels_files)} Dateien: {rels_files}\n")
                if os.path.exists(dest_ws_rels_dir):
                    shutil.rmtree(dest_ws_rels_dir)
                shutil.copytree(orig_ws_rels_dir, dest_ws_rels_dir)
                fixed_count += 1
        else:
            sys.stderr.write(f"[restore_ext] xl/worksheets/_rels im Original NICHT vorhanden\n")
        
        # =====================================================================
        # [Content_Types].xml SELEKTIV wiederherstellen
        # MUSS NACH allen Datei-Kopien laufen! Override-Einträge prüfen ob die
        # Datei existiert. Wenn dieser Block VOR dem Kopieren von drawings/media/
        # richData läuft, fehlen die Dateien → Override übersprungen →
        # "Entfernter Teil: Zeichnungsform"
        # =====================================================================
        content_types_path = os.path.join(temp_dir, '[Content_Types].xml')
        orig_content_types_path = os.path.join(orig_temp_dir, '[Content_Types].xml')
        if os.path.exists(content_types_path) and os.path.exists(orig_content_types_path):
            with open(content_types_path, 'r', encoding='utf-8') as f:
                dest_ct_content = f.read()
            with open(orig_content_types_path, 'r', encoding='utf-8') as f:
                orig_ct_content = f.read()
            
            ct_modified = False
            
            # A. Fehlende <Default Extension="..."> Einträge ergänzen
            # KRITISCH für VML-Zeichnungen (.vml), Metafiles (.emf, .wmf) etc.
            dest_extensions = set(re.findall(r'<Default\s+Extension="([^"]+)"', dest_ct_content))
            missing_defaults = []
            for df_match in re.finditer(r'<Default\s[^>]*/>', orig_ct_content):
                df_el = df_match.group(0)
                ext_m = re.search(r'Extension="([^"]+)"', df_el)
                if ext_m and ext_m.group(1) not in dest_extensions:
                    missing_defaults.append(df_el)
                    sys.stderr.write(f"[restore_ext] ContentTypes Default: Extension=\"{ext_m.group(1)}\" ergänzt\n")
            
            if missing_defaults:
                first_override = re.search(r'<Override\s', dest_ct_content)
                if first_override:
                    insert_str = '\n'.join(missing_defaults) + '\n'
                    dest_ct_content = dest_ct_content[:first_override.start()] + insert_str + dest_ct_content[first_override.start():]
                else:
                    insert_str = '\n'.join(missing_defaults)
                    dest_ct_content = dest_ct_content.replace('</Types>', insert_str + '\n</Types>')
                ct_modified = True
            
            # B. Fehlende <Override PartName="..."> Einträge ergänzen
            dest_parts = set(re.findall(r'PartName="([^"]+)"', dest_ct_content))
            missing_overrides = []
            for ov_match in re.finditer(r'<Override\s[^>]*/>', orig_ct_content):
                ov_el = ov_match.group(0)
                pn_m = re.search(r'PartName="([^"]+)"', ov_el)
                if pn_m and pn_m.group(1) not in dest_parts:
                    part_name = pn_m.group(1)
                    # Bei structural_change: Slicer/Pivot-Overrides NICHT ergänzen!
                    # SlicerCaches/Slicers und PivotTables/PivotCaches sind nach
                    # Spaltenoperationen invalide → Excel-Reparatur.
                    if structural_change and 'slicer' in part_name.lower():
                        sys.stderr.write(f"[restore_ext] ContentTypes Override ÜBERSPRUNGEN (structural_change): {part_name}\n")
                        continue
                    if structural_change and 'pivot' in part_name.lower():
                        sys.stderr.write(f"[restore_ext] ContentTypes Override ÜBERSPRUNGEN (structural_change): {part_name}\n")
                        continue
                    # Nur ergänzen wenn die Datei tatsächlich existiert
                    rel_path_ct = part_name.lstrip('/')
                    dest_file_check = os.path.join(temp_dir, rel_path_ct.replace('/', os.sep))
                    if os.path.exists(dest_file_check):
                        missing_overrides.append(ov_el)
                        sys.stderr.write(f"[restore_ext] ContentTypes Override: {part_name} ergänzt\n")
            
            if missing_overrides:
                insert_str = '\n'.join(missing_overrides)
                dest_ct_content = dest_ct_content.replace('</Types>', insert_str + '\n</Types>')
                ct_modified = True
            
            if ct_modified:
                with open(content_types_path, 'w', encoding='utf-8') as f:
                    f.write(dest_ct_content)
                fixed_count += 1
                sys.stderr.write(f"[restore_ext] [Content_Types].xml: {len(missing_defaults)} Default + {len(missing_overrides)} Override Einträge ergänzt\n")
        
        # KRITISCH: <drawing> und <legacyDrawing> Elemente in Worksheet-XMLs wiederherstellen
        # openpyxl ENTFERNT diese Elemente beim Speichern wenn Pillow nicht installiert ist.
        # Ohne <drawing r:id="..."/> im Worksheet-XML zeigt Excel keine Bilder an,
        # selbst wenn xl/media/, xl/drawings/ und _rels wiederhergestellt wurden.
        orig_ws_dir = os.path.join(orig_temp_dir, 'xl', 'worksheets')
        dest_ws_dir = os.path.join(temp_dir, 'xl', 'worksheets')
        if os.path.exists(orig_ws_dir) and os.path.exists(dest_ws_dir):
            for ws_file in os.listdir(orig_ws_dir):
                if not ws_file.endswith('.xml'):
                    continue
                orig_ws_file = os.path.join(orig_ws_dir, ws_file)
                dest_ws_file = os.path.join(dest_ws_dir, ws_file)
                if not os.path.exists(dest_ws_file):
                    continue
                
                with open(orig_ws_file, 'r', encoding='utf-8') as f:
                    orig_ws_content = f.read()
                with open(dest_ws_file, 'r', encoding='utf-8') as f:
                    dest_ws_content = f.read()
                
                ws_modified = False
                
                if structural_change:
                    # === MERGE-Modus ===
                    # tableParts und pageSetup NICHT ersetzen — openpyxl hat korrekte rIds.
                    # ABER: <drawing>, <legacyDrawing>, <picture> IMMER aus Original
                    # wiederherstellen (mit gemappten rIds), weil:
                    #   1. xl/drawings/ wird IMMER aus dem Original kopiert
                    #   2. openpyxl kann ohne Pillow die Drawing-Rels ändern/entfernen
                    #   3. Bei Spaltenops verschiebt openpyxl Anker → rId-Mismatch
                    # Nur der gemappte Original-rId zeigt zum richtigen Drawing.
                    rels_fn = f"{ws_file}.rels"
                    mapping = ws_rid_mappings.get(rels_fn, {})
                    
                    def _map_rid(element_str, _mapping=mapping):
                        """Ersetzt rId-Referenzen im Element mit gemappten rIds."""
                        def _rid_replacer(m):
                            orig_rid = m.group(1)
                            mapped = _mapping.get(orig_rid, orig_rid)
                            return f'r:id="{mapped}"'
                        return re.sub(r'r:id="([^"]+)"', _rid_replacer, element_str)
                    
                    # <drawing> IMMER aus Original wiederherstellen (mit gemapptem rId)
                    # Grund: xl/drawings/ kommt aus dem Original, also muss der rId
                    # zum Original-Drawing passen (nicht zu openpyxl's möglicherweise
                    # geändertem/entferntem Drawing-Rel).
                    # ABER: Wenn die drawing*.xml nach Slicer-Stripping leer ist
                    # (keine Anchors/Shapes mehr), kein <drawing> einfügen.
                    drawing_match = re.search(r'<drawing\s+[^>]*/\s*>', orig_ws_content)
                    if not drawing_match:
                        drawing_match = re.search(r'<drawing\s+[^>]*>.*?</drawing>', orig_ws_content, re.DOTALL)
                    if drawing_match:
                        # Prüfe ob die Ziel-drawing*.xml nach Slicer-Stripping noch Inhalt hat
                        drawing_el = _map_rid(drawing_match.group(0))
                        skip_drawing = False
                        dr_rid_m = re.search(r'r:id="([^"]+)"', drawing_el)
                        if dr_rid_m and structural_change:
                            mapped_rid = dr_rid_m.group(1)
                            # Finde den Target im dest-Rels
                            dest_rels_fp_check = os.path.join(dest_ws_rels_dir, f"{ws_file}.rels")
                            if os.path.exists(dest_rels_fp_check):
                                with open(dest_rels_fp_check, 'r', encoding='utf-8') as f:
                                    rels_check = f.read()
                                target_m = re.search(r'Id="' + re.escape(mapped_rid) + r'"[^>]*Target="([^"]+)"', rels_check)
                                if target_m:
                                    drawing_target = target_m.group(1)
                                    drawing_file_check = os.path.normpath(os.path.join(temp_dir, 'xl', 'worksheets', drawing_target))
                                    if os.path.exists(drawing_file_check):
                                        with open(drawing_file_check, 'r', encoding='utf-8') as f:
                                            drawing_content_check = f.read()
                                        has_anchors = bool(re.search(r'<(?:xdr:)?(?:twoCellAnchor|oneCellAnchor|absoluteAnchor)[\s>]', drawing_content_check))
                                        has_slicer_uri = any(uri in drawing_content_check for uri in [
                                            'schemas.microsoft.com/office/drawing/2010/slicer',
                                            'schemas.microsoft.com/office/drawing/2014/slicer'])
                                        if not has_anchors and not has_slicer_uri:
                                            skip_drawing = True
                                            sys.stderr.write(f"[restore] {ws_file}: <drawing> übersprungen (drawing*.xml leer nach Slicer-Stripping)\n")
                                        elif has_slicer_uri and not has_anchors:
                                            skip_drawing = True
                                            sys.stderr.write(f"[restore] {ws_file}: <drawing> übersprungen (nur Slicer-Reste in drawing*.xml)\n")
                        
                        if not skip_drawing:
                            if re.search(r'<drawing[\s>]', dest_ws_content):
                                # Ersetze openpyxl's <drawing> mit Original (gemappter rId)
                                dest_ws_content = re.sub(r'<drawing\s+[^>]*/\s*>', drawing_el, dest_ws_content)
                                dest_ws_content = re.sub(r'<drawing\s+[^>]*>.*?</drawing>', drawing_el, dest_ws_content, flags=re.DOTALL)
                            else:
                                dest_ws_content = _insert_ws_element(dest_ws_content, drawing_el, 'drawing')
                            ws_modified = True
                            sys.stderr.write(f"[restore] {ws_file}: <drawing> aus Original wiederhergestellt: {drawing_el}\n")
                        else:
                            # Drawing leer → auch openpyxl's <drawing> aus Sheet entfernen
                            if re.search(r'<drawing[\s>]', dest_ws_content):
                                dest_ws_content = re.sub(r'<drawing\s+[^>]*/\s*>', '', dest_ws_content)
                                dest_ws_content = re.sub(r'<drawing\s+[^>]*>.*?</drawing>', '', dest_ws_content, flags=re.DOTALL)
                                ws_modified = True
                                sys.stderr.write(f"[restore] {ws_file}: <drawing> aus Sheet-XML entfernt (leere drawing*.xml)\n")
                            # Drawing-Rel aus Worksheet-Rels entfernen
                            if dr_rid_m and os.path.exists(dest_rels_fp_check):
                                with open(dest_rels_fp_check, 'r', encoding='utf-8') as f:
                                    ws_rels_for_drawing = f.read()
                                ws_rels_before = ws_rels_for_drawing
                                drawing_rel_pat = r'<Relationship\s[^>]*Id="' + re.escape(mapped_rid) + r'"[^>]*/>'
                                ws_rels_for_drawing = re.sub(drawing_rel_pat, '', ws_rels_for_drawing)
                                if ws_rels_for_drawing != ws_rels_before:
                                    ws_rels_for_drawing = re.sub(r'\n\s*\n', '\n', ws_rels_for_drawing)
                                    with open(dest_rels_fp_check, 'w', encoding='utf-8') as f:
                                        f.write(ws_rels_for_drawing)
                                    sys.stderr.write(f"[restore] {ws_file}: Drawing-Rel {mapped_rid} aus Worksheet-Rels entfernt\n")
                            # Drawing-Datei NICHT löschen: Content_Types-Konsistenzprüfung
                            # würde sie aus dem Original zurückkopieren (mit Slicer-Shapes!).
                            # Die leere drawing*.xml bleibt, ist aber harmlos ohne Referenz.
                    
                    # <legacyDrawing> IMMER aus Original wiederherstellen
                    legacy_match = re.search(r'<legacyDrawing\s+[^>]*/\s*>', orig_ws_content)
                    if not legacy_match:
                        legacy_match = re.search(r'<legacyDrawing\s+[^>]*>.*?</legacyDrawing>', orig_ws_content, re.DOTALL)
                    if legacy_match:
                        legacy_el = _map_rid(legacy_match.group(0))
                        if re.search(r'<legacyDrawing[\s>]', dest_ws_content):
                            dest_ws_content = re.sub(r'<legacyDrawing\s+[^>]*/\s*>', legacy_el, dest_ws_content)
                            dest_ws_content = re.sub(r'<legacyDrawing\s+[^>]*>.*?</legacyDrawing>', legacy_el, dest_ws_content, flags=re.DOTALL)
                        else:
                            dest_ws_content = _insert_ws_element(dest_ws_content, legacy_el, 'legacyDrawing')
                        ws_modified = True
                        sys.stderr.write(f"[restore] {ws_file}: <legacyDrawing> aus Original wiederhergestellt\n")
                    
                    # <picture> IMMER aus Original wiederherstellen (Hintergrundbilder)
                    picture_match = re.search(r'<picture\s+[^>]*/\s*>', orig_ws_content)
                    if picture_match:
                        picture_el = _map_rid(picture_match.group(0))
                        if re.search(r'<picture[\s>]', dest_ws_content):
                            dest_ws_content = re.sub(r'<picture\s+[^>]*/\s*>', picture_el, dest_ws_content)
                        else:
                            dest_ws_content = _insert_ws_element(dest_ws_content, picture_el, 'picture')
                        ws_modified = True
                    
                    # <tableParts> aus Original wiederherstellen (mit gemappten rIds)
                    # KRITISCH: openpyxl's interne rIds stimmen nach MERGE nicht mehr,
                    # wenn openpyxl keine Rels-Datei für ein Sheet geschrieben hat
                    # (alle Rels wurden aus dem Original mit neuen rIds ergänzt).
                    # → tablePart r:id="rId2" zeigt dann auf Drawing statt Table
                    # → Excel repariert → "Entfernter Teil: Zeichnungsform"
                    orig_tp_match = re.search(r'<tableParts\b[^>]*>.*?</tableParts>', orig_ws_content, re.DOTALL)
                    if not orig_tp_match:
                        orig_tp_match = re.search(r'<tableParts\b[^/]*/>', orig_ws_content)
                    if orig_tp_match:
                        orig_tp = _map_rid(orig_tp_match.group(0))
                        dest_ws_content = re.sub(r'<tableParts\b[^>]*>.*?</tableParts>', '', dest_ws_content, flags=re.DOTALL)
                        dest_ws_content = re.sub(r'<tableParts\b[^/]*/>', '', dest_ws_content)
                        dest_ws_content = _insert_ws_element(dest_ws_content, orig_tp, 'tableParts')
                        ws_modified = True
                        sys.stderr.write(f"[restore] {ws_file}: <tableParts> aus Original wiederhergestellt (mapped rIds): {orig_tp}\n")
                    
                    # <pageSetup> r:id aus Original wiederherstellen (mit gemapptem rId)
                    # openpyxl verliert oft die r:id-Referenz zu printerSettings
                    orig_ps_match = re.search(r'<pageSetup\s[^>]*/\s*>', orig_ws_content)
                    if orig_ps_match:
                        orig_ps = _map_rid(orig_ps_match.group(0))
                        dest_ps_match = re.search(r'<pageSetup\s[^>]*/\s*>', dest_ws_content)
                        if dest_ps_match:
                            dest_ws_content = dest_ws_content.replace(dest_ps_match.group(0), orig_ps)
                            ws_modified = True
                            sys.stderr.write(f"[restore] {ws_file}: <pageSetup> r:id aus Original wiederhergestellt\n")
                    
                    # KEIN <extLst> im MERGE-Modus wiederherstellen!
                    # Grund: Bei structural_change (Spaltenoperationen) ändern sich die
                    # Tabellenspalten. SlicerCaches referenzieren diese Spalten → werden
                    # invalidiert → Excel kaskadiert: ungültiger Cache → Slicer entfernt
                    # → Zeichnungsform entfernt. Ohne extLst-Referenzen ignoriert Excel
                    # die orphaned SlicerCache/Slicer-Dateien stillschweigend.
                    # (extLst wird nur im REPLACE-Modus wiederhergestellt, s.u.)
                
                else:
                    # === REPLACE-Modus: Alle Elemente aus Original übernehmen ===
                    # Da die Rels komplett aus dem Original kommen, müssen auch
                    # die rId-Referenzen im Sheet-XML aus dem Original stammen.
                    
                    # <drawing r:id="rIdX"/> Element vom Original wiederherstellen
                    drawing_match = re.search(r'<drawing\s+[^>]*/\s*>', orig_ws_content)
                    if not drawing_match:
                        drawing_match = re.search(r'<drawing\s+[^>]*>.*?</drawing>', orig_ws_content, re.DOTALL)
                    if drawing_match:
                        drawing_el = drawing_match.group(0)
                        if re.search(r'<drawing[\s>]', dest_ws_content):
                            dest_ws_content = re.sub(r'<drawing\s+[^>]*/\s*>', drawing_el, dest_ws_content)
                            dest_ws_content = re.sub(r'<drawing\s+[^>]*>.*?</drawing>', drawing_el, dest_ws_content, flags=re.DOTALL)
                        else:
                            dest_ws_content = _insert_ws_element(dest_ws_content, drawing_el, 'drawing')
                        ws_modified = True
                        sys.stderr.write(f"[restore] {ws_file}: <drawing> Element wiederhergestellt: {drawing_el}\n")
                    
                    # <legacyDrawing r:id="..."/> Element vom Original wiederherstellen
                    legacy_match = re.search(r'<legacyDrawing\s+[^>]*/\s*>', orig_ws_content)
                    if not legacy_match:
                        legacy_match = re.search(r'<legacyDrawing\s+[^>]*>.*?</legacyDrawing>', orig_ws_content, re.DOTALL)
                    if legacy_match:
                        legacy_el = legacy_match.group(0)
                        if re.search(r'<legacyDrawing[\s>]', dest_ws_content):
                            dest_ws_content = re.sub(r'<legacyDrawing\s+[^>]*/\s*>', legacy_el, dest_ws_content)
                            dest_ws_content = re.sub(r'<legacyDrawing\s+[^>]*>.*?</legacyDrawing>', legacy_el, dest_ws_content, flags=re.DOTALL)
                        else:
                            dest_ws_content = _insert_ws_element(dest_ws_content, legacy_el, 'legacyDrawing')
                        ws_modified = True
                        sys.stderr.write(f"[restore] {ws_file}: <legacyDrawing> Element wiederhergestellt\n")
                    
                    # <picture r:id="..."/> Element vom Original wiederherstellen
                    picture_match = re.search(r'<picture\s+[^>]*/\s*>', orig_ws_content)
                    if picture_match:
                        picture_el = picture_match.group(0)
                        if re.search(r'<picture[\s>]', dest_ws_content):
                            dest_ws_content = re.sub(r'<picture\s+[^>]*/\s*>', picture_el, dest_ws_content)
                        else:
                            dest_ws_content = _insert_ws_element(dest_ws_content, picture_el, 'picture')
                        ws_modified = True
                    
                    # KRITISCH: <tableParts> vom Original wiederherstellen (rId-Konsistenz)
                    orig_tp_match = re.search(r'<tableParts\b[^>]*>.*?</tableParts>', orig_ws_content, re.DOTALL)
                    if not orig_tp_match:
                        orig_tp_match = re.search(r'<tableParts\b[^/]*/>', orig_ws_content)
                    if orig_tp_match:
                        orig_tp = orig_tp_match.group(0)
                        dest_ws_content = re.sub(r'<tableParts\b[^>]*>.*?</tableParts>', '', dest_ws_content, flags=re.DOTALL)
                        dest_ws_content = re.sub(r'<tableParts\b[^/]*/>', '', dest_ws_content)
                        dest_ws_content = _insert_ws_element(dest_ws_content, orig_tp, 'tableParts')
                        ws_modified = True
                        sys.stderr.write(f"[restore] {ws_file}: <tableParts> vom Original wiederhergestellt (rId-Konsistenz)\n")
                    
                    # <pageSetup> r:id vom Original wiederherstellen
                    orig_ps_match = re.search(r'<pageSetup\s[^>]*/\s*>', orig_ws_content)
                    if orig_ps_match:
                        orig_ps = orig_ps_match.group(0)
                        dest_ps_match = re.search(r'<pageSetup\s[^>]*/\s*>', dest_ws_content)
                        if dest_ps_match:
                            dest_ws_content = dest_ws_content.replace(dest_ps_match.group(0), orig_ps)
                            ws_modified = True
                            sys.stderr.write(f"[restore] {ws_file}: <pageSetup> vom Original wiederhergestellt\n")
                    
                    # <extLst> aus Original wiederherstellen
                    # openpyxl entfernt "Unknown extensions" (Slicers etc.)
                    # ABER: openpyxl BEHÄLT CF-Extensions (x14:conditionalFormattings)
                    # mit NEUEN GUIDs die zu den <cfRule>-GUIDs im Hauptbereich passen.
                    # Wenn wir die gesamte <extLst> aus dem Original kopieren, haben die
                    # CF-Extensions die ALTEN GUIDs → GUID-Mismatch → Excel-Reparatur
                    # → Slicer werden als Kollateralschaden entfernt.
                    # FIX: Selektives Mergen — openpyxl's CF-Extensions behalten,
                    # nur fehlende Extensions (Slicers etc.) aus dem Original ergänzen.
                    orig_ws_end_pos = orig_ws_content.rfind('</worksheet>')
                    if orig_ws_end_pos >= 0:
                        orig_extlst_start = orig_ws_content.rfind('<extLst', 0, orig_ws_end_pos)
                        if orig_extlst_start >= 0:
                            orig_extlst_end = orig_ws_content.find('</extLst>', orig_extlst_start)
                            if orig_extlst_end >= 0:
                                after_orig = orig_ws_content[orig_extlst_end + len('</extLst>'):].strip()
                                if after_orig.startswith('</worksheet>'):
                                    orig_extlst = orig_ws_content[orig_extlst_start:orig_extlst_end + len('</extLst>')]
                                    
                                    # Parse Extension-URIs aus Original und Dest
                                    orig_ext_by_uri = {}
                                    for ext_m in re.finditer(r'(<ext\s+uri="([^"]+)"[^>]*>.*?</ext>)', orig_extlst, re.DOTALL):
                                        orig_ext_by_uri[ext_m.group(2)] = ext_m.group(1)
                                    
                                    dest_ws_end_pos2 = dest_ws_content.rfind('</worksheet>')
                                    if dest_ws_end_pos2 >= 0:
                                        dest_extlst_start = dest_ws_content.rfind('<extLst', 0, dest_ws_end_pos2)
                                        if dest_extlst_start >= 0:
                                            dest_extlst_end = dest_ws_content.find('</extLst>', dest_extlst_start)
                                            if dest_extlst_end >= 0:
                                                after_dest = dest_ws_content[dest_extlst_end + len('</extLst>'):].strip()
                                                if after_dest.startswith('</worksheet>'):
                                                    dest_extlst = dest_ws_content[dest_extlst_start:dest_extlst_end + len('</extLst>')]
                                                    # Finde URIs die openpyxl geschrieben hat
                                                    dest_uris = set(re.findall(r'<ext\s+uri="([^"]+)"', dest_extlst))
                                                    # Ergänze fehlende Extensions aus dem Original
                                                    missing_exts = []
                                                    for uri, ext_block in orig_ext_by_uri.items():
                                                        if uri not in dest_uris:
                                                            missing_exts.append(ext_block)
                                                            sys.stderr.write(f"[restore] {ws_file}: <ext uri=\"{uri}\"> aus Original ergänzt (fehlte in openpyxl)\n")
                                                        else:
                                                            sys.stderr.write(f"[restore] {ws_file}: <ext uri=\"{uri}\"> von openpyxl beibehalten\n")
                                                    if missing_exts:
                                                        # Füge fehlende Extensions vor </extLst> ein
                                                        insert_str = '\n'.join(missing_exts)
                                                        dest_ws_content = dest_ws_content[:dest_extlst_end] + insert_str + dest_ws_content[dest_extlst_end:]
                                                        ws_modified = True
                                                        sys.stderr.write(f"[restore] {ws_file}: {len(missing_exts)} fehlende <ext>-Blöcke in <extLst> ergänzt\n")
                                        else:
                                            # openpyxl hat KEINE <extLst> geschrieben → komplett aus Original
                                            dest_ws_content = _insert_ws_element(dest_ws_content, orig_extlst, 'extLst')
                                            ws_modified = True
                                            sys.stderr.write(f"[restore] {ws_file}: <extLst> komplett aus Original eingefügt (openpyxl hatte keine)\n")
                
                # KRITISCH: Namespace-Deklarationen vom Original-Worksheet wiederherstellen
                # openpyxl schreibt nur minimal: xmlns="..." xmlns:r="..."
                # Excel benötigt aber: xmlns:mc, mc:Ignorable, xmlns:x14ac, xmlns:xr etc.
                # Ohne diese Namespaces erkennt Excel vm-Attribute und andere Erweiterungen nicht.
                # re.search statt re.match weil XML-Header vor <worksheet> stehen kann
                orig_root_match = re.search(r'(<worksheet\b[^>]+>)', orig_ws_content)
                dest_root_match = re.search(r'(<worksheet\b[^>]+>)', dest_ws_content)
                if orig_root_match and dest_root_match:
                    orig_root = orig_root_match.group(1)
                    dest_root = dest_root_match.group(1)
                    if orig_root != dest_root:
                        dest_ws_content = dest_ws_content.replace(dest_root, orig_root, 1)
                        ws_modified = True
                        sys.stderr.write(f"[restore] {ws_file}: Worksheet-Namespaces vom Original wiederhergestellt\n")
                
                # KRITISCH: vm-Attribute auf Zellen wiederherstellen (Excel 365 Zellbilder)
                # openpyxl kennt vm-Attribute NICHT und entfernt sie komplett beim Speichern.
                # vm="N" auf <c>-Elementen verweist auf xl/metadata.xml → xl/richData/ → Bilder
                # Ohne vm-Attribute werden Zellbilder trotz vorhandener richData nicht angezeigt.
                if 'vm=' in orig_ws_content:
                    # Sammle alle Zellen mit vm-Attribut aus dem Original
                    vm_cells = {}
                    # Zwei mögliche Attribut-Reihenfolgen: r="..." vm="..." oder vm="..." r="..."
                    for vm_match in re.finditer(r'<c\s[^>]*?r="([A-Z]+\d+)"[^>]*?\bvm="(\d+)"', orig_ws_content):
                        vm_cells[vm_match.group(1)] = vm_match.group(2)
                    for vm_match in re.finditer(r'<c\s[^>]*?\bvm="(\d+)"[^>]*?r="([A-Z]+\d+)"', orig_ws_content):
                        if vm_match.group(2) not in vm_cells:
                            vm_cells[vm_match.group(2)] = vm_match.group(1)
                    
                    if vm_cells:
                        vm_restored = 0
                        vm_created = 0
                        for cell_ref, vm_val in vm_cells.items():
                            # Finde die Zelle in der Zieldatei und füge vm-Attribut hinzu
                            cell_pattern = re.compile(r'(<c\s[^>]*?r="' + re.escape(cell_ref) + '"[^>]*?)(/?>)')
                            match = cell_pattern.search(dest_ws_content)
                            if match and 'vm=' not in match.group(0):
                                new_attr = match.group(1) + f' vm="{vm_val}"' + match.group(2)
                                dest_ws_content = dest_ws_content[:match.start()] + new_attr + dest_ws_content[match.end():]
                                vm_restored += 1
                                ws_modified = True
                            elif not match:
                                # Zelle existiert nicht in Zieldatei (openpyxl hat sie entfernt)
                                # Erstelle das Zell-Element in der passenden Zeile
                                row_num_match = re.search(r'(\d+)$', cell_ref)
                                if row_num_match:
                                    row_num = row_num_match.group(1)
                                    row_pattern = re.compile(r'(<row\s[^>]*?\br="' + re.escape(row_num) + '"[^>]*?>)')
                                    row_match = row_pattern.search(dest_ws_content)
                                    if row_match:
                                        cell_el = f'<c r="{cell_ref}" vm="{vm_val}"/>'
                                        insert_pos = row_match.end()
                                        dest_ws_content = dest_ws_content[:insert_pos] + cell_el + dest_ws_content[insert_pos:]
                                        vm_created += 1
                                        ws_modified = True
                                    else:
                                        # Zeile existiert auch nicht → Zeile UND Zelle in sheetData erstellen
                                        # Dies passiert wenn openpyxl oder fix_xlsx_relationships leere Zeilen entfernt hat
                                        sheet_data_end = re.search(r'</sheetData>', dest_ws_content)
                                        if sheet_data_end:
                                            row_el = f'<row r="{row_num}"><c r="{cell_ref}" vm="{vm_val}"/></row>'
                                            insert_pos = sheet_data_end.start()
                                            dest_ws_content = dest_ws_content[:insert_pos] + row_el + '\n' + dest_ws_content[insert_pos:]
                                            vm_created += 1
                                            ws_modified = True
                                            sys.stderr.write(f"[restore] {ws_file}: Zeile {row_num} + Zelle {cell_ref} mit vm={vm_val} neu erstellt\n")
                        if vm_restored > 0 or vm_created > 0:
                            sys.stderr.write(f"[restore] {ws_file}: {vm_restored} vm-Attribute wiederhergestellt, {vm_created} vm-Zellen neu erstellt (von {len(vm_cells)} im Original)\n")
                
                if ws_modified:
                    with open(dest_ws_file, 'w', encoding='utf-8') as f:
                        f.write(dest_ws_content)
                    fixed_count += 1
        
        # =====================================================================
        # CONTENT_TYPES KONSISTENZ: Fehlende referenzierte Dateien nachkopieren
        # openpyxl erzeugt nicht alle Dateien des Originals (z.B. calcChain.xml).
        # Wenn [Content_Types].xml vom Original kopiert wird aber Dateien fehlen,
        # löst Excel den "Reparatur"-Modus aus → richData/Bilder werden entfernt!
        # =====================================================================
        ct_file = os.path.join(temp_dir, '[Content_Types].xml')
        if os.path.exists(ct_file):
            with open(ct_file, 'r', encoding='utf-8') as f:
                ct_content = f.read()
            
            def _ct_override_fixer(m):
                nonlocal fixed_count
                part_name = m.group(1)  # z.B. "/xl/calcChain.xml"
                rel_path = part_name.lstrip('/')
                dest_file_ct = os.path.join(temp_dir, rel_path.replace('/', os.sep))
                if os.path.exists(dest_file_ct):
                    return m.group(0)  # Datei existiert, Eintrag behalten
                # externalLinks NIEMALS aus Original kopieren (stale cached values!)
                if 'externalLinks/' in part_name:
                    sys.stderr.write(f"[restore_ext] ContentTypes-Konsistenz: {part_name} entfernt (externalLink, nicht kopieren)\n")
                    return ''
                # Bei structural_change: Slicer-Dateien NICHT aus Original kopieren
                # (SlicerCaches/Slicers sind nach Spaltenoperationen invalide)
                if structural_change and 'slicer' in part_name.lower():
                    sys.stderr.write(f"[restore_ext] ContentTypes-Konsistenz: {part_name} entfernt (Slicer bei structural_change)\n")
                    return ''
                # Datei fehlt — aus Original kopieren
                orig_file_ct = os.path.join(orig_temp_dir, rel_path.replace('/', os.sep))
                if os.path.exists(orig_file_ct):
                    os.makedirs(os.path.dirname(dest_file_ct), exist_ok=True)
                    shutil.copy2(orig_file_ct, dest_file_ct)
                    sys.stderr.write(f"[restore_ext] ContentTypes-Konsistenz: {part_name} aus Original kopiert\n")
                    fixed_count += 1
                    return m.group(0)  # Datei jetzt vorhanden, Eintrag behalten
                else:
                    # Datei existiert nirgends — Eintrag entfernen
                    sys.stderr.write(f"[restore_ext] ContentTypes-Konsistenz: {part_name} entfernt (nirgends vorhanden)\n")
                    return ''
            
            new_ct = re.sub(r'<Override\s+PartName="(/[^"]+)"[^>]*/>\s*', _ct_override_fixer, ct_content)
            if new_ct != ct_content:
                with open(ct_file, 'w', encoding='utf-8') as f:
                    f.write(new_ct)
                sys.stderr.write(f"[restore_ext] [Content_Types].xml bereinigt\n")
        
        # WORKBOOK.XML.RELS KONSISTENZ: Fehlende referenzierte Dateien nachkopieren
        wb_rels_file = os.path.join(temp_dir, 'xl', '_rels', 'workbook.xml.rels')
        if os.path.exists(wb_rels_file):
            with open(wb_rels_file, 'r', encoding='utf-8') as f:
                wb_rels_ct = f.read()
            
            def _rels_fixer(m):
                nonlocal fixed_count
                full_el = m.group(0)
                target = m.group(1)
                # Externe URLs und TargetMode="External" überspringen
                if 'TargetMode="External"' in full_el:
                    return full_el
                if target.startswith('http://') or target.startswith('https://') or target.startswith('mailto:'):
                    return full_el
                # Relativer Pfad: relativ zu xl/
                target_file_r = os.path.normpath(os.path.join(temp_dir, 'xl', target))
                if os.path.exists(target_file_r):
                    return full_el  # Datei existiert
                # externalLinks NIEMALS aus Original kopieren (stale cached values!)
                if 'externalLinks/' in target:
                    sys.stderr.write(f"[restore_ext] Rels-Konsistenz: xl/{target} entfernt (externalLink, nicht kopieren)\n")
                    return ''
                # Bei structural_change: Slicer-Dateien NICHT aus Original kopieren
                if structural_change and 'slicer' in target.lower():
                    sys.stderr.write(f"[restore_ext] Rels-Konsistenz: xl/{target} entfernt (Slicer bei structural_change)\n")
                    return ''
                orig_target_r = os.path.normpath(os.path.join(orig_temp_dir, 'xl', target))
                if os.path.exists(orig_target_r):
                    os.makedirs(os.path.dirname(target_file_r), exist_ok=True)
                    shutil.copy2(orig_target_r, target_file_r)
                    sys.stderr.write(f"[restore_ext] Rels-Konsistenz: xl/{target} aus Original kopiert\n")
                    fixed_count += 1
                    return full_el
                else:
                    sys.stderr.write(f"[restore_ext] Rels-Konsistenz: xl/{target} entfernt (nicht gefunden)\n")
                    return ''
            
            new_rels = re.sub(r'<Relationship\s[^>]*?Target="([^"]+)"[^>]*/>\s*', _rels_fixer, wb_rels_ct)
            if new_rels != wb_rels_ct:
                with open(wb_rels_file, 'w', encoding='utf-8') as f:
                    f.write(new_rels)
                sys.stderr.write(f"[restore_ext] workbook.xml.rels bereinigt\n")
        
        # WORKSHEET RELS KONSISTENZ: Fehlende referenzierte Dateien prüfen
        # Worksheet-Rels wurden aus dem Original kopiert und referenzieren evtl.
        # Dateien (printerSettings, comments, ctrlProps etc.) die im Output fehlen.
        ws_rels_check_dir = os.path.join(temp_dir, 'xl', 'worksheets', '_rels')
        if os.path.exists(ws_rels_check_dir):
            for ws_rels_fn in os.listdir(ws_rels_check_dir):
                if not ws_rels_fn.endswith('.rels'):
                    continue
                ws_rels_fp = os.path.join(ws_rels_check_dir, ws_rels_fn)
                with open(ws_rels_fp, 'r', encoding='utf-8') as f:
                    ws_rels_content = f.read()
                
                def _ws_rels_fixer(m, _fn=ws_rels_fn):
                    nonlocal fixed_count
                    full_el = m.group(0)
                    target = m.group(1)
                    if 'TargetMode="External"' in full_el:
                        return full_el
                    if target.startswith('http://') or target.startswith('https://') or target.startswith('mailto:'):
                        return full_el
                    # Bei structural_change: Slicer-Targets komplett entfernen (Sicherheitsnetz)
                    if structural_change and 'slicer' in target.lower():
                        sys.stderr.write(f"[restore_ext] WS-Rels-Konsistenz: {_fn} → {target} entfernt (Slicer bei structural_change)\n")
                        return ''
                    # Targets sind relativ zu xl/worksheets/
                    target_file = os.path.normpath(os.path.join(temp_dir, 'xl', 'worksheets', target))
                    if os.path.exists(target_file):
                        return full_el
                    # Aus Original kopieren
                    orig_target = os.path.normpath(os.path.join(orig_temp_dir, 'xl', 'worksheets', target))
                    if os.path.exists(orig_target):
                        os.makedirs(os.path.dirname(target_file), exist_ok=True)
                        shutil.copy2(orig_target, target_file)
                        sys.stderr.write(f"[restore_ext] WS-Rels-Konsistenz: {_fn} → {target} aus Original kopiert\n")
                        fixed_count += 1
                        return full_el
                    else:
                        sys.stderr.write(f"[restore_ext] WS-Rels-Konsistenz: {_fn} → {target} entfernt (nicht gefunden)\n")
                        return ''
                
                new_ws_rels = re.sub(r'<Relationship\s[^>]*?Target="([^"]+)"[^>]*/>\s*', _ws_rels_fixer, ws_rels_content)
                if new_ws_rels != ws_rels_content:
                    with open(ws_rels_fp, 'w', encoding='utf-8') as f:
                        f.write(new_ws_rels)
                    sys.stderr.write(f"[restore_ext] {ws_rels_fn} bereinigt\n")
        
        sys.stderr.write(f"[restore_ext] fixed_count={fixed_count}\n")
        
        # =================================================================
        # ROBUST RE-ZIP: ZIP-to-ZIP Ansatz ab dem ORIGINAL
        #
        # Statt nur Dateien aus temp_dir einzupacken (wo Dateien fehlen
        # können die nie explizit kopiert wurden), starten wir vom
        # ORIGINAL-ZIP und ersetzen nur Einträge, für die eine
        # modifizierte Version in temp_dir existiert.
        #
        # Dadurch bleiben ALLE Original-Dateien erhalten:
        # - xl/drawings/ (Drawing-XMLs + _rels)
        # - xl/media/ (Bilder)
        # - xl/embeddings/ (eingebettete Objekte)
        # - xl/ctrlProps/ (Formular-Steuerelemente)
        # - xl/charts/ (Diagramme)
        # - xl/richData/ (Excel 365 Zellbilder)
        # - xl/activeX/ (ActiveX-Controls)
        # - xl/diagrams/ (SmartArt)
        # und alle anderen Dateien, die openpyxl ohne Pillow
        # nicht korrekt verarbeitet.
        #
        # → Beseitigt "Entfernter Teil: Zeichnungsform" definitiv.
        # =================================================================
        
        # Sammle alle Dateien in temp_dir (modifizierte Versionen)
        temp_files = {}
        for root, dirs, files_list in os.walk(temp_dir):
            dirs[:] = [d for d in dirs if d != '__MACOSX']
            for fname in files_list:
                if fname == 'restored.xlsx' or fname == '.DS_Store' or fname.startswith('._'):
                    continue
                full_path = os.path.join(root, fname)
                arc_name = os.path.relpath(full_path, temp_dir).replace('\\', '/')
                # Bei structural_change: Slicer-Dateien NICHT ins ZIP aufnehmen.
                # Sicherheitsnetz: Falls ein Code-Pfad Slicer-Dateien nach temp_dir
                # kopiert hat, werden sie hier gefiltert.
                if structural_change and 'slicer' in arc_name.lower():
                    sys.stderr.write(f"[re-zip] Slicer in temp_dir gefiltert: {arc_name}\n")
                    continue
                temp_files[arc_name] = full_path
        
        written = set()
        with zipfile.ZipFile(temp_xlsx, 'w', zipfile.ZIP_DEFLATED) as new_zf:
            # Phase 1: Alle Einträge aus dem ORIGINAL-ZIP durchgehen.
            # Wenn eine modifizierte Version in temp_dir existiert → diese nehmen,
            # sonst Original-Bytes 1:1 übernehmen (insbesondere drawings, media etc.)
            with zipfile.ZipFile(original_path, 'r') as orig_zf:
                for item in orig_zf.infolist():
                    if item.filename.endswith('/'):
                        continue  # Verzeichnis-Einträge überspringen
                    name = item.filename
                    if name.startswith('__MACOSX') or name.endswith('.DS_Store') or \
                       name.split('/')[-1].startswith('._'):
                        continue
                    
                    # Bei structural_change: Slicer-Dateien NICHT aus Original übernehmen.
                    # openpyxl unterstützt keine Slicers — nach Spaltenoperationen sind
                    # die SlicerCache-Referenzen auf Tabellenspalten invalide.
                    # Diese Dateien dürfen weder kopiert noch aus dem Original übernommen werden.
                    if structural_change and 'slicer' in name.lower():
                        sys.stderr.write(f"[re-zip] Slicer-Datei übersprungen: {name}\n")
                        continue
                    
                    # KRITISCH: Worksheet-Dateien die nicht im Output (temp_dir) sind,
                    # wurden absichtlich entfernt (z.B. nicht-ausgewählte Sheets beim Export).
                    # Diese dürfen NICHT aus dem Original zurückkopiert werden!
                    # Betrifft: xl/worksheets/sheet*.xml und xl/worksheets/_rels/sheet*.xml.rels
                    if name not in temp_files and (
                        re.match(r'^xl/worksheets/sheet\d+\.xml$', name) or
                        re.match(r'^xl/worksheets/_rels/sheet\d+\.xml\.rels$', name)
                    ):
                        sys.stderr.write(f"[re-zip] Entferntes Worksheet übersprungen: {name}\n")
                        continue
                    
                    if name in temp_files:
                        # Modifizierte Version aus temp_dir verwenden
                        new_zf.write(temp_files[name], name)
                    else:
                        # Original-Bytes 1:1 übernehmen (drawings, embeddings etc.)
                        data = orig_zf.read(name)
                        info = item
                        info.compress_type = zipfile.ZIP_DEFLATED
                        new_zf.writestr(info, data)
                    written.add(name)
            
            # Phase 2: Neue Dateien aus temp_dir die NICHT im Original waren
            # (z.B. neue Tables, neue Sheets, geänderte sharedStrings)
            for arc_name, full_path in temp_files.items():
                if arc_name not in written:
                    new_zf.write(full_path, arc_name)
                    written.add(arc_name)
        
        shutil.copy2(temp_xlsx, output_path)
        sys.stderr.write(f"[restore_ext] XLSX wiederhergestellt (ZIP-to-ZIP, {len(written)} Einträge)\n")
        
        # DIAGNOSE: Dump des endgültigen ZIP-Inhalts für Image-Debugging
        # Läuft IMMER (auch wenn fixed_count == 0), um den Endzustand zu zeigen
        try:
            with zipfile.ZipFile(output_path, 'r') as zf:
                all_names = zf.namelist()
                img_related = [n for n in all_names if any(k in n.lower() for k in 
                    ['draw', 'media', 'image', 'picture', 'vml', 'richdata', 'rdrichvalue', 'metadata', '_rels/sheet'])]
                sys.stderr.write(f"[DIAGNOSE] === FINAL ZIP STATE ===\n")
                sys.stderr.write(f"[DIAGNOSE] Image-relevante Dateien ({len(img_related)}):\n")
                for n in sorted(img_related):
                    size = zf.getinfo(n).file_size
                    sys.stderr.write(f"[DIAGNOSE]   {n} ({size} bytes)\n")
                
                # Slicer-/Drawing-Kette analysieren
                slicer_related = [n for n in all_names if any(k in n.lower() for k in 
                    ['slicer', 'ctrlprop'])]
                if slicer_related:
                    sys.stderr.write(f"[DIAGNOSE] Slicer-relevante Dateien ({len(slicer_related)}):\n")
                    for n in sorted(slicer_related):
                        size = zf.getinfo(n).file_size
                        sys.stderr.write(f"[DIAGNOSE]   {n} ({size} bytes)\n")
                
                # Drawing-XMLs analysieren
                for n in sorted(all_names):
                    if n.startswith('xl/drawings/') and n.endswith('.xml') and '/_rels/' not in n:
                        drawing_xml = zf.read(n).decode('utf-8', errors='replace')
                        has_slicer_uri = any(uri in drawing_xml for uri in [
                            'schemas.microsoft.com/office/drawing/2010/slicer',
                            'schemas.microsoft.com/office/drawing/2014/slicer'])
                        anchor_count = len(re.findall(r'<(?:xdr:)?(?:twoCellAnchor|oneCellAnchor)', drawing_xml))
                        mc_count = len(re.findall(r'<mc:AlternateContent', drawing_xml))
                        sys.stderr.write(f"[DIAGNOSE] {n}: {len(drawing_xml)} bytes, "
                                       f"anchors={anchor_count}, mc:AC={mc_count}, "
                                       f"slicer_uri={has_slicer_uri}\n")
                        # Drawing rels
                        dr_rels_name = n.replace('xl/drawings/', 'xl/drawings/_rels/') + '.rels'
                        if dr_rels_name in all_names:
                            dr_rels = zf.read(dr_rels_name).decode('utf-8', errors='replace')
                            dr_rels_entries = re.findall(r'Id="([^"]+)"[^>]*Target="([^"]+)"', dr_rels)
                            sys.stderr.write(f"[DIAGNOSE]   rels: {dr_rels_entries}\n")
                        else:
                            sys.stderr.write(f"[DIAGNOSE]   KEINE drawing rels ({dr_rels_name})\n")
                
                # Sheet XMLs prüfen
                for n in sorted(all_names):
                    if n.startswith('xl/worksheets/') and n.endswith('.xml') and '/_rels/' not in n:
                        sheet_xml = zf.read(n).decode('utf-8', errors='replace')
                        has_drawing = bool(re.search(r'<drawing[\s>]', sheet_xml))
                        has_tp = bool(re.search(r'<tableParts', sheet_xml))
                        has_vm = bool(re.search(r'\bvm=', sheet_xml))
                        has_mc = 'mc:Ignorable' in sheet_xml
                        
                        # Check worksheet-level <extLst> (last one before </worksheet>)
                        has_extlst = False
                        ws_end_pos = sheet_xml.rfind('</worksheet>')
                        if ws_end_pos >= 0:
                            ext_start = sheet_xml.rfind('<extLst', 0, ws_end_pos)
                            if ext_start >= 0:
                                ext_end = sheet_xml.find('</extLst>', ext_start)
                                if ext_end >= 0:
                                    after_ext = sheet_xml[ext_end + len('</extLst>'):].strip()
                                    has_extlst = after_ext.startswith('</worksheet>')
                        
                        # rId-Werte extrahieren
                        dr_rid = re.search(r'<drawing r:id="(rId\d+)"', sheet_xml)
                        tp_rids = re.findall(r'<tablePart[^>]*r:id="(rId\d+)"', sheet_xml)
                        ps_rid = re.search(r'<pageSetup[^>]*r:id="(rId\d+)"', sheet_xml)
                        vm_vals = re.findall(r'\bvm="(\d+)"', sheet_xml)
                        
                        # Namespace im Root
                        root_m = re.search(r'(<worksheet\b[^>]{0,100})', sheet_xml)
                        root_preview = root_m.group(1) if root_m else 'N/A'
                        
                        sys.stderr.write(f"[DIAGNOSE] {n}: drawing={has_drawing}({dr_rid.group(1) if dr_rid else '-'}), "
                                        f"tableParts={has_tp}({','.join(tp_rids) if tp_rids else '-'}), "
                                        f"vm={has_vm}({','.join(vm_vals[:3]) if vm_vals else '-'}), "
                                        f"mc:Ignorable={has_mc}, "
                                        f"pageSetup_rid={ps_rid.group(1) if ps_rid else '-'}, "
                                        f"extLst={has_extlst}\n")
                        sys.stderr.write(f"[DIAGNOSE]   root: {root_preview}...\n")
                        
                        # Rels für dieses Sheet prüfen
                        sheet_name_only = n.split('/')[-1]
                        rels_name = f"xl/worksheets/_rels/{sheet_name_only}.rels"
                        if rels_name in all_names:
                            rels_xml = zf.read(rels_name).decode('utf-8', errors='replace')
                            rels_entries = re.findall(r'Id="(rId\d+)"[^>]*Type="[^"]*?/(\w+)"[^>]*Target="([^"]*)"', rels_xml)
                            sys.stderr.write(f"[DIAGNOSE]   rels ({rels_name}):\n")
                            for rid, rtype, target in rels_entries:
                                sys.stderr.write(f"[DIAGNOSE]     {rid} -> {rtype} ({target})\n")
                        else:
                            sys.stderr.write(f"[DIAGNOSE]   KEINE RELS DATEI für {sheet_name_only}\n")
                
                # Content_Types Slicer-Einträge prüfen
                if '[Content_Types].xml' in all_names:
                    ct_xml = zf.read('[Content_Types].xml').decode('utf-8', errors='replace')
                    ct_slicer = [m.group(0) for m in re.finditer(r'<Override[^>]*slicer[^>]*/>', ct_xml, re.IGNORECASE)]
                    if ct_slicer:
                        sys.stderr.write(f"[DIAGNOSE] [Content_Types].xml hat {len(ct_slicer)} Slicer-Override(s):\n")
                        for s in ct_slicer:
                            sys.stderr.write(f"[DIAGNOSE]   {s}\n")
                    else:
                        sys.stderr.write(f"[DIAGNOSE] [Content_Types].xml: KEINE Slicer-Overrides\n")
                
                # workbook.xml.rels Slicer-Einträge prüfen
                wb_rels_name = 'xl/_rels/workbook.xml.rels'
                if wb_rels_name in all_names:
                    wb_rels_xml = zf.read(wb_rels_name).decode('utf-8', errors='replace')
                    wb_slicer_rels = re.findall(r'<Relationship[^>]*(?:slicer|Slicer)[^>]*/>', wb_rels_xml)
                    if wb_slicer_rels:
                        sys.stderr.write(f"[DIAGNOSE] workbook.xml.rels hat {len(wb_slicer_rels)} Slicer-Rel(s):\n")
                        for r in wb_slicer_rels:
                            sys.stderr.write(f"[DIAGNOSE]   {r}\n")
                    else:
                        sys.stderr.write(f"[DIAGNOSE] workbook.xml.rels: KEINE Slicer-Rels\n")
                
                sys.stderr.write(f"[DIAGNOSE] === END ===\n")
        except Exception as diag_err:
            sys.stderr.write(f"[DIAGNOSE] Fehler: {diag_err}\n")
    
    finally:
        if temp_dir:
            shutil.rmtree(temp_dir, ignore_errors=True)
        if orig_temp_dir:
            shutil.rmtree(orig_temp_dir, ignore_errors=True)


def apply_tint(rgb_hex, tint):
    """
    Wendet einen Tint auf eine RGB-Farbe an.
    Tint > 0: heller (Richtung weiß)
    Tint < 0: dunkler (Richtung schwarz)
    """
    if not rgb_hex or len(rgb_hex) < 6:
        return rgb_hex
    
    # Parse RGB
    r = int(rgb_hex[0:2], 16)
    g = int(rgb_hex[2:4], 16)
    b = int(rgb_hex[4:6], 16)
    
    if tint > 0:
        # Aufhellen (Richtung weiß)
        r = int(r + (255 - r) * tint)
        g = int(g + (255 - g) * tint)
        b = int(b + (255 - b) * tint)
    elif tint < 0:
        # Abdunkeln (Richtung schwarz)
        r = int(r * (1 + tint))
        g = int(g * (1 + tint))
        b = int(b * (1 + tint))
    
    # Clamp to 0-255
    r = max(0, min(255, r))
    g = max(0, min(255, g))
    b = max(0, min(255, b))
    
    return f'{r:02X}{g:02X}{b:02X}'

def theme_color_to_rgb(color, workbook=None):
    """
    Konvertiert eine Theme-Farbe zu RGB.
    
    Args:
        color: openpyxl Color Objekt
        workbook: Workbook für Theme-Lookup (optional)
    
    Returns:
        RGB Hex-String (z.B. 'FF0000') oder None
    """
    if not color:
        return None
    
    color_type = getattr(color, 'type', None)
    
    if color_type == 'rgb':
        rgb = color.rgb
        if isinstance(rgb, str) and len(rgb) >= 6:
            # Entferne Alpha wenn vorhanden (ARGB -> RGB)
            if len(rgb) == 8:
                return rgb[2:]
            return rgb
        return None
    
    if color_type == 'theme':
        theme_idx = color.theme
        tint = getattr(color, 'tint', 0) or 0
        
        # Hole Basis-Farbe aus Theme
        if theme_idx is not None and 0 <= theme_idx < len(THEME_COLORS):
            base_rgb = THEME_COLORS[theme_idx]
            # Wende Tint an
            return apply_tint(base_rgb, tint)
        return None
    
    if color_type == 'indexed':
        # Indexed colors - verwende Standard-Palette
        # Für einfache Fälle
        indexed = getattr(color, 'indexed', None)
        if indexed == 9:  # Weiß
            return 'FFFFFF'
        elif indexed == 8:  # Schwarz
            return '000000'
        # Andere indexed colors erstmal ignorieren
        return None
    
    return None

def convert_fill_to_rgb(fill):
    """
    Konvertiert ein Fill-Objekt mit Theme-Farben zu einem Fill mit RGB-Farben.
    Dies ist nötig weil openpyxl Theme-Farben nicht korrekt schreibt.
    
    WICHTIG: Pattern-Typen wie gray125 mit Theme-Farben werden zu solid konvertiert,
    da das Muster sonst nicht korrekt dargestellt wird.
    """
    if not fill or fill.patternType is None:
        return fill
    
    fg_rgb = None
    bg_rgb = None
    
    if fill.fgColor:
        fg_rgb = theme_color_to_rgb(fill.fgColor)
    if fill.bgColor:
        bg_rgb = theme_color_to_rgb(fill.bgColor)
    
    # Wenn keine Konvertierung nötig (schon RGB und solid), gib Original zurück
    fg_type = getattr(fill.fgColor, 'type', None) if fill.fgColor else None
    bg_type = getattr(fill.bgColor, 'type', None) if fill.bgColor else None
    
    if fg_type == 'rgb' and (bg_type == 'rgb' or bg_type is None) and fill.patternType == 'solid':
        return fill
    
    # Pattern-Typ: gray125 oder andere Muster mit Theme-Farben -> solid
    # Denn das Muster-Rendering hängt von der Theme-Definition ab
    pattern_type = fill.patternType
    if pattern_type and pattern_type != 'solid' and fg_type == 'theme':
        pattern_type = 'solid'  # Konvertiere zu solid fill
    
    # Erstelle neues Fill mit RGB-Farben
    new_fill = PatternFill(
        patternType=pattern_type,
        fgColor=Color(rgb='FF' + fg_rgb) if fg_rgb else None,
        bgColor=Color(rgb='FF' + bg_rgb) if bg_rgb else None
    )
    
    return new_fill

def convert_font_to_rgb(font):
    """
    Konvertiert ein Font-Objekt mit Theme-Farben zu einem Font mit RGB-Farben.
    """
    if not font:
        return font
    
    if not font.color:
        return font
    
    color_type = getattr(font.color, 'type', None)
    if color_type == 'rgb':
        return font  # Schon RGB
    
    rgb = theme_color_to_rgb(font.color)
    if not rgb:
        return font  # Konnte nicht konvertieren
    
    # Erstelle neuen Font mit RGB-Farbe
    new_font = Font(
        name=font.name,
        size=font.size,
        bold=font.bold,
        italic=font.italic,
        underline=font.underline,
        strike=font.strike,
        color=Color(rgb='FF' + rgb)
    )
    
    return new_font

# xlwings-Unterstützung (optional, für strukturelle Änderungen mit CF-Erhalt)
try:
    from excel_utils import is_excel_installed, structural_change_with_excel
    XLWINGS_AVAILABLE = True
except ImportError:
    XLWINGS_AVAILABLE = False
    def is_excel_installed():
        return False
    def structural_change_with_excel(*args, **kwargs):
        return False


def hex_to_argb(hex_color):
    """Konvertiert Hex ('#FF0000') zu ARGB ('FFFF0000')"""
    if not hex_color:
        return None
    if hex_color.startswith('#'):
        hex_color = hex_color[1:]
    if len(hex_color) == 6:
        return 'FF' + hex_color.upper()
    return hex_color.upper()


def shift_cell_reference(cell_ref, deleted_col_indices, inserted_cols=None):
    """
    Verschiebt eine Zell-Referenz basierend auf gelöschten/eingefügten Spalten.
    
    Args:
        cell_ref: Zell-Referenz wie 'A1' oder 'AB123'
        deleted_col_indices: Liste der gelöschten Spalten-Indices (0-basiert)
        inserted_cols: Dict mit {position: count} für eingefügte Spalten
    
    Returns:
        Neue Zell-Referenz oder None wenn die Zelle gelöscht wurde
    """
    if not cell_ref:
        return cell_ref
    
    # Parse Zell-Referenz
    match = re.match(r'^([A-Z]+)(\d+)$', cell_ref.upper())
    if not match:
        return cell_ref
    
    col_letter = match.group(1)
    row_num = match.group(2)
    col_idx = column_index_from_string(col_letter) - 1  # 0-basiert
    
    # Prüfe ob Spalte gelöscht wurde
    if deleted_col_indices and col_idx in deleted_col_indices:
        return None
    
    # Berechne Verschiebung
    shift = 0
    
    # Verschiebung durch gelöschte Spalten (die VOR dieser Spalte lagen)
    if deleted_col_indices:
        for del_idx in sorted(deleted_col_indices):
            if del_idx < col_idx:
                shift -= 1
    
    # Verschiebung durch eingefügte Spalten
    if inserted_cols:
        for pos, count in inserted_cols.items():
            if pos <= col_idx:
                shift += count
    
    new_col_idx = col_idx + shift
    if new_col_idx < 0:
        return None
    
    new_col_letter = get_column_letter(new_col_idx + 1)
    return f"{new_col_letter}{row_num}"


def shift_range_reference(range_ref, deleted_col_indices, inserted_cols=None):
    """
    Verschiebt einen Bereichs-Referenz wie 'A1:C10'.
    
    Returns:
        Neuen Bereich oder None wenn der Bereich komplett gelöscht wurde
    """
    if not range_ref:
        return range_ref
    
    # Handle mehrere Bereiche (z.B. "A1:B2 C3:D4")
    parts = range_ref.split()
    new_parts = []
    
    for part in parts:
        if ':' in part:
            # Bereich wie A1:C10
            start, end = part.split(':')
            new_start = shift_cell_reference(start, deleted_col_indices, inserted_cols)
            new_end = shift_cell_reference(end, deleted_col_indices, inserted_cols)
            
            if new_start and new_end:
                new_parts.append(f"{new_start}:{new_end}")
        else:
            # Einzelne Zelle
            new_ref = shift_cell_reference(part, deleted_col_indices, inserted_cols)
            if new_ref:
                new_parts.append(new_ref)
    
    return ' '.join(new_parts) if new_parts else None


def adjust_tables(ws, deleted_col_indices, inserted_cols=None, new_headers=None):
    """
    Passt alle Excel-Tabellen (Tables) an wenn Spalten gelöscht/eingefügt werden.
    
    WICHTIG: openpyxl's insert_cols/delete_cols passt Table-Ranges NICHT automatisch an!
    
    Args:
        ws: Worksheet
        deleted_col_indices: Liste der gelöschten Spalten-Indices (0-basiert)
        inserted_cols: Dict mit {position: count} für eingefügte Spalten
        new_headers: Liste der neuen Header (falls vorhanden, für Column-Update)
    """
    if not deleted_col_indices and not inserted_cols:
        return
    
    from openpyxl.worksheet.table import TableColumn
    from openpyxl.utils.cell import range_boundaries
    
    for table_name in ws.tables:
        table = ws.tables[table_name]
        old_ref = table.ref
        
        # Parse die alte Range
        min_col, min_row, max_col, max_row = range_boundaries(old_ref)
        
        # Berechne neue Spaltenanzahl
        old_col_count = max_col - min_col + 1
        deleted_count = len(deleted_col_indices) if deleted_col_indices else 0
        inserted_count = sum(inserted_cols.values()) if inserted_cols else 0
        new_col_count = old_col_count - deleted_count + inserted_count
        
        if new_col_count <= 0:
            continue
        
        # Table startet immer bei Spalte A (openpyxl verschiebt die Daten)
        # Nach delete_cols() ist die erste Spalte immer A1
        new_max_col = min_col + new_col_count - 1
        new_ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(new_max_col)}{max_row}"
        
        table.ref = new_ref
        
        # Auch den AutoFilter der Tabelle anpassen
        if table.autoFilter and table.autoFilter.ref:
            table.autoFilter.ref = new_ref
        
        # TABLE COLUMNS ANPASSEN
        # Die tableColumns müssen zur neuen Spaltenanzahl passen
        old_columns = list(table.tableColumns)
        
        # SCHRITT 1: Gelöschte Spalten aus tableColumns entfernen
        if deleted_col_indices:
            # Sortiere absteigend um Indexverschiebungen zu vermeiden
            for del_idx in sorted(deleted_col_indices, reverse=True):
                if del_idx < len(old_columns):
                    removed = old_columns.pop(del_idx)
        
        # SCHRITT 2: Neue Spalten einfügen
        if inserted_cols and new_headers:
            for pos, count in sorted(inserted_cols.items()):
                insert_idx = pos
                for i in range(count):
                    new_col_id = len(old_columns) + i + 1
                    new_col_name = new_headers[insert_idx + i] if insert_idx + i < len(new_headers) else f"Column{new_col_id}"
                    new_column = TableColumn(id=new_col_id, name=new_col_name)
                    old_columns.insert(insert_idx + i, new_column)
        
        # SCHRITT 3: Aktualisiere alle Column IDs (müssen 1, 2, 3, ... sein)
        # WICHTIG: Namen NICHT mit new_headers überschreiben bei delete!
        # Die Namen bleiben korrekt wenn wir nur die gelöschte Column entfernen.
        for idx, col in enumerate(old_columns):
            col.id = idx + 1
        
        # Setze die neuen Columns
        table.tableColumns = old_columns


def adjust_conditional_formatting(ws, deleted_col_indices, inserted_cols=None):
    """
    Passt alle bedingten Formatierungen an wenn Spalten gelöscht/eingefügt werden.
    
    WICHTIG: openpyxl's delete_cols() macht das NICHT automatisch!
    
    Args:
        ws: Worksheet
        deleted_col_indices: Liste der gelöschten Spalten-Indices (0-basiert)
        inserted_cols: Dict mit {position: count} für eingefügte Spalten
    """
    if not deleted_col_indices and not inserted_cols:
        return
    
    
    # Sammle alle CF-Regeln
    old_rules = list(ws.conditional_formatting._cf_rules.items())
    
    # Lösche alle CF-Regeln
    ws.conditional_formatting = ConditionalFormattingList()
    
    # Füge angepasste Regeln wieder hinzu
    for cf_obj, rules in old_rules:
        old_sqref = str(cf_obj.sqref)
        new_sqref = shift_range_reference(old_sqref, deleted_col_indices, inserted_cols)
        
        
        if new_sqref:
            # Füge Regel mit neuem Bereich hinzu
            for rule in rules:
                ws.conditional_formatting.add(new_sqref, rule)


def adjust_cf_for_row_changes(ws, row_mapping, original_row_count):
    """
    Passt alle bedingten Formatierungen an wenn Zeilen gelöscht/verschoben werden.
    
    Args:
        ws: Worksheet
        row_mapping: Liste wo row_mapping[new_pos] = original_data_row_idx (0-basiert)
        original_row_count: Ursprüngliche Anzahl der Datenzeilen
    """
    import re
    import sys
    
    if not row_mapping:
        return
    
    new_row_count = len(row_mapping)
    
    # Wenn keine Änderung in der Anzahl, nichts zu tun
    rows_deleted = original_row_count - new_row_count
    if rows_deleted <= 0:
        return
    
    sys.stderr.write(f"[CF ROW ADJUST] {rows_deleted} Zeilen gelöscht, passe CF an...\n")
    
    # Sammle alle CF-Regeln
    old_rules = list(ws.conditional_formatting._cf_rules.items())
    
    # Lösche alle CF-Regeln
    ws.conditional_formatting = ConditionalFormattingList()
    
    def adjust_cell_ref(cell_ref, deleted_count, new_max_row):
        """Passt eine Zellreferenz an (z.B. H2404 -> H2403)"""
        match = re.match(r'^(\$?)([A-Z]+)(\$?)(\d+)$', cell_ref.upper())
        if not match:
            return cell_ref
        
        col_abs = match.group(1)
        col_letter = match.group(2)
        row_abs = match.group(3)
        row_num = int(match.group(4))
        
        # Header-Zeile (1) nicht anpassen
        if row_num == 1:
            return cell_ref
        
        # Datenzeilen: Zeile 2 = Datenzeile 0
        # Nach Löschen: Neue max Zeile = new_max_row + 1 (Header)
        new_row = row_num - deleted_count
        
        # Nicht unter Zeile 2 gehen
        if new_row < 2:
            new_row = 2
        
        # Nicht über die neue maximale Zeile hinaus
        max_excel_row = new_max_row + 1  # +1 für Header
        if new_row > max_excel_row:
            new_row = max_excel_row
        
        return f"{col_abs}{col_letter}{row_abs}{new_row}"
    
    def adjust_range(range_str, deleted_count, new_max_row):
        """Passt einen Bereich an (z.B. H2:H2404 -> H2:H2403)"""
        # Kann mehrere Bereiche enthalten, getrennt durch Leerzeichen
        parts = range_str.split(' ')
        adjusted_parts = []
        
        for part in parts:
            if ':' in part:
                # Bereich wie H2:H2404
                start, end = part.split(':')
                new_start = adjust_cell_ref(start, deleted_count, new_max_row)
                new_end = adjust_cell_ref(end, deleted_count, new_max_row)
                adjusted_parts.append(f"{new_start}:{new_end}")
            else:
                # Einzelne Zelle wie I458
                adjusted_parts.append(adjust_cell_ref(part, deleted_count, new_max_row))
        
        return ' '.join(adjusted_parts)
    
    adjusted_count = 0
    # Füge angepasste Regeln wieder hinzu
    for cf_obj, rules in old_rules:
        old_sqref = str(cf_obj.sqref)
        new_sqref = adjust_range(old_sqref, rows_deleted, new_row_count)
        
        if new_sqref != old_sqref:
            adjusted_count += 1
        
        if new_sqref:
            for rule in rules:
                ws.conditional_formatting.add(new_sqref, rule)
    
    sys.stderr.write(f"[CF ROW ADJUST] {adjusted_count} CF-Bereiche angepasst\n")


def transform_cf_range(range_ref, column_mapping, deleted_set, target_col_count):
    """
    Transformiert CF-Bereiche basierend auf dem Spalten-Mapping.
    
    Args:
        range_ref: Original-Bereich wie 'A1:C10' oder 'A1:B2 C3:D4'
        column_mapping: Dict {new_col_idx: original_col_idx} (-1 für neue Spalten)
        deleted_set: Set der gelöschten Original-Spalten
        target_col_count: Anzahl der Zielspalten
    
    Returns:
        Transformierter Bereich oder None
    """
    if not range_ref:
        return None
    
    # Baue reverse mapping: original_col -> new_col
    reverse_mapping = {}
    for new_col, orig_col in column_mapping.items():
        if orig_col >= 0:  # Nicht neue Spalten
            reverse_mapping[orig_col] = new_col
    
    def transform_cell_ref(cell_ref):
        """Transformiert eine einzelne Zellreferenz"""
        match = re.match(r'^([A-Z]+)(\d+)$', cell_ref.upper())
        if not match:
            return None
        
        col_letter = match.group(1)
        row_num = match.group(2)
        orig_col_idx = column_index_from_string(col_letter) - 1  # 0-basiert
        
        # Spalte gelöscht?
        if orig_col_idx in deleted_set:
            return None
        
        # Finde neue Position
        if orig_col_idx in reverse_mapping:
            new_col_idx = reverse_mapping[orig_col_idx]
            new_col_letter = get_column_letter(new_col_idx + 1)
            return f"{new_col_letter}{row_num}"
        else:
            # Spalte nicht im Mapping - behalte Original (falls im Zielbereich)
            if orig_col_idx < target_col_count:
                return cell_ref
            return None
    
    # Handle mehrere Bereiche
    parts = range_ref.split()
    new_parts = []
    
    for part in parts:
        if ':' in part:
            start, end = part.split(':')
            new_start = transform_cell_ref(start)
            new_end = transform_cell_ref(end)
            
            if new_start and new_end:
                new_parts.append(f"{new_start}:{new_end}")
        else:
            new_ref = transform_cell_ref(part)
            if new_ref:
                new_parts.append(new_ref)
    
    return ' '.join(new_parts) if new_parts else None


def apply_cell_value(cell, value):
    """
    Setzt den Wert einer Zelle mit korrektem Typ.
    OPTIMIERT für Performance bei großen Datenmengen.
    Überspringt MergedCell-Objekte (nur die obere linke Zelle ist beschreibbar).
    Überspringt Bild-Platzhalter ('🖼️ Bild') damit Original-Zellwerte erhalten bleiben.
    """
    from datetime import date
    from openpyxl.cell.cell import MergedCell
    import re
    
    # MergedCell überspringen - nur die obere linke Zelle einer Merged-Region ist beschreibbar
    if isinstance(cell, MergedCell):
        return
    
    # Bild-Platzhalter überspringen — der Reader setzt '🖼️ Bild' für Zellen mit
    # richData/vm-Bildern. Wenn wir diesen Wert schreiben, ändert sich der Zell-Typ
    # von t="e" (error) zu t="inlineStr", was die vm-Attribut-Wiederherstellung und
    # damit die Bildanzeige in Excel zerstört. Original-Zellwert beibehalten!
    if isinstance(value, str) and '🖼️' in value:
        return
    
    # Schnelle Typchecks zuerst
    if value is None or value == '':
        cell.value = None
        return
    
    value_type = type(value)
    
    if value_type is bool:
        cell.value = value
    elif value_type in (int, float):
        cell.value = value
    elif value_type is datetime:
        cell.value = value
    elif value_type is date:
        cell.value = datetime.combine(value, datetime.min.time())
    elif value_type is str:
        # Versuche Datum-Strings zurück zu datetime zu konvertieren
        # Format vom Reader: '30.06.2013 00:00:00' oder '30.06.2013'
        parsed_date = None
        if len(value) >= 10:
            # Versuche verschiedene Datumsformate
            for fmt in ['%d.%m.%Y %H:%M:%S', '%d.%m.%Y', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d']:
                try:
                    parsed_date = datetime.strptime(value, fmt)
                    break
                except ValueError:
                    continue
        
        if parsed_date:
            cell.value = parsed_date
        else:
            cell.value = value
    else:
        cell.value = str(value)


def _filter_rows_xml_regex(sheet_content, row_mapping, new_max_row, hidden_rows=None):
    """
    Filtert und renummeriert Zeilen im Sheet-XML per Regex.
    
    Im Gegensatz zum lxml-Ansatz bleibt das EXAKTE XML-Format erhalten:
    - Namespace-Deklarationen, Attribut-Reihenfolge, Whitespace
    - Alle Zell-Attribute (s, t, vm, cm, ph, etc.)
    - Alle Sub-Elemente (<f>, <is>, <extLst>, etc.)
    
    Args:
        sheet_content: Sheet-XML als String
        row_mapping: Liste wo row_mapping[new_idx] = original_idx (0-basiert, ohne Header)
        new_max_row: Neue maximale Zeilenzahl (inkl. Header)
        hidden_rows: Liste versteckter Zeilen (0-basiert) oder None
    
    Returns:
        Modifizierter Sheet-XML String
    """
    import re
    
    # 1. Alle <row>...</row> und <row .../> Elemente extrahieren
    row_pattern = re.compile(
        r'(<row\s[^>]*?\br="(\d+)"[^>]*?)(\s*/>|>(.*?)</row\s*>)',
        re.DOTALL
    )
    
    original_rows = {}  # excel_row_num → full_match_string
    for m in row_pattern.finditer(sheet_content):
        row_num = int(m.group(2))
        original_rows[row_num] = m.group(0)
    
    sys.stderr.write(f"[REGEX-FILTER] {len(original_rows)} Zeilen im Original gefunden\n")
    
    # 2. Rows-to-keep und Renumber-Map aufbauen
    rows_to_keep = {1}  # Header immer behalten
    row_renumber = {}    # old_excel_row → new_excel_row
    
    for new_idx, orig_idx in enumerate(row_mapping):
        if orig_idx >= 0:
            orig_excel_row = orig_idx + 2
            new_excel_row = new_idx + 2
            rows_to_keep.add(orig_excel_row)
            row_renumber[orig_excel_row] = new_excel_row
    
    # 3. Zeilen filtern und renummerieren
    new_rows_xml = []
    
    # Header (Zeile 1) bleibt unverändert
    if 1 in original_rows:
        new_rows_xml.append(original_rows[1])
    
    # Datenzeilen gemäß row_mapping
    for new_idx, orig_idx in enumerate(row_mapping):
        if orig_idx < 0:
            continue  # Eingefügte Zeile — bei Filter nicht relevant
        orig_excel_row = orig_idx + 2
        new_excel_row = new_idx + 2
        
        if orig_excel_row in original_rows:
            row_xml = original_rows[orig_excel_row]
            
            if orig_excel_row != new_excel_row:
                # Zeilen-Nummer im <row r="..."> ändern
                row_xml = re.sub(
                    r'(<row\s[^>]*?\br=")' + str(orig_excel_row) + r'"',
                    r'\g<1>' + str(new_excel_row) + '"',
                    row_xml,
                    count=1
                )
                # Zell-Referenzen in der Zeile ändern: r="AB123" → r="AB456"
                row_xml = re.sub(
                    r'(r="[A-Z]{1,3})' + str(orig_excel_row) + r'"',
                    r'\g<1>' + str(new_excel_row) + '"',
                    row_xml
                )
            
            # Hidden-Attribut setzen/entfernen
            if hidden_rows is not None:
                row_data_idx = new_idx  # 0-basiert
                if row_data_idx in set(hidden_rows):
                    # hidden="1" setzen
                    if 'hidden="' in row_xml:
                        row_xml = re.sub(r'hidden="[^"]*"', 'hidden="1"', row_xml, count=1)
                    else:
                        row_xml = re.sub(r'(<row\s)', r'\1hidden="1" ', row_xml, count=1)
                else:
                    # hidden entfernen falls vorhanden
                    if 'hidden="1"' in row_xml:
                        row_xml = re.sub(r'\s*hidden="1"', '', row_xml, count=1)
            
            new_rows_xml.append(row_xml)
    
    sys.stderr.write(f"[REGEX-FILTER] {len(new_rows_xml)} Zeilen behalten (inkl. Header)\n")
    
    # 4. sheetData-Block ersetzen
    new_sheet_data = '<sheetData>' + ''.join(new_rows_xml) + '</sheetData>'
    sheet_content = re.sub(
        r'<sheetData[^>]*>.*?</sheetData\s*>',
        new_sheet_data,
        sheet_content,
        flags=re.DOTALL
    )
    
    # 5. dimension aktualisieren
    sheet_content = re.sub(
        r'(<dimension\s+ref="[A-Z]+\d+:[A-Z]+)\d+"',
        r'\g<1>' + str(new_max_row) + '"',
        sheet_content
    )
    
    # 6. autoFilter-Range aktualisieren (im Sheet-XML)
    sheet_content = re.sub(
        r'(<autoFilter\s[^>]*?ref="[A-Z]+\d+:[A-Z]+)\d+"',
        r'\g<1>' + str(new_max_row) + '"',
        sheet_content
    )
    
    # 6b. filterColumn und sortState aus autoFilter entfernen
    # Nach physischem Filtern/Umordnen sind die Filter-Kriterien bereits angewandt.
    # Stale filterColumn-Einträge können dazu führen, dass Excel die Daten
    # erneut filtert und dadurch Zeilen falsch versteckt.
    sheet_content = re.sub(
        r'<filterColumn\s[^>]*/>\s*', '', sheet_content)
    sheet_content = re.sub(
        r'<filterColumn\s[^>]*>.*?</filterColumn>\s*',
        '', sheet_content, flags=re.DOTALL)
    sheet_content = re.sub(
        r'<sortState[^>]*>.*?</sortState>\s*',
        '', sheet_content, flags=re.DOTALL)
    # Leere autoFilter bereinigen (nur noch ref, keine Kinder)
    sheet_content = re.sub(
        r'(<autoFilter\s[^>]*?)>\s*</autoFilter>', r'\1/>', sheet_content)
    
    # 7. sqref-Bereiche renummerieren (Conditional Formatting, DataValidation, etc.)
    # Ohne Renummerierung zeigen CF-Regeln auf falsche Zeilen nach dem Filtern
    # → falsche Farben, falsche Validierungen.
    def _renumber_sqref(m):
        prefix = m.group(1)   # z.B. 'sqref="'
        sqref_val = m.group(2)  # z.B. 'A2:A2404 B2:B2404'
        parts = sqref_val.split()
        new_parts = []
        for part in parts:
            rm = re.match(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', part)
            if rm:
                sc, sr, ec, er = rm.group(1), int(rm.group(2)), rm.group(3), int(rm.group(4))
                # Zeile 1 (Header) bleibt immer — Endzeile auf new_max_row kürzen
                if sr == 1:
                    new_er = new_max_row if er > 1 else er
                    new_parts.append(f"{sc}1:{ec}{new_er}")
                    continue
                # Startzeile renummerieren
                new_sr = row_renumber.get(sr)
                if new_sr is None:
                    # Startzeile wurde weggefiltert — nächste verfügbare gemappte Zeile suchen
                    # WICHTIG: Nur innerhalb des Original-Bereichs [sr, er] suchen!
                    # Sonst werden CF-Regeln auf Zellen angewandt, die nie im Bereich waren.
                    for probe in range(sr, er + 1):
                        if probe in row_renumber:
                            new_sr = row_renumber[probe]
                            break
                if new_sr is None:
                    continue  # Bereich komplett weggefiltert → weglassen
                # Endzeile renummerieren
                new_er = row_renumber.get(er)
                if new_er is None:
                    # Endzeile nicht im Mapping → nächste verfügbare von hinten suchen
                    # WICHTIG: Nur innerhalb des Original-Bereichs [sr, er] suchen!
                    for probe in range(er, sr - 1, -1):
                        if probe in row_renumber:
                            new_er = row_renumber[probe]
                            break
                if new_er is None:
                    continue  # Keine gültige Endzeile im Bereich → weglassen
                if new_sr > new_max_row:
                    continue
                new_er = min(new_er, new_max_row)
                if new_sr > new_er:
                    continue  # Ungültiger Bereich nach Renummerierung → weglassen
                new_parts.append(f"{sc}{new_sr}:{ec}{new_er}")
            else:
                # Einzelne Zelle wie "A5"
                cm = re.match(r'([A-Z]+)(\d+)$', part)
                if cm:
                    col_str = cm.group(1)
                    row_num = int(cm.group(2))
                    if row_num == 1:
                        new_parts.append(part)  # Header
                    elif row_num in row_renumber:
                        new_parts.append(f"{col_str}{row_renumber[row_num]}")
                    # Sonst: Zeile weggefiltert → weglassen
                elif not cm:
                    new_parts.append(part)  # Unbekanntes Format → behalten
        if new_parts:
            return prefix + ' '.join(new_parts) + '"'
        return prefix + '"'  # Leeres sqref (Excel ignoriert es)
    
    sheet_content = re.sub(
        r'(sqref=")([^"]+)"',
        _renumber_sqref,
        sheet_content
    )
    
    # 7b. conditionalFormatting-Elemente mit leerem sqref entfernen
    # Diese entstehen wenn alle Zeilen eines CF-Bereichs weggefiltert wurden.
    sheet_content = re.sub(
        r'<conditionalFormatting\s+sqref=""[^>]*>.*?</conditionalFormatting>\s*',
        '', sheet_content, flags=re.DOTALL)
    sheet_content = re.sub(
        r'<conditionalFormatting\s+sqref=""[^>]*/>\s*',
        '', sheet_content)
    
    # 8. <mergeCells> renummerieren
    # Merge-Referenzen zeigen auf alte Zeilennummern → müssen aktualisiert werden
    def _renumber_merge_cell(m):
        full_match = m.group(0)
        ref = m.group(1)  # z.B. "A5:C10"
        rm = re.match(r'([A-Z]+)(\d+):([A-Z]+)(\d+)', ref)
        if not rm:
            return full_match  # Unbekanntes Format → behalten
        sc, sr, ec, er = rm.group(1), int(rm.group(2)), rm.group(3), int(rm.group(4))
        # Header-Merges (Zeile 1) bleiben unverändert
        if sr == 1 and er == 1:
            return full_match
        # Prüfe ob Start- und Endzeile im Mapping sind
        new_sr = row_renumber.get(sr, sr if sr == 1 else None)
        new_er = row_renumber.get(er, er if er == 1 else None)
        if new_sr is None or new_er is None:
            # Mindestens eine Zeile wurde weggefiltert → Merge entfernen
            return ''
        return f'<mergeCell ref="{sc}{new_sr}:{ec}{new_er}"/>'
    
    sheet_content = re.sub(
        r'<mergeCell\s+ref="([^"]+)"\s*/>',
        _renumber_merge_cell,
        sheet_content
    )
    # Leere <mergeCells> entfernen (falls alle Merges entfernt wurden)
    sheet_content = re.sub(r'<mergeCells\s+count="\d+">\s*</mergeCells>', '', sheet_content)
    # mergeCells count aktualisieren
    remaining_merges = len(re.findall(r'<mergeCell\s', sheet_content))
    if remaining_merges > 0:
        sheet_content = re.sub(
            r'<mergeCells\s+count="\d+"',
            f'<mergeCells count="{remaining_merges}"',
            sheet_content
        )
    
    # 9. <hyperlinks> renummerieren
    # Hyperlink-Referenzen zeigen auf alte Zellpositionen
    def _renumber_hyperlink_ref(m):
        full_match = m.group(0)
        prefix = m.group(1)  # alles vor der Zelladresse
        col_str = m.group(2)   # z.B. "A"
        row_num = int(m.group(3))  # z.B. 5
        suffix = m.group(4)  # alles nach der Zelladresse
        if row_num == 1:
            return full_match  # Header
        new_row = row_renumber.get(row_num)
        if new_row is None:
            return ''  # Zeile weggefiltert → Hyperlink entfernen
        return f'{prefix}{col_str}{new_row}{suffix}'
    
    sheet_content = re.sub(
        r'(<hyperlink\s[^>]*?ref=")([A-Z]+)(\d+)("[^>]*?/>)',
        _renumber_hyperlink_ref,
        sheet_content
    )
    # Leere <hyperlinks> entfernen
    sheet_content = re.sub(r'<hyperlinks>\s*</hyperlinks>', '', sheet_content)
    
    sys.stderr.write(f"[REGEX-FILTER] mergeCell={remaining_merges}, sqref und hyperlinks renummeriert\n")
    
    # 10. dataValidation sqref — schon durch Schritt 7 abgedeckt (sqref= global gematcht)
    
    return sheet_content


def _filter_table_xml_regex(table_content, new_max_row):
    """
    Aktualisiert Table-XML (ref und autoFilter) per Regex.
    Erhält exaktes XML-Format im Gegensatz zu lxml.
    """
    import re
    
    # Table ref aktualisieren
    table_content = re.sub(
        r'(ref="[A-Z]+\d+:[A-Z]+)\d+"',
        r'\g<1>' + str(new_max_row) + '"',
        table_content
    )
    
    # autoFilter innerhalb der Table aktualisieren  
    table_content = re.sub(
        r'(<autoFilter\s[^>]*?ref="[A-Z]+\d+:[A-Z]+)\d+"',
        r'\g<1>' + str(new_max_row) + '"',
        table_content
    )
    
    return table_content


def _apply_vm_cell_map_to_xlsx(xlsx_path, sheet_name, vm_cell_map):
    """Setzt vm-Attribute für kopierte Bild-Zellen im Output-XLSX.
    
    Beim Copy&Paste von Zellen mit eingebetteten Bildern (vm-Attribut) wird
    das Bild an der neuen Position nur angezeigt, wenn das vm-Attribut vorhanden ist.
    restore_external_links_from_original stellt nur ORIGINAL-Positionen wieder her.
    Diese Funktion ergänzt vm-Attribute für NEUE Positionen aus dem Frontend vmCellMap.
    
    vm_cell_map: Dict mit Keys "row-col" (0-basiert) und Values vm-String, z.B. {"19-6": "1"}
    """
    import zipfile, io, re
    
    if not vm_cell_map:
        return
    
    # Frontend-Keys "row-col" (row ist 1-basiert wie styleKey, col ist 0-basiert) → Excel-Zellreferenzen "G20"
    vm_by_ref = {}
    for key, vm_val in vm_cell_map.items():
        parts = str(key).split('-')
        if len(parts) != 2:
            continue
        try:
            row_1based = int(parts[0])  # 1-basiert (styleKey-Format)
            col_0based = int(parts[1])  # 0-basiert
            # Excel row = row_1based + 1 (Header-Zeile berücksichtigen)
            # Excel col = col_0based + 1 (get_column_letter ist 1-basiert)
            cell_ref = f"{get_column_letter(col_0based + 1)}{row_1based + 1}"
            vm_by_ref[cell_ref] = str(vm_val)
        except (ValueError, TypeError):
            continue
    
    if not vm_by_ref:
        return
    
    # Sheet-XML-Pfad ermitteln (analog zu _apply_auto_filter_xml)
    with zipfile.ZipFile(xlsx_path, 'r') as zin:
        wb_xml = zin.read('xl/workbook.xml').decode('utf-8')
        
        sheet_pattern = re.compile(
            r'<sheet\s[^>]*name="' + re.escape(sheet_name) + r'"[^>]*r:id="(rId\d+)"',
            re.IGNORECASE
        )
        m = sheet_pattern.search(wb_xml)
        if not m:
            # Versuche alternative Attribut-Reihenfolge
            sheet_pattern2 = re.compile(
                r'<sheet\s[^>]*r:id="(rId\d+)"[^>]*name="' + re.escape(sheet_name) + r'"',
                re.IGNORECASE
            )
            m = sheet_pattern2.search(wb_xml)
        if not m:
            sys.stderr.write(f"[VM_CELL_MAP] Sheet '{sheet_name}' nicht in workbook.xml gefunden\n")
            return
        
        r_id = m.group(1)
        
        rels_xml = zin.read('xl/_rels/workbook.xml.rels').decode('utf-8')
        # Attribut-Reihenfolge variiert je nach Writer (openpyxl schreibt Target vor Id)
        rel_pattern = re.compile(
            r'<Relationship\s[^>]*Id="' + re.escape(r_id) + r'"[^>]*Target="([^"]+)"',
            re.IGNORECASE
        )
        m2 = rel_pattern.search(rels_xml)
        if not m2:
            rel_pattern2 = re.compile(
                r'<Relationship\s[^>]*Target="([^"]+)"[^>]*Id="' + re.escape(r_id) + r'"',
                re.IGNORECASE
            )
            m2 = rel_pattern2.search(rels_xml)
        if not m2:
            sys.stderr.write(f"[VM_CELL_MAP] Relationship '{r_id}' nicht in workbook.xml.rels gefunden\n")
            return
        
        sheet_target = m2.group(1)
        sheet_xml_path = 'xl/' + sheet_target if not sheet_target.startswith('xl/') else sheet_target
        
        sheet_xml = zin.read(sheet_xml_path).decode('utf-8')
    
    # Bestehende vm-Zellen ermitteln + Style-Index der Originale sammeln
    existing_vm = set()
    vm_val_to_style = {}
    for vm_match in re.finditer(r'<c\s[^>]*?r="([A-Z]+\d+)"[^>]*?\bvm="(\d+)"', sheet_xml):
        existing_vm.add(vm_match.group(1))
        vv = vm_match.group(2)
        if vv not in vm_val_to_style:
            s_m = re.search(r'\bs="(\d+)"', vm_match.group(0))
            if s_m:
                vm_val_to_style[vv] = s_m.group(1)
    for vm_match in re.finditer(r'<c\s[^>]*?\bvm="(\d+)"[^>]*?r="([A-Z]+\d+)"', sheet_xml):
        existing_vm.add(vm_match.group(2))
        vv = vm_match.group(1)
        if vv not in vm_val_to_style:
            s_m = re.search(r'\bs="(\d+)"', vm_match.group(0))
            if s_m:
                vm_val_to_style[vv] = s_m.group(1)
    
    # Nur neue vm-Zellen (nicht schon vom Original-Restore gesetzt)
    new_vm_refs = {ref: val for ref, val in vm_by_ref.items() if ref not in existing_vm}
    if not new_vm_refs:
        sys.stderr.write(f"[VM_CELL_MAP] Alle {len(vm_by_ref)} vm-Zellen bereits vorhanden\n")
        return
    
    # vm-Attribute setzen — wie v1.8.2: t="e" + <v>#VALUE!</v> + Style vom Original
    vm_applied = 0
    for cell_ref, vm_val in new_vm_refs.items():
        orig_style_idx = vm_val_to_style.get(vm_val)
        
        cell_pattern = re.compile(r'(<c\s[^>]*?r="' + re.escape(cell_ref) + r'"[^>]*?)(/?>)')
        match = cell_pattern.search(sheet_xml)
        if match and 'vm=' not in match.group(0):
            # vm-Attribut hinzufügen
            c_tag = match.group(1) + f' vm="{vm_val}"' + match.group(2)
            
            # Zelltyp auf t="e" (Error) setzen — Excel erwartet #VALUE! für Bild-Zellen
            c_tag = re.sub(r'\bt="[^"]*"', 't="e"', c_tag)
            if ' t="' not in c_tag:
                c_tag = c_tag.replace('<c ', '<c t="e" ', 1)
            
            # Style vom Original übernehmen
            if orig_style_idx and f's="{orig_style_idx}"' not in c_tag:
                if re.search(r'\bs="\d+"', c_tag):
                    c_tag = re.sub(r'\bs="\d+"', f's="{orig_style_idx}"', c_tag)
                else:
                    c_tag = c_tag.replace('<c ', f'<c s="{orig_style_idx}" ', 1)
            
            # Value auf #VALUE! setzen
            rest_start = match.end()
            rest_of_cell = sheet_xml[rest_start:]
            close_c = rest_of_cell.find('</c>')
            if close_c >= 0 and not match.group(2).endswith('/>'):
                cell_content = rest_of_cell[:close_c]
                after_cell = rest_of_cell[close_c:]
                cell_content = re.sub(r'<v>[^<]*</v>', '<v>#VALUE!</v>', cell_content)
                if '<v>' not in cell_content:
                    cell_content = '<v>#VALUE!</v>'
                sheet_xml = sheet_xml[:match.start()] + c_tag + cell_content + after_cell
            elif match.group(2) == '/>':
                c_tag = c_tag[:-2] + '><v>#VALUE!</v></c>'
                sheet_xml = sheet_xml[:match.start()] + c_tag + sheet_xml[match.end():]
            else:
                sheet_xml = sheet_xml[:match.start()] + c_tag + sheet_xml[match.end():]
            vm_applied += 1
        elif not match:
            # Zelle existiert nicht — in passender Zeile erstellen
            style_attr = f' s="{orig_style_idx}"' if orig_style_idx else ''
            row_num_match = re.search(r'(\d+)$', cell_ref)
            if row_num_match:
                row_num = row_num_match.group(1)
                row_pattern = re.compile(r'(<row\s[^>]*?\br="' + re.escape(row_num) + r'"[^>]*?>)')
                row_match = row_pattern.search(sheet_xml)
                if row_match:
                    cell_el = f'<c r="{cell_ref}"{style_attr} t="e" vm="{vm_val}"><v>#VALUE!</v></c>'
                    insert_pos = row_match.end()
                    sheet_xml = sheet_xml[:insert_pos] + cell_el + sheet_xml[insert_pos:]
                    vm_applied += 1
                else:
                    sheet_data_end = re.search(r'</sheetData>', sheet_xml)
                    if sheet_data_end:
                        row_el = f'<row r="{row_num}"><c r="{cell_ref}"{style_attr} t="e" vm="{vm_val}"><v>#VALUE!</v></c></row>\n'
                        insert_pos = sheet_data_end.start()
                        sheet_xml = sheet_xml[:insert_pos] + row_el + sheet_xml[insert_pos:]
                        vm_applied += 1
    
    if vm_applied > 0:
        # ZIP neu schreiben mit geändertem Sheet-XML
        buf = io.BytesIO()
        with zipfile.ZipFile(xlsx_path, 'r') as zin:
            with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    if item.filename == sheet_xml_path:
                        zout.writestr(item, sheet_xml.encode('utf-8'))
                    else:
                        zout.writestr(item, zin.read(item.filename))
        
        with open(xlsx_path, 'wb') as f:
            f.write(buf.getvalue())
        
        sys.stderr.write(f"[VM_CELL_MAP] {vm_applied} neue vm-Attribute für Copy&Paste-Bilder in '{sheet_name}' gesetzt\n")


def _apply_auto_filter_xml(xlsx_path, sheet_name, auto_filter_range):
    """Setzt oder entfernt den autoFilter in einer XLSX-Datei per ZIP/XML.
    
    auto_filter_range: z.B. "A1:F100" oder None/leer zum Entfernen.
    Arbeitet direkt auf der fertigen ZIP-Datei — unabhängig vom Export-Pfad.
    """
    import zipfile, io, re
    
    if not auto_filter_range and auto_filter_range is not None:
        # Leerer String → entfernen
        auto_filter_range = None
    
    # Sheet-XML-Pfad ermitteln
    with zipfile.ZipFile(xlsx_path, 'r') as zin:
        # workbook.xml lesen um Sheet-Index zu finden
        wb_xml = zin.read('xl/workbook.xml').decode('utf-8')
        
        # Sheet-Name zu rId mapping
        sheet_pattern = re.compile(
            r'<sheet\s[^>]*name="' + re.escape(sheet_name) + r'"[^>]*r:id="(rId\d+)"',
            re.IGNORECASE
        )
        m = sheet_pattern.search(wb_xml)
        if not m:
            sys.stderr.write(f"[AUTO_FILTER] Sheet '{sheet_name}' nicht in workbook.xml gefunden\n")
            return
        
        r_id = m.group(1)
        
        # rId → Dateiname aus workbook.xml.rels
        rels_xml = zin.read('xl/_rels/workbook.xml.rels').decode('utf-8')
        # Attribut-Reihenfolge variiert je nach Writer (openpyxl schreibt Target vor Id)
        rel_pattern = re.compile(
            r'<Relationship\s[^>]*Id="' + re.escape(r_id) + r'"[^>]*Target="([^"]+)"',
            re.IGNORECASE
        )
        m2 = rel_pattern.search(rels_xml)
        if not m2:
            rel_pattern2 = re.compile(
                r'<Relationship\s[^>]*Target="([^"]+)"[^>]*Id="' + re.escape(r_id) + r'"',
                re.IGNORECASE
            )
            m2 = rel_pattern2.search(rels_xml)
        if not m2:
            sys.stderr.write(f"[AUTO_FILTER] rId '{r_id}' nicht in rels gefunden\n")
            return
        
        sheet_target = m2.group(1)
        # Relativer Pfad: worksheets/sheet1.xml → xl/worksheets/sheet1.xml
        if not sheet_target.startswith('xl/'):
            sheet_xml_path = 'xl/' + sheet_target
        else:
            sheet_xml_path = sheet_target
        
        sheet_xml = zin.read(sheet_xml_path).decode('utf-8')
    
    # Bestehenden autoFilter prüfen
    existing_af = re.search(r'<autoFilter\s[^>]*ref="([^"]*)"[^/]*/>', sheet_xml) or \
                  re.search(r'<autoFilter\s[^>]*ref="([^"]*)"[^>]*>.*?</autoFilter>', sheet_xml, re.DOTALL) or \
                  re.search(r'<autoFilter\s[^>]*ref="([^"]*)"[^>]*/>', sheet_xml)
    
    if auto_filter_range is None:
        # Kein autoFilter gewünscht — nichts tun (Original beibehalten)
        # Nur entfernen wenn explizit "" gesendet wurde (was oben abgefangen wird)
        return
    
    if existing_af and existing_af.group(1) == auto_filter_range:
        # Bereits korrekt — nichts ändern
        sys.stderr.write(f"[AUTO_FILTER] Bereits korrekt: {auto_filter_range}\n")
        return
    
    new_af_element = f'<autoFilter ref="{auto_filter_range}"/>'
    
    if existing_af:
        # Ersetze bestehenden autoFilter (komplett, inkl. filterColumn-Kinder)
        full_af = re.search(r'<autoFilter\s[^>]*>.*?</autoFilter>', sheet_xml, re.DOTALL)
        if full_af:
            sheet_xml = sheet_xml[:full_af.start()] + new_af_element + sheet_xml[full_af.end():]
        else:
            # Self-closing tag
            af_self = re.search(r'<autoFilter\s[^/]*/>', sheet_xml)
            if af_self:
                sheet_xml = sheet_xml[:af_self.start()] + new_af_element + sheet_xml[af_self.end():]
    else:
        # Kein autoFilter vorhanden — einfügen nach <sheetData>...</sheetData>
        sd_end = re.search(r'</sheetData>', sheet_xml)
        if sd_end:
            sheet_xml = sheet_xml[:sd_end.end()] + new_af_element + sheet_xml[sd_end.end():]
    
    # ZIP neu schreiben mit geändertem Sheet-XML
    buf = io.BytesIO()
    with zipfile.ZipFile(xlsx_path, 'r') as zin:
        with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                if item.filename == sheet_xml_path:
                    zout.writestr(item, sheet_xml.encode('utf-8'))
                else:
                    zout.writestr(item, zin.read(item.filename))
    
    with open(xlsx_path, 'wb') as f:
        f.write(buf.getvalue())
    
    sys.stderr.write(f"[AUTO_FILTER] autoFilter gesetzt: {auto_filter_range} in {sheet_name}\n")


def _direct_xml_cell_edit(file_path, output_path, sheet_name, real_edits,
                          hidden_columns=None, hidden_rows=None,
                          row_highlights=None, source_bytes=None,
                          return_bytes=False):
    """
    Direkte XML-Bearbeitung von Zellwerten OHNE openpyxl-Roundtrip.
    
    Kopiert die Original-XLSX 1:1 und ändert nur die betroffenen Zellwerte
    direkt im Worksheet-XML. Dadurch bleiben ALLE Strukturen intakt:
    - Relationships (rIds) 
    - SharedStrings-Indizes
    - Namespaces (mc:Ignorable, xr, x14ac, etc.)
    - Drawings, Media, RichData
    - Tables, Slicers, External Links
    - Conditional Formatting, Styles, Fonts
    
    Verwendet ZIP-to-ZIP Entry-Kopie: Liest Einträge direkt aus dem Original-ZIP
    und schreibt sie in ein neues ZIP. Nur modifizierte XMLs werden ersetzt.
    Keine Dateisystem-Extraktion nötig → keine .DS_Store, ._files, Permission-Probleme.
    
    Kein fix_xlsx_relationships, kein restore_table_xml, kein restore_external_links nötig.
    
    Args:
        file_path: Quelldatei
        output_path: Zieldatei
        sheet_name: Name des Sheets
        real_edits: Dict mit "rowIdx-colIdx" → value (0-basiert, ohne Header)
        hidden_columns: Liste versteckter Spalten (0-basiert) oder None
        hidden_rows: Liste versteckter Zeilen (0-basiert) oder None
        row_highlights: Dict mit "rowIdx" → hexColor (0-basiert, ohne Header) oder None
    """
    import zipfile
    import tempfile
    import shutil
    from xml.etree import ElementTree as ET
    
    sys.stderr.write(f"[DIRECT_XML] Start: {len(real_edits)} Edits für Sheet '{sheet_name}'\n")
    
    # Namespaces die in Excel-Dateien vorkommen
    NS = {
        '': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
        'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
        'mc': 'http://schemas.openxmlformats.org/markup-compatibility/2006',
        'x14ac': 'http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac',
        'xr': 'http://schemas.microsoft.com/office/spreadsheetml/2014/revision',
        'xr3': 'http://schemas.microsoft.com/office/spreadsheetml/2016/revision3',
        'xr6': 'http://schemas.microsoft.com/office/spreadsheetml/2014/revision6',
        'xr10': 'http://schemas.microsoft.com/office/spreadsheetml/2014/revision10',
    }
    MAIN_NS = NS['']
    
    # Registriere alle Namespaces um sie beim Schreiben nicht zu verlieren
    for prefix, uri in NS.items():
        if prefix:
            ET.register_namespace(prefix, uri)
    ET.register_namespace('', MAIN_NS)
    
    # =========================================================================
    # ZIP-to-ZIP Ansatz: Lese direkt aus dem ZIP, modifiziere im Speicher,
    # schreibe in neues ZIP. Keine Dateisystem-Extraktion!
    # =========================================================================
    
    # Temporäre Ausgabedatei (wird am Ende umbenannt)
    temp_output = output_path + '.tmp' if not return_bytes else None
    
    try:
        src_stream = source_bytes if source_bytes is not None else file_path
        with zipfile.ZipFile(src_stream, 'r') as src_zip:
            # 1. Finde Sheet-XML Pfad aus workbook.xml + workbook.xml.rels
            wb_xml = src_zip.read('xl/workbook.xml').decode('utf-8')
            wb_root = ET.fromstring(wb_xml)
            
            sheet_rid = None
            for sheet_el in wb_root.iter(f'{{{MAIN_NS}}}sheet'):
                if sheet_el.get('name') == sheet_name:
                    sheet_rid = sheet_el.get(f'{{{NS["r"]}}}id')
                    break
            
            if not sheet_rid:
                raise ValueError(f"Sheet '{sheet_name}' nicht in workbook.xml gefunden")
            
            sys.stderr.write(f"[DIRECT_XML] Sheet '{sheet_name}' hat rId={sheet_rid}\n")
            
            rels_xml = src_zip.read('xl/_rels/workbook.xml.rels').decode('utf-8')
            rels_root = ET.fromstring(rels_xml)
            rels_ns = 'http://schemas.openxmlformats.org/package/2006/relationships'
            
            sheet_file = None
            for rel_el in rels_root.iter(f'{{{rels_ns}}}Relationship'):
                if rel_el.get('Id') == sheet_rid:
                    sheet_file = rel_el.get('Target')
                    break
            
            if not sheet_file:
                raise ValueError(f"Relationship {sheet_rid} nicht in workbook.xml.rels gefunden")
            
            # Sheet-Pfad normalisieren (Target ist relativ zu xl/)
            sheet_zip_path = 'xl/' + sheet_file.lstrip('/')
            # Normalisiere ../ etc.
            parts = sheet_zip_path.split('/')
            normalized = []
            for p in parts:
                if p == '..':
                    if normalized:
                        normalized.pop()
                elif p != '.':
                    normalized.append(p)
            sheet_zip_path = '/'.join(normalized)
            
            sys.stderr.write(f"[DIRECT_XML] Sheet-ZIP-Pfad: {sheet_zip_path}\n")
            
            # 2. Lese SharedStrings (nur zum Referenzieren)
            shared_strings = []
            has_shared_strings = 'xl/sharedStrings.xml' in src_zip.namelist()
            ss_content = None
            if has_shared_strings:
                ss_content = src_zip.read('xl/sharedStrings.xml').decode('utf-8')
                ss_root = ET.fromstring(ss_content)
                for si in ss_root.iter(f'{{{MAIN_NS}}}si'):
                    texts = []
                    for t in si.iter(f'{{{MAIN_NS}}}t'):
                        if t.text:
                            texts.append(t.text)
                    shared_strings.append(''.join(texts))
            
            # 3. Lese Sheet-XML
            sheet_content = src_zip.read(sheet_zip_path).decode('utf-8')
            
            # 4. Konvertiere Edits zu Excel-Koordinaten
            edits_by_ref = {}
            for key, value in real_edits.items():
                parts = key.split('-')
                if len(parts) != 2:
                    continue
                row_idx = int(parts[0])
                col_idx = int(parts[1])
                excel_row = row_idx + 2
                excel_col = col_idx + 1
                col_letter = get_column_letter(excel_col)
                cell_ref = f"{col_letter}{excel_row}"
                edits_by_ref[cell_ref] = value
            
            sys.stderr.write(f"[DIRECT_XML] Zell-Referenzen: {len(edits_by_ref)} Zellen\n")
            
            # 5. BATCH: Alle Zellen in EINEM Durchlauf ersetzen (statt 1000x Einzel-Regex)
            modified = False
            if edits_by_ref:
                sheet_content, was_modified = _batch_replace_cells_in_xml(
                    sheet_content, edits_by_ref, MAIN_NS, shared_strings, has_shared_strings
                )
                if was_modified:
                    modified = True
            
            # 6. Hidden Columns/Rows direkt im XML setzen
            if hidden_columns is not None:
                sheet_content = _set_hidden_cols_in_xml(sheet_content, hidden_columns, MAIN_NS)
                modified = True
            if hidden_rows is not None:
                sheet_content = _set_hidden_rows_in_xml(sheet_content, hidden_rows, MAIN_NS)
                modified = True
            
            # 6b. Row-Highlights direkt im XML setzen (ohne openpyxl!)
            styles_content_mod = None
            styles_modified = False
            if row_highlights:
                # styles.xml lesen
                styles_zip_path = 'xl/styles.xml'
                if styles_zip_path in src_zip.namelist():
                    styles_raw = src_zip.read(styles_zip_path).decode('utf-8')
                    result = _apply_row_highlights_xml(sheet_content, styles_raw, row_highlights)
                    if result is not None:
                        sheet_content, styles_content_mod = result
                        styles_modified = True
                        modified = True
                        sys.stderr.write(f"[DIRECT_XML] Row-Highlights via XML angewendet\n")
                    else:
                        sys.stderr.write(f"[DIRECT_XML] WARNUNG: Row-Highlights XML-Anwendung fehlgeschlagen\n")
                else:
                    sys.stderr.write(f"[DIRECT_XML] WARNUNG: styles.xml nicht gefunden\n")
            
            # 7. SharedStrings aktualisieren falls neue Strings hinzugefügt wurden
            ss_modified = False
            if has_shared_strings and hasattr(_replace_cell_value_in_xml, '_new_strings'):
                new_strings = _replace_cell_value_in_xml._new_strings
                if new_strings and ss_content:
                    new_count = len(shared_strings) + len(new_strings)
                    ss_content = re.sub(r'count="\d+"', f'count="{new_count}"', ss_content)
                    ss_content = re.sub(r'uniqueCount="\d+"', f'uniqueCount="{new_count}"', ss_content)
                    
                    new_si_xml = ''
                    for s in new_strings:
                        escaped = s.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                        new_si_xml += f'<si><t>{escaped}</t></si>'
                    ss_content = ss_content.replace('</sst>', new_si_xml + '</sst>')
                    ss_modified = True
                    sys.stderr.write(f"[DIRECT_XML] {len(new_strings)} neue SharedStrings hinzugefügt\n")
                
                _replace_cell_value_in_xml._new_strings = []
            
            if not modified:
                # Keine Änderungen → Original einfach kopieren
                if return_bytes:
                    if source_bytes is not None:
                        source_bytes.seek(0)
                    return {'success': True, 'outputPath': output_path, 'method': 'direct-xml',
                            'zip_bytes': source_bytes}
                if os.path.normpath(file_path) != os.path.normpath(output_path):
                    shutil.copy2(file_path, output_path)
                sys.stderr.write(f"[DIRECT_XML] Keine Änderungen nötig\n")
                return {'success': True, 'outputPath': output_path, 'method': 'direct-xml'}
            
            # 8. ZIP-to-ZIP: Kopiere alle Einträge, ersetze nur modifizierte
            import io as _io
            dst_target = _io.BytesIO() if return_bytes else temp_output
            with zipfile.ZipFile(dst_target, 'w') as dst_zip:
                for item in src_zip.infolist():
                    # Überspringe macOS-Artefakte
                    if item.filename.startswith('__MACOSX') or \
                       item.filename.endswith('.DS_Store') or \
                       '/.DS_Store' in item.filename or \
                       item.filename.split('/')[-1].startswith('._'):
                        continue
                    
                    if item.filename == sheet_zip_path:
                        # Modifiziertes Sheet-XML schreiben
                        item.compress_type = zipfile.ZIP_DEFLATED
                        dst_zip.writestr(item, sheet_content.encode('utf-8'))
                    elif item.filename == 'xl/sharedStrings.xml' and ss_modified:
                        # Modifizierte SharedStrings schreiben
                        item.compress_type = zipfile.ZIP_DEFLATED
                        dst_zip.writestr(item, ss_content.encode('utf-8'))
                    elif item.filename == 'xl/styles.xml' and styles_modified:
                        # Modifizierte Styles schreiben (Row-Highlights)
                        item.compress_type = zipfile.ZIP_DEFLATED
                        dst_zip.writestr(item, styles_content_mod.encode('utf-8'))
                    else:
                        # Original-Bytes 1:1 kopieren
                        data = src_zip.read(item.filename)
                        dst_zip.writestr(item, data)
        
        if return_bytes:
            dst_target.seek(0)
            sys.stderr.write(f"[DIRECT_XML] Erfolgreich (in-memory, {dst_target.getbuffer().nbytes} bytes)\n")
            return {'success': True, 'outputPath': output_path, 'method': 'direct-xml',
                    'zip_bytes': dst_target}
        
        # 9. Temporäre Datei an Zielort verschieben
        if os.path.exists(output_path):
            os.remove(output_path)
        shutil.move(temp_output, output_path)
        sys.stderr.write(f"[DIRECT_XML] Erfolgreich gespeichert: {output_path}\n")
        
        return {'success': True, 'outputPath': output_path, 'method': 'direct-xml'}
    
    except Exception:
        # Aufräumen bei Fehler
        if temp_output and os.path.exists(temp_output):
            os.remove(temp_output)
        raise


def _prepare_cell_value(value, cell_ref, shared_strings, has_shared_strings, new_strings_list):
    """
    Bereitet einen Zellwert für die XML-Ersetzung vor.
    Returns (new_type, new_val) oder None wenn übersprungen werden soll.
    """
    from datetime import datetime
    
    if isinstance(value, str) and '🖼️' in value:
        return None
    
    if value is None or value == '':
        return ('empty', '')
    elif isinstance(value, bool):
        return ('bool', '1' if value else '0')
    elif isinstance(value, (int, float)):
        val = str(value)
        if isinstance(value, float) and value == int(value):
            val = str(int(value))
        return ('number', val)
    elif isinstance(value, str):
        # Prüfe ob es ein Datum ist
        parsed_date = None
        if len(value) >= 10:
            for fmt in ['%d.%m.%Y %H:%M:%S', '%d.%m.%Y', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d']:
                try:
                    parsed_date = datetime.strptime(value, fmt)
                    break
                except ValueError:
                    continue
        
        if parsed_date:
            excel_epoch = datetime(1899, 12, 30)
            delta = parsed_date - excel_epoch
            serial = delta.days + delta.seconds / 86400.0
            val = str(int(serial)) if delta.seconds == 0 else str(serial)
            return ('number', val)
        else:
            return ('string', value)
    else:
        return ('string', str(value))


def _build_cell_xml(cell_ref, new_type, new_val, style_attr, vm_attr,
                    shared_strings, has_shared_strings, new_strings_list):
    """Baut das XML-Element für eine Zelle."""
    if new_type == 'empty':
        return f'<c r="{cell_ref}"{style_attr}{vm_attr}/>'
    elif new_type == 'number':
        return f'<c r="{cell_ref}"{style_attr}{vm_attr}><v>{new_val}</v></c>'
    elif new_type == 'bool':
        return f'<c r="{cell_ref}"{style_attr}{vm_attr} t="b"><v>{new_val}</v></c>'
    elif new_type == 'string':
        if has_shared_strings:
            new_idx = len(shared_strings) + len(new_strings_list)
            new_strings_list.append(new_val)
            return f'<c r="{cell_ref}"{style_attr}{vm_attr} t="s"><v>{new_idx}</v></c>'
        else:
            escaped = new_val.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
            return f'<c r="{cell_ref}"{style_attr}{vm_attr} t="inlineStr"><is><t>{escaped}</t></is></c>'
    return None


def _batch_replace_cells_in_xml(sheet_content, edits_by_ref, main_ns, shared_strings, has_shared_strings):
    """
    Ersetzt ALLE Zellwerte in EINEM Durchlauf durch das Sheet-XML.
    
    Statt 1000x Einzel-Regex (O(n*m) bei n Zellen und m=XML-Größe)
    wird das XML nur 1x durchlaufen: Alle <c r="..."> werden gematcht,
    und wenn die Zelle in edits_by_ref ist, wird sie ersetzt.
    
    Nicht gefundene Zellen (neue Zellen) werden am Ende in die passenden
    Zeilen eingefügt.
    
    Returns (new_content, was_modified)
    """
    import re
    
    new_strings_list = []
    
    # Vorbereitung: Werte für alle Zellen berechnen
    prepared = {}
    for cell_ref, value in edits_by_ref.items():
        result = _prepare_cell_value(value, cell_ref, shared_strings, has_shared_strings, new_strings_list)
        if result is not None:
            prepared[cell_ref] = result
    
    if not prepared:
        return sheet_content, False
    
    # Set für schnelles Lookup
    refs_to_edit = set(prepared.keys())
    found_refs = set()
    
    # EIN Regex-Durchlauf: Matche ALLE <c r="..."> Elemente
    cell_pattern = re.compile(
        r'(<c\s[^>]*?r="([A-Z]{1,3}\d+)"[^>]*?)(/\s*>|>(.*?)</c>)',
        re.DOTALL
    )
    
    def _replace_match(m):
        cell_open = m.group(1)
        cell_ref = m.group(2)
        
        if cell_ref not in refs_to_edit:
            return m.group(0)  # Nicht editiert → unverändert lassen
        
        found_refs.add(cell_ref)
        new_type, new_val = prepared[cell_ref]
        
        # Style-Index extrahieren und beibehalten  
        style_attr = ''
        s_match = re.search(r'\bs="(\d+)"', cell_open)
        if s_match:
            style_attr = f' s="{s_match.group(1)}"'
        
        # vm-Attribut beibehalten (Zellbilder)
        vm_attr = ''
        vm_match = re.search(r'\bvm="(\d+)"', cell_open)
        if vm_match:
            vm_attr = f' vm="{vm_match.group(1)}"'
        
        return _build_cell_xml(cell_ref, new_type, new_val, style_attr, vm_attr,
                               shared_strings, has_shared_strings, new_strings_list)
    
    sheet_content = cell_pattern.sub(_replace_match, sheet_content)
    
    # Nicht gefundene Zellen: Müssen in passende Zeilen eingefügt werden
    missing_refs = refs_to_edit - found_refs
    if missing_refs:
        # Gruppiere nach Zeilennummer
        rows_to_insert = {}
        for cell_ref in missing_refs:
            new_type, new_val = prepared[cell_ref]
            if new_type == 'empty':
                continue  # Leere nicht-existierende Zelle → überspringen
            row_num = re.search(r'(\d+)$', cell_ref).group(1)
            if row_num not in rows_to_insert:
                rows_to_insert[row_num] = []
            cell_xml = _build_cell_xml(cell_ref, new_type, new_val, '', '',
                                       shared_strings, has_shared_strings, new_strings_list)
            if cell_xml:
                rows_to_insert[row_num].append(cell_xml)
        
        for row_num, cells_xml in rows_to_insert.items():
            if not cells_xml:
                continue
            cells_str = ''.join(cells_xml)
            
            # Suche die Zeile
            row_pattern = re.compile(
                r'(<row\s[^>]*?\br="' + re.escape(row_num) + r'"[^>]*?>)',
                re.DOTALL
            )
            row_match = row_pattern.search(sheet_content)
            
            if row_match:
                # Nach dem <row ...> Tag einfügen
                insert_pos = row_match.end()
                sheet_content = sheet_content[:insert_pos] + cells_str + sheet_content[insert_pos:]
            else:
                # Zeile existiert nicht → vor </sheetData> erstellen
                new_row = f'<row r="{row_num}">{cells_str}</row>\n'
                sheet_content = sheet_content.replace('</sheetData>', new_row + '</sheetData>')
    
    # SharedStrings-Tracker aktualisieren (für Kompatibilität mit altem Code)
    if not hasattr(_replace_cell_value_in_xml, '_new_strings'):
        _replace_cell_value_in_xml._new_strings = []
    _replace_cell_value_in_xml._new_strings.extend(new_strings_list)
    
    modified = len(found_refs) > 0 or len(missing_refs - {r for r in missing_refs if prepared[r][0] == 'empty'}) > 0
    sys.stderr.write(f"[BATCH_XML] {len(found_refs)} Zellen ersetzt, {len(missing_refs)} neue Zellen eingefuegt\n")
    
    return sheet_content, modified


def _replace_cell_value_in_xml(sheet_content, cell_ref, value, main_ns, shared_strings, has_shared_strings):
    """
    Ersetzt den Wert einer einzelnen Zelle im rohen Worksheet-XML.
    
    Strategien:
    - Zelle existiert mit t="s" (SharedString): Neuen String zur SharedStrings-Tabelle
      hinzufügen und Index aktualisieren
    - Zelle existiert mit t="inlineStr": Text direkt ersetzen
    - Zelle existiert ohne Typ (Zahl): Wert direkt ersetzen
    - Zelle existiert nicht: Zelle in passender Zeile einfügen
    
    Returns (new_content, was_modified)
    """
    import re
    from datetime import datetime, date
    
    # Initialisiere _new_strings Tracker (für SharedStrings-Updates)
    if not hasattr(_replace_cell_value_in_xml, '_new_strings'):
        _replace_cell_value_in_xml._new_strings = []
    
    # Bestimme den neuen Wert und Typ
    if value is None or value == '':
        new_type = 'empty'
        new_val = ''
    elif isinstance(value, bool):
        new_type = 'bool'
        new_val = '1' if value else '0'
    elif isinstance(value, (int, float)):
        new_type = 'number'
        new_val = str(value)
        # Ganzzahlen ohne .0 
        if isinstance(value, float) and value == int(value):
            new_val = str(int(value))
    elif isinstance(value, str):
        # Prüfe ob es ein Datum ist
        parsed_date = None
        if len(value) >= 10:
            for fmt in ['%d.%m.%Y %H:%M:%S', '%d.%m.%Y', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d']:
                try:
                    parsed_date = datetime.strptime(value, fmt)
                    break
                except ValueError:
                    continue
        
        if parsed_date:
            # Excel-Datumsserial (Tage seit 1899-12-30)
            from datetime import timedelta
            excel_epoch = datetime(1899, 12, 30)
            delta = parsed_date - excel_epoch
            serial = delta.days + delta.seconds / 86400.0
            new_type = 'number'
            new_val = str(int(serial)) if delta.seconds == 0 else str(serial)
        else:
            new_type = 'string'
            new_val = value
    else:
        new_type = 'string'
        new_val = str(value)
    
    # Suche die Zelle im XML
    # Mögliche Formate:
    # <c r="A1" t="s"><v>5</v></c>
    # <c r="A1" t="inlineStr"><is><t>text</t></is></c>
    # <c r="A1" s="3"><v>42</v></c>
    # <c r="A1"/>
    # <c r="A1" s="3"/>
    
    # Pattern für die Zelle (selbstschließend oder mit Inhalt)
    cell_pattern = re.compile(
        r'(<c\s[^>]*?r="' + re.escape(cell_ref) + r'"[^>]*?)(/\s*>|>(.*?)</c>)',
        re.DOTALL
    )
    
    match = cell_pattern.search(sheet_content)
    
    if match:
        cell_open = match.group(1)   # <c r="A1" t="s" s="3"
        cell_close = match.group(2)  # /> oder >...</c>
        
        # Extrahiere das s="..." Attribut (Style-Index) — muss erhalten bleiben!
        style_attr = ''
        s_match = re.search(r'\bs="(\d+)"', cell_open)
        if s_match:
            style_attr = f' s="{s_match.group(1)}"'
        
        # Extrahiere vm="..." Attribut — muss erhalten bleiben (Zellbilder)
        vm_attr = ''
        vm_match = re.search(r'\bvm="(\d+)"', cell_open)
        if vm_match:
            vm_attr = f' vm="{vm_match.group(1)}"'
        
        # Baue neues Zell-Element
        if new_type == 'empty':
            new_cell = f'<c r="{cell_ref}"{style_attr}{vm_attr}/>'
        elif new_type == 'number':
            new_cell = f'<c r="{cell_ref}"{style_attr}{vm_attr}><v>{new_val}</v></c>'
        elif new_type == 'bool':
            new_cell = f'<c r="{cell_ref}"{style_attr}{vm_attr} t="b"><v>{new_val}</v></c>'
        elif new_type == 'string':
            if has_shared_strings:
                # Neuen SharedString hinzufügen und Index verwenden
                new_idx = len(shared_strings) + len(_replace_cell_value_in_xml._new_strings)
                _replace_cell_value_in_xml._new_strings.append(new_val)
                new_cell = f'<c r="{cell_ref}"{style_attr}{vm_attr} t="s"><v>{new_idx}</v></c>'
            else:
                # Inline String
                escaped = new_val.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                new_cell = f'<c r="{cell_ref}"{style_attr}{vm_attr} t="inlineStr"><is><t>{escaped}</t></is></c>'
        
        sheet_content = sheet_content[:match.start()] + new_cell + sheet_content[match.end():]
        return sheet_content, True
    
    else:
        # Zelle existiert nicht → in passender Zeile einfügen
        row_num = re.search(r'(\d+)$', cell_ref).group(1)
        
        # Finde die Zeile
        row_pattern = re.compile(
            r'(<row\s[^>]*?\br="' + re.escape(row_num) + r'"[^>]*?>)',
            re.DOTALL
        )
        row_match = row_pattern.search(sheet_content)
        
        if row_match:
            # Baue neues Zell-Element
            if new_type == 'empty':
                return sheet_content, False  # Leere Zelle die nicht existiert → nichts tun
            elif new_type == 'number':
                new_cell = f'<c r="{cell_ref}"><v>{new_val}</v></c>'
            elif new_type == 'bool':
                new_cell = f'<c r="{cell_ref}" t="b"><v>{new_val}</v></c>'
            elif new_type == 'string':
                if has_shared_strings:
                    new_idx = len(shared_strings) + len(_replace_cell_value_in_xml._new_strings)
                    _replace_cell_value_in_xml._new_strings.append(new_val)
                    new_cell = f'<c r="{cell_ref}" t="s"><v>{new_idx}</v></c>'
                else:
                    escaped = new_val.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                    new_cell = f'<c r="{cell_ref}" t="inlineStr"><is><t>{escaped}</t></is></c>'
            
            insert_pos = row_match.end()
            sheet_content = sheet_content[:insert_pos] + new_cell + sheet_content[insert_pos:]
            return sheet_content, True
        else:
            # Zeile existiert nicht → vor </sheetData> erstellen
            if new_type == 'empty':
                return sheet_content, False
            
            if new_type == 'number':
                new_cell = f'<c r="{cell_ref}"><v>{new_val}</v></c>'
            elif new_type == 'bool':
                new_cell = f'<c r="{cell_ref}" t="b"><v>{new_val}</v></c>'
            elif new_type == 'string':
                if has_shared_strings:
                    new_idx = len(shared_strings) + len(_replace_cell_value_in_xml._new_strings)
                    _replace_cell_value_in_xml._new_strings.append(new_val)
                    new_cell = f'<c r="{cell_ref}" t="s"><v>{new_idx}</v></c>'
                else:
                    escaped = new_val.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
                    new_cell = f'<c r="{cell_ref}" t="inlineStr"><is><t>{escaped}</t></is></c>'
            
            new_row = f'<row r="{row_num}">{new_cell}</row>\n'
            sheet_content = sheet_content.replace('</sheetData>', new_row + '</sheetData>')
            return sheet_content, True


def _set_hidden_cols_in_xml(sheet_content, hidden_columns, main_ns):
    """Setzt hidden-Attribute auf <col> Elemente im Worksheet-XML.
    Splittet <col>-Ranges wenn nötig, um gemischte hidden/visible Bereiche korrekt abzubilden.
    Erstellt fehlende <col>-Elemente für Spalten ohne bestehende Definition.
    """
    import re
    
    if not hidden_columns:
        return sheet_content
    
    hidden_set = set(hidden_columns)  # 0-basiert
    
    # Finde <cols>...</cols> Bereich
    cols_match = re.search(r'(<cols>)(.*?)(</cols>)', sheet_content, re.DOTALL)
    if not cols_match:
        # Kein <cols> vorhanden - erstelle neue <col>-Elemente für versteckte Spalten
        sorted_hidden = sorted(hidden_set)
        if not sorted_hidden:
            return sheet_content
        ranges = _group_consecutive_indices(sorted_hidden)
        col_els = ''.join(f'<col min="{s+1}" max="{e+1}" hidden="1"/>' for s, e in ranges)
        cols_xml = f'<cols>{col_els}</cols>'
        sheet_content = re.sub(r'(<sheetData)', cols_xml + r'\1', sheet_content, count=1)
        return sheet_content
    
    cols_inner = cols_match.group(2)
    col_elements = list(re.finditer(r'<col\s[^>]*/>', cols_inner))
    covered_cols = set()  # 0-basiert
    new_col_strs = []
    
    for m_el in col_elements:
        el_str = m_el.group(0)
        min_m = re.search(r'min="(\d+)"', el_str)
        max_m = re.search(r'max="(\d+)"', el_str)
        if not min_m or not max_m:
            new_col_strs.append(el_str)
            continue
        
        col_min = int(min_m.group(1))
        col_max = int(max_m.group(1))
        
        for c in range(col_min, col_max + 1):
            covered_cols.add(c - 1)
        
        # Base-Attribute extrahieren (ohne min, max, hidden)
        inner = el_str[4:-2].strip()  # '<col ' ... '/>'
        inner = re.sub(r'\bmin="[^"]*"', '', inner)
        inner = re.sub(r'\bmax="[^"]*"', '', inner)
        inner = re.sub(r'\bhidden="[^"]*"', '', inner)
        base_attrs = ' '.join(inner.split())  # Whitespace normalisieren
        if base_attrs:
            base_attrs = ' ' + base_attrs
        
        all_hidden = all((c - 1) in hidden_set for c in range(col_min, col_max + 1))
        none_hidden = not any((c - 1) in hidden_set for c in range(col_min, col_max + 1))
        
        if all_hidden:
            new_col_strs.append(f'<col min="{col_min}" max="{col_max}"{base_attrs} hidden="1"/>')
        elif none_hidden:
            new_col_strs.append(f'<col min="{col_min}" max="{col_max}"{base_attrs}/>')
        else:
            # Gemischt - Range splitten in zusammenhängende hidden/visible Teilbereiche
            run_start = col_min
            run_hidden = (col_min - 1) in hidden_set
            for c in range(col_min + 1, col_max + 1):
                c_hidden = (c - 1) in hidden_set
                if c_hidden != run_hidden:
                    # Aktuellen Run abschließen
                    h_attr = ' hidden="1"' if run_hidden else ''
                    new_col_strs.append(f'<col min="{run_start}" max="{c - 1}"{base_attrs}{h_attr}/>')
                    run_start = c
                    run_hidden = c_hidden
            # Letzten Run abschließen
            h_attr = ' hidden="1"' if run_hidden else ''
            new_col_strs.append(f'<col min="{run_start}" max="{col_max}"{base_attrs}{h_attr}/>')
    
    # Fehlende <col>-Elemente für versteckte Spalten ohne bestehende Definition
    uncovered_hidden = sorted(h for h in hidden_set if h not in covered_cols)
    if uncovered_hidden:
        ranges = _group_consecutive_indices(uncovered_hidden)
        for s, e in ranges:
            new_col_strs.append(f'<col min="{s+1}" max="{e+1}" hidden="1"/>')
    
    new_cols_inner = ''.join(new_col_strs)
    sheet_content = sheet_content[:cols_match.start(2)] + new_cols_inner + sheet_content[cols_match.end(2):]
    return sheet_content


def _group_consecutive_indices(sorted_indices):
    """Gruppiert aufeinanderfolgende 0-basierte Indizes in (start, end) Ranges."""
    if not sorted_indices:
        return []
    ranges = []
    start = sorted_indices[0]
    end = start
    for h in sorted_indices[1:]:
        if h == end + 1:
            end = h
        else:
            ranges.append((start, end))
            start = h
            end = h
    ranges.append((start, end))
    return ranges


def _set_hidden_rows_in_xml(sheet_content, hidden_rows, main_ns):
    """Setzt hidden-Attribute auf <row> Elemente im Worksheet-XML.
    
    Robuste Variante: Entfernt ERST alle existierenden hidden-Attribute
    (egal ob "1", "0", "true", "false"), dann setzt hidden="1" neu.
    Vermeidet doppelte Attribute bei unerwarteten hidden-Werten.
    """
    import re
    
    if not hidden_rows:
        return sheet_content
    
    hidden_set = set(hidden_rows)
    _counts = [0, 0, 0]  # [added, removed, unchanged]
    
    def _fix_row(m):
        row_tag = m.group(0)
        r_m = re.search(r'\br="(\d+)"', row_tag)
        if not r_m:
            return row_tag
        row_num = int(r_m.group(1))
        # 0-basiert in hidden_rows, Datenzeilen ab row 2 (row 1 = Header)
        row_idx = row_num - 2
        
        # ERST alle existierenden hidden-Attribute entfernen (alle Varianten)
        cleaned = re.sub(r'\s+hidden="[^"]*"', '', row_tag)
        
        if row_idx in hidden_set:
            # hidden="1" einfügen
            if cleaned.rstrip().endswith('/>'):
                base = cleaned.rstrip()[:-2].rstrip()
                result = base + ' hidden="1"/>'
            else:
                result = cleaned.rstrip('>') + ' hidden="1">'
            if result != row_tag:
                _counts[0] += 1
            else:
                _counts[2] += 1
            return result
        else:
            if cleaned != row_tag:
                _counts[1] += 1
            else:
                _counts[2] += 1
            return cleaned
    
    # <row ...> Opening-Tags UND selbstschließende <row .../> matchen
    sheet_content = re.sub(r'<row\s[^>]*?/?>', _fix_row, sheet_content)
    sys.stderr.write(f"[HIDDEN_ROWS] Rows: hidden_added={_counts[0]}, "
                     f"unhidden={_counts[1]}, unchanged={_counts[2]}\n")
    return sheet_content


def _apply_hidden_rows_to_xlsx(xlsx_path, sheet_name, hidden_rows, source_bytes=None, return_bytes=False):
    """
    Wendet Hidden-Row-Attribute auf eine fertige XLSX-Datei an (ZIP-to-ZIP).
    
    Setzt hidden="1" auf Zeilen die versteckt sein sollen und entfernt
    das Attribut von Zeilen die sichtbar sein sollen. ZIP-to-ZIP Ansatz
    ohne openpyxl → alle Strukturen bleiben intakt.
    
    Optimierung: Wenn die XML-Inhalte unverändert bleiben (Source hat bereits
    korrekte hidden-Attribute), wird der ZIP-Neubau übersprungen.
    """
    import zipfile
    import shutil
    from xml.etree import ElementTree as ET

    MAIN_NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
    RELS_NS = 'http://schemas.openxmlformats.org/package/2006/relationships'
    R_NS = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'

    temp_output = xlsx_path + '.hr_tmp' if not return_bytes else None

    try:
        src_stream = source_bytes if source_bytes is not None else xlsx_path
        with zipfile.ZipFile(src_stream, 'r') as src_zip:
            wb_xml = src_zip.read('xl/workbook.xml').decode('utf-8')
            wb_root = ET.fromstring(wb_xml)

            sheet_rid = None
            for sheet_el in wb_root.iter(f'{{{MAIN_NS}}}sheet'):
                if sheet_el.get('name') == sheet_name:
                    sheet_rid = sheet_el.get(f'{{{R_NS}}}id')
                    break

            if not sheet_rid:
                sys.stderr.write(f"[HIDDEN_ROWS] Sheet '{sheet_name}' nicht gefunden\n")
                if return_bytes:
                    if source_bytes is not None:
                        source_bytes.seek(0)
                    return source_bytes
                return

            rels_xml = src_zip.read('xl/_rels/workbook.xml.rels').decode('utf-8')
            rels_root = ET.fromstring(rels_xml)

            sheet_file = None
            for rel_el in rels_root.iter(f'{{{RELS_NS}}}Relationship'):
                if rel_el.get('Id') == sheet_rid:
                    sheet_file = rel_el.get('Target')
                    break

            if not sheet_file:
                sys.stderr.write(f"[HIDDEN_ROWS] Relationship {sheet_rid} nicht gefunden\n")
                if return_bytes:
                    if source_bytes is not None:
                        source_bytes.seek(0)
                    return source_bytes
                return

            sheet_zip_path = 'xl/' + sheet_file.lstrip('/')
            parts = sheet_zip_path.split('/')
            normalized = []
            for p in parts:
                if p == '..':
                    if normalized:
                        normalized.pop()
                elif p != '.':
                    normalized.append(p)
            sheet_zip_path = '/'.join(normalized)

            original_content = src_zip.read(sheet_zip_path).decode('utf-8')
            sheet_content = _set_hidden_rows_in_xml(original_content, hidden_rows, MAIN_NS)

            # Optimierung: Wenn keine Änderungen, ZIP-Neubau überspringen
            if sheet_content == original_content:
                sys.stderr.write(f"[HIDDEN_ROWS] Keine Änderungen nötig — ZIP-Neubau übersprungen\n")
                if return_bytes:
                    if source_bytes is not None:
                        source_bytes.seek(0)
                    return source_bytes
                return

            sys.stderr.write(f"[HIDDEN_ROWS] Änderungen erkannt, schreibe neues ZIP...\n")

            import io as _io
            dst_target = _io.BytesIO() if return_bytes else temp_output
            with zipfile.ZipFile(dst_target, 'w', zipfile.ZIP_DEFLATED) as dst_zip:
                for item in src_zip.infolist():
                    if item.filename.endswith('/'):
                        continue
                    if item.filename.startswith('__MACOSX') or item.filename.endswith('.DS_Store'):
                        continue
                    if item.filename == sheet_zip_path:
                        dst_zip.writestr(item.filename, sheet_content.encode('utf-8'))
                    else:
                        dst_zip.writestr(item.filename, src_zip.read(item.filename))

        if return_bytes:
            dst_target.seek(0)
            sys.stderr.write(f"[HIDDEN_ROWS] {len(hidden_rows)} hidden rows angewendet (in-memory)\n")
            return dst_target

        os.remove(xlsx_path)
        shutil.move(temp_output, xlsx_path)
        sys.stderr.write(f"[HIDDEN_ROWS] {len(hidden_rows)} hidden rows angewendet auf '{sheet_name}'\n")

    except Exception as e:
        if temp_output and os.path.exists(temp_output):
            os.remove(temp_output)
        raise


def _apply_highlights_to_xlsx(xlsx_path, sheet_name, row_highlights, source_bytes=None, return_bytes=False):
    """
    Wendet Row-Highlights auf eine fertige XLSX-Datei an (ZIP-to-ZIP).
    
    Liest Sheet-XML und styles.xml, ruft _apply_row_highlights_xml auf,
    schreibt die modifizierten Dateien zurück ins ZIP.
    Kein openpyxl → Slicers/Drawings bleiben intakt.
    """
    import zipfile
    import shutil
    import tempfile
    from xml.etree import ElementTree as ET
    
    MAIN_NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
    RELS_NS = 'http://schemas.openxmlformats.org/package/2006/relationships'
    R_NS = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
    
    temp_output = xlsx_path + '.hl_tmp' if not return_bytes else None
    
    src_stream = source_bytes if source_bytes is not None else xlsx_path
    with zipfile.ZipFile(src_stream, 'r') as src_zip:
        # Sheet-ZIP-Pfad finden
        wb_xml = src_zip.read('xl/workbook.xml').decode('utf-8')
        wb_root = ET.fromstring(wb_xml)
        
        sheet_rid = None
        for sheet_el in wb_root.iter(f'{{{MAIN_NS}}}sheet'):
            if sheet_el.get('name') == sheet_name:
                sheet_rid = sheet_el.get(f'{{{R_NS}}}id')
                break
        
        if not sheet_rid:
            raise ValueError(f"Sheet '{sheet_name}' nicht gefunden")
        
        rels_xml = src_zip.read('xl/_rels/workbook.xml.rels').decode('utf-8')
        rels_root = ET.fromstring(rels_xml)
        
        sheet_file = None
        for rel_el in rels_root.iter(f'{{{RELS_NS}}}Relationship'):
            if rel_el.get('Id') == sheet_rid:
                sheet_file = rel_el.get('Target')
                break
        
        if not sheet_file:
            raise ValueError(f"Relationship {sheet_rid} nicht gefunden")
        
        sheet_zip_path = 'xl/' + sheet_file.lstrip('/')
        parts = sheet_zip_path.split('/')
        normalized = []
        for p in parts:
            if p == '..':
                if normalized:
                    normalized.pop()
            elif p != '.':
                normalized.append(p)
        sheet_zip_path = '/'.join(normalized)
        
        # Sheet-XML und styles.xml lesen
        sheet_content = src_zip.read(sheet_zip_path).decode('utf-8')
        styles_content = src_zip.read('xl/styles.xml').decode('utf-8')
        
        # Highlights anwenden
        result = _apply_row_highlights_xml(sheet_content, styles_content, row_highlights)
        if result is None:
            sys.stderr.write(f"[HL_XLSX] _apply_row_highlights_xml fehlgeschlagen\n")
            if return_bytes:
                if source_bytes is not None:
                    source_bytes.seek(0)
                return source_bytes
            return
        
        new_sheet, new_styles = result
        
        # ZIP-to-ZIP: Alle Einträge kopieren, Sheet+Styles ersetzen
        import io as _io
        dst_target = _io.BytesIO() if return_bytes else temp_output
        with zipfile.ZipFile(dst_target, 'w') as dst_zip:
            for item in src_zip.infolist():
                if item.filename.startswith('__MACOSX') or item.filename.endswith('.DS_Store'):
                    continue
                
                if item.filename == sheet_zip_path:
                    item.compress_type = zipfile.ZIP_DEFLATED
                    dst_zip.writestr(item, new_sheet.encode('utf-8'))
                elif item.filename == 'xl/styles.xml':
                    item.compress_type = zipfile.ZIP_DEFLATED
                    dst_zip.writestr(item, new_styles.encode('utf-8'))
                else:
                    dst_zip.writestr(item, src_zip.read(item.filename))
    
    if return_bytes:
        dst_target.seek(0)
        sys.stderr.write(f"[HL_XLSX] Row-Highlights erfolgreich angewendet (in-memory)\n")
        return dst_target

    # Ersetze Original
    os.remove(xlsx_path)
    shutil.move(temp_output, xlsx_path)
    sys.stderr.write(f"[HL_XLSX] Row-Highlights erfolgreich angewendet auf {xlsx_path}\n")


def _clear_row_highlights_xml(xlsx_path, sheet_name, cleared_row_indices, source_bytes=None, return_bytes=False):
    """
    Entfernt Row-Highlights direkt im XML, ohne openpyxl-Roundtrip.
    
    Setzt die Zellen der angegebenen Zeilen auf ihren Original-Style zurück,
    indem fillId auf 0 (kein Fill) gesetzt wird über neue XF-Einträge.
    """
    import zipfile
    import shutil
    import re
    from xml.etree import ElementTree as ET
    
    MAIN_NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
    R_NS = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'
    RELS_NS = 'http://schemas.openxmlformats.org/package/2006/relationships'
    
    if not cleared_row_indices:
        if return_bytes:
            if source_bytes is not None:
                source_bytes.seek(0)
            return source_bytes
        return
    
    temp_output = xlsx_path + '.clr_tmp' if not return_bytes else None
    
    src_stream = source_bytes if source_bytes is not None else xlsx_path
    with zipfile.ZipFile(src_stream, 'r') as src_zip:
        wb_xml = src_zip.read('xl/workbook.xml').decode('utf-8')
        wb_root = ET.fromstring(wb_xml)
        
        sheet_rid = None
        for sheet_el in wb_root.iter(f'{{{MAIN_NS}}}sheet'):
            if sheet_el.get('name') == sheet_name:
                sheet_rid = sheet_el.get(f'{{{R_NS}}}id')
                break
        
        if not sheet_rid:
            sys.stderr.write(f"[CLR_HL] Sheet '{sheet_name}' nicht gefunden\n")
            if return_bytes:
                if source_bytes is not None:
                    source_bytes.seek(0)
                return source_bytes
            return
        
        rels_xml = src_zip.read('xl/_rels/workbook.xml.rels').decode('utf-8')
        rels_root = ET.fromstring(rels_xml)
        
        sheet_file = None
        for rel_el in rels_root.iter(f'{{{RELS_NS}}}Relationship'):
            if rel_el.get('Id') == sheet_rid:
                sheet_file = rel_el.get('Target')
                break
        
        if not sheet_file:
            sys.stderr.write(f"[CLR_HL] Relationship {sheet_rid} nicht gefunden\n")
            if return_bytes:
                if source_bytes is not None:
                    source_bytes.seek(0)
                return source_bytes
            return
        
        sheet_zip_path = 'xl/' + sheet_file.lstrip('/')
        parts = sheet_zip_path.split('/')
        normalized = [p for p in parts if p != '.' and p != '..']
        sheet_zip_path = '/'.join(normalized)
        
        sheet_content = src_zip.read(sheet_zip_path).decode('utf-8')
        styles_content = src_zip.read('xl/styles.xml').decode('utf-8')
        
        # XF-Einträge aus cellXfs parsen
        cellxfs_match = re.search(r'<cellXfs\s+count="(\d+)">(.*?)</cellXfs>', styles_content, re.DOTALL)
        if not cellxfs_match:
            sys.stderr.write(f"[CLR_HL] <cellXfs> nicht gefunden\n")
            return
        
        cellxfs_count = int(cellxfs_match.group(1))
        cellxfs_inner = cellxfs_match.group(2)
        xf_entries = re.findall(r'<xf\b[^>]*?(?:/>|>(?:.*?)</xf>)', cellxfs_inner, re.DOTALL)
        
        cleared_excel_rows = set()
        for row_idx in cleared_row_indices:
            cleared_excel_rows.add(int(row_idx) + 2)
        
        styles_needed = set()
        for excel_row in cleared_excel_rows:
            row_pattern = re.compile(
                r'<row\s[^>]*?\br="' + str(excel_row) + r'"[^>]*?>(.*?)</row>',
                re.DOTALL
            )
            row_match = row_pattern.search(sheet_content)
            if row_match:
                row_inner = row_match.group(1)
                for cell_m in re.finditer(r'<c\s[^>]*?(?:/?>)', row_inner):
                    cell_el = cell_m.group(0)
                    s_m = re.search(r'\bs="(\d+)"', cell_el)
                    if s_m:
                        current_s = int(s_m.group(1))
                        if current_s < len(xf_entries):
                            xf = xf_entries[current_s]
                            fill_m = re.search(r'fillId="(\d+)"', xf)
                            if fill_m and int(fill_m.group(1)) > 0:
                                styles_needed.add(current_s)
        
        if not styles_needed:
            sys.stderr.write(f"[CLR_HL] Keine Styles zum Zurücksetzen gefunden\n")
            if return_bytes:
                if source_bytes is not None:
                    source_bytes.seek(0)
                return source_bytes
            return
        
        style_map = {}
        new_xfs_xml = ''
        
        for current_s in sorted(styles_needed):
            old_xf = xf_entries[current_s]
            new_xf = re.sub(r'fillId="\d+"', 'fillId="0"', old_xf)
            if 'applyFill="' in new_xf:
                new_xf = re.sub(r'applyFill="\d+"', 'applyFill="0"', new_xf)
            
            new_s = cellxfs_count
            style_map[current_s] = new_s
            new_xfs_xml += new_xf
            cellxfs_count += 1
        
        if new_xfs_xml:
            styles_content = styles_content.replace('</cellXfs>', new_xfs_xml + '</cellXfs>')
            styles_content = re.sub(r'(<cellXfs\s+)count="\d+"', f'\\1count="{cellxfs_count}"', styles_content)
        
        for excel_row in sorted(cleared_excel_rows):
            row_pattern = re.compile(
                r'(<row\s[^>]*?\br="' + str(excel_row) + r'"[^>]*?>)(.*?)(</row>)',
                re.DOTALL
            )
            row_match = row_pattern.search(sheet_content)
            if row_match:
                row_open = row_match.group(1)
                row_inner = row_match.group(2)
                row_close = row_match.group(3)
                
                def _clear_cell_style(cell_m, _style_map=style_map):
                    cell_el = cell_m.group(0)
                    s_m = re.search(r'\bs="(\d+)"', cell_el)
                    if s_m:
                        current_s = int(s_m.group(1))
                        new_s = _style_map.get(current_s)
                        if new_s is not None:
                            return re.sub(r'\bs="\d+"', f's="{new_s}"', cell_el)
                    return cell_el
                
                new_inner = re.sub(r'<c\s[^>]*?(?:/?>)', _clear_cell_style, row_inner)
                sheet_content = (sheet_content[:row_match.start()] +
                               row_open + new_inner + row_close +
                               sheet_content[row_match.end():])
        
        sys.stderr.write(f"[CLR_HL] {len(cleared_excel_rows)} Zeilen von Highlights befreit, {len(style_map)} Styles angepasst\n")
        
        import io as _io
        dst_target = _io.BytesIO() if return_bytes else temp_output
        with zipfile.ZipFile(dst_target, 'w') as dst_zip:
            for item in src_zip.infolist():
                if item.filename.startswith('__MACOSX') or item.filename.endswith('.DS_Store'):
                    continue
                if item.filename == sheet_zip_path:
                    item.compress_type = zipfile.ZIP_DEFLATED
                    dst_zip.writestr(item, sheet_content.encode('utf-8'))
                elif item.filename == 'xl/styles.xml':
                    item.compress_type = zipfile.ZIP_DEFLATED
                    dst_zip.writestr(item, styles_content.encode('utf-8'))
                else:
                    dst_zip.writestr(item, src_zip.read(item.filename))
    
    if return_bytes:
        dst_target.seek(0)
        sys.stderr.write(f"[CLR_HL] Highlights erfolgreich entfernt (in-memory)\n")
        return dst_target

    os.remove(xlsx_path)
    shutil.move(temp_output, xlsx_path)
    sys.stderr.write(f"[CLR_HL] Highlights erfolgreich entfernt aus {xlsx_path}\n")


def _apply_row_highlights_xml(sheet_content, styles_content, row_highlights):
    """
    Wendet Zeilen-Markierungen direkt im Sheet-XML und styles.xml an.
    
    Vermeidet den openpyxl-Roundtrip komplett → SlicerCaches, Drawings,
    RichData, externalLinks etc. bleiben 100% intakt aus dem Original.
    
    Strategie:
    1. Für jede Highlight-Farbe einen neuen <fill> in styles.xml erstellen
    2. Für jede (alter_style, farbe) Kombination einen neuen <xf> in <cellXfs> erstellen
    3. Zell-Attribute s="..." im Sheet-XML auf neue <xf>-Indizes setzen
    
    Args:
        sheet_content: Sheet-XML als String
        styles_content: styles.xml als String
        row_highlights: dict {row_idx_str: color} (0-basiert ohne Header)
    
    Returns:
        (new_sheet_content, new_styles_content) oder None bei Parse-Fehler
    """
    import re
    
    if not row_highlights:
        return sheet_content, styles_content
    
    # ---- Schritt 1: Fill-Einträge für Highlight-Farben ----
    unique_colors = sorted(set(row_highlights.values()))
    
    # Farbnamen → ARGB Mapping (gleiche Zuordnung wie _apply_row_highlights)
    highlight_colors = {
        'green': 'FF90EE90',
        'yellow': 'FFFFFF00',
        'orange': 'FFFFA500',
        'red': 'FFFF6B6B',
        'blue': 'FF87CEEB',
        'purple': 'FFDDA0DD'
    }
    
    # Parse <fills> Sektion
    fills_match = re.search(r'<fills\s+count="(\d+)">(.*?)</fills>', styles_content, re.DOTALL)
    if not fills_match:
        sys.stderr.write(f"[HL_XML] FEHLER: <fills> nicht gefunden in styles.xml\n")
        return None
    
    fills_count = int(fills_match.group(1))
    fills_inner = fills_match.group(2)
    
    # Existierende Fills durchzählen um Indizes korrekt zu vergeben
    existing_fills = re.findall(r'<fill\b[^/]*?>.*?</fill>', fills_inner, re.DOTALL)
    existing_fills += re.findall(r'<fill\s*/>', fills_inner)
    # fills_count sollte == len(existing_fills), aber wir verwenden fills_count als Basis
    
    color_to_fill_id = {}
    new_fills_xml = ''
    
    for color in unique_colors:
        # Farbnamen in ARGB konvertieren (oder Hex direkt verwenden)
        if color in highlight_colors:
            hex_color = highlight_colors[color]
        elif isinstance(color, str) and color.startswith('#'):
            hex_color = color.lstrip('#').upper()
            if len(hex_color) == 6:
                hex_color = 'FF' + hex_color  # Alpha-Kanal hinzufügen
        else:
            hex_color = 'FFFFFF00'  # Fallback: Gelb
            sys.stderr.write(f"[HL_XML] WARNUNG: Unbekannte Farbe '{color}', verwende Gelb\n")
        
        # Prüfe ob dieser Fill bereits existiert
        existing_idx = None
        for idx, fill_xml in enumerate(existing_fills):
            if f'rgb="{hex_color}"' in fill_xml and 'patternType="solid"' in fill_xml:
                existing_idx = idx
                break
        
        if existing_idx is not None:
            color_to_fill_id[color] = existing_idx
            sys.stderr.write(f"[HL_XML] Farbe {color} → existierender fillId={existing_idx}\n")
        else:
            new_fill = f'<fill><patternFill patternType="solid"><fgColor rgb="{hex_color}"/><bgColor indexed="64"/></patternFill></fill>'
            new_fills_xml += new_fill
            color_to_fill_id[color] = fills_count
            sys.stderr.write(f"[HL_XML] Farbe {color} → neuer fillId={fills_count}\n")
            fills_count += 1
    
    # Fills in styles.xml aktualisieren
    if new_fills_xml:
        styles_content = styles_content.replace('</fills>', new_fills_xml + '</fills>')
        styles_content = re.sub(r'(<fills\s+)count="\d+"', f'\\1count="{fills_count}"', styles_content)
    
    # ---- Schritt 2: cellXfs parsen und neue XF-Einträge erstellen ----
    cellxfs_match = re.search(r'<cellXfs\s+count="(\d+)">(.*?)</cellXfs>', styles_content, re.DOTALL)
    if not cellxfs_match:
        sys.stderr.write(f"[HL_XML] FEHLER: <cellXfs> nicht gefunden in styles.xml\n")
        return None
    
    cellxfs_count = int(cellxfs_match.group(1))
    cellxfs_inner = cellxfs_match.group(2)
    
    # XF-Einträge parsen (self-closing und non-self-closing)
    xf_entries = re.findall(r'<xf\b[^>]*(?:/>|>(?:.*?)</xf>)', cellxfs_inner, re.DOTALL)
    
    sys.stderr.write(f"[HL_XML] {len(xf_entries)} bestehende XF-Einträge, count={cellxfs_count}\n")
    
    # Highlighted Excel-Zeilen ermitteln (0-basiert → Excel-Zeile = idx + 2)
    highlighted_excel_rows = set()
    row_color_map = {}  # excel_row → color
    for row_idx_str, color in row_highlights.items():
        excel_row = int(row_idx_str) + 2
        highlighted_excel_rows.add(excel_row)
        row_color_map[excel_row] = color
    
    # Sammle alle (alter_style, farbe) Kombinationen in EINEM Durchlauf
    styles_needed = set()
    _row_re = re.compile(r'<row\s[^>]*?\br="(\d+)"[^>]*?>(.*?)</row>', re.DOTALL)
    for rm in _row_re.finditer(sheet_content):
        excel_row = int(rm.group(1))
        if excel_row not in highlighted_excel_rows:
            continue
        color = row_color_map[excel_row]
        for cell_m in re.finditer(r'<c\s[^>]*?(?:/?>)', rm.group(2)):
            cell_el = cell_m.group(0)
            s_m = re.search(r'\bs="(\d+)"', cell_el)
            current_s = int(s_m.group(1)) if s_m else 0
            styles_needed.add((current_s, color))
    
    sys.stderr.write(f"[HL_XML] {len(styles_needed)} einzigartige (style, farbe) Kombinationen\n")
    
    # Für jede Kombination einen neuen XF-Eintrag erstellen
    style_map = {}  # (alter_s, farbe) → neuer_s
    new_xfs_xml = ''
    
    for current_s, color in sorted(styles_needed):
        fill_id = color_to_fill_id[color]
        
        if current_s < len(xf_entries):
            old_xf = xf_entries[current_s]
        else:
            # Fallback: Standard-XF
            old_xf = '<xf numFmtId="0" fontId="0" fillId="0" borderId="0"/>'
        
        # fillId im XF ersetzen
        if 'fillId="' in old_xf:
            new_xf = re.sub(r'fillId="\d+"', f'fillId="{fill_id}"', old_xf)
        else:
            new_xf = old_xf.replace('<xf ', f'<xf fillId="{fill_id}" ', 1)
        
        # applyFill="1" setzen
        if 'applyFill="' in new_xf:
            new_xf = re.sub(r'applyFill="\d+"', 'applyFill="1"', new_xf)
        else:
            new_xf = new_xf.replace('<xf ', '<xf applyFill="1" ', 1)
        
        new_s = cellxfs_count
        style_map[(current_s, color)] = new_s
        new_xfs_xml += new_xf
        cellxfs_count += 1
    
    # cellXfs in styles.xml aktualisieren
    if new_xfs_xml:
        styles_content = styles_content.replace('</cellXfs>', new_xfs_xml + '</cellXfs>')
        styles_content = re.sub(r'(<cellXfs\s+)count="\d+"', f'\\1count="{cellxfs_count}"', styles_content)
    
    sys.stderr.write(f"[HL_XML] {len(style_map)} neue XF-Einträge erstellt, cellXfs count={cellxfs_count}\n")
    
    # ---- Schritt 3: Zell-Styles im Sheet-XML aktualisieren (EINZELDURCHLAUF) ----
    def _process_row_highlight(row_match):
        """Verarbeitet eine <row>...</row> und setzt ggf. Highlight-Styles."""
        row_tag = row_match.group(0)
        r_m = re.search(r'\br="(\d+)"', row_tag)
        if not r_m:
            return row_tag
        excel_row = int(r_m.group(1))
        if excel_row not in highlighted_excel_rows:
            return row_tag
        color = row_color_map[excel_row]
        
        def _update_cell_style(cell_m):
            cell_el = cell_m.group(0)
            s_m_inner = re.search(r'\bs="(\d+)"', cell_el)
            current_s = int(s_m_inner.group(1)) if s_m_inner else 0
            new_s = style_map.get((current_s, color))
            if new_s is None:
                return cell_el
            if s_m_inner:
                return re.sub(r'\bs="\d+"', f's="{new_s}"', cell_el)
            else:
                return cell_el.replace('<c ', f'<c s="{new_s}" ', 1)
        
        return re.sub(r'<c\s[^>]*?(?:/?>)', _update_cell_style, row_tag)
    
    sheet_content = re.sub(r'<row\s[^>]*?>.*?</row>', _process_row_highlight, sheet_content, flags=re.DOTALL)
    
    sys.stderr.write(f"[HL_XML] {len(highlighted_excel_rows)} Zeilen markiert\n")
    return sheet_content, styles_content


def write_sheet(file_path, output_path, sheet_name, changes, original_path=None):
    """
    Schreibt Änderungen in ein Excel-Sheet
    
    WICHTIG: Bei strukturellen Änderungen (fullRewrite=True) werden die 
    NEUEN Daten geschrieben. Die Original-Struktur wird beibehalten wo möglich.
    
    Args:
        file_path: Pfad zur Arbeitsdatei (kopierte Datei)
        output_path: Pfad zur Ausgabe-Datei
        sheet_name: Name des Sheets
        changes: Dict mit allen Änderungen
        original_path: Pfad zur Original-Datei (für restore_table_xml)
    
    Returns:
        Dict mit success und ggf. error
    """
    import sys
    
    # Wenn kein original_path gegeben, verwende file_path (Legacy-Kompatibilität)
    if original_path is None:
        original_path = file_path
    
    # KRITISCH: Wenn original_path == file_path == output_path (Speichern in gleicher Datei),
    # muss eine Backup-Kopie erstellt werden BEVOR openpyxl die Datei überschreibt.
    # Sonst kann restore_external_links_from_original nichts wiederherstellen.
    # WINDOWS: normpath normalisiert Pfade (Slashes, Groß/Klein, trailing sep)
    
    # Diagnostik: Was kommt vom Frontend?
    sys.stderr.write(f"[WRITE_SHEET] === ENTRY === sheet={sheet_name}\n")
    sys.stderr.write(f"[WRITE_SHEET] fullRewrite={changes.get('fullRewrite')}, structuralChange={changes.get('structuralChange')}, fromFile={changes.get('fromFile')}\n")
    sys.stderr.write(f"[WRITE_SHEET] deletedRowIndices={len(changes.get('deletedRowIndices', []))}, rowOrder={'ja' if changes.get('rowOrder') else 'nein'}, insRows={'ja' if changes.get('insertedRowInfo') else 'nein'}\n")
    sys.stderr.write(f"[WRITE_SHEET] deletedColumns={len(changes.get('deletedColumns', []))}, columnOrder={'ja' if changes.get('columnOrder') else 'nein'}, insColsType={type(changes.get('insertedColumns')).__name__}\n")
    sys.stderr.write(f"[WRITE_SHEET] editedCells={len({k: v for k, v in changes.get('editedCells', {}).items() if not k.startswith('_')})}, rowMapping={'ja' if changes.get('rowMapping') else 'nein'}\n")
    
    _backup_file = None
    if os.path.normpath(original_path) == os.path.normpath(output_path):
        import tempfile
        _backup_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        _backup_path = _backup_file.name
        _backup_file.close()
        import shutil
        shutil.copy2(original_path, _backup_path)
        sys.stderr.write(f"[WRITE_SHEET] Backup erstellt: {_backup_path} (original==output)\\n")
        # Speichere Backup-Pfad als Funktionsattribut für restore_external_links_from_original
        restore_external_links_from_original._backup_original_path = _backup_path
    
    # =========================================================================
    # FAST-PATH: FALL 3a Pre-Check OHNE openpyxl
    # Wenn nur einfache Zell-Edits / Visibility / Highlights vorliegen und
    # KEINE strukturellen Änderungen → direkte XML-Bearbeitung via ZIP.
    # Umgeht load_workbook() komplett → vermeidet openpyxl Nested.from_tree Bug.
    # =========================================================================
    _fp_edited_cells = changes.get('editedCells', {})
    _fp_real_edits = {k: v for k, v in _fp_edited_cells.items() if not k.startswith('_')} if _fp_edited_cells else {}
    _fp_from_file = changes.get('fromFile', False)
    _fp_full_rewrite = changes.get('fullRewrite', False)
    _fp_structural = changes.get('structuralChange', False)
    _fp_del_cols = changes.get('deletedColumns', [])
    _fp_ins_cols = changes.get('insertedColumns')
    _fp_col_order = changes.get('columnOrder')
    _fp_del_rows = changes.get('deletedRowIndices', [])
    _fp_ins_rows = changes.get('insertedRowInfo')
    _fp_row_order = changes.get('rowOrder')
    _fp_row_mapping = changes.get('rowMapping')
    _fp_hidden_cols = changes.get('hiddenColumns', [])
    _fp_hidden_rows = changes.get('hiddenRows', [])
    _fp_row_highlights = changes.get('rowHighlights', {})
    _fp_cleared_highlights = changes.get('clearedRowHighlights', [])
    _fp_affected_rows = changes.get('affectedRows', [])
    _fp_has_format = changes.get('hasFormatChanges', False)
    _fp_cell_fonts = changes.get('cellFonts', {})
    _fp_rich_text = changes.get('richTextCells', {})
    _fp_auto_filter = changes.get('autoFilterRange')  # AutoFilter für Fast-Path
    _fp_vm_cell_map = changes.get('vmCellMap', {})  # vm-Attribute für kopierte Bild-Zellen
    
    # Keine strukturellen Operationen?
    # rowMapping (Filter) und affectedRows blockieren NICHT:
    # - rowMapping bei Filter → hiddenRows reicht, kein Zeilen-Umbau nötig
    # - affectedRows → nur für openpyxl Style-Reset, XML bewahrt Styles automatisch
    _fp_no_structural = (
        not _fp_from_file
        and not _fp_full_rewrite
        and not _fp_structural
        and not _fp_del_cols
        and not _fp_ins_cols
        and not _fp_col_order
        and not _fp_del_rows
        and not _fp_ins_rows
        and not _fp_row_order
    )
    
    # FALL 3a Gate-Logik spiegeln
    _fp_has_highlight_changes = bool(_fp_cleared_highlights)
    _fp_has_visibility = bool(_fp_hidden_rows) or bool(_fp_hidden_cols)
    _fp_has_extra = _fp_has_format
    if _fp_has_extra and _fp_row_highlights and not _fp_has_highlight_changes:
        if not _fp_cell_fonts and not _fp_rich_text:
            _fp_has_extra = False
    _fp_has_add_highlights = bool(_fp_row_highlights)
    
    _fp_is_fall3a = (
        _fp_no_structural
        and (_fp_real_edits or _fp_has_visibility or _fp_has_add_highlights)
        and not _fp_has_extra
        and not _fp_has_highlight_changes
    )
    
    if _fp_is_fall3a:
        # Sheet-Existenz via ZIP prüfen (ohne openpyxl)
        import zipfile
        from xml.etree import ElementTree as _ET
        try:
            with zipfile.ZipFile(file_path, 'r') as _zf:
                _wb_xml = _zf.read('xl/workbook.xml')
            _wb_root = _ET.fromstring(_wb_xml)
            _ns = {'s': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}
            _sheet_names = [s.get('name') for s in _wb_root.findall('.//s:sheet', _ns)]
        except Exception:
            _sheet_names = []
        
        if sheet_name not in _sheet_names:
            return {'success': False, 'error': f'Sheet "{sheet_name}" nicht gefunden'}
        
        sys.stderr.write(f"[FAST-PATH] FALL 3a direkt ohne openpyxl: {len(_fp_real_edits)} Edits, visibility={_fp_has_visibility}, highlights={_fp_has_add_highlights}\n")
        try:
            result = _direct_xml_cell_edit(
                file_path, output_path, sheet_name, _fp_real_edits,
                _fp_hidden_cols, _fp_hidden_rows,
                row_highlights=_fp_row_highlights,
                source_bytes=None,
                return_bytes=True
            )
            # EINZIGER Disk-Write am Ende (wie FAST-PATH pipeline)
            _fp_zip_bytes = result.get('zip_bytes')
            if _fp_zip_bytes is not None:
                with open(output_path, 'wb') as _fp_out_f:
                    _fp_out_f.write(_fp_zip_bytes.getvalue())
                sys.stderr.write(f"[FAST-PATH] FALL 3a Erfolgreich (in-memory pipeline)\n")
            else:
                sys.stderr.write(f"[FAST-PATH] FALL 3a Erfolgreich (keine Änderungen)\n")
            # AutoFilter vom Frontend anwenden (z.B. bei GUI-Filter → hiddenRows + Dropdown-Pfeile)
            if _fp_auto_filter:
                _apply_auto_filter_xml(output_path, sheet_name, _fp_auto_filter)
            # Backup aufräumen (wurde oben erstellt, wird für FALL 3a nicht gebraucht)
            if _backup_file is not None:
                try:
                    os.remove(_backup_path)
                    sys.stderr.write(f"[FAST-PATH] Backup entfernt: {_backup_path}\n")
                except Exception:
                    pass
            return {'success': True, 'outputPath': output_path, 'method': 'direct-xml'}
        except Exception as xml_err:
            sys.stderr.write(f"[FAST-PATH] XML-Edit fehlgeschlagen: {xml_err}, weiter mit openpyxl...\n")
            # Weiter zu normalem Pfad — load_workbook wird unten versucht
    
    # =========================================================================
    # FAST-PATH-COMBINED: Row + Col Ops → XML-DIREKT in EINEM Python-Aufruf
    # Vermeidet den KOMBINIERT-Branch (zwei Python-Subprozesse + 3× ZIP I/O).
    # Reihenfolge: Row-Ops → Col-Ops → Cell-Edits → Highlights
    # =========================================================================
    _fp_has_any_col_ops = (_fp_col_order and len(_fp_col_order) > 0) or bool(_fp_del_cols) or bool(_fp_ins_cols)
    _fp_has_any_row_ops_c = bool(_fp_del_rows) or bool(_fp_ins_rows) or (_fp_row_order and len(_fp_row_order) > 0)
    
    if _fp_has_any_col_ops and _fp_has_any_row_ops_c and not _fp_from_file:
        _fp_combined_eligible = True  # Row-Insert jetzt auch via XML-Direkt
        
        sys.stderr.write(f"[FAST-PATH-COMBINED] eligible={_fp_combined_eligible}, "
                         f"ins_rows={bool(_fp_ins_rows)}, cell_edits={len(_fp_real_edits)}\n")
        
        if _fp_combined_eligible:
            sys.stderr.write(f"[XML-DIREKT-COMBINED] Verwende XML-Direkt-Weg für Row+Col Ops\n")
            try:
                _fp_script_dir = os.path.dirname(os.path.abspath(__file__))
                if _fp_script_dir not in sys.path:
                    sys.path.insert(0, _fp_script_dir)
                from excel_xml_ops import direct_xml_row_operations, direct_xml_column_operations
                import io as _io_combined
                
                # Quelldatei EINMAL in Speicher lesen
                with open(file_path, 'rb') as _fp_src_f:
                    _fp_zip_bytes = _io_combined.BytesIO(_fp_src_f.read())
                sys.stderr.write(f"[XML-DIREKT-COMBINED] Quelldatei in Speicher geladen\n")
                
                # SCHRITT 1: Zeilen-Operationen (in-memory)
                _fp_row_result = direct_xml_row_operations(
                    file_path=file_path,
                    output_path=output_path,
                    sheet_name=sheet_name,
                    deleted_rows=_fp_del_rows if _fp_del_rows else None,
                    row_order=_fp_row_order,
                    hidden_rows=_fp_hidden_rows if _fp_hidden_rows else None,
                    inserted_rows=_fp_ins_rows,
                    source_bytes=_fp_zip_bytes,
                    return_bytes=True
                )
                _fp_zip_bytes = _fp_row_result['zip_bytes']
                sys.stderr.write(f"[XML-DIREKT-COMBINED] Row-Ops OK: {_fp_row_result.get('method', '?')}\n")
                
                # SCHRITT 2: Spalten-Operationen auf dem Ergebnis (in-memory)
                _fp_headers = changes.get('headers', [])
                _fp_data = changes.get('data', [])
                
                _fp_col_result = direct_xml_column_operations(
                    file_path=output_path,
                    output_path=output_path,
                    sheet_name=sheet_name,
                    deleted_columns=_fp_del_cols if _fp_del_cols else None,
                    inserted_columns=_fp_ins_cols,
                    column_order=_fp_col_order,
                    hidden_columns=_fp_hidden_cols,
                    headers=_fp_headers,
                    data=_fp_data,
                    strip_row_hidden=False,
                    hidden_rows=None,
                    source_bytes=_fp_zip_bytes,
                    return_bytes=True
                )
                _fp_zip_bytes = _fp_col_result['zip_bytes']
                sys.stderr.write(f"[XML-DIREKT-COMBINED] Col-Ops OK: {_fp_col_result.get('method', '?')}\n")
                
                # SCHRITT 3: Cell-Edits via XML (in-memory)
                if _fp_real_edits:
                    sys.stderr.write(f"[XML-DIREKT-COMBINED] {len(_fp_real_edits)} Cell-Edits via XML anwenden\n")
                    try:
                        _fp_cell_result = _direct_xml_cell_edit(
                            output_path, output_path, sheet_name, _fp_real_edits,
                            hidden_columns=None,
                            hidden_rows=None,
                            source_bytes=_fp_zip_bytes,
                            return_bytes=True
                        )
                        _fp_zip_bytes = _fp_cell_result['zip_bytes']
                    except Exception as edit_err:
                        _fp_zip_bytes.seek(0)
                        sys.stderr.write(f"[XML-DIREKT-COMBINED] WARNUNG: Cell-Edit Fehler: {edit_err}\n")
                
                # Slicer-Strip bei Delete/Reorder (in-memory)
                _fp_has_slicers = _fp_col_result.get('has_slicers', True)
                _fp_is_insert_only = bool(_fp_ins_cols) and not _fp_del_cols
                if _fp_has_slicers and not _fp_is_insert_only:
                    try:
                        _fp_zip_bytes = _strip_slicers_from_zip(output_path, source_bytes=_fp_zip_bytes, return_bytes=True)
                    except Exception as slicer_err:
                        _fp_zip_bytes.seek(0)
                        sys.stderr.write(f"[XML-DIREKT-COMBINED] WARNUNG: Slicer-Strip Fehler: {slicer_err}\n")
                
                # PivotTable-Strip (in-memory)
                if not _fp_is_insert_only:
                    try:
                        _fp_zip_bytes = _strip_pivot_tables_for_sheet(output_path, sheet_name, source_bytes=_fp_zip_bytes, return_bytes=True)
                    except Exception as pivot_err:
                        _fp_zip_bytes.seek(0)
                        sys.stderr.write(f"[XML-DIREKT-COMBINED] WARNUNG: PivotTable-Strip Fehler: {pivot_err}\n")
                
                # Row Highlights (in-memory)
                if _fp_row_highlights:
                    try:
                        _fp_zip_bytes = _apply_highlights_to_xlsx(output_path, sheet_name, _fp_row_highlights, source_bytes=_fp_zip_bytes, return_bytes=True)
                    except Exception as hl_err:
                        _fp_zip_bytes.seek(0)
                        sys.stderr.write(f"[XML-DIREKT-COMBINED] WARNUNG: row_highlights Fehler: {hl_err}\n")
                
                # Cleared Row Highlights (in-memory)
                if _fp_cleared_highlights:
                    try:
                        _fp_zip_bytes = _clear_row_highlights_xml(output_path, sheet_name, _fp_cleared_highlights, source_bytes=_fp_zip_bytes, return_bytes=True)
                    except Exception as cl_err:
                        _fp_zip_bytes.seek(0)
                        sys.stderr.write(f"[XML-DIREKT-COMBINED] WARNUNG: cleared_highlights Fehler: {cl_err}\n")
                
                # EINZIGER Disk-Write am Ende
                with open(output_path, 'wb') as _fp_out_f:
                    _fp_out_f.write(_fp_zip_bytes.getvalue())
                
                sys.stderr.write(f"[XML-DIREKT-COMBINED] Erfolgreich (in-memory pipeline)\n")
                # AutoFilter vom Frontend anwenden
                if _fp_auto_filter:
                    _apply_auto_filter_xml(output_path, sheet_name, _fp_auto_filter)
                # Backup aufräumen (wird für XML-DIREKT nicht gebraucht)
                if _backup_file is not None:
                    try:
                        os.remove(_backup_path)
                    except Exception:
                        pass
                return {'success': True, 'outputPath': output_path, 'method': 'xml-combined-fast'}
            
            except Exception as comb_err:
                sys.stderr.write(f"[XML-DIREKT-COMBINED] FEHLER: {comb_err}\n")
                import traceback
                traceback.print_exc(file=sys.stderr)
                return {'success': False, 'error': f'XML-DIREKT-COMBINED fehlgeschlagen: {comb_err}', 'method': 'xml-combined-fast'}
    
    # =========================================================================
    # FAST-PATH: Spalten-only Ops (Delete/Insert/Reorder/Hide) → XML-DIREKT
    # ohne openpyxl-Load. Spart den kompletten wb.load/save Roundtrip und
    # vermeidet openpyxl-bedingte Korruption (Namespace-Verlust, Slicer-Bruch).
    # Identisch zu v1.8.2 "XML-DIREKT-FAST" — prüft NICHT has_copy_styles,
    # da cellStyles/mergedCells für reine Spaltenoperationen irrelevant sind.
    # =========================================================================
    
    if _fp_has_any_col_ops and not _fp_from_file:
        _fp_has_cell_edits = bool(_fp_real_edits)
        _fp_has_row_ops = _fp_del_rows or _fp_ins_rows or (_fp_row_order and len(_fp_row_order) > 0)
        
        # Data Join Erkennung: Wenn ALLE Zell-Edits in eingefügten Spalten liegen,
        # sind sie bereits in data[] enthalten und werden von XML-DIREKT geschrieben.
        # → Dann sind die editedCells redundant und blockieren den Fast Path nicht.
        _fp_edits_in_inserted_only = False
        if _fp_has_cell_edits and _fp_ins_cols:
            _fp_ins_ops = _fp_ins_cols.get('operations', [])
            if not _fp_ins_ops and _fp_ins_cols.get('position') is not None:
                _fp_ins_ops = [{'position': _fp_ins_cols['position'], 'count': _fp_ins_cols.get('count', 1)}]
            if _fp_ins_ops:
                _fp_ins_col_set = set()
                for _fp_op in _fp_ins_ops:
                    _fp_pos = _fp_op['position']
                    _fp_cnt = _fp_op.get('count', 1)
                    for _fp_i in range(_fp_cnt):
                        _fp_ins_col_set.add(_fp_pos + _fp_i)
                _fp_edits_in_inserted_only = True
                for _fp_k in _fp_edited_cells:
                    if _fp_k.startswith('_'):
                        continue
                    _fp_parts = _fp_k.split('-')
                    if len(_fp_parts) >= 2:
                        try:
                            _fp_edit_col = int(_fp_parts[1])
                            if _fp_edit_col not in _fp_ins_col_set:
                                _fp_edits_in_inserted_only = False
                                break
                        except ValueError:
                            _fp_edits_in_inserted_only = False
                            break
                    else:
                        _fp_edits_in_inserted_only = False
                        break
        
        _fp_row_mapping_ok = True
        if _fp_row_mapping:
            for i, val in enumerate(_fp_row_mapping):
                if val != i:
                    _fp_row_mapping_ok = False
                    break
        
        _fp_col_eligible = (not _fp_has_row_ops and
                           _fp_row_mapping_ok)
                           # Cell-Edits: werden NACH Col-Ops per _direct_xml_cell_edit angewendet
                           # edits_in_inserted_only → schon in data[] enthalten, werden trotzdem harmlos nochmal applied
                           # affected_rows: nur für openpyxl Style-Reset nötig
        
        sys.stderr.write(f"[FAST-PATH-COL] eligible={_fp_col_eligible}, has_row_ops={_fp_has_row_ops}, "
                         f"cell_edits={len(_fp_real_edits)}, edits_in_inserted_only={_fp_edits_in_inserted_only}, "
                         f"row_mapping_ok={_fp_row_mapping_ok}\n")
        
        if _fp_col_eligible:
            sys.stderr.write(f"[XML-DIREKT-FAST] Verwende XML-Direkt-Weg OHNE openpyxl-Load\n")
            try:
                _fp_script_dir = os.path.dirname(os.path.abspath(__file__))
                if _fp_script_dir not in sys.path:
                    sys.path.insert(0, _fp_script_dir)
                from excel_xml_ops import direct_xml_column_operations
                import io as _io_col
                
                _fp_headers = changes.get('headers', [])
                _fp_data = changes.get('data', [])
                
                # Quelldatei EINMAL in Speicher lesen
                with open(file_path, 'rb') as _fp_src_f:
                    _fp_zip_bytes = _io_col.BytesIO(_fp_src_f.read())
                sys.stderr.write(f"[XML-DIREKT-FAST] Quelldatei in Speicher geladen\n")
                
                # Hidden Rows: NICHT strip/re-apply!
                # Der FAST-PATH arbeitet auf der ORIGINAL-Datei, die bereits
                # die korrekten hidden-Attribute hat. Strip+Re-Apply verursacht
                # "Zellinformationen"-Reparaturfehler in Excel.
                # Die originalen hidden-Attribute werden 1:1 durchgereicht.
                _fp_result = direct_xml_column_operations(
                    file_path=file_path,
                    output_path=output_path,
                    sheet_name=sheet_name,
                    deleted_columns=_fp_del_cols if _fp_del_cols else None,
                    inserted_columns=_fp_ins_cols,
                    column_order=_fp_col_order,
                    hidden_columns=_fp_hidden_cols,
                    headers=_fp_headers,
                    data=_fp_data,
                    strip_row_hidden=False,
                    hidden_rows=None,
                    source_bytes=_fp_zip_bytes,
                    return_bytes=True
                )
                _fp_zip_bytes = _fp_result['zip_bytes']
                
                # Slicer-Strip bei Reorder/Delete (in-memory)
                _fp_has_slicers = _fp_result.get('has_slicers', True)
                _fp_is_insert_only = bool(_fp_ins_cols) and not _fp_del_cols
                if _fp_has_slicers and not _fp_is_insert_only:
                    try:
                        _fp_zip_bytes = _strip_slicers_from_zip(output_path, source_bytes=_fp_zip_bytes, return_bytes=True)
                    except Exception as slicer_err:
                        _fp_zip_bytes.seek(0)
                        sys.stderr.write(f"[XML-DIREKT-FAST] WARNUNG: Slicer-Strip Fehler: {slicer_err}\n")
                else:
                    reason = "keine Slicers" if not _fp_has_slicers else "Insert-only (Slicers bleiben valide)"
                    sys.stderr.write(f"[XML-DIREKT-FAST] Slicer-Strip übersprungen: {reason}\n")
                
                # PivotTable-Strip bei Delete/Reorder (in-memory)
                if not _fp_is_insert_only:
                    try:
                        _fp_zip_bytes = _strip_pivot_tables_for_sheet(output_path, sheet_name, source_bytes=_fp_zip_bytes, return_bytes=True)
                    except Exception as pivot_err:
                        _fp_zip_bytes.seek(0)
                        sys.stderr.write(f"[XML-DIREKT-FAST] WARNUNG: PivotTable-Strip Fehler: {pivot_err}\n")
                
                # Hidden Rows: Originale Attribute aus Quelldatei durchgereicht
                if _fp_hidden_rows:
                    sys.stderr.write(f"[XML-DIREKT-FAST] {len(_fp_hidden_rows)} hidden rows aus Original-XML beibehalten (kein strip/re-apply)\n")
                
                # Cell-Edits via XML anwenden (in-memory, nach Col-Ops)
                if _fp_has_cell_edits and not _fp_edits_in_inserted_only:
                    sys.stderr.write(f"[XML-DIREKT-FAST] {len(_fp_real_edits)} Cell-Edits via XML anwenden\n")
                    try:
                        _fp_cell_result = _direct_xml_cell_edit(
                            output_path, output_path, sheet_name, _fp_real_edits,
                            hidden_columns=None,
                            hidden_rows=None,
                            source_bytes=_fp_zip_bytes,
                            return_bytes=True
                        )
                        _fp_zip_bytes = _fp_cell_result['zip_bytes']
                    except Exception as edit_err:
                        _fp_zip_bytes.seek(0)
                        sys.stderr.write(f"[XML-DIREKT-FAST] WARNUNG: Cell-Edit Fehler: {edit_err}\n")
                
                # Row Highlights (in-memory)
                if _fp_row_highlights:
                    sys.stderr.write(f"[XML-DIREKT-FAST] {len(_fp_row_highlights)} row_highlights via XML anwenden\n")
                    try:
                        _fp_zip_bytes = _apply_highlights_to_xlsx(output_path, sheet_name, _fp_row_highlights, source_bytes=_fp_zip_bytes, return_bytes=True)
                    except Exception as hl_err:
                        _fp_zip_bytes.seek(0)
                        sys.stderr.write(f"[XML-DIREKT-FAST] WARNUNG: row_highlights Fehler: {hl_err}\n")
                
                # Cleared Row Highlights (in-memory)
                if _fp_cleared_highlights:
                    try:
                        _fp_zip_bytes = _clear_row_highlights_xml(output_path, sheet_name, _fp_cleared_highlights, source_bytes=_fp_zip_bytes, return_bytes=True)
                    except Exception as cl_err:
                        _fp_zip_bytes.seek(0)
                        sys.stderr.write(f"[XML-DIREKT-FAST] WARNUNG: cleared_row_highlights Fehler: {cl_err}\n")
                
                # EINZIGER Disk-Write am Ende
                with open(output_path, 'wb') as _fp_out_f:
                    _fp_out_f.write(_fp_zip_bytes.getvalue())
                
                sys.stderr.write(f"[XML-DIREKT-FAST] Erfolgreich (in-memory pipeline): {_fp_result.get('method', 'unknown')}\n")
                # AutoFilter vom Frontend anwenden
                if _fp_auto_filter:
                    _apply_auto_filter_xml(output_path, sheet_name, _fp_auto_filter)
                # Backup aufräumen (wird für XML-DIREKT nicht gebraucht)
                if _backup_file is not None:
                    try:
                        os.remove(_backup_path)
                    except Exception:
                        pass
                return {'success': True, 'outputPath': output_path, 'method': _fp_result.get('method', 'xml-col-ops-fast')}
            
            except Exception as fp_err:
                sys.stderr.write(f"[XML-DIREKT-FAST] FEHLER: {fp_err}\n")
                import traceback
                traceback.print_exc(file=sys.stderr)
                return {'success': False, 'error': f'XML-DIREKT-FAST fehlgeschlagen: {fp_err}', 'method': 'xml-col-ops-fast'}
    
    # =========================================================================
    # FAST-PATH: Zeilen-only Ops (Delete/Reorder/Hide) → XML-DIREKT
    # ohne openpyxl-Load. Analog zum Spalten-FAST-PATH.
    # Spart den kompletten wb.load/save Roundtrip + Zelle-für-Zelle Backup/Restore.
    # =========================================================================
    _fp_has_any_row_ops = bool(_fp_del_rows) or bool(_fp_ins_rows) or (_fp_row_order and len(_fp_row_order) > 0)
    _fp_has_any_col_ops_check = (_fp_col_order and len(_fp_col_order) > 0) or bool(_fp_del_cols) or bool(_fp_ins_cols)
    
    if _fp_has_any_row_ops and not _fp_from_file and not _fp_has_any_col_ops_check:
        # Cell-Edits blockieren NICHT mehr den Fast-Path:
        # Nach den Row-Ops werden sie per _direct_xml_cell_edit angewendet.
        # Koordinaten passen 1:1: visuellerRowIdx + 2 = neue Excel-Zeile,
        # weil direct_xml_row_operations die überlebenden Zeilen konsekutiv ab 2 nummeriert.
        _fp_changed_cells = changes.get('changedCells', {})
        _fp_all_edits = dict(_fp_real_edits)
        if _fp_changed_cells:
            _fp_all_edits.update({k: v for k, v in _fp_changed_cells.items() if not k.startswith('_')})
        
        _fp_row_eligible = (
            True  # Row-Insert jetzt auch via XML-Direkt
            # Cell-Edits: werden NACH Row-Ops per _direct_xml_cell_edit angewendet
            # affectedRows: nur für openpyxl Style-Reset nötig, XML bewahrt Styles automatisch
            # row_mapping: direct_xml_row_operations nutzt deleted_rows/row_order, nicht rowMapping
        )
        
        sys.stderr.write(f"[FAST-PATH-ROW] eligible={_fp_row_eligible}, cell_edits={len(_fp_all_edits)}, "
                         f"ins_rows={bool(_fp_ins_rows)}, affected_rows={len(_fp_affected_rows)}\n")
        
        if _fp_row_eligible:
            sys.stderr.write(f"[XML-DIREKT-ROW] Verwende XML-Direkt-Weg für Zeilenoperationen\n")
            try:
                _fp_script_dir = os.path.dirname(os.path.abspath(__file__))
                if _fp_script_dir not in sys.path:
                    sys.path.insert(0, _fp_script_dir)
                from excel_xml_ops import direct_xml_row_operations
                import io as _io_row
                
                # Quelldatei EINMAL in Speicher lesen
                with open(file_path, 'rb') as _fp_src_f:
                    _fp_zip_bytes = _io_row.BytesIO(_fp_src_f.read())
                sys.stderr.write(f"[XML-DIREKT-ROW] Quelldatei in Speicher geladen\n")
                
                _fp_row_result = direct_xml_row_operations(
                    file_path=file_path,
                    output_path=output_path,
                    sheet_name=sheet_name,
                    deleted_rows=_fp_del_rows if _fp_del_rows else None,
                    row_order=_fp_row_order,
                    hidden_rows=_fp_hidden_rows if _fp_hidden_rows else None,
                    inserted_rows=_fp_ins_rows,
                    source_bytes=_fp_zip_bytes,
                    return_bytes=True
                )
                _fp_zip_bytes = _fp_row_result['zip_bytes']
                
                # Cell-Edits via XML anwenden (in-memory, nach Row-Ops)
                if _fp_all_edits:
                    sys.stderr.write(f"[XML-DIREKT-ROW] {len(_fp_all_edits)} Cell-Edits via XML anwenden\n")
                    try:
                        _fp_cell_result = _direct_xml_cell_edit(
                            output_path, output_path, sheet_name, _fp_all_edits,
                            hidden_columns=_fp_hidden_cols if _fp_hidden_cols else None,
                            hidden_rows=None,
                            source_bytes=_fp_zip_bytes,
                            return_bytes=True
                        )
                        _fp_zip_bytes = _fp_cell_result['zip_bytes']
                    except Exception as edit_err:
                        sys.stderr.write(f"[XML-DIREKT-ROW] FEHLER: Cell-Edit Fehler: {edit_err}\n")
                        import traceback
                        traceback.print_exc(file=sys.stderr)
                        return {'success': False, 'error': f'XML-DIREKT-ROW Cell-Edit fehlgeschlagen: {edit_err}', 'method': 'xml-row-ops-fast'}
                
                if True:  # Cell-Edits erfolgreich oder keine vorhanden
                    # Row Highlights (in-memory)
                    if _fp_row_highlights:
                        sys.stderr.write(f"[XML-DIREKT-ROW] {len(_fp_row_highlights)} row_highlights via XML anwenden\n")
                        try:
                            _fp_zip_bytes = _apply_highlights_to_xlsx(output_path, sheet_name, _fp_row_highlights, source_bytes=_fp_zip_bytes, return_bytes=True)
                        except Exception as hl_err:
                            _fp_zip_bytes.seek(0)
                            sys.stderr.write(f"[XML-DIREKT-ROW] WARNUNG: row_highlights Fehler: {hl_err}\n")
                    
                    # Cleared Row Highlights (in-memory)
                    if _fp_cleared_highlights:
                        try:
                            _fp_zip_bytes = _clear_row_highlights_xml(output_path, sheet_name, _fp_cleared_highlights, source_bytes=_fp_zip_bytes, return_bytes=True)
                        except Exception as cl_err:
                            _fp_zip_bytes.seek(0)
                            sys.stderr.write(f"[XML-DIREKT-ROW] WARNUNG: cleared_row_highlights Fehler: {cl_err}\n")
                    
                    # EINZIGER Disk-Write am Ende
                    with open(output_path, 'wb') as _fp_out_f:
                        _fp_out_f.write(_fp_zip_bytes.getvalue())
                    
                    sys.stderr.write(f"[XML-DIREKT-ROW] Erfolgreich (in-memory pipeline): {_fp_row_result.get('method', 'unknown')}\n")
                    # AutoFilter vom Frontend anwenden
                    if _fp_auto_filter:
                        _apply_auto_filter_xml(output_path, sheet_name, _fp_auto_filter)
                    # Backup aufräumen (wird für XML-DIREKT nicht gebraucht)
                    if _backup_file is not None:
                        try:
                            os.remove(_backup_path)
                        except Exception:
                            pass
                    return {'success': True, 'outputPath': output_path, 'method': _fp_row_result.get('method', 'xml-row-ops-fast')}
            
            except Exception as fp_row_err:
                sys.stderr.write(f"[XML-DIREKT-ROW] FEHLER: {fp_row_err}\n")
                import traceback
                traceback.print_exc(file=sys.stderr)
                return {'success': False, 'error': f'XML-DIREKT-ROW fehlgeschlagen: {fp_row_err}', 'method': 'xml-row-ops-fast'}
    
    try:
        # Original-Workbook laden
        # Workaround für openpyxl Bug mit extLst in PatternFill
        # rich_text=True damit CellRichText-Objekte erhalten bleiben
        _load_rich_text = True
        try:
            wb = load_workbook(file_path, rich_text=True)
        except TypeError as e:
            err_str = str(e)
            if 'extLst' in err_str:
                # openpyxl kann diese Datei nicht verarbeiten - Fallback-Fehler
                return {
                    'success': False, 
                    'error': f'Diese Datei enthält erweiterte Formatierungen die openpyxl nicht unterstützt. Bitte Excel/xlwings verwenden.',
                    'requiresXlwings': True
                }
            elif 'from_tree' in err_str or 'Nested' in err_str or 'node' in err_str:
                # Nested.from_tree Bug — Fallback: ohne rich_text laden
                sys.stderr.write(f"[LOAD] rich_text=True fehlgeschlagen ({e}), versuche ohne rich_text...\n")
                try:
                    wb = load_workbook(file_path, rich_text=False)
                    _load_rich_text = False
                    sys.stderr.write(f"[LOAD] Erfolgreich ohne rich_text geladen\n")
                except TypeError as e2:
                    err_str2 = str(e2)
                    if 'extLst' in err_str2:
                        return {
                            'success': False,
                            'error': f'Diese Datei enthält erweiterte Formatierungen die openpyxl nicht unterstützt. Bitte Excel/xlwings verwenden.',
                            'requiresXlwings': True
                        }
                    else:
                        raise
            else:
                raise
        
        if sheet_name not in wb.sheetnames:
            return {'success': False, 'error': f'Sheet "{sheet_name}" nicht gefunden'}
        
        ws = wb[sheet_name]
        
        # Parameter extrahieren
        headers = changes.get('headers', [])
        data = changes.get('data', [])
        edited_cells = changes.get('editedCells', {})
        cell_styles = changes.get('cellStyles', {})
        row_highlights = changes.get('rowHighlights', {})
        deleted_columns = changes.get('deletedColumns', [])
        inserted_columns = changes.get('insertedColumns')
        column_order = changes.get('columnOrder')  # [neuIdx] = altIdx
        hidden_columns = changes.get('hiddenColumns', [])
        hidden_rows = changes.get('hiddenRows', [])
        row_mapping = changes.get('rowMapping')
        from_file = changes.get('fromFile', False)
        full_rewrite = changes.get('fullRewrite', False)
        structural_change = changes.get('structuralChange', False)
        frontend_auto_filter = changes.get('autoFilterRange')  # AutoFilter vom Frontend
        
        cleared_row_highlights = changes.get('clearedRowHighlights', [])
        affected_rows = changes.get('affectedRows', [])
        
        # Zeilen-Operationen (analog zu Spalten-Operationen)
        deleted_rows = changes.get('deletedRowIndices', [])
        inserted_rows = changes.get('insertedRowInfo')
        row_order = changes.get('rowOrder')  # [neuIdx] = altIdx
        
        # DEBUG: Zeige alle relevanten Flags
        import sys
        sys.stderr.write(f"[WRITE_SHEET] row_highlights={row_highlights}, cleared_row_highlights={cleared_row_highlights}\n")
        sys.stderr.write(f"[WRITE_SHEET] row_mapping={bool(row_mapping)}, structural_change={structural_change}, full_rewrite={full_rewrite}\n")
        sys.stderr.write(f"[WRITE_SHEET] deleted_rows={deleted_rows}, inserted_rows={bool(inserted_rows)}, row_order={bool(row_order)}\n")
        sys.stderr.write(f"[WRITE_SHEET] deleted_columns={deleted_columns}, inserted_columns={bool(inserted_columns)}, column_order={bool(column_order)}\n")
        
        # DEBUG: mergedCells vom Frontend
        imported_mc_debug = changes.get('mergedCells', [])
        sys.stderr.write(f"[WRITE_SHEET] mergedCells count={len(imported_mc_debug)}\n")
        if imported_mc_debug:
            for i, mc in enumerate(imported_mc_debug[:5]):
                sys.stderr.write(f"[WRITE_SHEET] mergedCells[{i}]: startRow={mc.get('startRow')}, startCol={mc.get('startCol')}, endRow={mc.get('endRow')}, endCol={mc.get('endCol')}\n")
            if len(imported_mc_debug) > 5:
                sys.stderr.write(f"[WRITE_SHEET] ... und {len(imported_mc_debug) - 5} weitere\n")
        
        # =====================================================================
        # FALL 1: fromFile - Nur versteckte Spalten/Zeilen setzen
        # =====================================================================
        if from_file:
            _apply_hidden_columns(ws, hidden_columns)
            _apply_hidden_rows(ws, hidden_rows)
            wb.save(output_path)
            wb.close()
            fix_xlsx_relationships(output_path)
            # WICHTIG: Auch bei fromFile richData/Bilder/Namespaces wiederherstellen!
            # openpyxl verliert beim Speichern richData, metadata, vm-Attribute etc.
            restore_table_xml_from_original(output_path, original_path, table_changes=None)
            restore_external_links_from_original(output_path, original_path)
            # vm-Attribute für kopierte Bild-Zellen setzen
            if _fp_vm_cell_map:
                _apply_vm_cell_map_to_xlsx(output_path, sheet_name, _fp_vm_cell_map)
            # AutoFilter vom Frontend anwenden
            if frontend_auto_filter:
                _apply_auto_filter_xml(output_path, sheet_name, frontend_auto_filter)
            return {'success': True, 'outputPath': output_path}
        
        # =====================================================================
        # FALL 1.X: UNIVERSELLE PIPELINE für Spalten- UND Zeilen-Operationen
        # Führt alle Operationen STRIKT SEQUENTIELL aus:
        # 1-4. Zeilen-Operationen (alle Daten zuerst speichern, dann rekonstruieren)
        #      1. Alle Original-Zeilen speichern
        #      2. Finale Zeilen-Reihenfolge berechnen (Löschen + Verschieben)
        #      3. Überschüssige Zeilen entfernen
        #      4. Zeilen in neuer Reihenfolge schreiben
        # 5. Zeilen einfügen
        # 6. Zeilen verstecken (NACH allen strukturellen Änderungen)
        # 7. Spalten löschen (von hinten nach vorne)
        # 8. Spalten einfügen (von vorne nach hinten)
        # 9. Spalten verschieben/reorder
        # 10. Spalten verstecken
        # 11. Row Highlights
        # 12. Tables reparieren
        # 13. Einmal speichern
        # 14. XML restore
        # =====================================================================
        
        # Prüfe ob rowMapping nur die Identität ist (keine echte Änderung)
        row_mapping_is_identity = True
        if row_mapping:
            for i, val in enumerate(row_mapping):
                if val != i:
                    row_mapping_is_identity = False
                    break
        
        # Prüfe ob wir Zeilen-Operationen haben
        has_row_operations = deleted_rows or inserted_rows or (row_order and len(row_order) > 0)
        
        # Prüfe ob wir den Pipeline-Pfad nutzen können
        # (Spalten- ODER Zeilen-Operationen, aber NUR wenn rowMapping Identität ist)
        has_column_operations = deleted_columns or inserted_columns or (column_order and len(column_order) > 0)
        can_use_pipeline = (has_column_operations or has_row_operations) and row_mapping_is_identity and not affected_rows
        
        if can_use_pipeline:
            # =====================================================================
            # XML-DIREKT-WEG: Spaltenoperationen ohne openpyxl-Roundtrip
            # Wenn NUR Spaltenoperationen (+ hidden) aber KEINE Zeilenoperationen,
            # verwende den direkten XML-Weg → kein Slicer/Drawing/Namespace-Verlust
            # =====================================================================
            import sys
            
            # Prüfe ob echte Zell-Edits vorhanden sind (Keys ohne _ Prefix)
            has_cell_edits = bool(edited_cells and any(not k.startswith('_') for k in edited_cells))
            has_copy_styles = bool(changes.get('cellStyles')) or bool(changes.get('cellFonts')) or bool(changes.get('richTextCells')) or bool(changes.get('mergedCells', []))
            
            only_column_ops = has_column_operations and not has_row_operations and not inserted_rows and not has_cell_edits and not has_copy_styles
            
            if only_column_ops:
                sys.stderr.write(f"[XML-DIREKT] Verwende XML-Direkt-Weg für Spaltenoperationen\n")
                sys.stderr.write(f"[XML-DIREKT] deleted={deleted_columns}, inserted={inserted_columns is not None}, "
                                 f"reorder={column_order is not None}, hidden={hidden_columns}\n")
                
                try:
                    wb.close()  # openpyxl-Workbook schließen — wir brauchen es nicht
                    
                    from excel_xml_ops import direct_xml_column_operations
                    import io as _io_fallback
                    
                    # Quelldatei EINMAL in Speicher lesen
                    with open(file_path, 'rb') as _fb_src_f:
                        _fb_zip_bytes = _io_fallback.BytesIO(_fb_src_f.read())
                    
                    result = direct_xml_column_operations(
                        file_path=file_path,
                        output_path=output_path,
                        sheet_name=sheet_name,
                        deleted_columns=deleted_columns,
                        inserted_columns=inserted_columns,
                        column_order=column_order,
                        hidden_columns=hidden_columns,
                        headers=headers,
                        data=data,
                        source_bytes=_fb_zip_bytes,
                        return_bytes=True
                    )
                    _fb_zip_bytes = result['zip_bytes']
                    
                    # Slicer-Infrastruktur entfernen (in-memory)
                    has_slicers = result.get('has_slicers', True)
                    is_insert_only = bool(inserted_columns) and not deleted_columns
                    if has_slicers and not is_insert_only:
                        try:
                            _fb_zip_bytes = _strip_slicers_from_zip(output_path, source_bytes=_fb_zip_bytes, return_bytes=True)
                        except Exception as slicer_err:
                            _fb_zip_bytes.seek(0)
                            sys.stderr.write(f"[XML-DIREKT] WARNUNG: Slicer-Strip Fehler: {slicer_err}\n")
                    else:
                        reason = "keine Slicers" if not has_slicers else "Insert-only (Slicers bleiben valide)"
                        sys.stderr.write(f"[XML-DIREKT] Slicer-Strip übersprungen: {reason}\n")
                    
                    # PivotTable-Strip (in-memory)
                    if not is_insert_only:
                        try:
                            _fb_zip_bytes = _strip_pivot_tables_for_sheet(output_path, sheet_name, source_bytes=_fb_zip_bytes, return_bytes=True)
                        except Exception as pivot_err:
                            _fb_zip_bytes.seek(0)
                            sys.stderr.write(f"[XML-DIREKT] WARNUNG: PivotTable-Strip Fehler: {pivot_err}\n")
                    
                    # Row Highlights (in-memory)
                    if row_highlights:
                        sys.stderr.write(f"[XML-DIREKT] {len(row_highlights)} row_highlights via XML anwenden\n")
                        try:
                            _fb_zip_bytes = _apply_highlights_to_xlsx(output_path, sheet_name, row_highlights, source_bytes=_fb_zip_bytes, return_bytes=True)
                        except Exception as hl_err:
                            _fb_zip_bytes.seek(0)
                            sys.stderr.write(f"[XML-DIREKT] WARNUNG: row_highlights Fehler: {hl_err}\n")
                    
                    # Hidden Rows (in-memory)
                    if hidden_rows:
                        sys.stderr.write(f"[XML-DIREKT] {len(hidden_rows)} hidden rows via XML anwenden\n")
                        try:
                            _fb_zip_bytes = _apply_hidden_rows_to_xlsx(output_path, sheet_name, hidden_rows, source_bytes=_fb_zip_bytes, return_bytes=True)
                        except Exception as hr_err:
                            _fb_zip_bytes.seek(0)
                            sys.stderr.write(f"[XML-DIREKT] WARNUNG: hidden rows Fehler: {hr_err}\n")
                    
                    # Cleared Row Highlights (in-memory)
                    if cleared_row_highlights:
                        try:
                            _fb_zip_bytes = _clear_row_highlights_xml(output_path, sheet_name, cleared_row_highlights, source_bytes=_fb_zip_bytes, return_bytes=True)
                        except Exception as cl_err:
                            _fb_zip_bytes.seek(0)
                            sys.stderr.write(f"[XML-DIREKT] WARNUNG: cleared_row_highlights Fehler: {cl_err}\n")
                    
                    # EINZIGER Disk-Write am Ende
                    with open(output_path, 'wb') as _fb_out_f:
                        _fb_out_f.write(_fb_zip_bytes.getvalue())
                    
                    sys.stderr.write(f"[XML-DIREKT] Erfolgreich (in-memory pipeline): {result.get('method', 'unknown')}\n")
                    # Backup aufräumen (wird für XML-DIREKT nicht gebraucht)
                    if _backup_file is not None:
                        try:
                            os.remove(_backup_path)
                        except Exception:
                            pass
                    return {'success': True, 'outputPath': output_path, 'method': result.get('method', 'xml-col-ops')}
                
                except Exception as xml_err:
                    sys.stderr.write(f"[XML-DIREKT] Fehler: {xml_err} — Fallback auf openpyxl-Pipeline\n")
                    import traceback
                    traceback.print_exc(file=sys.stderr)
                    # Bei Fehler: Workbook neu öffnen und normalen Pipeline-Pfad nehmen
                    wb = _safe_load_workbook(original_path, rich_text=False)
                    ws = wb[sheet_name]
            
            # =====================================================================
            # OPENPYXL-PIPELINE (Fallback, oder wenn Zeilenoperationen dabei sind)
            # =====================================================================
            from openpyxl.worksheet.table import TableColumn
            from openpyxl.utils.cell import range_boundaries
            from openpyxl.cell.cell import MergedCell
            
            sys.stderr.write(f"[PIPELINE] Starte: deleted_rows={deleted_rows}, row_order={row_order is not None}, hidden_rows={hidden_rows}, deleted_columns={deleted_columns}, inserted_columns={inserted_columns is not None}, column_order={column_order is not None}\n")
            
            # =====================================================================
            # ZEILEN-OPERATIONEN: Alle Daten ZUERST speichern, dann rekonstruieren
            # =====================================================================
            
            has_any_row_change = deleted_rows or (row_order and len(row_order) > 0)
            
            if has_any_row_change:
                max_col = ws.max_column
                original_max_row = ws.max_row
                
                deleted_set = set(deleted_rows) if deleted_rows else set()
                is_pure_delete = bool(deleted_set) and not (row_order and len(row_order) > 0)
                
                # SCHRITT 1: Original-Zeilen speichern (vor jeder Änderung!)
                # OPTIMIERUNG: Bei reinem Löschen (Filter) nur die behaltenen Zeilen sichern
                if is_pure_delete:
                    keep_set = set(range(original_max_row - 1)) - deleted_set
                    sys.stderr.write(f"[PIPELINE] Schritt 1: Speichere {len(keep_set)} von {original_max_row - 1} Zeilen (nur behaltene)\n")
                else:
                    keep_set = None  # Alle sichern
                    sys.stderr.write(f"[PIPELINE] Schritt 1: Speichere alle {original_max_row - 1} Original-Zeilen\n")
                
                all_rows_backup = {}
                row_heights_backup = {}
                for excel_row in range(2, original_max_row + 1):  # Ab Zeile 2 (nach Header)
                    row_idx = excel_row - 2  # 0-basierter Index
                    
                    # Bei reinem Löschen: Überspringe Zeilen die gelöscht werden
                    if keep_set is not None and row_idx not in keep_set:
                        continue
                    all_rows_backup[row_idx] = {}
                    
                    # Zeilenhöhe sichern
                    if excel_row in ws.row_dimensions:
                        rd = ws.row_dimensions[excel_row]
                        if rd.height is not None:
                            row_heights_backup[row_idx] = rd.height
                    
                    for col in range(1, max_col + 1):
                        cell = ws.cell(row=excel_row, column=col)
                        if isinstance(cell, MergedCell):
                            continue
                        all_rows_backup[row_idx][col] = {
                            'value': cell.value,
                            'fill': copy(cell.fill),
                            'font': copy(cell.font),
                            'alignment': copy(cell.alignment),
                            'border': copy(cell.border),
                            'number_format': cell.number_format,
                            'hyperlink': cell.hyperlink.target if cell.hyperlink else None
                        }
                
                # SCHRITT 2: Bestimme finale Zeilen-Reihenfolge
                # row_order enthält: [neuIdx] = altIdx (nach Löschen!)
                # deleted_rows enthält: Original-Indizes der gelöschten Zeilen
                
                if row_order and len(row_order) > 0:
                    # row_order gibt die neue Reihenfolge vor
                    # Die Indizes in row_order beziehen sich auf Zeilen NACH dem Löschen
                    # Wir müssen sie zurück auf Original-Indizes mappen
                    
                    # Erstelle Mapping: Index nach Löschen → Original-Index
                    remaining_original_indices = []
                    for orig_idx in range(original_max_row - 1):
                        if orig_idx not in deleted_set:
                            remaining_original_indices.append(orig_idx)
                    
                    # row_order[new_pos] = after_delete_idx → wir brauchen original_idx
                    final_row_order = []
                    for new_pos, after_delete_idx in enumerate(row_order):
                        if after_delete_idx < len(remaining_original_indices):
                            original_idx = remaining_original_indices[after_delete_idx]
                            final_row_order.append(original_idx)
                    
                    sys.stderr.write(f"[PIPELINE] Schritt 2: Finale Zeilen-Reihenfolge (Original-Indizes): {final_row_order[:10]}...\n")
                else:
                    # Keine Verschiebung, nur Löschen - behalte Reihenfolge der nicht-gelöschten
                    final_row_order = sorted(all_rows_backup.keys())
                    sys.stderr.write(f"[PIPELINE] Schritt 2: Nur Löschen, behalte {len(final_row_order)} Zeilen\n")
                
                # SCHRITT 3: Überschüssige Zeilen löschen (von hinten)
                target_row_count = len(final_row_order)
                current_data_rows = original_max_row - 1  # Ohne Header
                
                if current_data_rows > target_row_count:
                    rows_to_delete = current_data_rows - target_row_count
                    sys.stderr.write(f"[PIPELINE] Schritt 3: Lösche {rows_to_delete} überschüssige Zeilen (bulk)\n")
                    # BULK-Löschung: Alle Zeilen in EINEM Aufruf löschen (statt einzeln)
                    # Einzelne delete_rows() Aufrufe sind extrem langsam bei openpyxl
                    first_delete_row = target_row_count + 2  # +2 wegen Header
                    ws.delete_rows(first_delete_row, rows_to_delete)
                
                # SCHRITT 4: Zeilen in neuer Reihenfolge schreiben
                sys.stderr.write(f"[PIPELINE] Schritt 4: Schreibe {len(final_row_order)} Zeilen in neuer Reihenfolge\n")
                for new_idx, original_idx in enumerate(final_row_order):
                    new_excel_row = new_idx + 2
                    
                    if original_idx not in all_rows_backup:
                        continue
                    
                    for col, data_item in all_rows_backup[original_idx].items():
                        cell = ws.cell(row=new_excel_row, column=col)
                        if isinstance(cell, MergedCell):
                            continue
                        cell.value = data_item['value']
                        cell.fill = data_item['fill']
                        cell.font = data_item['font']
                        cell.alignment = data_item['alignment']
                        cell.border = data_item['border']
                        cell.number_format = data_item['number_format'] or 'General'
                        if data_item['hyperlink']:
                            cell.hyperlink = data_item['hyperlink']
                        elif cell.hyperlink:
                            cell.hyperlink = None
                    
                    # Zeilenhöhe wiederherstellen
                    if original_idx in row_heights_backup:
                        ws.row_dimensions[new_excel_row].height = row_heights_backup[original_idx]
                    elif new_excel_row in ws.row_dimensions:
                        ws.row_dimensions[new_excel_row].height = None
            
            # ===== SCHRITT 5: Zeilen EINFÜGEN =====
            if inserted_rows:
                operations = inserted_rows.get('operations', [])
                operations.sort(key=lambda x: x['position'])
                sys.stderr.write(f"[PIPELINE] Schritt 5: Füge Zeilen ein {[op['position'] for op in operations]}\n")
                
                for op in operations:
                    position = op['position']
                    count = op.get('count', 1)
                    excel_row = position + 2
                    
                    for i in range(count):
                        ws.insert_rows(excel_row + i, 1)
                        
                        # Formatierung von Zeile darüber kopieren
                        if excel_row + i > 2:
                            source_row = excel_row + i - 1
                            for col in range(1, ws.max_column + 1):
                                source_cell = ws.cell(row=source_row, column=col)
                                target_cell = ws.cell(row=excel_row + i, column=col)
                                if source_cell.fill:
                                    target_cell.fill = copy(source_cell.fill)
                                if source_cell.font:
                                    target_cell.font = copy(source_cell.font)
                                if source_cell.alignment:
                                    target_cell.alignment = copy(source_cell.alignment)
                                if source_cell.border:
                                    target_cell.border = copy(source_cell.border)
                                if source_cell.number_format:
                                    target_cell.number_format = source_cell.number_format
            
            # ===== SCHRITT 6: Zeilen VERSTECKEN (NACH allen strukturellen Änderungen) =====
            sys.stderr.write(f"[PIPELINE] Schritt 6: Zeilen verstecken, hidden_rows={hidden_rows}\n")
            _apply_hidden_rows(ws, hidden_rows)
            
            # ===== SCHRITT 7: Spalten LÖSCHEN (von hinten nach vorne) =====
            if deleted_columns:
                sorted_deleted = sorted(deleted_columns, reverse=True)
                sys.stderr.write(f"[PIPELINE] Schritt 7: Lösche Spalten {sorted_deleted}\n")
                
                for col_idx in sorted_deleted:
                    excel_col = col_idx + 1
                    max_col = ws.max_column
                    
                    # Spaltenbreiten speichern
                    saved_widths = {}
                    for col in range(excel_col + 1, max_col + 1):
                        col_letter = get_column_letter(col)
                        if col_letter in ws.column_dimensions:
                            saved_widths[col] = ws.column_dimensions[col_letter].width
                    
                    # Spalte löschen
                    ws.delete_cols(excel_col, 1)
                    
                    # Spaltenbreiten wiederherstellen
                    for old_col, width in saved_widths.items():
                        if width:
                            new_letter = get_column_letter(old_col - 1)
                            ws.column_dimensions[new_letter].width = width
                    
                    # CF anpassen
                    adjust_conditional_formatting(ws, [col_idx], None)
            
            # ===== SCHRITT 8: Spalten EINFÜGEN (von vorne nach hinten) =====
            if inserted_columns:
                operations = inserted_columns.get('operations', [])
                if not operations and inserted_columns.get('position') is not None:
                    operations = [{
                        'position': inserted_columns['position'],
                        'count': inserted_columns.get('count', 1),
                        'sourceColumn': inserted_columns.get('sourceColumn')
                    }]
                
                operations.sort(key=lambda x: x['position'])
                sys.stderr.write(f"[PIPELINE] Schritt 8: Füge Spalten ein\n")
                
                for op_idx, op in enumerate(operations):
                    position = op['position']
                    count = op.get('count', 1)
                    source_column = op.get('sourceColumn')
                    excel_col = position + 1
                    
                    for i in range(count):
                        insert_at = excel_col + i
                        
                        # Formatierung der Referenzspalte speichern
                        source_format = {}
                        source_width = None
                        if source_column is not None:
                            source_excel_col = source_column + 1
                            for prev_op in operations[:op_idx]:
                                if source_column >= prev_op['position']:
                                    source_excel_col += prev_op.get('count', 1)
                            
                            col_letter = get_column_letter(source_excel_col)
                            if col_letter in ws.column_dimensions:
                                source_width = ws.column_dimensions[col_letter].width
                            
                            for row in range(1, ws.max_row + 1):
                                cell = ws.cell(row=row, column=source_excel_col)
                                source_format[row] = {
                                    'fill': copy(cell.fill) if cell.fill else None,
                                    'font': copy(cell.font) if cell.font else None,
                                    'alignment': copy(cell.alignment) if cell.alignment else None,
                                    'border': copy(cell.border) if cell.border else None,
                                    'number_format': cell.number_format
                                }
                        
                        # Spaltenbreiten speichern
                        saved_widths = {}
                        for col in range(insert_at, ws.max_column + 1):
                            col_letter = get_column_letter(col)
                            if col_letter in ws.column_dimensions:
                                saved_widths[col] = ws.column_dimensions[col_letter].width
                        
                        # Spalte einfügen
                        ws.insert_cols(insert_at, 1)
                        
                        # Spaltenbreiten wiederherstellen
                        for old_col, width in saved_widths.items():
                            if width:
                                new_letter = get_column_letter(old_col + 1)
                                ws.column_dimensions[new_letter].width = width
                        
                        # CF anpassen
                        inserted_cols_for_cf = {insert_at - 1: 1}
                        adjust_conditional_formatting(ws, [], inserted_cols_for_cf)
                        
                        # Formatierung anwenden
                        if source_width:
                            ws.column_dimensions[get_column_letter(insert_at)].width = source_width
                        
                        for row, fmt in source_format.items():
                            cell = ws.cell(row=row, column=insert_at)
                            if fmt['fill']:
                                cell.fill = fmt['fill']
                            if fmt['font']:
                                cell.font = fmt['font']
                            if fmt['alignment']:
                                cell.alignment = fmt['alignment']
                            if fmt['border']:
                                cell.border = fmt['border']
                            if fmt.get('number_format'):
                                cell.number_format = fmt['number_format']
                    
                    # Header setzen
                    op_headers = op.get('headers', [])
                    for i, header in enumerate(op_headers):
                        ws.cell(row=1, column=excel_col + i).value = header
                    
                    # Daten schreiben
                    if data and headers:
                        for i in range(count):
                            col_idx = position + i
                            if col_idx < len(headers):
                                for row_idx, row_data in enumerate(data):
                                    if col_idx < len(row_data):
                                        cell = ws.cell(row=row_idx + 2, column=excel_col + i)
                                        apply_cell_value(cell, row_data[col_idx])
            
            # ===== SCHRITT 9: Spalten VERSCHIEBEN/REORDER =====
            sys.stderr.write(f"[PIPELINE] Schritt 9: Spalten verschieben\n")
            if column_order and len(column_order) > 0:
                columns_changed = False
                for new_idx, old_idx in enumerate(column_order):
                    if new_idx != old_idx:
                        columns_changed = True
                        break
                
                if columns_changed:
                    num_cols = len(column_order)
                    max_row = ws.max_row
                    
                    # Spaltenbreiten sichern
                    col_widths_backup = {}
                    for col_idx in range(num_cols):
                        col_letter = get_column_letter(col_idx + 1)
                        if col_letter in ws.column_dimensions:
                            dim = ws.column_dimensions[col_letter]
                            col_widths_backup[col_idx] = {
                                'width': dim.width,
                                'hidden': dim.hidden,
                                'bestFit': dim.bestFit,
                                'customWidth': dim.customWidth
                            }
                    
                    # Alle Spalten in temp_columns speichern (inkl. Formatierung!)
                    temp_columns = {}
                    for old_col_idx in range(num_cols):
                        old_excel_col = old_col_idx + 1
                        temp_columns[old_col_idx] = {}
                        
                        for row in range(1, max_row + 1):
                            cell = ws.cell(row=row, column=old_excel_col)
                            if isinstance(cell, MergedCell):
                                continue
                            temp_columns[old_col_idx][row] = {
                                'value': cell.value,
                                'fill': copy(cell.fill),
                                'font': copy(cell.font),
                                'alignment': copy(cell.alignment),
                                'border': copy(cell.border),
                                'number_format': cell.number_format,
                                'hyperlink': cell.hyperlink.target if cell.hyperlink else None,
                            }
                    
                    # Spalten in neuer Reihenfolge schreiben
                    for new_col_idx, old_col_idx in enumerate(column_order):
                        new_excel_col = new_col_idx + 1
                        
                        if old_col_idx not in temp_columns:
                            continue
                        
                        for row, data_item in temp_columns[old_col_idx].items():
                            cell = ws.cell(row=row, column=new_excel_col)
                            if isinstance(cell, MergedCell):
                                continue
                            cell.value = data_item['value']
                            cell.fill = data_item['fill']
                            cell.font = data_item['font']
                            cell.alignment = data_item['alignment']
                            cell.border = data_item['border']
                            cell.number_format = data_item['number_format'] or 'General'
                            if data_item['hyperlink']:
                                cell.hyperlink = data_item['hyperlink']
                            elif cell.hyperlink:
                                cell.hyperlink = None
                    
                    # Spaltenbreiten in neuer Reihenfolge wiederherstellen
                    for new_col_idx, old_col_idx in enumerate(column_order):
                        new_col_letter = get_column_letter(new_col_idx + 1)
                        if old_col_idx in col_widths_backup:
                            wb_data = col_widths_backup[old_col_idx]
                            ws.column_dimensions[new_col_letter].width = wb_data['width']
                            if wb_data.get('hidden'):
                                ws.column_dimensions[new_col_letter].hidden = wb_data['hidden']
                            if wb_data.get('bestFit'):
                                ws.column_dimensions[new_col_letter].bestFit = wb_data['bestFit']
                            if wb_data.get('customWidth'):
                                ws.column_dimensions[new_col_letter].customWidth = wb_data['customWidth']
            
            # ===== SCHRITT 10: Versteckte Spalten =====
            sys.stderr.write(f"[PIPELINE] Schritt 10: Spalten verstecken\n")
            _apply_hidden_columns(ws, hidden_columns)
            
            # ===== SCHRITT 11: Row Highlights =====
            sys.stderr.write(f"[PIPELINE] Schritt 11: Row Highlights\n")
            if row_highlights:
                _apply_row_highlights(ws, row_highlights, len(headers) if headers else 0)
            
            # ===== SCHRITT 11b: Zell-Edits und Copy-Paste-Daten anwenden =====
            # Bei kombinierten Operationen (Zeilen+Spalten+Edits) werden editedCells,
            # cellStyles, cellFonts, richTextCells und mergedCells hier angewendet.
            # Der Pipeline-Pfad wurde bisher ohne Edits verlassen — sie gingen verloren.
            real_edits = {k: v for k, v in edited_cells.items() if not k.startswith('_')} if edited_cells else {}
            if real_edits:
                sys.stderr.write(f"[PIPELINE] Schritt 11b: {len(real_edits)} Zell-Edits anwenden\n")
                for key, value in real_edits.items():
                    parts = key.split('-')
                    if len(parts) != 2:
                        continue
                    row_idx = int(parts[0])
                    col_idx = int(parts[1])
                    cell = ws.cell(row=row_idx + 2, column=col_idx + 1)
                    apply_cell_value(cell, value)
            
            imported_cell_styles_p = changes.get('cellStyles', {})
            if imported_cell_styles_p:
                sys.stderr.write(f"[PIPELINE] Schritt 11b: {len(imported_cell_styles_p)} cellStyles anwenden\n")
                _apply_imported_cell_styles(ws, imported_cell_styles_p)
            
            cell_fonts_p = changes.get('cellFonts', {})
            if cell_fonts_p:
                sys.stderr.write(f"[PIPELINE] Schritt 11b: {len(cell_fonts_p)} cellFonts anwenden\n")
                _apply_cell_fonts(ws, cell_fonts_p)
            
            imported_rich_text_p = changes.get('richTextCells', {})
            if imported_rich_text_p:
                sys.stderr.write(f"[PIPELINE] Schritt 11b: {len(imported_rich_text_p)} richTextCells anwenden\n")
                _apply_imported_rich_text(ws, imported_rich_text_p)
            
            imported_merged_cells_p = changes.get('mergedCells', [])
            if imported_merged_cells_p:
                sys.stderr.write(f"[PIPELINE] Schritt 11b: {len(imported_merged_cells_p)} mergedCells anwenden\n")
                _apply_imported_merged_cells(ws, imported_merged_cells_p)
            
            # ===== SCHRITT 12: Tables reparieren =====
            sys.stderr.write(f"[PIPELINE] Schritt 12: Tables reparieren\n")
            new_max_row = ws.max_row  # Aktuelle max_row NACH Zeilen-Löschung
            table_changes = {}
            for table_name in ws.tables:
                table = ws.tables[table_name]
                min_col, min_row, max_col, max_row = range_boundaries(table.ref)
                
                new_max_col = ws.max_column
                # WICHTIG: max_row auf aktuelle Zeilenanzahl anpassen (nach Zeilen-Löschung)
                adjusted_max_row = min(max_row, new_max_row) if has_any_row_change else max_row
                new_ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(new_max_col)}{adjusted_max_row}"
                table.ref = new_ref
                if table.autoFilter:
                    table.autoFilter.ref = new_ref
                sys.stderr.write(f"[PIPELINE] Table '{table_name}': ref={new_ref} (max_row: {max_row} -> {adjusted_max_row})\n")
                
                # tableColumns aus Header-Zellen neu aufbauen
                new_columns = []
                for col_idx in range(min_col, new_max_col + 1):
                    header_cell = ws.cell(row=min_row, column=col_idx)
                    col_name = str(header_cell.value) if header_cell.value else f"Column{col_idx}"
                    new_columns.append(TableColumn(id=col_idx - min_col + 1, name=col_name))
                
                table.tableColumns = new_columns
                table_changes[table_name] = {'ref': table.ref, 'columns': [col.name for col in new_columns]}
            
            # Sheet-level AutoFilter anpassen (nach Zeilen-Löschung)
            if has_any_row_change and ws.auto_filter and ws.auto_filter.ref:
                af_min_col, af_min_row, af_max_col, af_max_row = range_boundaries(ws.auto_filter.ref)
                adjusted_af_max_row = min(af_max_row, new_max_row)
                new_af_ref = f"{get_column_letter(af_min_col)}{af_min_row}:{get_column_letter(af_max_col)}{adjusted_af_max_row}"
                ws.auto_filter.ref = new_af_ref
                sys.stderr.write(f"[PIPELINE] AutoFilter: {ws.auto_filter.ref} (max_row: {af_max_row} -> {adjusted_af_max_row})\n")
            
            # ===== SCHRITT 13: EINMAL speichern =====
            sys.stderr.write(f"[PIPELINE] Schritt 13: Speichern\n")
            wb.save(output_path)
            wb.close()
            _check_zip_drawings(output_path, "nach wb.save()")
            fix_xlsx_relationships(output_path)
            _check_zip_drawings(output_path, "nach fix_xlsx_relationships()")
            
            # ===== SCHRITT 14: XML restore =====
            sys.stderr.write(f"[PIPELINE] Schritt 14: XML restore\n")
            if table_changes:
                restore_table_xml_from_original(output_path, original_path, table_changes)
                _check_zip_drawings(output_path, "nach restore_table_xml()")
            
            restore_external_links_from_original(output_path, original_path, structural_change=True)
            # vm-Attribute für kopierte Bild-Zellen setzen
            if _fp_vm_cell_map:
                _apply_vm_cell_map_to_xlsx(output_path, sheet_name, _fp_vm_cell_map)
            _check_zip_drawings(output_path, "nach restore_external_links()")
            
            # Sicherheitsnetz: Slicer-Infrastruktur entfernen falls restore_external_links
            # trotz structural_change noch Artefakte hinterlassen hat
            try:
                _strip_slicers_from_zip(output_path)
            except Exception as slicer_err:
                sys.stderr.write(f"[PIPELINE] WARNUNG: Slicer-Strip Fehler: {slicer_err}\n")
            
            # PivotTables entfernen falls Spalten gelöscht wurden
            # MUSS NACH restore_external_links laufen (sonst werden Pivot-Dateien vom Original re-kopiert)
            if deleted_columns:
                try:
                    _strip_pivot_tables_for_sheet(output_path, sheet_name)
                except Exception as pivot_err:
                    sys.stderr.write(f"[PIPELINE] WARNUNG: PivotTable-Strip Fehler: {pivot_err}\n")
            
            # Cleared Row Highlights per XML (falls im Frontend Highlights entfernt wurden)
            if cleared_row_highlights:
                sys.stderr.write(f"[PIPELINE] {len(cleared_row_highlights)} cleared_row_highlights via XML entfernen\n")
                try:
                    _clear_row_highlights_xml(output_path, sheet_name, cleared_row_highlights)
                except Exception as cl_err:
                    sys.stderr.write(f"[PIPELINE] WARNUNG: cleared_row_highlights Fehler: {cl_err}\n")
            
            # AutoFilter vom Frontend anwenden
            if frontend_auto_filter:
                _apply_auto_filter_xml(output_path, sheet_name, frontend_auto_filter)
            
            return {'success': True, 'outputPath': output_path, 'method': 'openpyxl-pipeline'}
        
        # =====================================================================
        # LEGACY FALLBACK: Alte Einzel-FÄLLe für Kompatibilität
        # (werden nur noch erreicht wenn can_use_pipeline = False)
        # =====================================================================
        
        # LEGACY: Bei Spalten-Insert IMMER FALL 1.5 verwenden!
        # WICHTIG: NUR wenn keine Zeilen-Operationen - sonst FALL 2 (ZIP-ANSATZ + XML-DIREKT)
        only_column_insert = inserted_columns and not deleted_columns and not has_row_operations
        
        if only_column_insert:
            
            operations = inserted_columns.get('operations', [])
            if not operations and inserted_columns.get('position') is not None:
                operations = [{
                    'position': inserted_columns['position'],
                    'count': inserted_columns.get('count', 1),
                    'sourceColumn': inserted_columns.get('sourceColumn')
                }]
            
            # Sortiere aufsteigend - so kompensiert jede Einfügung die nächste automatisch
            operations.sort(key=lambda x: x['position'])
            
            from openpyxl.worksheet.table import TableColumn
            from openpyxl.utils.cell import range_boundaries
            
            # Alle Operationen im Speicher durchführen
            # Die Positionen vom Frontend sind die FINALEN Positionen (nach allen Einfügungen)
            # Wenn wir aufsteigend einfügen, brauchen wir keinen Offset!
            
            for op_idx, op in enumerate(operations):
                position = op['position']
                count = op.get('count', 1)
                source_column = op.get('sourceColumn')
                excel_col = position + 1  # 0-basiert → 1-basiert, KEIN Offset nötig!
                
                
                for i in range(count):
                    insert_at = excel_col + i
                    
                    # Speichere Formatierung der Referenzspalte (im aktuellen Zustand des Worksheets)
                    source_format = {}
                    source_width = None
                    if source_column is not None:
                        # source_column muss auch im aktuellen Worksheet-Zustand gefunden werden
                        # Nach vorherigen Einfügungen könnte die Position verschoben sein
                        source_excel_col = source_column + 1
                        # Korrigiere für bereits eingefügte Spalten
                        for prev_op in operations[:op_idx]:
                            if source_column >= prev_op['position']:
                                source_excel_col += prev_op.get('count', 1)
                        
                        col_letter = get_column_letter(source_excel_col)
                        if col_letter in ws.column_dimensions:
                            source_width = ws.column_dimensions[col_letter].width
                        
                        for row in range(1, ws.max_row + 1):
                            cell = ws.cell(row=row, column=source_excel_col)
                            source_format[row] = {
                                'fill': copy(cell.fill) if cell.fill else None,
                                'font': copy(cell.font) if cell.font else None,
                                'alignment': copy(cell.alignment) if cell.alignment else None,
                                'border': copy(cell.border) if cell.border else None,
                                'number_format': cell.number_format
                            }
                    
                    # Spaltenbreiten speichern
                    saved_widths = {}
                    for col in range(insert_at, ws.max_column + 1):
                        col_letter = get_column_letter(col)
                        if col_letter in ws.column_dimensions:
                            saved_widths[col] = ws.column_dimensions[col_letter].width
                    
                    # Spalte einfügen
                    ws.insert_cols(insert_at, 1)
                    
                    # Spaltenbreiten wiederherstellen
                    for old_col, width in saved_widths.items():
                        if width:
                            new_letter = get_column_letter(old_col + 1)
                            ws.column_dimensions[new_letter].width = width
                    
                    # CF anpassen
                    inserted_cols_for_cf = {insert_at - 1: 1}
                    adjust_conditional_formatting(ws, [], inserted_cols_for_cf)
                    
                    # Formatierung auf neue Spalte anwenden
                    if source_width:
                        ws.column_dimensions[get_column_letter(insert_at)].width = source_width
                    
                    for row, fmt in source_format.items():
                        cell = ws.cell(row=row, column=insert_at)
                        if fmt['fill']:
                            cell.fill = fmt['fill']
                        if fmt['font']:
                            cell.font = fmt['font']
                        if fmt['alignment']:
                            cell.alignment = fmt['alignment']
                        if fmt['border']:
                            cell.border = fmt['border']
                        if fmt.get('number_format'):
                            cell.number_format = fmt['number_format']
                
                # Header für neue Spalten setzen
                op_headers = op.get('headers', [])
                for i, header in enumerate(op_headers):
                    ws.cell(row=1, column=excel_col + i).value = header
                
                # Daten für diese Spalten schreiben
                if data and headers:
                    for i in range(count):
                        col_idx = position + i
                        if col_idx < len(headers):
                            for row_idx, row_data in enumerate(data):
                                if col_idx < len(row_data):
                                    cell = ws.cell(row=row_idx + 2, column=excel_col + i)
                                    apply_cell_value(cell, row_data[col_idx])
            
            # Versteckte Spalten/Zeilen
            _apply_hidden_columns(ws, hidden_columns)
            _apply_hidden_rows(ws, hidden_rows)
            
            # Row Highlights (FALL 1.5 - Spalten einfügen)
            if row_highlights:
                _apply_row_highlights(ws, row_highlights, ws.max_column)
            
            # Tables reparieren: Am Ende EINMAL aus Header-Zellen neu aufbauen
            for table_name in ws.tables:
                table = ws.tables[table_name]
                min_col, min_row, max_col, max_row = range_boundaries(table.ref)
                
                new_max_col = ws.max_column
                new_ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(new_max_col)}{max_row}"
                table.ref = new_ref
                if table.autoFilter:
                    table.autoFilter.ref = new_ref
                
                # Baue tableColumns aus den Header-Zellen
                new_columns = []
                for col_idx in range(min_col, new_max_col + 1):
                    header_cell = ws.cell(row=min_row, column=col_idx)
                    col_name = str(header_cell.value) if header_cell.value else f"Column{col_idx}"
                    new_columns.append(TableColumn(id=col_idx - min_col + 1, name=col_name))
                
                table.tableColumns = new_columns
            
            # Einmal speichern
            wb.save(output_path)
            wb.close()
            fix_xlsx_relationships(output_path)
            
            # Table-Infos für XML restore sammeln
            table_changes = {}
            wb_temp = load_workbook(output_path, rich_text=True)
            ws_temp = wb_temp[sheet_name]
            for table_name in ws_temp.tables:
                table = ws_temp.tables[table_name]
                col_names = [col.name for col in table.tableColumns]
                table_changes[table_name] = {'ref': table.ref, 'columns': col_names}
            wb_temp.close()
            
            # Original-Table-XML wiederherstellen (xr:uid etc.)
            if table_changes:
                restore_table_xml_from_original(output_path, original_path, table_changes)
            
            restore_external_links_from_original(output_path, original_path, structural_change=True)
            # vm-Attribute für kopierte Bild-Zellen setzen
            if _fp_vm_cell_map:
                _apply_vm_cell_map_to_xlsx(output_path, sheet_name, _fp_vm_cell_map)
            
            # Sicherheitsnetz: Slicer-Artefakte entfernen
            try:
                _strip_slicers_from_zip(output_path)
            except Exception as slicer_err:
                sys.stderr.write(f"[FALL 1.5] WARNUNG: Slicer-Strip Fehler: {slicer_err}\n")
            
            # PivotTables sicherheitshalber entfernen (Insert ändert Spaltenanzahl)
            try:
                _strip_pivot_tables_for_sheet(output_path, sheet_name)
            except Exception as pivot_err:
                sys.stderr.write(f"[FALL 1.5] WARNUNG: PivotTable-Strip Fehler: {pivot_err}\n")
            
            # AutoFilter vom Frontend anwenden
            if frontend_auto_filter:
                _apply_auto_filter_xml(output_path, sheet_name, frontend_auto_filter)
            
            return {'success': True, 'outputPath': output_path, 'method': 'openpyxl-insert-only'}
        
        # =====================================================================
        # FALL 1.9: Spalten LÖSCHEN UND EINFÜGEN kombiniert
        # SERIELL im Speicher - so bleibt die Formatierung erhalten!
        # =====================================================================
        column_delete_and_insert = deleted_columns and inserted_columns and row_mapping_is_identity
        
        if column_delete_and_insert:
            
            from openpyxl.worksheet.table import TableColumn
            from openpyxl.utils.cell import range_boundaries
            
            # ===== SCHRITT 1: Erst alle Spalten LÖSCHEN (von hinten nach vorne) =====
            sorted_deleted = sorted(deleted_columns, reverse=True)
            
            for col_idx in sorted_deleted:
                excel_col = col_idx + 1  # 0-basiert → 1-basiert
                max_col = ws.max_column
                
                # Spaltenbreiten speichern (rechts von der zu löschenden Spalte)
                saved_widths = {}
                for col in range(excel_col + 1, max_col + 1):
                    col_letter = get_column_letter(col)
                    if col_letter in ws.column_dimensions:
                        saved_widths[col] = ws.column_dimensions[col_letter].width
                
                # Spalte löschen
                ws.delete_cols(excel_col, 1)
                
                # Spaltenbreiten wiederherstellen (um 1 nach links verschoben)
                for old_col, width in saved_widths.items():
                    if width:
                        new_letter = get_column_letter(old_col - 1)
                        ws.column_dimensions[new_letter].width = width
                
                # CF anpassen
                adjust_conditional_formatting(ws, [col_idx], None)
            
            # ===== SCHRITT 2: Dann alle Spalten EINFÜGEN =====
            operations = inserted_columns.get('operations', [])
            if not operations and inserted_columns.get('position') is not None:
                operations = [{
                    'position': inserted_columns['position'],
                    'count': inserted_columns.get('count', 1),
                    'sourceColumn': inserted_columns.get('sourceColumn')
                }]
            
            # Sortiere aufsteigend
            operations.sort(key=lambda x: x['position'])
            
            for op_idx, op in enumerate(operations):
                position = op['position']
                count = op.get('count', 1)
                source_column = op.get('sourceColumn')
                excel_col = position + 1  # 0-basiert → 1-basiert
                
                for i in range(count):
                    insert_at = excel_col + i
                    
                    # Formatierung der Referenzspalte speichern
                    source_format = {}
                    source_width = None
                    if source_column is not None:
                        source_excel_col = source_column + 1
                        # Korrigiere für bereits eingefügte Spalten
                        for prev_op in operations[:op_idx]:
                            if source_column >= prev_op['position']:
                                source_excel_col += prev_op.get('count', 1)
                        
                        col_letter = get_column_letter(source_excel_col)
                        if col_letter in ws.column_dimensions:
                            source_width = ws.column_dimensions[col_letter].width
                        
                        for row in range(1, ws.max_row + 1):
                            cell = ws.cell(row=row, column=source_excel_col)
                            source_format[row] = {
                                'fill': copy(cell.fill) if cell.fill else None,
                                'font': copy(cell.font) if cell.font else None,
                                'alignment': copy(cell.alignment) if cell.alignment else None,
                                'border': copy(cell.border) if cell.border else None,
                                'number_format': cell.number_format
                            }
                    
                    # Spaltenbreiten speichern
                    saved_widths = {}
                    for col in range(insert_at, ws.max_column + 1):
                        col_letter = get_column_letter(col)
                        if col_letter in ws.column_dimensions:
                            saved_widths[col] = ws.column_dimensions[col_letter].width
                    
                    # Spalte einfügen
                    ws.insert_cols(insert_at, 1)
                    
                    # Spaltenbreiten wiederherstellen
                    for old_col, width in saved_widths.items():
                        if width:
                            new_letter = get_column_letter(old_col + 1)
                            ws.column_dimensions[new_letter].width = width
                    
                    # CF anpassen
                    inserted_cols_for_cf = {insert_at - 1: 1}
                    adjust_conditional_formatting(ws, [], inserted_cols_for_cf)
                    
                    # Formatierung auf neue Spalte anwenden
                    if source_width:
                        ws.column_dimensions[get_column_letter(insert_at)].width = source_width
                    
                    for row, fmt in source_format.items():
                        cell = ws.cell(row=row, column=insert_at)
                        if fmt['fill']:
                            cell.fill = fmt['fill']
                        if fmt['font']:
                            cell.font = fmt['font']
                        if fmt['alignment']:
                            cell.alignment = fmt['alignment']
                        if fmt['border']:
                            cell.border = fmt['border']
                        if fmt.get('number_format'):
                            cell.number_format = fmt['number_format']
                
                # Header für neue Spalten setzen
                op_headers = op.get('headers', [])
                for i, header in enumerate(op_headers):
                    ws.cell(row=1, column=excel_col + i).value = header
                
                # Daten für diese Spalten schreiben
                if data and headers:
                    for i in range(count):
                        col_idx = position + i
                        if col_idx < len(headers):
                            for row_idx, row_data in enumerate(data):
                                if col_idx < len(row_data):
                                    cell = ws.cell(row=row_idx + 2, column=excel_col + i)
                                    apply_cell_value(cell, row_data[col_idx])
            
            # Versteckte Spalten/Zeilen
            _apply_hidden_columns(ws, hidden_columns)
            _apply_hidden_rows(ws, hidden_rows)
            
            # Row Highlights (FALL 1.9 - Spalten löschen und einfügen)
            if row_highlights:
                _apply_row_highlights(ws, row_highlights, ws.max_column)
            
            # Tables reparieren: Am Ende EINMAL aus Header-Zellen neu aufbauen
            for table_name in ws.tables:
                table = ws.tables[table_name]
                min_col, min_row, max_col, max_row = range_boundaries(table.ref)
                
                new_max_col = ws.max_column
                new_ref = f"{get_column_letter(min_col)}{min_row}:{get_column_letter(new_max_col)}{max_row}"
                table.ref = new_ref
                if table.autoFilter:
                    table.autoFilter.ref = new_ref
                
                # Baue tableColumns aus den Header-Zellen
                new_columns = []
                for col_idx in range(min_col, new_max_col + 1):
                    header_cell = ws.cell(row=min_row, column=col_idx)
                    col_name = str(header_cell.value) if header_cell.value else f"Column{col_idx}"
                    new_columns.append(TableColumn(id=col_idx - min_col + 1, name=col_name))
                
                table.tableColumns = new_columns
            
            # Einmal speichern
            wb.save(output_path)
            wb.close()
            fix_xlsx_relationships(output_path)
            
            # Table-Infos für XML restore sammeln
            table_changes = {}
            wb_temp = load_workbook(output_path, rich_text=True)
            ws_temp = wb_temp[sheet_name]
            for table_name in ws_temp.tables:
                table = ws_temp.tables[table_name]
                col_names = [col.name for col in table.tableColumns]
                table_changes[table_name] = {'ref': table.ref, 'columns': col_names}
            wb_temp.close()
            
            # Original-Table-XML wiederherstellen (xr:uid etc.)
            if table_changes:
                restore_table_xml_from_original(output_path, original_path, table_changes)
            
            restore_external_links_from_original(output_path, original_path, structural_change=True)
            # vm-Attribute für kopierte Bild-Zellen setzen
            if _fp_vm_cell_map:
                _apply_vm_cell_map_to_xlsx(output_path, sheet_name, _fp_vm_cell_map)
            
            # Sicherheitsnetz: Slicer-Artefakte entfernen
            try:
                _strip_slicers_from_zip(output_path)
            except Exception as slicer_err:
                sys.stderr.write(f"[FALL 1.9] WARNUNG: Slicer-Strip Fehler: {slicer_err}\n")
            
            # PivotTables entfernen (Spalten gelöscht → location ref invalide)
            try:
                _strip_pivot_tables_for_sheet(output_path, sheet_name)
            except Exception as pivot_err:
                sys.stderr.write(f"[FALL 1.9] WARNUNG: PivotTable-Strip Fehler: {pivot_err}\n")
            
            # AutoFilter vom Frontend anwenden
            if frontend_auto_filter:
                _apply_auto_filter_xml(output_path, sheet_name, frontend_auto_filter)
            
            return {'success': True, 'outputPath': output_path, 'method': 'openpyxl-delete-and-insert'}
        
        # =====================================================================
        # FALL 1.6: Nur Spalten LÖSCHEN (keine anderen strukturellen Änderungen)
        # Analog zu FALL 1.5 - nutzt openpyxl's delete_cols() direkt
        # OHNE alle Daten neu zu schreiben - das erhält Table-Styles!
        # =====================================================================
        only_column_delete = deleted_columns and not inserted_columns and row_mapping_is_identity
        
        if only_column_delete:
            
            # Sortiere absteigend (von hinten nach vorne löschen)
            sorted_deleted = sorted(deleted_columns, reverse=True)
            
            for col_idx in sorted_deleted:
                excel_col = col_idx + 1  # 0-basiert → 1-basiert
                
                max_col = ws.max_column
                
                # 1. SPALTENBREITEN SPEICHERN (rechts von der zu löschenden Spalte)
                saved_widths = {}
                for col in range(excel_col + 1, max_col + 1):
                    col_letter = get_column_letter(col)
                    if col_letter in ws.column_dimensions:
                        saved_widths[col] = ws.column_dimensions[col_letter].width
                
                # 2. SPALTE LÖSCHEN (openpyxl verschiebt alles automatisch)
                ws.delete_cols(excel_col, 1)
                
                # 3. SPALTENBREITEN WIEDERHERSTELLEN (um 1 nach links verschoben)
                for old_col, width in saved_widths.items():
                    if width:
                        new_letter = get_column_letter(old_col - 1)
                        ws.column_dimensions[new_letter].width = width
                
                # 4. CF anpassen
                adjust_conditional_formatting(ws, [col_idx], None)
                
                # 5. Tables anpassen
                adjust_tables(ws, [col_idx], None, headers)
            
            # Versteckte Spalten/Zeilen
            _apply_hidden_columns(ws, hidden_columns)
            _apply_hidden_rows(ws, hidden_rows)
            
            # Row Highlights (FALL 1.6 - Spalten löschen)
            if row_highlights:
                _apply_row_highlights(ws, row_highlights, ws.max_column)
            
            # Sammle Table-Infos für restore
            table_changes = {}
            for table_name in ws.tables:
                table = ws.tables[table_name]
                col_names = [col.name for col in table.tableColumns]
                table_changes[table_name] = {
                    'ref': table.ref,
                    'columns': col_names
                }
            
            wb.save(output_path)
            wb.close()
            fix_xlsx_relationships(output_path)
            
            # Stelle Original-Table-XML wieder her (mit korrekten xr:uid etc.)
            if table_changes:
                restore_table_xml_from_original(output_path, original_path, table_changes)
            
            # Stelle externalLinks aus Original wieder her (openpyxl verliert Namespaces)
            restore_external_links_from_original(output_path, original_path, structural_change=True)
            # vm-Attribute für kopierte Bild-Zellen setzen
            if _fp_vm_cell_map:
                _apply_vm_cell_map_to_xlsx(output_path, sheet_name, _fp_vm_cell_map)
            
            # Sicherheitsnetz: Slicer-Artefakte entfernen
            try:
                _strip_slicers_from_zip(output_path)
            except Exception as slicer_err:
                sys.stderr.write(f"[FALL 1.6] WARNUNG: Slicer-Strip Fehler: {slicer_err}\n")
            
            # PivotTables entfernen (Spalten gelöscht → location ref invalide)
            try:
                _strip_pivot_tables_for_sheet(output_path, sheet_name)
            except Exception as pivot_err:
                sys.stderr.write(f"[FALL 1.6] WARNUNG: PivotTable-Strip Fehler: {pivot_err}\n")
            
            # AutoFilter vom Frontend anwenden
            if frontend_auto_filter:
                _apply_auto_filter_xml(output_path, sheet_name, frontend_auto_filter)
            
            return {'success': True, 'outputPath': output_path, 'method': 'openpyxl-delete-only'}
        
        # =====================================================================
        # FALL 1.7: NUR Spaltenreihenfolge ändern (ohne Insert/Delete)
        # Dieser Pfad ordnet Spalten physisch um, OHNE alle Zellen neu zu schreiben.
        # Das erhält Table-Styles (Zebra-Muster) perfekt!
        # =====================================================================
        only_column_order = (column_order and len(column_order) > 0 and 
                            not inserted_columns and not deleted_columns and 
                            row_mapping_is_identity and not affected_rows)
        
        if only_column_order:
            
            # Prüfe ob sich die Spaltenreihenfolge wirklich geändert hat
            columns_changed = False
            for new_idx, old_idx in enumerate(column_order):
                if new_idx != old_idx:
                    columns_changed = True
                    break
            
            if not columns_changed:
                pass  # Keine Änderung nötig
            else:
                # Physische Spaltenumordnung durch Swap-Operationen
                # column_order[neue_position] = alte_position
                
                from openpyxl.cell.cell import MergedCell
                
                num_cols = len(column_order)
                max_row = ws.max_row
                
                # Temporärer Speicher für alle Spalten (Werte + Hyperlinks)
                temp_columns = {}
                
                # SCHRITT 1: Alle Spalten in temp_columns speichern
                for old_col_idx in range(num_cols):
                    old_excel_col = old_col_idx + 1
                    temp_columns[old_col_idx] = {}
                    
                    for row in range(1, max_row + 1):
                        cell = ws.cell(row=row, column=old_excel_col)
                        if isinstance(cell, MergedCell):
                            continue
                        temp_columns[old_col_idx][row] = {
                            'value': cell.value,
                            'hyperlink': cell.hyperlink.target if cell.hyperlink else None,
                        }
                
                # SCHRITT 2: Spalten in neuer Reihenfolge schreiben
                for new_col_idx, old_col_idx in enumerate(column_order):
                    new_excel_col = new_col_idx + 1
                    
                    if old_col_idx not in temp_columns:
                        continue
                    
                    for row, data_item in temp_columns[old_col_idx].items():
                        cell = ws.cell(row=row, column=new_excel_col)
                        if isinstance(cell, MergedCell):
                            continue
                        
                        # Nur Wert und Hyperlink setzen - KEINE Formatierung!
                        # So bleibt das Table-Style-Zebra-Muster erhalten
                        cell.value = data_item['value']
                        if data_item['hyperlink']:
                            cell.hyperlink = data_item['hyperlink']
                
            
            # Versteckte Spalten/Zeilen anwenden
            _apply_hidden_columns(ws, hidden_columns, len(headers))
            _apply_hidden_rows(ws, hidden_rows, len(data) if data else 0)
            
            # Row Highlights
            if row_highlights:
                _apply_row_highlights(ws, row_highlights, len(headers))
            
            # WICHTIG: Bei Spalten-Verschieben die tableColumns AKTUALISIEREN!
            # Die Spalten wurden physisch umgeordnet, also müssen die Column-Namen
            # aus den Header-Zellen neu gelesen werden.
            from openpyxl.worksheet.table import TableColumn
            from openpyxl.utils.cell import range_boundaries
            
            table_changes = {}
            for table_name in ws.tables:
                table = ws.tables[table_name]
                min_col, min_row, max_col, max_row = range_boundaries(table.ref)
                
                # Baue tableColumns aus den Header-Zellen (die sind jetzt umgeordnet)
                new_columns = []
                for col_idx in range(min_col, max_col + 1):
                    header_cell = ws.cell(row=min_row, column=col_idx)
                    col_name = str(header_cell.value) if header_cell.value else f"Column{col_idx}"
                    new_columns.append(TableColumn(id=col_idx - min_col + 1, name=col_name))
                
                table.tableColumns = new_columns
                
                col_names = [col.name for col in new_columns]
                table_changes[table_name] = {
                    'ref': table.ref,
                    'columns': col_names
                }
            
            wb.save(output_path)
            wb.close()
            fix_xlsx_relationships(output_path)
            
            # Stelle Table-XML aus Original wieder her MIT der neuen Spaltenreihenfolge
            if table_changes:
                restore_table_xml_from_original(output_path, original_path, table_changes)
            
            # Stelle externalLinks aus Original wieder her
            restore_external_links_from_original(output_path, original_path, structural_change=True)
            # vm-Attribute für kopierte Bild-Zellen setzen
            if _fp_vm_cell_map:
                _apply_vm_cell_map_to_xlsx(output_path, sheet_name, _fp_vm_cell_map)
            
            # Sicherheitsnetz: Slicer-Artefakte entfernen
            try:
                _strip_slicers_from_zip(output_path)
            except Exception as slicer_err:
                sys.stderr.write(f"[FALL 1.7] WARNUNG: Slicer-Strip Fehler: {slicer_err}\n")
            
            # PivotTables sicherheitshalber entfernen (Spaltenreihenfolge geändert)
            try:
                _strip_pivot_tables_for_sheet(output_path, sheet_name)
            except Exception as pivot_err:
                sys.stderr.write(f"[FALL 1.7] WARNUNG: PivotTable-Strip Fehler: {pivot_err}\n")
            
            # AutoFilter vom Frontend anwenden
            if frontend_auto_filter:
                _apply_auto_filter_xml(output_path, sheet_name, frontend_auto_filter)
            
            return {'success': True, 'outputPath': output_path, 'method': 'openpyxl-column-order'}
        
        # =====================================================================
        # FALL 3: Nur Zell-Edits (keine strukturellen Änderungen)
        # =====================================================================
        
        # Prüfe ob wir echte Zell-Edits haben (nicht nur Highlights)
        real_edits = {k: v for k, v in edited_cells.items() if not k.startswith('_')} if edited_cells else {}
        
        # Prüfe ob zusätzliche Änderungen neben reinen Zell-Edits vorliegen
        imported_cell_styles = changes.get('cellStyles', {})
        cell_fonts = changes.get('cellFonts', {})
        imported_rich_text = changes.get('richTextCells', {})
        imported_merged_cells = changes.get('mergedCells', [])
        
        # WICHTIG: FALL 3a (direkte XML-Bearbeitung) kopiert die Original-Datei 1:1
        # und ändert NUR Zellwerte. Alle Styles, Fonts, MergedCells, Highlights,
        # Drawings etc. bleiben aus dem Original erhalten.
        #
        # Das Frontend sendet IMMER die Original-Daten mit:
        # - cellStyles: Original-Styles der editierten Zellen (zum Wiederherstellen in openpyxl)
        # - cellFonts: Original-Fonts der editierten Zellen
        # - richTextCells: Original-RichText der editierten Zellen
        # - mergedCells: ALLE pre-existierenden Merged Cells
        # - rowHighlights: ALLE aktuellen Highlights (auch pre-existierende)
        #
        # Diese Daten sind KEINE neuen Änderungen - sie sind pre-existierend!
        # FALL 3a preserviert sie automatisch aus dem Original.
        #
        # NUR wenn das Frontend _hasFormatChanges gesetzt hat (Paste-mit-Format,
        # Data Join Styles, oder Highlight-Fill-Löschung), sind echte
        # Formatierungsänderungen vorhanden → FALL 3b nötig.
        has_format_flag = changes.get('hasFormatChanges', False)
        has_extra_changes = has_format_flag
        has_highlight_changes = bool(cleared_row_highlights)  # Nur wenn Highlights explizit entfernt
        has_visibility_changes = bool(hidden_rows) or bool(hidden_columns)
        
        # Safety-Net: Wenn hasFormatChanges NUR wegen Row-Highlights gesetzt wurde
        # (keine echten Format-Änderungen wie Paste-Styles, Fonts, RichText),
        # dann können wir den XML-Pfad nutzen → Slicers bleiben intakt.
        if has_extra_changes and row_highlights and not has_highlight_changes:
            # Prüfe ob echte Format-Änderungen vorliegen
            if not cell_fonts and not imported_rich_text:
                has_extra_changes = False
                sys.stderr.write(f"[GATE] hasFormatChanges nur wegen Highlights → überschrieben für XML-Pfad\n")
        
        sys.stderr.write(f"[GATE] hasFormatChanges={has_format_flag}, has_extra_changes={has_extra_changes}, has_highlight_changes={has_highlight_changes}\n")
        sys.stderr.write(f"[GATE] real_edits={len(real_edits)}, cellStyles={len(imported_cell_styles)}, mergedCells={len(imported_merged_cells)}\n")
        sys.stderr.write(f"[GATE] has_visibility_changes={has_visibility_changes} (hidden_rows={len(hidden_rows) if hidden_rows else 0}, hidden_cols={len(hidden_columns) if hidden_columns else 0})\n")
        
        # =====================================================================
        # FALL 3a: Zell-Edits ODER Visibility-Änderungen ODER Row-Highlights
        # → Direkte XML-Bearbeitung (kein openpyxl-Roundtrip)
        # Dies vermeidet das Überschreiben von Rels, Namespaces, SharedStrings etc.
        # Die Original-Datei bleibt zu 100% intakt, nur die Zellwerte,
        # hidden-Attribute und Highlight-Fills werden geändert.
        #
        # AUCH für reine Visibility-Änderungen (hideRow/hideColumn ohne Edits):
        # openpyxl-Roundtrip verliert Drawings ohne Pillow → "Zeichnungsform entfernt".
        # ZIP-to-ZIP preserviert ALLES aus dem Original.
        #
        # NEU: Row-Highlights werden direkt in styles.xml + Sheet-XML gesetzt.
        # Dadurch bleiben SlicerCaches, Slicers, Drawings etc. 100% intakt.
        # =====================================================================
        has_add_highlights = bool(row_highlights)
        if (real_edits or has_visibility_changes or has_add_highlights) and not has_extra_changes and not has_highlight_changes:
            wb.close()  # openpyxl Workbook nicht mehr benötigt
            sys.stderr.write(f"[FALL 3a] Direkte XML-Bearbeitung für {len(real_edits)} Zell-Edits, visibility={has_visibility_changes}, highlights={has_add_highlights}\n")
            
            try:
                result = _direct_xml_cell_edit(
                    file_path, output_path, sheet_name, real_edits,
                    hidden_columns, hidden_rows,
                    row_highlights=row_highlights
                )
                # AutoFilter vom Frontend anwenden
                if frontend_auto_filter and result.get('success'):
                    _apply_auto_filter_xml(output_path, sheet_name, frontend_auto_filter)
                return result
            except Exception as xml_err:
                sys.stderr.write(f"[FALL 3a] Fehler bei direkter XML-Bearbeitung: {xml_err}\n")
                sys.stderr.write(f"[FALL 3a] Fallback auf openpyxl-Pfad...\n")
                # Fallback: openpyxl-Pfad (FALL 3b)
                wb = _safe_load_workbook(file_path, rich_text=True)
                ws = wb[sheet_name]
        
        # =====================================================================
        # FALL 3b: Zell-Edits MIT zusätzlichen Änderungen (Highlights, Styles, etc.)
        # Hier muss openpyxl verwendet werden, da XML-Bearbeitung zu komplex wäre.
        # =====================================================================
        
        # Wenn NUR Highlights (keine echten Edits), lade von Original-Datei neu (falls verfügbar)
        # Das stellt sicher dass alte Highlights nicht erhalten bleiben
        if row_highlights is not None and not real_edits:
            if original_path and original_path != file_path and os.path.exists(original_path):
                wb.close()
                import shutil
                shutil.copy2(original_path, output_path)
                wb = _safe_load_workbook(output_path, rich_text=True)
                ws = wb[sheet_name]
            else:
                # Kein Original verfügbar - entferne alle Fills in Zeilen die NICHT markiert sind
                # Das ist nicht perfekt (verliert Zebra-Muster), aber besser als alte Highlights zu behalten
                _clear_all_row_fills_except(ws, row_highlights)
        
        if real_edits:
            for key, value in real_edits.items():
                parts = key.split('-')
                if len(parts) != 2:
                    continue
                row_idx = int(parts[0])
                col_idx = int(parts[1])
                cell = ws.cell(row=row_idx + 2, column=col_idx + 1)
                apply_cell_value(cell, value)
        
        # Versteckte Spalten/Zeilen setzen
        _apply_hidden_columns(ws, hidden_columns)
        _apply_hidden_rows(ws, hidden_rows)
        
        # Kopierte Zell-Hintergründe anwenden (aus Copy-Paste)
        if imported_cell_styles:
            _apply_imported_cell_styles(ws, imported_cell_styles)
        
        # Kopierte Schriftformatierungen anwenden (aus Copy-Paste)
        if cell_fonts:
            _apply_cell_fonts(ws, cell_fonts)
        
        # Kopierte RichText-Formatierung anwenden (aus Copy-Paste)
        if imported_rich_text:
            _apply_imported_rich_text(ws, imported_rich_text)
        
        # Kopierte Merged Cells anwenden (aus Copy-Paste)
        if imported_merged_cells:
            _apply_imported_merged_cells(ws, imported_merged_cells)
        
        # Row Highlights
        if row_highlights:
            _apply_row_highlights(ws, row_highlights, ws.max_column)
        
        # Cleared Row Highlights (Markierungen entfernen)
        if cleared_row_highlights:
            sys.stderr.write(f"[FALL 3b] Entferne {len(cleared_row_highlights)} Row Highlights\n")
            for row_idx in cleared_row_highlights:
                excel_row = row_idx + 2  # 0-basiert nach 1-basiert + Header
                for col_idx in range(1, ws.max_column + 1):
                    cell = ws.cell(row=excel_row, column=col_idx)
                    cell.fill = PatternFill()  # Keine Füllung
        
        wb.save(output_path)
        wb.close()
        fix_xlsx_relationships(output_path)
        
        # WICHTIG: Table-XML vom Original wiederherstellen!
        restore_table_xml_from_original(output_path, original_path, table_changes=None)
        
        # WICHTIG: Auch workbook.xml, slicerCaches, etc. vom Original wiederherstellen!
        restore_external_links_from_original(output_path, original_path)
        # vm-Attribute für kopierte Bild-Zellen setzen
        if _fp_vm_cell_map:
            _apply_vm_cell_map_to_xlsx(output_path, sheet_name, _fp_vm_cell_map)
        
        # AutoFilter vom Frontend anwenden
        if frontend_auto_filter:
            _apply_auto_filter_xml(output_path, sheet_name, frontend_auto_filter)
        
        return {'success': True, 'outputPath': output_path}
        
    except Exception as e:
        import traceback
        error_msg = str(e)
        tb = traceback.format_exc()
        print(f"[Python Writer] ERROR: {error_msg}", file=sys.stderr)
        print(f"[Python Writer] Traceback: {tb}", file=sys.stderr)
        return {
            'success': False, 
            'error': error_msg,
            'traceback': tb
        }
    finally:
        # Backup-Datei aufräumen (erstellt für Speichern-in-gleicher-Datei Szenario)
        if _backup_file is not None:
            try:
                os.unlink(_backup_path)
                sys.stderr.write(f"[WRITE_SHEET] Backup gelöscht: {_backup_path}\\n")
            except Exception:
                pass
            # Funktionsattribut zurücksetzen
            restore_external_links_from_original._backup_original_path = None


def _apply_hidden_columns(ws, hidden_columns, max_cols=None):
    """Setzt versteckte Spalten"""
    if hidden_columns is None:
        return
    
    hidden_set = set(hidden_columns)
    max_col = max_cols if max_cols else ws.max_column
    
    for col_idx in range(max_col):
        col_letter = get_column_letter(col_idx + 1)
        ws.column_dimensions[col_letter].hidden = col_idx in hidden_set


def _apply_hidden_rows(ws, hidden_rows, max_rows=None):
    """Setzt versteckte Zeilen"""
    if hidden_rows is None:
        return
    
    hidden_set = set(hidden_rows)
    max_row = max_rows if max_rows else (ws.max_row - 1)  # Ohne Header
    
    for row_idx in range(max_row):
        excel_row = row_idx + 2  # +2 für Header
        ws.row_dimensions[excel_row].hidden = row_idx in hidden_set


def _clear_all_row_fills_except(ws, row_highlights):
    """
    Entfernt Fills von allen Zeilen AUSSER den in row_highlights angegebenen.
    Wird verwendet wenn kein Original verfügbar ist und Highlights entfernt werden sollen.
    """
    # Sammle die Zeilen die Highlights behalten sollen
    highlighted_rows = set()
    if row_highlights:
        for row_idx_str in row_highlights.keys():
            highlighted_rows.add(int(row_idx_str) + 2)  # +2 für Excel-Row (1-basiert + Header)
    
    # Durchgehe alle Datenzeilen und entferne Fills die nicht Highlights sind
    max_row = ws.max_row


def _apply_number_formats(ws, number_formats):
    """Wendet Zahlenformate aus dem Frontend auf Zellen an"""
    if not number_formats:
        return
    
    for key, fmt in number_formats.items():
        try:
            parts = key.split('-')
            if len(parts) != 2:
                continue
            row_idx = int(parts[0])
            col_idx = int(parts[1])
            cell = ws.cell(row=row_idx + 2, column=col_idx + 1)  # +2 für Header, +1 für 1-basiert
            cell.number_format = fmt
        except Exception:
            pass


def _apply_cell_fonts(ws, cell_fonts):
    """Wendet Font-Formatierungen aus dem Frontend auf Zellen an"""
    if not cell_fonts:
        return
    
    sys.stderr.write(f"[CellFonts] Starte: {len(cell_fonts)} Font-Einträge zu verarbeiten\n")
    applied = 0
    
    for key, font_info in cell_fonts.items():
        try:
            parts = key.split('-')
            if len(parts) != 2:
                continue
            row_idx = int(parts[0])
            col_idx = int(parts[1])
            cell = ws.cell(row=row_idx + 2, column=col_idx + 1)
            
            # Bestehende Font-Eigenschaften beibehalten und nur überschreiben was wir haben
            existing_font = cell.font
            font_kwargs = {
                'name': existing_font.name,
                'size': existing_font.size,
                'bold': existing_font.bold,
                'italic': existing_font.italic,
                'color': existing_font.color,
                'underline': existing_font.underline,
                'strikethrough': existing_font.strikethrough,
            }
            
            # Neue Werte überschreiben
            if font_info.get('name'):
                font_kwargs['name'] = font_info['name']
            if font_info.get('size'):
                font_kwargs['size'] = font_info['size']
            if 'bold' in font_info:
                font_kwargs['bold'] = font_info['bold']
            if 'italic' in font_info:
                font_kwargs['italic'] = font_info['italic']
            if font_info.get('color'):
                color_val = font_info['color']
                # #RRGGBB → RRGGBB konvertieren (openpyxl erwartet kein #)
                if isinstance(color_val, str) and color_val.startswith('#'):
                    color_val = color_val[1:]
                # Zu ARGB konvertieren wenn nötig
                if isinstance(color_val, str) and len(color_val) == 6:
                    color_val = 'FF' + color_val
                font_kwargs['color'] = color_val
            
            cell.font = Font(**font_kwargs)
            applied += 1
        except Exception as e:
            sys.stderr.write(f"[CellFonts] Fehler bei {key}: {e}\n")
    
    sys.stderr.write(f"[CellFonts] Fertig: {applied} von {len(cell_fonts)} Fonts angewendet\n")


def _apply_imported_cell_styles(ws, cell_styles):
    """
    Wendet KOMPLETTE Zell-Formatierung aus dem Frontend auf Zellen an.
    Unterstützt sowohl altes Format (String = nur Hintergrundfarbe)
    als auch neues Format (Object = komplette Formatierung inkl. Font, Fill, Alignment).
    
    Neues Format-Objekt kann enthalten:
    - fill: Hintergrundfarbe (#RRGGBB oder ARGB)
    - bold, italic, underline, strikethrough: Font-Eigenschaften
    - fontSize, fontName, fontColor: Font-Details
    - textAlign: Horizontale Ausrichtung (left, center, right)
    - verticalAlign: Vertikale Ausrichtung (top, center, bottom)
    - wrapText: Textumbruch (true/false)
    """
    if not cell_styles:
        return
    
    applied = 0
    for key, style_data in cell_styles.items():
        try:
            parts = key.split('-')
            if len(parts) != 2:
                continue
            row_idx = int(parts[0])
            col_idx = int(parts[1])
            cell = ws.cell(row=row_idx + 2, column=col_idx + 1)
            
            # Altes Format: String = nur Hintergrundfarbe
            if isinstance(style_data, str):
                color = style_data
                if color.startswith('#'):
                    argb = hex_to_argb(color)
                else:
                    argb = color if len(color) == 8 else f'FF{color}'
                cell.fill = PatternFill(start_color=argb, end_color=argb, fill_type='solid')
                applied += 1
                continue
            
            # Neues Format: Komplettes Style-Objekt
            if not isinstance(style_data, dict):
                continue
            
            # === FILL (Hintergrundfarbe) ===
            fill_color = style_data.get('fill')
            if fill_color and fill_color != 'transparent':
                if isinstance(fill_color, str):
                    if fill_color.startswith('#'):
                        argb = hex_to_argb(fill_color)
                    else:
                        argb = fill_color if len(fill_color) == 8 else f'FF{fill_color}'
                    cell.fill = PatternFill(start_color=argb, end_color=argb, fill_type='solid')
            
            # === FONT (Schriftformatierung) ===
            has_font_info = any(style_data.get(k) is not None for k in 
                              ['bold', 'italic', 'underline', 'strikethrough', 'fontSize', 'fontName', 'fontColor'])
            if has_font_info:
                # Bestehende Font-Eigenschaften als Basis verwenden
                existing_font = cell.font
                font_kwargs = {
                    'name': existing_font.name,
                    'size': existing_font.size,
                    'bold': existing_font.bold,
                    'italic': existing_font.italic,
                    'underline': existing_font.underline,
                    'strikethrough': existing_font.strikethrough,
                    'color': existing_font.color,
                }
                
                if style_data.get('fontName'):
                    font_kwargs['name'] = style_data['fontName']
                if style_data.get('fontSize'):
                    font_kwargs['size'] = style_data['fontSize']
                if 'bold' in style_data:
                    font_kwargs['bold'] = bool(style_data['bold'])
                if 'italic' in style_data:
                    font_kwargs['italic'] = bool(style_data['italic'])
                if 'underline' in style_data:
                    font_kwargs['underline'] = 'single' if style_data['underline'] else None
                if 'strikethrough' in style_data:
                    font_kwargs['strikethrough'] = bool(style_data['strikethrough'])
                if style_data.get('fontColor'):
                    color_val = style_data['fontColor']
                    if isinstance(color_val, str) and color_val.startswith('#'):
                        color_val = color_val[1:]
                    if isinstance(color_val, str) and len(color_val) == 6:
                        color_val = 'FF' + color_val
                    font_kwargs['color'] = color_val
                
                cell.font = Font(**font_kwargs)
            
            # === ALIGNMENT (Ausrichtung) ===
            has_alignment = any(style_data.get(k) is not None for k in 
                              ['textAlign', 'verticalAlign', 'wrapText'])
            if has_alignment:
                align_kwargs = {}
                if style_data.get('textAlign'):
                    align_kwargs['horizontal'] = style_data['textAlign']
                if style_data.get('verticalAlign'):
                    align_kwargs['vertical'] = style_data['verticalAlign']
                if 'wrapText' in style_data:
                    align_kwargs['wrap_text'] = bool(style_data['wrapText'])
                if align_kwargs:
                    cell.alignment = Alignment(**align_kwargs)
            
            applied += 1
        except Exception as e:
            sys.stderr.write(f"[CellStyles] Fehler bei {key}: {e}\n")
    
    if applied > 0:
        sys.stderr.write(f"[CellStyles] {applied} von {len(cell_styles)} Zell-Styles angewendet\n")


def _apply_imported_merged_cells(ws, merged_cells_list):
    """
    Wendet Merged Cells aus dem Frontend auf das Worksheet an.
    Verwendet differentiellen Ansatz: nur Änderungen anwenden statt alles neu zu mergen.
    So bleibt die Textformatierung bestehender Merges erhalten.
    
    Frontend-Format: [{ startRow, startCol, endRow, endCol, rowSpan, colSpan }]
    startRow/startCol sind 0-basiert (0 = Excel-Zeile 1 = Header).
    """
    if not merged_cells_list:
        sys.stderr.write(f"[MergedCells] Liste leer - übersprungen\n")
        return
    
    sys.stderr.write(f"[MergedCells] Starte: {len(merged_cells_list)} Merges zu verarbeiten\n")
    
    # Bestehende Merges als String-Set
    existing_ranges = set(str(mr) for mr in ws.merged_cells.ranges)
    sys.stderr.write(f"[MergedCells] Bestehende Merges im Sheet: {len(existing_ranges)}\n")
    for mr_str in list(existing_ranges)[:5]:
        sys.stderr.write(f"[MergedCells] Bestehend: {mr_str}\n")
    
    # Ziel-Merges als String-Set aufbauen
    target_ranges = set()
    for merge in merged_cells_list:
        start_row = merge.get('startRow', 0) + 1  # 0-basiert -> 1-basiert
        start_col = merge.get('startCol', 0) + 1
        end_row = merge.get('endRow', 0) + 1
        end_col = merge.get('endCol', 0) + 1
        
        if start_row > 0 and start_col > 0 and end_row >= start_row and end_col >= start_col:
            cell_range = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(end_col)}{end_row}"
            target_ranges.add(cell_range)
        else:
            sys.stderr.write(f"[MergedCells] Übersprungen (ungültig): ({start_row},{start_col}):({end_row},{end_col})\n")
    
    # Differenz berechnen
    to_remove = existing_ranges - target_ranges  # Merges die entfernt werden müssen
    to_add = target_ranges - existing_ranges      # Neue Merges die hinzugefügt werden
    unchanged = existing_ranges & target_ranges   # Merges die bleiben (nicht anfassen!)
    
    sys.stderr.write(f"[MergedCells] Differenz: {len(unchanged)} unverändert, {len(to_remove)} entfernen, {len(to_add)} hinzufügen\n")
    
    # Nur nicht mehr benötigte Merges entfernen
    for mr_str in to_remove:
        try:
            ws.unmerge_cells(mr_str)
            sys.stderr.write(f"[MergedCells] Entfernt: {mr_str}\n")
        except Exception as e:
            sys.stderr.write(f"[MergedCells] Fehler beim Unmerge {mr_str}: {e}\n")
    
    # Nur neue Merges hinzufügen
    for mr_str in to_add:
        try:
            ws.merge_cells(mr_str)
            sys.stderr.write(f"[MergedCells] Hinzugefügt: {mr_str}\n")
        except Exception as e:
            sys.stderr.write(f"[MergedCells] Fehler beim Mergen {mr_str}: {e}\n")
    
    sys.stderr.write(f"[MergedCells] Fertig: {len(unchanged) + len(to_add)} Merged Cells im Ergebnis\n")


def _apply_imported_rich_text(ws, rich_text_cells):
    """
    Wendet RichText-Formatierung aus dem Frontend auf Zellen an.
    Wird für kopierte RichText-Zellen verwendet.
    
    Frontend-Format: { "row-col": [{ text: "...", styles: { bold, italic, ... } }] }
    Keys sind 0-basiert (wie editedCells).
    """
    if not rich_text_cells:
        return
    
    try:
        from openpyxl.cell.rich_text import CellRichText, TextBlock
        from openpyxl.cell.text import InlineFont
    except ImportError:
        sys.stderr.write("[RichText] openpyxl CellRichText nicht verfügbar - überspringe RichText\n")
        return
    
    applied = 0
    for key, parts in rich_text_cells.items():
        try:
            key_parts = key.split('-')
            if len(key_parts) != 2:
                continue
            row_idx = int(key_parts[0])
            col_idx = int(key_parts[1])
            cell = ws.cell(row=row_idx + 2, column=col_idx + 1)
            
            if not isinstance(parts, list) or len(parts) == 0:
                continue
            
            # Konvertiere Frontend-Format zu CellRichText
            text_blocks = []
            for part in parts:
                text = part.get('text', '')
                styles = part.get('styles', {})
                
                font_kwargs = {}
                if styles.get('bold'):
                    font_kwargs['b'] = True
                if styles.get('italic'):
                    font_kwargs['i'] = True
                if styles.get('underline'):
                    font_kwargs['u'] = 'single'
                if styles.get('strikethrough'):
                    font_kwargs['strike'] = True
                if styles.get('fontSize'):
                    font_kwargs['sz'] = styles['fontSize']
                if styles.get('fontName'):
                    font_kwargs['rFont'] = styles['fontName']
                if styles.get('color'):
                    color_val = styles['color']
                    if isinstance(color_val, str) and color_val.startswith('#'):
                        font_kwargs['color'] = color_val[1:]  # Entferne #
                    elif isinstance(color_val, str):
                        font_kwargs['color'] = color_val
                
                if font_kwargs:
                    inline_font = InlineFont(**font_kwargs)
                    text_blocks.append(TextBlock(inline_font, text))
                else:
                    text_blocks.append(text)
            
            cell.value = CellRichText(*text_blocks)
            applied += 1
        except Exception as e:
            sys.stderr.write(f"[RichText] Fehler bei {key}: {e}\n")
    
    if applied > 0:
        sys.stderr.write(f"[RichText] {applied} RichText-Zellen angewendet\n")


def _apply_row_highlights(ws, row_highlights, num_columns):
    """Wendet Zeilen-Highlights an"""
    highlight_colors = {
        'green': 'FF90EE90',
        'yellow': 'FFFFFF00',
        'orange': 'FFFFA500',
        'red': 'FFFF6B6B',
        'blue': 'FF87CEEB',
        'purple': 'FFDDA0DD'
    }
    
    for row_idx_str, color in row_highlights.items():
        row_idx = int(row_idx_str)
        excel_row = row_idx + 2  # +2 für 1-basiert und Header
        
        if isinstance(color, str) and color.startswith('#'):
            argb = hex_to_argb(color)
        else:
            argb = highlight_colors.get(color, 'FFFFFF00')
        
        # Alle Zellen in der Zeile färben
        for col_idx in range(1, num_columns + 1):
            cell = ws.cell(row=excel_row, column=col_idx)
            cell.fill = PatternFill(start_color=argb, end_color=argb, fill_type='solid')


def main():
    """Hauptfunktion - liest Befehle von stdin oder Argumenten"""
    # Auf Windows: Stelle sicher dass stdin/stdout UTF-8 verwenden
    import io
    if sys.platform == 'win32':
        sys.stdin = io.TextIOWrapper(sys.stdin.buffer, encoding='utf-8')
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    
    if len(sys.argv) < 2:
        print(json.dumps({'success': False, 'error': 'Kein Befehl angegeben'}))
        sys.exit(1)
    
    command = sys.argv[1]
    
    if command == 'write_sheet':
        # Daten von stdin lesen (für große Datenmengen)
        input_data = sys.stdin.read()
        try:
            params = json.loads(input_data)
        except json.JSONDecodeError as e:
            print(json.dumps({'success': False, 'error': f'JSON Parse Error: {str(e)}'}))
            sys.exit(1)
        
        result = write_sheet(
            params.get('filePath'),
            params.get('outputPath'),
            params.get('sheetName'),
            params.get('changes', {}),
            params.get('originalPath')  # NEU: Original-Datei für restore_table_xml
        )
        print(json.dumps(result, ensure_ascii=False))
    
    elif command == 'check_excel':
        # Prüft ob Microsoft Excel verfügbar ist
        excel_available = is_excel_installed()
        result = {
            'success': True,
            'excelAvailable': excel_available,
            'xlwingsAvailable': XLWINGS_AVAILABLE,
            'message': 'Excel verfügbar - strukturelle Änderungen mit CF-Erhalt möglich' if excel_available else 'Excel nicht verfügbar - CF-Erhalt bei strukturellen Änderungen eingeschränkt'
        }
        print(json.dumps(result, ensure_ascii=False))
    
    else:
        print(json.dumps({'success': False, 'error': f'Unbekannter Befehl: {command}'}))
        sys.exit(1)


if __name__ == '__main__':
    main()
