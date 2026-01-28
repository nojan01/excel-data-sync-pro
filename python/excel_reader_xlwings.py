#!/usr/bin/env python3
"""
Excel Reader für Excel Data Sync Pro - xlwings Version
Verwendet xlwings für native Excel-Kompatibilität und perfekte Format-Erhaltung

Vorteile von xlwings:
- Native Excel-Integration (Excel macht alles)
- Perfekte Erhaltung von CF, Styles, Formeln
- Zuverlässige Datentyp-Konvertierung
"""

import json
import sys
import os
import platform
import subprocess
from datetime import datetime, date
import xlwings as xw
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def kill_excel_instances():
    """Beendet alle laufenden Excel-Instanzen - plattformübergreifend"""
    import time
    system = platform.system()
    
    # Methode 1: xlwings Apps beenden (funktioniert auf allen Plattformen)
    try:
        for app in xw.apps:
            try:
                app.quit()
            except:
                pass
    except:
        pass
    
    if system == 'Darwin':  # macOS
        # Methode 2: AppleScript quit (ohne waiting)
        try:
            subprocess.run(['osascript', '-e', 
                'tell application "Microsoft Excel" to quit saving no'], 
                capture_output=True, timeout=2)
        except Exception:
            pass
        
        # Kurz warten
        time.sleep(0.2)
        
        # Methode 3: Prüfen ob Excel noch läuft und mit pkill beenden
        try:
            result = subprocess.run(['pgrep', '-x', 'Microsoft Excel'], 
                                    capture_output=True, timeout=2)
            if result.returncode == 0:
                # Excel läuft noch - sofort killen!
                subprocess.run(['pkill', '-9', 'Microsoft Excel'], 
                               capture_output=True, timeout=2)
                time.sleep(0.2)
        except:
            pass
        
        # Methode 4: Falls immer noch da, nochmal versuchen
        try:
            result = subprocess.run(['pgrep', '-x', 'Microsoft Excel'], 
                                    capture_output=True, timeout=1)
            if result.returncode == 0:
                subprocess.run(['killall', '-9', 'Microsoft Excel'], 
                               capture_output=True, timeout=2)
        except:
            pass
    
    elif system == 'Windows':  # Windows
        time.sleep(0.2)
        
        # Methode 2: Mit taskkill Excel beenden
        try:
            subprocess.run(['taskkill', '/F', '/IM', 'EXCEL.EXE'], 
                          capture_output=True, timeout=5)
            time.sleep(0.3)
        except:
            pass


def hide_excel():
    """Versteckt Excel - plattformübergreifend"""
    system = platform.system()
    
    if system == 'Darwin':  # macOS
        try:
            subprocess.run(['osascript', '-e', 
                'tell application "System Events" to set visible of process "Microsoft Excel" to false'], 
                capture_output=True, timeout=2)
        except:
            pass
    
    elif system == 'Windows':  # Windows
        # Auf Windows: Excel-Fenster verstecken via xlwings
        try:
            for app in xw.apps:
                try:
                    app.visible = False
                except:
                    pass
        except:
            pass


def argb_to_hex(argb):
    """Konvertiert ARGB (z.B. 'FF00FF00') zu Hex ('#00FF00')"""
    if not argb or argb == '00000000':
        return None
    if len(argb) == 8:
        return '#' + argb[2:].upper()
    elif len(argb) == 6:
        return '#' + argb.upper()
    return None


def rgb_to_hex(r, g, b):
    """Konvertiert RGB-Werte zu Hex ('#RRGGBB')"""
    return f'#{r:02X}{g:02X}{b:02X}'


def serialize_value(value):
    """Konvertiert Python-Werte zu JSON-kompatiblen Werten"""
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.strftime('%d.%m.%Y %H:%M:%S')
    if isinstance(value, date):
        return value.strftime('%d.%m.%Y')
    if isinstance(value, (int, float)):
        return value
    return str(value)


def get_cell_fill_color_xlwings(cell):
    """
    Extrahiert die Hintergrundfarbe einer Zelle via xlwings
    Gibt Hex-String zurück oder None
    """
    try:
        color = cell.color
        if color:
            # color ist ein Tuple (R, G, B)
            r, g, b = int(color[0]), int(color[1]), int(color[2])
            if r == 255 and g == 255 and b == 255:
                return None  # Weiß = keine Farbe
            return rgb_to_hex(r, g, b)
    except Exception:
        pass
    return None


def get_font_info_xlwings(cell):
    """
    Extrahiert Font-Informationen via xlwings
    """
    try:
        font = cell.font
        info = {}
        
        if font.bold:
            info['bold'] = True
        if font.italic:
            info['italic'] = True
        if hasattr(font, 'underline') and font.underline:
            info['underline'] = True
        
        # Font-Farbe
        if font.color:
            try:
                r, g, b = int(font.color[0]), int(font.color[1]), int(font.color[2])
                if not (r == 0 and g == 0 and b == 0):
                    info['color'] = rgb_to_hex(r, g, b)
            except:
                pass
        
        # Font-Größe (nur wenn vom Default abweichend)
        if font.size and font.size != 11:
            info['size'] = font.size
        
        # Font-Name (nur wenn vom Default abweichend)
        if font.name and font.name.lower() not in ['calibri', 'arial']:
            info['name'] = font.name
        
        return info if info else None
    except Exception:
        pass
    return None


def read_sheet_xlwings(file_path, sheet_name=None, options=None):
    """
    Liest ein Excel-Sheet mit xlwings und gibt Daten + Metadaten zurück
    
    Args:
        file_path: Pfad zur Excel-Datei
        sheet_name: Name des Sheets (None = aktives Sheet)
        options: Dict mit Optionen (extractStyles, etc.)
    
    Returns:
        Dict mit headers, data, styles, etc.
    """
    options = options or {}
    extract_styles = options.get('extractStyles', True)
    
    # NICHT am Anfang beenden - das verursacht Probleme!
    # Excel wird nur am Ende beendet
    
    import time
    
    app = None
    wb = None
    
    try:
        # Excel-App starten (ohne with-Kontext, um Cleanup-Probleme zu vermeiden)
        app = xw.App(visible=False, add_book=False)
        app.display_alerts = False
        app.screen_updating = False
        
        # Verstecke Excel sofort nach dem Start
        hide_excel()
        
        # Workbook öffnen (read_only für schnelleres Lesen)
        wb = app.books.open(file_path, read_only=True)
        
        # Sheet auswählen
        if sheet_name:
            sheet_names = [s.name for s in wb.sheets]
            if sheet_name not in sheet_names:
                wb.close()
                return {'success': False, 'error': f'Sheet "{sheet_name}" nicht gefunden'}
            ws = wb.sheets[sheet_name]
        else:
            ws = wb.sheets.active
        
        # Bereich ermitteln (used range)
        used_range = ws.used_range
        if not used_range:
            wb.close()
            return {
                'success': True,
                'headers': [],
                'data': [],
                'sheetName': ws.name,
                'rowCount': 0,
                'columnCount': 0
            }
        
        max_row = used_range.last_cell.row
        max_col = used_range.last_cell.column
        actual_sheet_name = ws.name
        
        # Alle Daten auf einmal lesen (Performance!)
        all_data = ws.range((1, 1), (max_row, max_col)).value
        
        # Falls nur eine Zeile/Spalte, in Liste konvertieren
        if max_row == 1:
            all_data = [all_data] if not isinstance(all_data, list) else [all_data]
        if max_col == 1:
            all_data = [[row] if not isinstance(row, list) else row for row in all_data]
        
        # Headers (erste Zeile)
        headers = [serialize_value(v) for v in all_data[0]] if all_data else []
        
        # Daten (ab Zeile 2)
        data = []
        for row in all_data[1:]:
            if isinstance(row, list):
                data.append([serialize_value(v) for v in row])
            else:
                data.append([serialize_value(row)])
        
        # Column Widths - nur bei kleineren Dateien (sonst zu langsam)
        col_widths = {}
        if max_col <= 100:  # Nur bei max 100 Spalten
            for col_idx in range(1, max_col + 1):
                try:
                    width = ws.range((1, col_idx)).column_width
                    if width:
                        col_widths[str(col_idx - 1)] = width
                except:
                    pass
        
        # Formeln überspringen - werden von openpyxl gelesen (schneller)
        cell_formulas = {}
        
        # Workbook schließen (Excel bleibt laufen für weitere Operationen)
        wb.close()
        
        # Jetzt Struktur-Metadaten mit openpyxl lesen (ohne Excel zu öffnen!)
        result = {
            'success': True,
            'headers': headers,
            'data': data,
            'sheetName': actual_sheet_name,
            'rowCount': max_row - 1,
            'columnCount': max_col,
            'columnWidths': col_widths,
            'cellFormulas': cell_formulas
        }
        
        # Styles, Merged Cells, AutoFilter etc. mit openpyxl lesen (kein Excel nötig!)
        try:
            wb_xl = load_workbook(file_path, read_only=False, data_only=True)
            ws_xl = wb_xl[actual_sheet_name] if actual_sheet_name else wb_xl.active
            
            # Merged Cells
            merged = [str(r) for r in ws_xl.merged_cells.ranges]
            result['mergedCells'] = merged
            
            # AutoFilter
            if ws_xl.auto_filter and ws_xl.auto_filter.ref:
                result['autoFilterRange'] = ws_xl.auto_filter.ref
            
            # Hidden Columns
            hidden_cols = []
            for col_idx in range(1, max_col + 1):
                try:
                    col_letter = get_column_letter(col_idx)
                    col_dim = ws_xl.column_dimensions.get(col_letter)
                    if col_dim and col_dim.hidden:
                        hidden_cols.append(col_idx - 1)
                except:
                    pass
            result['hiddenColumns'] = hidden_cols
            
            # Hidden Rows
            hidden_rows = []
            for row_idx in range(2, max_row + 1):
                try:
                    row_dim = ws_xl.row_dimensions.get(row_idx)
                    if row_dim and row_dim.hidden:
                        hidden_rows.append(row_idx - 2)
                except:
                    pass
            result['hiddenRows'] = hidden_rows
            
            wb_xl.close()
        except Exception:
            result['mergedCells'] = []
            result['hiddenColumns'] = []
            result['hiddenRows'] = []
        
        # Styles mit openpyxl extrahieren (kein Excel nötig, schneller!)
        if extract_styles:
            try:
                wb_styles = load_workbook(file_path, read_only=False, data_only=False)
                ws_styles = wb_styles[actual_sheet_name] if actual_sheet_name else wb_styles.active
                
                cell_styles = {}
                cell_fonts = {}
                
                for row_idx in range(1, max_row + 1):
                    for col_idx in range(1, max_col + 1):
                        key = f"{row_idx - 1}-{col_idx - 1}"
                        cell = ws_styles.cell(row=row_idx, column=col_idx)
                        
                        # Fill Color
                        if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                            rgb = cell.fill.fgColor.rgb
                            if isinstance(rgb, str) and len(rgb) >= 6 and rgb != '00000000':
                                hex_color = argb_to_hex(rgb)
                                if hex_color:
                                    cell_styles[key] = hex_color
                        
                        # Font Info
                        font = cell.font
                        if font:
                            font_info = {}
                            if font.bold:
                                font_info['bold'] = True
                            if font.italic:
                                font_info['italic'] = True
                            if font.color and font.color.rgb and font.color.rgb != '00000000':
                                hex_color = argb_to_hex(font.color.rgb)
                                if hex_color and hex_color != '#000000':
                                    font_info['color'] = hex_color
                            if font.size and font.size != 11:
                                font_info['size'] = font.size
                            if font.name and font.name.lower() not in ['calibri', 'arial']:
                                font_info['name'] = font.name
                            if font_info:
                                cell_fonts[key] = font_info
                
                result['cellStyles'] = cell_styles
                result['cellFonts'] = cell_fonts
                result['defaultFont'] = {'name': 'Calibri', 'size': 11}
                
                wb_styles.close()
            except Exception:
                result['cellStyles'] = {}
                result['cellFonts'] = {}
        
        # Excel bleibt laufen für weitere Sheet-Operationen
        return result
        
    except Exception as e:
        import traceback
        return {'success': False, 'error': str(e), 'traceback': traceback.format_exc()}


def list_sheets_xlwings(file_path):
    """Listet alle Sheets in einer Excel-Datei mit xlwings"""
    # NICHT am Anfang beenden!
    
    try:
        # Excel-App starten (ohne with-Kontext)
        app = xw.App(visible=False, add_book=False)
        app.display_alerts = False
        app.screen_updating = False
        
        # Verstecke Excel sofort
        hide_excel()
        
        wb = app.books.open(file_path, read_only=True)
        sheets = [s.name for s in wb.sheets]
        wb.close()
        
        # Excel bleibt laufen für weitere Operationen
        return {'success': True, 'sheets': sheets}
    except Exception as e:
        return {'success': False, 'error': str(e)}


def main():
    """Hauptfunktion - liest Befehle von Argumenten"""
    if len(sys.argv) < 2:
        print(json.dumps({'success': False, 'error': 'Kein Befehl angegeben'}))
        sys.exit(1)
    
    command = sys.argv[1]
    
    if command == 'list_sheets':
        if len(sys.argv) < 3:
            print(json.dumps({'success': False, 'error': 'Kein Dateipfad angegeben'}))
            sys.exit(1)
        result = list_sheets_xlwings(sys.argv[2])
        print(json.dumps(result, ensure_ascii=False))
    
    elif command == 'read_sheet':
        if len(sys.argv) < 3:
            print(json.dumps({'success': False, 'error': 'Kein Dateipfad angegeben'}))
            sys.exit(1)
        file_path = sys.argv[2]
        sheet_name = sys.argv[3] if len(sys.argv) > 3 else None
        options = json.loads(sys.argv[4]) if len(sys.argv) > 4 else {}
        result = read_sheet_xlwings(file_path, sheet_name, options)
        print(json.dumps(result, ensure_ascii=False))
    
    else:
        print(json.dumps({'success': False, 'error': f'Unbekannter Befehl: {command}'}))
        sys.exit(1)


if __name__ == '__main__':
    main()
