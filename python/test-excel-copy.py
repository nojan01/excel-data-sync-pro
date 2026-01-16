#!/usr/bin/env python3
"""
Test: Excel-Datei lesen und wieder speichern mit openpyxl
Ziel: Prüfen ob alle Formatierungen erhalten bleiben
"""

import sys
import os
from openpyxl import load_workbook
from datetime import datetime

def copy_excel_file(input_path, output_path):
    """
    Kopiert eine Excel-Datei mit openpyxl.
    Behält ALLE Formatierungen bei, da wir einfach speichern.
    """
    print(f"Lese Datei: {input_path}")
    start = datetime.now()
    
    # Workbook laden (nicht data_only, um Formeln zu behalten)
    wb = load_workbook(input_path)
    
    load_time = (datetime.now() - start).total_seconds()
    print(f"Geladen in {load_time:.2f}s")
    
    print(f"Sheets: {wb.sheetnames}")
    
    # Jedes Sheet analysieren
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\nSheet '{sheet_name}':")
        print(f"  Zeilen: {ws.max_row}")
        print(f"  Spalten: {ws.max_column}")
        
        # Merged Cells
        if ws.merged_cells.ranges:
            print(f"  Merged Cells: {len(ws.merged_cells.ranges)}")
        
        # AutoFilter
        if ws.auto_filter and ws.auto_filter.ref:
            print(f"  AutoFilter: {ws.auto_filter.ref}")
        
        # Conditional Formatting
        if ws.conditional_formatting:
            print(f"  Conditional Formatting: {len(ws.conditional_formatting)} Regeln")
        
        # Tables
        if hasattr(ws, 'tables') and ws.tables:
            print(f"  Tables: {len(ws.tables)}")
    
    # Speichern
    print(f"\nSpeichere nach: {output_path}")
    save_start = datetime.now()
    wb.save(output_path)
    save_time = (datetime.now() - save_start).total_seconds()
    print(f"Gespeichert in {save_time:.2f}s")
    
    # Dateigröße vergleichen
    input_size = os.path.getsize(input_path)
    output_size = os.path.getsize(output_path)
    print(f"\nDateigrößen:")
    print(f"  Original: {input_size:,} bytes")
    print(f"  Kopie:    {output_size:,} bytes")
    print(f"  Differenz: {output_size - input_size:+,} bytes ({(output_size/input_size*100-100):+.1f}%)")
    
    return True


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python test-excel-copy.py <input.xlsx> [output.xlsx]")
        sys.exit(1)
    
    input_file = sys.argv[1]
    
    if len(sys.argv) >= 3:
        output_file = sys.argv[2]
    else:
        # Automatisch Output-Namen generieren
        base, ext = os.path.splitext(input_file)
        output_file = f"{base}_python_copy{ext}"
    
    if not os.path.exists(input_file):
        print(f"Fehler: Datei nicht gefunden: {input_file}")
        sys.exit(1)
    
    try:
        copy_excel_file(input_file, output_file)
        print(f"\n✅ Erfolgreich! Ausgabedatei: {output_file}")
    except Exception as e:
        print(f"\n❌ Fehler: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
