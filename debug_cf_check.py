#!/usr/bin/env python3
"""Prüfe CF-Bereiche in exportierter Datei"""
from openpyxl import load_workbook
import os

desktop = '/Users/nojan/Desktop'
exports = [f for f in os.listdir(desktop) if f.startswith('Export_') and f.endswith('.xlsx')]
if exports:
    exports.sort(key=lambda x: os.path.getmtime(os.path.join(desktop, x)), reverse=True)
    export_file = os.path.join(desktop, exports[0])
    print(f'Neueste Export-Datei: {export_file}')
    
    wb = load_workbook(export_file)
    ws = wb.active
    
    print(f'\n=== Conditional Formatting Bereiche ===')
    cf_count = 0
    for cf_range in ws.conditional_formatting:
        print(f'Range: {cf_range}')
        cf_count += 1
        for rule in ws.conditional_formatting[cf_range]:
            print(f'  Rule type: {rule.type}')
            if hasattr(rule, 'formula') and rule.formula:
                print(f'  Formula: {rule.formula}')
    
    print(f'\nGesamt: {cf_count} CF-Bereiche')
    
    # Vergleiche mit Original
    print('\n=== Vergleich mit Original ===')
    orig_file = '/Users/nojan/Desktop/test-styles-exceljs.xlsx'
    if os.path.exists(orig_file):
        wb_orig = load_workbook(orig_file)
        ws_orig = wb_orig.active
        orig_count = 0
        for cf_range in ws_orig.conditional_formatting:
            orig_count += 1
        print(f'Original: {orig_count} CF-Bereiche')
        wb_orig.close()
    
    wb.close()
else:
    print('Keine Export-Datei gefunden')
