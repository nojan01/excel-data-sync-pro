#!/usr/bin/env python3
from openpyxl import load_workbook
import os

# Finde die Excel-Datei
desktop = os.path.expanduser('~/Desktop')
files = [f for f in os.listdir(desktop) if f.endswith('.xlsx') and 'MVMS' in f]
print(f'Excel-Dateien auf Desktop: {files}')

if files:
    file_path = os.path.join(desktop, files[0])
    print(f'Analysiere: {file_path}')
    
    wb = load_workbook(file_path, rich_text=True)
    ws = wb.active
    print(f'Aktives Sheet: {ws.title}')
    print(f'Max Row: {ws.max_row}, Max Col: {ws.max_column}')
    
    # Tables prüfen
    print(f'\n=== TABLES ===')
    if hasattr(ws, '_tables'):
        print(f'Anzahl Tables: {len(ws._tables)}')
        for table in ws._tables:
            print(f'  Table: {table.name}, Bereich: {table.ref}')
            if table.tableStyleInfo:
                print(f'    Style: {table.tableStyleInfo.name}')
                print(f'    showRowStripes: {table.tableStyleInfo.showRowStripes}')
    
    # Prüfe Zellformatierungen in den ersten Zeilen
    print(f'\n=== ZELL-FORMATIERUNGEN (Zeile 2-5, Spalte 1-10) ===')
    for row in range(2, 6):
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col)
            if cell.fill and cell.fill.patternType and cell.fill.patternType != 'none':
                print(f'  Zeile {row}, Spalte {col}: fill={cell.fill.patternType}, fgColor={cell.fill.fgColor.rgb if cell.fill.fgColor else None}')
    
    # Prüfe ob es Theme-Farben gibt
    print(f'\n=== THEME-FARBEN CHECK ===')
    for row in range(2, 4):
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            if cell.fill and cell.fill.fgColor:
                theme = cell.fill.fgColor.theme
                tint = cell.fill.fgColor.tint
                if theme is not None:
                    print(f'  Zeile {row}, Spalte {col}: theme={theme}, tint={tint}')
    
    wb.close()
else:
    print('Keine MVMS Excel-Datei gefunden!')
