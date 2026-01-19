#!/usr/bin/env python3
"""Test: Spalte einfügen mit CF-Anpassung - mit echter Datei"""
import sys
import os
import json
sys.path.insert(0, 'python')

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from copy import copy
from excel_writer import adjust_conditional_formatting, adjust_tables

# Echte Test-Datei
INPUT_FILE = '/Users/nojan/Desktop/2025-08-31 DEFENCE&SPACE MVMS Master Asset List GER.xlsx'
OUTPUT_FILE = '/Users/nojan/Desktop/DEBUG_Insert_GER.xlsx'
SHEET_NAME = 'DEFENCE&SPACE Aug-2025'

print(f"Lade: {INPUT_FILE}")
print(f"Sheet: {SHEET_NAME}")

wb = load_workbook(INPUT_FILE, rich_text=True)
ws = wb[SHEET_NAME]

print(f"\n=== VOR Insert ===")
print(f"Spalten: {ws.max_column}, Zeilen: {ws.max_row}")

# Zeige erste 10 Spalten Header
print("\nHeader:")
headers = []
for col in range(1, min(11, ws.max_column + 1)):
    cell = ws.cell(row=1, column=col)
    headers.append(cell.value)
    print(f"  {col} ({get_column_letter(col)}): '{cell.value}'")

# Zeige CF-Bereiche
print(f"\n=== Conditional Formatting (erste 5) ===")
cf_count = 0
for cf_range in ws.conditional_formatting:
    print(f"Range: {cf_range}")
    cf_count += 1
    if cf_count >= 5:
        print(f"  ... und {len(list(ws.conditional_formatting)) - 5} weitere")
        break

# Simuliere Insert bei Position 3 (0-basiert) = Spalte D (1-basiert = 4)
# Das ist wie wenn der User rechtsklickt auf Spalte D und "Spalte einfügen" wählt
insert_position = 3  # 0-basiert (Frontend-Index)
source_column = 3    # 0-basiert (die Spalte von der Format kopiert wird)
insert_at = insert_position + 1  # 1-basiert für Excel

print(f"\n=== INSERT bei Position {insert_position} (Excel-Spalte {insert_at} = {get_column_letter(insert_at)}) ===")
print(f"Source Column für Format: {source_column} (Excel {source_column + 1} = {get_column_letter(source_column + 1)})")

# Speichere Formatierung von sourceColumn
source_excel_col = source_column + 1
source_format = {}
source_width = None

col_letter = get_column_letter(source_excel_col)
if col_letter in ws.column_dimensions:
    source_width = ws.column_dimensions[col_letter].width
    print(f"Source Spaltenbreite: {source_width}")

# Alle Zeilen der Referenzspalte speichern
for row in range(1, ws.max_row + 1):
    cell = ws.cell(row=row, column=source_excel_col)
    source_format[row] = {
        'fill': copy(cell.fill) if cell.fill else None,
        'font': copy(cell.font) if cell.font else None,
        'alignment': copy(cell.alignment) if cell.alignment else None,
        'border': copy(cell.border) if cell.border else None,
        'number_format': cell.number_format
    }

# Spaltenbreiten speichern
saved_widths = {}
max_col = ws.max_column
for col in range(insert_at, max_col + 1):
    col_letter = get_column_letter(col)
    if col_letter in ws.column_dimensions:
        saved_widths[col] = ws.column_dimensions[col_letter].width

# SPALTE EINFÜGEN
ws.insert_cols(insert_at, 1)
print(f"Spalte eingefügt bei {insert_at}")

# Spaltenbreiten wiederherstellen (um 1 nach rechts verschoben)
for old_col, width in saved_widths.items():
    if width:
        new_letter = get_column_letter(old_col + 1)
        ws.column_dimensions[new_letter].width = width

# CF ANPASSEN
inserted_cols_for_cf = {insert_at - 1: 1}  # 0-basiert für die Funktion
adjust_conditional_formatting(ws, [], inserted_cols_for_cf)
print("CF-Bereiche angepasst")

# Tables anpassen
adjust_tables(ws, [], inserted_cols_for_cf, headers)
print("Tables angepasst")

# Formatierung auf neue Spalte anwenden
if source_width:
    new_letter = get_column_letter(insert_at)
    ws.column_dimensions[new_letter].width = source_width

for row, fmt in source_format.items():
    cell = ws.cell(row=row, column=insert_at)
    if fmt['fill']:
        cell.fill = fmt['fill']
    if fmt['font']:
        cell.font = fmt['font']
    if fmt['alignment']:
        cell.alignment = fmt['alignment']
    if fmt['border']:
        cell.border = fmt['border']
    if fmt.get('number_format'):
        cell.number_format = fmt['number_format']

# Header für neue Spalte setzen
ws.cell(row=1, column=insert_at).value = "Neue Spalte"
print("Formatierung kopiert")

print(f"\n=== NACH Insert ===")
print(f"Spalten: {ws.max_column}")

# Zeige erste 10 Spalten Header
print("\nHeader nach Insert:")
for col in range(1, min(11, ws.max_column + 1)):
    cell = ws.cell(row=1, column=col)
    print(f"  {col} ({get_column_letter(col)}): '{cell.value}'")

# Zeige CF-Bereiche nach Insert
print(f"\n=== CF nach Insert (erste 5) ===")
cf_count = 0
for cf_range in ws.conditional_formatting:
    print(f"Range: {cf_range}")
    cf_count += 1
    if cf_count >= 5:
        break

# Speichern
wb.save(OUTPUT_FILE)
print(f"\n✅ Gespeichert: {OUTPUT_FILE}")
wb.close()

print("\nBitte vergleichen Sie diese Datei mit dem Export aus der App!")
