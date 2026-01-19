#!/usr/bin/env python3
"""Test-Export mit Spalte einfügen"""
import sys
sys.path.insert(0, 'python')

from excel_reader import read_sheet
from excel_writer import write_sheet

original_path = "/Users/nojan/Desktop/2025-08-31 DEFENCE&SPACE MVMS Master Asset List GER.xlsx"
output_path = "/Users/nojan/Desktop/DEBUG_Insert.xlsx"
sheet_name = 'DEFENCE&SPACE Aug-2025'

# 1. Zuerst die Datei lesen
print("Lese Original-Datei...")
read_result = read_sheet(original_path, sheet_name)
headers = read_result['headers']
data = read_result['data']

print(f"Gelesene Spalten: {len(headers)}")
print(f"Gelesene Zeilen: {len(data)}")
print(f"Erste 5 Header: {headers[:5]}")

# 2. Spalte an Position 2 einfügen (nach "Division's unique ID")
insert_pos = 2
new_header = "NEUE SPALTE"

# Header anpassen
new_headers = headers[:insert_pos] + [new_header] + headers[insert_pos:]

# Daten anpassen - leere Werte für neue Spalte
new_data = []
for row in data:
    new_row = row[:insert_pos] + [''] + row[insert_pos:]
    new_data.append(new_row)

print(f"\nNach Einfügen an Position {insert_pos}:")
print(f"Neue Spalten: {len(new_headers)}")
print(f"Erste 6 Header: {new_headers[:6]}")

# 3. Export mit korrekten Änderungen
changes = {
    'headers': new_headers,
    'data': new_data,
    'insertedColumns': {
        'operations': [{
            'position': insert_pos,
            'count': 1,
            'sourceColumn': insert_pos  # Formatierung von Spalte 2 kopieren
        }]
    },
    'editedCells': {},
    'fullRewrite': True
}

print("\nExportiere...")
result = write_sheet(original_path, output_path, sheet_name, changes)
print("Result:", result)

# Prüfe Ergebnis
if result.get('success'):
    from openpyxl import load_workbook
    
    exp = load_workbook(output_path)
    ws_exp = exp.active
    
    print(f"\n=== Export Prüfung ===")
    print(f"Spalten: {ws_exp.max_column}")
    
    # CF-Regeln prüfen
    print(f"CF-Regeln: {len(ws_exp.conditional_formatting._cf_rules)}")
    
    # Erste 3 CF-Regeln
    for cf_range, rules in list(ws_exp.conditional_formatting._cf_rules.items())[:3]:
        print(f"  {cf_range}")
    
    exp.close()
